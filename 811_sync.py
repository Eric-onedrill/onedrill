#!/usr/bin/env python3
"""
OneDrill - 811 Automation Script
Indiana 811 + Sunshine 811

Melhorias v2:
  - Retry genérico em TODAS as operações Supabase
  - Lock file pra evitar execuções simultâneas
  - Resumo de sync estruturado (cleared, reverted, private_locator)
  - Dedup e truncamento de notas automáticas
  - COMPANY_PHONE movido pro .env
  - Validação no parser de contatos
  - Scheduler com stagger e melhor controle de concorrência
  - Context manager pra Playwright (evita Chrome órfão)
  - Logs mais claros e consistentes

═══════════════════════════════════════════════════════════════════════════════
ÍNDICE DE SEÇÕES — use Ctrl+F com "│ SECTION" pra navegar rapidamente
═══════════════════════════════════════════════════════════════════════════════

CONFIG E UTILITÁRIOS
  │ SECTION: DEBUG_MODE       ~L91    │  ONEDRILL_DEBUG env flags
  │ SECTION: CONFIG           ~L97    │  Constantes do app (.env, URLs, paths)
  │ SECTION: LOGGING          ~L136   │  Setup do logger
  │ SECTION: CONSTANTS        ~L152   │  Constantes gerais (BATCH_SIZE, etc)
  │ SECTION: LOCK             ~L161   │  ProcessLock (class) — lock file anti-concorrência
  │ SECTION: CANCEL_CACHE     ~L229   │  Cache de tickets cancelados em disco
  │ SECTION: WAIT_HELPERS     ~L278   │  Smart waits pra Playwright
  │ SECTION: PLAYWRIGHT_CTX   ~L334   │  Context manager pra Playwright
  │ SECTION: SUPABASE_IO      ~L366   │  sb_get, sb_insert, sb_patch, sb_upsert, sb_delete

DOMÍNIO DO 811
  │ SECTION: AUTO_NOTES       ~L440   │  append_auto_note (dedup + truncamento)
  │ SECTION: CLASSIFY         ~L470   │  classify() — interpreta status code → Clear/Pending/Damage
  │ SECTION: LOCATION_TEXT    ~L669   │  extract_location_text — parse do body
  │ SECTION: EXPIRE_PARSE     ~L714   │  extract_expire_date + normalize_expire + _is_polluted_expire
  │ SECTION: CANCEL_DETECT    ~L883   │  is_ticket_canceled()
  │ SECTION: COORDS_ADJUST    ~L898   │  adjust_coords_by_location (alinha pro texto 811)
  │ SECTION: GEOCODING        ~L934   │  geocode_address (Nominatim — sleep 1.1s, fix #14)

PORTAIS — PLAYWRIGHT
  │ SECTION: AUTO_LOGIN       ~L971   │  auto_login — renovação manual em janela visível
  │ SECTION: PORTAL_NAV       ~L1019  │  goto_dashboard, filter_ticket, back_to_dashboard, ensure_login
  │ SECTION: SCRAPE           ~L1177  │  scrape() — função central paralela
  │ SECTION: SYNC_SUMMARY     ~L1535  │  SyncSummary (class) — stats estruturadas
  │ SECTION: BATCH_PATCH      ~L1579  │  _get_latest_response_date + sb_batch_patch (fix #5)
  │ SECTION: UNRECOGNIZED     ~L1685  │  save_unrecognized_responses + send_unrecognized_alert
  │ SECTION: SAVE             ~L1760  │  save_to_supabase — grava resultados do scrape

COMANDOS PRINCIPAIS
  │ SECTION: IMPORT           ~L2098  │  import_new_tickets (batch upsert, fix #4)
  │ SECTION: RESCRAPE         ~L2393  │  rescrape_notes — atualiza notes + expire
  │ SECTION: CLEANUP          ~L2486  │  cleanup_canceled — remove tickets cancelados
  │ SECTION: FILTER_SYNC      ~L2573  │  filter_tickets_for_sync — prioriza o que scrapear
  │ SECTION: SYNC             ~L2661  │  sync_state + sync_all + sync_and_import*

ILLINOIS — JULIE PÚBLICO
  │ SECTION: JULIE            ~L2787  │  scrape_julie_ticket + scrape_il + sync_il + save_ticket_pdfs_il

WISCONSIN — DIGGERS HOTLINE
  │ SECTION: WI               ~L3819  │  scrape_diggers_ticket + scrape_wi + sync_wi (público)
  │ SECTION: WI_IMPORT        ~L4245  │  import_wi + login_diggers + search_excavator (logado)

CONTATOS (FL)
  │ SECTION: CONTACTS_FL      ~L3161  │  parse_contact_table + scrape_contacts
  │ SECTION: UTILITY_HELPERS  ~L3509  │  get_contacts_for_utility, get_all_contacts

EXPORTS E DEBUG
  │ SECTION: EXPORT_EXCEL     ~L3525  │  export_excel
  │ SECTION: DEBUG_SCREENSHOT ~L3596  │  debug_screenshot
  │ SECTION: BACKFILL         ~L3620  │  backfill_history + fix_clear_dates
  │ SECTION: PDF_SAVE         ~L3814  │  save_ticket_pdfs (Print Dialog via hwnd)
  │ SECTION: BACKUP           ~L4093  │  backup_database

MANUTENÇÃO E TESTES
  │ SECTION: SELF_TESTS       ~L4186  │  run_self_tests — validação inline
  │ SECTION: FIX_RENEWALS     ~L4412  │  fix_renewals — conserta expire_old
  │ SECTION: FIX_EXPIRES      ~L4549  │  fix_expires — conserta formato antigo
  │ SECTION: AUDIT            ~L4668  │  audit_health + clean_ghost_utilities
  │ SECTION: DEBUG_EXPIRE     ~L4994  │  debug_expire — testa patterns em um ticket
  │ SECTION: CLI              ~L5117  │  Argparse + main entry point

═══════════════════════════════════════════════════════════════════════════════
"""

import os, sys, time, logging, logging.handlers, argparse, asyncio, re, urllib.parse, shutil, base64, glob
from datetime import datetime, timezone, timedelta
from contextlib import asynccontextmanager
from dotenv import load_dotenv
import requests
from playwright.async_api import async_playwright
import json as _json
import html as _html

# ── │ SECTION: DEBUG_MODE │ DEBUG MODE ────────────────────────────────────────
DEBUG_MODE = os.environ.get("ONEDRILL_DEBUG", "").lower() in ("1", "true", "yes")
DEBUG_TICKET = os.environ.get("ONEDRILL_DEBUG_TICKET", "").strip()  # Se setado, força debug desse ticket específico

load_dotenv()

# ── │ SECTION: CONFIG │ CONFIGURAÇÃO ──────────────────────────────────────────
SB_URL        = os.getenv("SB_URL",  "https://ofbqtaulvzeltfpqcjhh.supabase.co")
SB_KEY        = os.getenv("SB_KEY",  "")
IN_USER       = os.getenv("IN_USER", "")
IN_PASS       = os.getenv("IN_PASS", "")
FL_USER       = os.getenv("FL_USER", "")
FL_PASS       = os.getenv("FL_PASS", "")
WI_USER       = os.getenv("WI_USER", "")
WI_PASS       = os.getenv("WI_PASS", "")
IL_USER       = os.getenv("IL_USER", "")
IL_PASS       = os.getenv("IL_PASS", "")
COMPANY_PHONE = os.getenv("COMPANY_PHONE", "3219473131")

DIGGERS_CLIENT_URL = "https://geocall.diggershotline.com/geocall/client/login"

# Validação de configuração crítica
if not SB_KEY:
    print("⚠ AVISO: SB_KEY não definida no .env — operações Supabase vão falhar", file=sys.stderr)

PORTALS = {
    "IN": {
        "url":       "https://811.indiana811.org/login?returnUrl=%2Fhome",
        "home":      "https://811.indiana811.org/home",
        "dashboard": "https://811.indiana811.org/tickets/dashboard",
        "user":      lambda: IN_USER,
        "pass":      lambda: IN_PASS,
    },
    "FL": {
        "url":       "https://exactix.sunshine811.com/login",
        "home":      "https://exactix.sunshine811.com/home",
        "dashboard": "https://exactix.sunshine811.com/tickets/dashboard",
        "user":      lambda: FL_USER,
        "pass":      lambda: FL_PASS,
    },
    "WI": {
        "url":       DIGGERS_CLIENT_URL,
        "home":      "https://geocall.diggershotline.com/geocall/client",
        "dashboard": "https://geocall.diggershotline.com/geocall/client",
        "user":      lambda: WI_USER,
        "pass":      lambda: WI_PASS,
    },
}

JULIE_URL = "https://newtin.julie1call.com/responsedisplay/"
JULIE_TICKETENTRY_URL = "https://newtin.julie1call.com/newtinweb/julie_ticketentry.html"
DIGGERS_URL = "https://geocall.diggershotline.com/geocall/portal"

# Counties/places onde a ONEDRILL atua em IL — usado pelo import_il
IL_SEARCH_COUNTY = "COOK"
IL_ONEDRILL_VARIANTS = {"ONEDRILL", "ONE DRILL"}  # filtro company (case insensitive)

SB_H = {
    "apikey":        SB_KEY,
    "Authorization": f"Bearer {SB_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "return=representation"
}

# ── │ SECTION: LOGGING │ LOGGING ──────────────────────────────────────────────
_log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "811_sync.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.handlers.RotatingFileHandler(
            _log_file, maxBytes=5*1024*1024, backupCount=3, encoding="utf-8"
        ),
        logging.StreamHandler(
            stream=open(sys.stdout.fileno(), mode='w', encoding='utf-8', closefd=False)
        )
    ]
)
log = logging.getLogger(__name__)

# ── │ SECTION: CONSTANTS │ CONSTANTES ─────────────────────────────────────────
BASE_DIR           = os.path.dirname(os.path.abspath(__file__))
VALID_STATES       = {"FL", "IN", "IL", "WI"}
W_SAFETY           = 300      # fallback ms quando não há seletor confiável
NUM_TABS           = 5        # 5 abas paralelas (reduzir nao ajuda, aumentar de 3→5 corta tempo ~40%)
CLEAR_CACHE_HOURS  = 24       # re-verifica Clear tickets 1x por dia
MAX_AUTO_NOTES     = 10       # máximo de notas automáticas por ticket
BATCH_SIZE         = 200      # tamanho do lote para bulk upsert
MAX_IMPORT_PAGES   = 5        # com 100/pagina, 5 paginas = 500 tickets
LOCK_FILE          = os.path.join(BASE_DIR, "811_sync.lock")
TIMEOUT_PAGE       = 60000    # Playwright default timeout (ms)
TIMEOUT_STABLE     = 2000     # wait_stable / wait_tab_content / wait_filter_results (3s era conservador demais)
TIMEOUT_NAV        = 5000     # wait_for / wait_nav
TIMEOUT_CLICK      = 8000     # click_and_wait


def _profile_path(state):
    return os.path.join(BASE_DIR, f"chrome_profile_{state}")

# ── │ SECTION: LOCK │ LOCK FILE (evita execuções simultâneas) ────────────────
class ProcessLock:
    """Lock file simples pra evitar múltiplas instâncias do script."""

    def __init__(self, path=LOCK_FILE):
        self.path = path
        self.acquired = False

    def acquire(self):
        """Tenta adquirir o lock. Retorna True se conseguiu."""
        if os.path.exists(self.path):
            try:
                with open(self.path, "r") as f:
                    data = _json.load(f)
                pid = data.get("pid", 0)
                started = data.get("started", "")
                # Se o processo ainda existe, lock ativo.
                # Fix bug #16 Python: diferenciar PermissionError (processo existe mas de outro
                # usuário) de ProcessLookupError (processo realmente morreu). No Windows especialmente,
                # os.kill(pid, 0) pode levantar PermissionError pra processos de outros usuários.
                # Antes: ambos caíam no except OSError, marcando lock como stale incorretamente —
                # podia resultar em 2 instâncias do scraper rodando simultaneamente.
                try:
                    os.kill(pid, 0)  # Só checa se existe (não mata)
                    log.warning(f"Lock ativo — PID {pid} iniciado em {started}")
                    return False
                except PermissionError:
                    # Processo existe mas é de outro usuário — tratar como ativo (NÃO stale)
                    log.warning(f"Lock ativo — PID {pid} (outro usuário) iniciado em {started}")
                    return False
                except (ProcessLookupError, OSError):
                    # Processo realmente morreu — stale lock
                    log.warning(f"Lock stale detectado (PID {pid}) — removendo")
                    self.release()
            except Exception:
                # Lock corrompido — remove
                self.release()

        try:
            with open(self.path, "w") as f:
                _json.dump({
                    "pid": os.getpid(),
                    "started": datetime.now().isoformat()
                }, f)
            self.acquired = True
            return True
        except Exception as e:
            log.error(f"Erro ao criar lock file: {e}")
            return False

    def release(self):
        """Libera o lock."""
        try:
            if os.path.exists(self.path):
                os.remove(self.path)
        except Exception as e:
            log.warning(f"Erro ao remover lock file: {e}")
        self.acquired = False

    def __enter__(self):
        if not self.acquire():
            raise RuntimeError("Outra instância já está rodando (lock file ativo)")
        return self

    def __exit__(self, *args):
        self.release()


# ── │ SECTION: CANCEL_CACHE │ CACHE DE TICKETS CANCELADOS ────────────────────
_CANCELED_CACHE_FILE = os.path.join(BASE_DIR, "canceled_cache.json")
_canceled_cache_mem = None


def load_canceled_cache():
    """Carrega set de tickets cancelados do disco (com cache em memória)."""
    global _canceled_cache_mem
    if _canceled_cache_mem is not None:
        return _canceled_cache_mem
    try:
        if os.path.exists(_CANCELED_CACHE_FILE):
            with open(_CANCELED_CACHE_FILE, "r") as f:
                data = _json.load(f)
            _canceled_cache_mem = {state: set(nums) for state, nums in data.items()}
            return _canceled_cache_mem
    except Exception as e:
        log.warning(f"Erro ao carregar cache de cancelados: {e}")
    _canceled_cache_mem = {}
    return _canceled_cache_mem


def save_canceled_cache(cache=None):
    """Salva cache de cancelados no disco."""
    global _canceled_cache_mem
    if cache is None:
        cache = _canceled_cache_mem or {}
    try:
        data = {state: sorted(nums) for state, nums in cache.items()}
        with open(_CANCELED_CACHE_FILE, "w") as f:
            _json.dump(data, f, indent=2)
    except Exception as e:
        log.warning(f"Erro ao salvar cache de cancelados: {e}")


def add_to_canceled_cache(state, ticket_num):
    """Adiciona ticket ao cache de cancelados (salva imediatamente)."""
    cache = load_canceled_cache()
    if state not in cache:
        cache[state] = set()
    cache[state].add(ticket_num)
    save_canceled_cache(cache)


def get_canceled_set(state):
    """Retorna set de tickets cancelados para um estado."""
    return load_canceled_cache().get(state, set())


# ── │ SECTION: WAIT_HELPERS │ SMART WAIT HELPERS ──────────────────────────────
async def wait_stable(page, timeout=TIMEOUT_STABLE):
    """Espera a página estabilizar (rede ociosa). Fallback silencioso."""
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout)
    except Exception:
        await page.wait_for_timeout(W_SAFETY)


async def wait_for(page, selector, timeout=TIMEOUT_NAV, state="visible"):
    """Espera seletor aparecer. Retorna True se encontrou."""
    try:
        await page.wait_for_selector(selector, timeout=timeout, state=state)
        return True
    except Exception:
        return False


async def wait_nav(page, timeout=TIMEOUT_NAV):
    """Espera após navegação: load completo + rede ociosa."""
    try:
        await page.wait_for_load_state("domcontentloaded", timeout=timeout)
        await page.wait_for_load_state("networkidle", timeout=timeout)
    except Exception:
        await page.wait_for_timeout(W_SAFETY)


async def wait_tab_content(page, timeout=TIMEOUT_STABLE):
    """Espera conteúdo de aba carregar."""
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout)
    except Exception:
        await page.wait_for_timeout(W_SAFETY)


async def wait_filter_results(page, timeout=TIMEOUT_STABLE):
    """Espera resultados do filtro carregarem."""
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout)
    except Exception:
        await page.wait_for_timeout(W_SAFETY)


async def click_and_wait(page, locator, wait_type="stable", timeout=TIMEOUT_CLICK):
    """Clica num elemento e espera o resultado."""
    await locator.click()
    if wait_type == "nav":
        await wait_nav(page, timeout)
    elif wait_type == "tab":
        await wait_tab_content(page, timeout)
    elif wait_type == "filter":
        await wait_filter_results(page, timeout)
    else:
        await wait_stable(page, timeout)


async def _dismiss_dialog(page, heading_text, label=""):
    """Fecha modal Angular Material (Exactix FL) tentando 4 estratégias."""
    dialog = page.locator(f'h1:has-text("{heading_text}")')
    if not await dialog.count():
        return True
    for sel in [
        'button[mat-dialog-close]', 'button.mat-dialog-close',
        'button:has-text("Close")', 'button:has-text("OK")',
        'button:has-text("Dismiss")', 'button:has-text("Cancel")',
        'mat-dialog-container button.mat-icon-button',
        'mat-dialog-container button[aria-label*="close" i]',
        'mat-dialog-container button[aria-label*="dismiss" i]',
    ]:
        try:
            btn = page.locator(sel)
            if await btn.count():
                await btn.first.click(timeout=2000)
                await page.wait_for_timeout(400)
                if not await dialog.count():
                    return True
        except Exception:
            continue
    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(400)
        if not await dialog.count():
            return True
    except Exception:
        pass
    try:
        backdrop = page.locator('.cdk-overlay-backdrop').first
        if await backdrop.count():
            await backdrop.click(timeout=2000, force=True)
            await page.wait_for_timeout(400)
            if not await dialog.count():
                return True
    except Exception:
        pass
    try:
        await page.evaluate(
            "document.querySelectorAll('.cdk-overlay-container mat-dialog-container,"
            " .cdk-overlay-pane, .cdk-overlay-backdrop').forEach(el => el.remove())"
        )
        await page.wait_for_timeout(200)
        if not await dialog.count():
            log.warning(f"  {label}: Dialog '{heading_text}' removido via JS (fallback)")
            return True
    except Exception as e:
        log.warning(f"  {label}: Falha ao remover dialog via JS: {e}")
    log.warning(f"  {label}: NAO conseguiu fechar dialog '{heading_text}'")
    return False


# ── │ SECTION: PLAYWRIGHT_CTX │ PLAYWRIGHT CONTEXT MANAGER (evita Chrome órfão)
@asynccontextmanager
async def playwright_context(state, headless=True):
    """Context manager que garante fechamento do browser mesmo com crash.

    Uso:
        async with playwright_context("FL") as (p, ctx, page):
            ...
    """
    perfil = _profile_path(state)
    p = ctx = page = None
    try:
        p = await async_playwright().start()
        ctx = await p.chromium.launch_persistent_context(
            perfil, headless=headless, args=["--no-sandbox"]
        )
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)
        yield p, ctx, page
    finally:
        if ctx:
            try:
                await ctx.close()
            except Exception:
                pass
        if p:
            try:
                await p.stop()
            except Exception:
                pass


# ── │ SECTION: SUPABASE_IO │ SUPABASE (com retry genérico) ───────────────────
def _sb_request(method, url, retries=3, **kwargs):
    """Wrapper genérico com retry e backoff exponencial pra TODAS as operações Supabase."""
    kwargs.setdefault("timeout", 20)
    for attempt in range(retries):
        try:
            r = method(url, **kwargs)
            r.raise_for_status()
            return r
        except requests.exceptions.HTTPError as e:
            status_code = e.response.status_code if e.response is not None else 0
            # 4xx (exceto 429) = erro do cliente, não adianta retry
            if 400 <= status_code < 500 and status_code != 429:
                log.error(f"[Supabase] {status_code} — {e}")
                raise
            if attempt < retries - 1:
                wait = 2 ** (attempt + 1)
                log.warning(f"[Supabase] Retry {attempt+1}/{retries} em {wait}s: {e}")
                time.sleep(wait)
            else:
                log.error(f"[Supabase] Falha após {retries} tentativas: {e}")
                raise
        except Exception as e:
            if attempt < retries - 1:
                wait = 2 ** (attempt + 1)
                log.warning(f"[Supabase] Retry {attempt+1}/{retries} em {wait}s: {e}")
                time.sleep(wait)
            else:
                log.error(f"[Supabase] Falha após {retries} tentativas: {e}")
                raise


def _qv(val):
    """URL-encode um valor pra PostgREST query (safe='' escapa &, =, espaço, etc.)."""
    return urllib.parse.quote(str(val), safe="")


def sb_get(table, qs=""):
    r = _sb_request(requests.get, f"{SB_URL}/rest/v1/{table}?select=*{qs}", headers=SB_H)
    return r.json()


def sb_upsert(table, data, on_conflict="ticket_num,utility_name,state"):
    """Upsert com retry. Suporta dict único ou lista."""
    h = {**SB_H, "Prefer": "resolution=merge-duplicates,return=minimal"}
    _sb_request(requests.post, f"{SB_URL}/rest/v1/{table}?on_conflict={on_conflict}",
                headers=h, json=data, timeout=30)


def sb_insert(table, data):
    r = _sb_request(requests.post, f"{SB_URL}/rest/v1/{table}", headers=SB_H, json=data)
    return r.json()


def sb_patch(table, id_val, data):
    h = {**SB_H, "Prefer": "return=minimal"}
    _sb_request(requests.patch, f"{SB_URL}/rest/v1/{table}?id=eq.{id_val}", headers=h, json=data)


def sb_delete(table, qs):
    h = {**SB_H, "Prefer": "return=minimal"}
    _sb_request(requests.delete, f"{SB_URL}/rest/v1/{table}?{qs}", headers=h)


def log_start(state, by="manual"):
    res = sb_insert("sync_811_log", {"state": state, "status": "running", "triggered_by": by})
    return res[0]["id"] if res else None


def log_finish(lid, checked, updated, status="success", error=None):
    if not lid:
        return
    sb_patch("sync_811_log", lid, {
        "finished_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
        "tickets_checked": checked, "tickets_updated": updated,
        "status": status, "error_msg": error
    })


# ── │ SECTION: AUTO_NOTES │ NOTAS AUTOMÁTICAS (dedup + truncamento) ──────────
def append_auto_note(existing_notes, new_note, max_notes=MAX_AUTO_NOTES):
    """Adiciona nota automática ao campo notes com dedup e truncamento.

    - Não adiciona se a nota exata já existir
    - Remove notas auto mais antigas se ultrapassar max_notes
    - Retorna o campo notes atualizado
    """
    existing = (existing_notes or "").strip()
    lines = existing.split("\n") if existing else []

    # Dedup: se a nota exata já existe, não adiciona
    if new_note.strip() in [l.strip() for l in lines]:
        return existing

    lines.append(new_note.strip())

    # Separa notas automáticas das manuais
    auto_lines = [l for l in lines if l.strip().startswith("[AUTO 811]")]
    manual_lines = [l for l in lines if not l.strip().startswith("[AUTO 811]")]

    # Trunca auto-notas se exceder limite (mantém as mais recentes)
    if len(auto_lines) > max_notes:
        auto_lines = auto_lines[-max_notes:]

    # Reconstrói: manuais primeiro, depois auto
    result = "\n".join(manual_lines + auto_lines).strip()
    return result


# ── │ SECTION: CLASSIFY │ CLASSIFY STATUS ─────────────────────────────────────
def classify(status_text, response_text=""):
    """Classifica resposta de utility como Clear, Pending ou Damage.

    Retorna tupla (status, unrecognized):
      - status: "Clear", "Pending" ou "Damage"
      - unrecognized: True se nenhum padrão conhecido bateu (fallback genérico)
    """
    s = (status_text or "").strip().lower()
    r = (response_text or "").strip().lower()
    full = (s + " " + r).lower()

    # ── INDIANA 811 — OPEN CODES (PENDING: precisam de atualização) ──
    # 3C: Marking Delay — do not excavate
    # 3F: Line untonable — do not excavate
    # 3G: Ongoing — partially marked, do not excavate unmarked area
    # 6A: Joint Meet Conflict — meeting in conflict
    # 6B: Joint Meet Accepted — meeting accepted (aguardando)

    # ── INDIANA 811 — CLOSED CODES ──
    # 1:  Marked → Clear
    # 1A: Marked with Exceptions, Do Not Excavate, High-Profile → Clear (utility marcou)
    # 1B: Marked with Exceptions, High-Profile → Clear (MAY contact)
    # 1C: Work by Facility Owner → Clear
    # 2:  Clear — no underground facilities → Clear
    # 3A: Could Not Gain Access → PENDING (do not excavate until resolved)
    # 3B: Incorrect Address → PENDING (do not excavate until resolved)
    # 3D: Marking Instructions Unclear → PENDING (do not excavate until resolved)
    # 3E: Excavation Already Performed or Canceled → PENDING (utility NÃO marcou)
    # 4:  Private Line → Clear (private locator needed)
    # 5A: Design Notice Documents Provided → Clear
    # 5B: Design Notice Marked → Clear
    # 6C: Joint Meet Complete → Clear
    # 7:  Damage → Damage

    # ── FL SUNSHINE 811 CODES ──
    # 1: Marked → Clear
    # 2: Clear → Clear
    # 2E: Marked with Exceptions → Clear
    # 3F: Marking delay → Pending
    # 3H: Privately owned → Clear (needs private locator)
    # 3T: Extraordinary circumstances → Pending
    # 3U: Not service provider → Clear
    # 4: Clear No Facilities → Clear
    # 5: No Conflict → Clear
    # 6A: Active Facilities → Pending (DO NOT demolish)
    # 8: Ongoing Job → Pending

    # ── IL/JULIE: ENTRADAS DE AGENDAMENTO (não são respostas de status) ──
    # Idealmente filtradas antes de chegar aqui, mas safety net pra não disparar alerta
    SCHEDULE_SKIP = [
        "declined code 50", "accepted code 50",
        "locator and excavator agreed", "alternate marking schedule",
        "alternate date requested",
    ]
    for pat in SCHEDULE_SKIP:
        if pat in full:
            return "Pending", False  # Reconhecido: agendamento, não resposta real

    # ── DAMAGE ──
    if "7:" in full and "damage" in full:
        return "Damage", False

    # ── 1A: Marked with Exceptions — utility MARCOU, é Clear ──
    # "Do Not Excavate, High-Profile" é restrição operacional (Watch & Protect),
    # mas a utility respondeu e marcou → status é Clear.
    if "1a" in full and ("marked with exceptions" in full or "high-profile" in full):
        return "Clear", False

    # ── ALWAYS PENDING — "do not excavate/demolish" overrides everything ──
    # (1A já tratado acima — não cai aqui)
    if "do not excavate" in full and "3u" not in full and "1a" not in full:
        return "Pending", False
    if "do not demolish" in full:
        return "Pending", False

    # ── SPECIFIC CLEAR CODES (check before generic blockers) ──
    # WI / Diggers Hotline: "Not Participating" = utility não atende a área (similar ao 3U)
    if "3u" in full or "not service provider" in full or "not participating" in full:
        return "Clear", False
    # WI: "Closed by DHL" = Diggers Hotline fechou porque utility não respondeu
    # dentro do prazo legal (Wis. Stat. §182.0175).
    # DISTINÇÃO CRÍTICA:
    #   - Se Facilities diz "Not Participating" → utility não participa → Clear
    #   - Se NÃO diz "Not Participating" → utility ignorou o prazo → Cancel
    # A checagem de "not participating" acima (L633) já cobre o 1º caso.
    # Aqui trata só o 2º: prazo expirou sem resposta real.
    if "closed by dhl" in full or "closed by diggers" in full:
        return "Cancel", False
    if "3h" in full or "privately owned" in full or "private facility owner" in full:
        return "Clear", False
    if "3e" in full and ("already performed" in full or "canceled" in full):
        return "Pending", False  # Utility NÃO marcou — escavação aconteceu antes do locator
    # IL/JULIE: Code 60 — Watch and Protect (W&P): utility has critical facility,
    # rep must be present during excavation. Clear, but needs W&P coordination.
    if "watch and protect" in full:
        return "Clear", False
    # IL/JULIE: Code 21 — Re-mark Not Needed. É um ACK de extensão de prazo, NÃO uma
    # liberação real (Eric 2026-05-27: "não é clear real, é ack de extensão"). → Pending.
    # Se a utility marcou de verdade numa revisão anterior, o dedup em save_to_supabase
    # mantém aquela resposta real (Clear) sobre este ack (JULIE traz todas as revisões).
    if "re-mark not needed" in full or "remark not needed" in full:
        return "Pending", False
    # IL/JULIE: Code 22 — Re-mark Needed. Quando o ticket é estendido COM
    # solicitação de nova marcação, todas as utilities precisam remarcar.
    # RESET total: todas voltam a Pending até marcarem de novo.
    if "re-mark needed" in full or "remark needed" in full:
        return "Pending", False
    # IL/JULIE: Ticket aberto pra reportar facility exposed/damaged. Utility precisa
    # inspecionar antes de liberar escavação. Tratamos como Pending até resposta real.
    if "reporting of an exposed" in full or "reporting of a damaged" in full:
        return "Pending", False

    # ── SPECIFIC PENDING CODES ──
    if "6a" in full or "active facilities" in full:
        return "Pending", False
    if "6b" in full and "joint meet accepted" in full:
        return "Pending", False
    if "3a" in full and "could not gain access" in full:
        return "Pending", False
    if "3b" in full and "incorrect address" in full:
        return "Pending", False
    if "3d" in full and ("instructions" in full or "unclear" in full):
        return "Pending", False
    if "3g" in full and ("ongoing" in full or "partially" in full):
        return "Pending", False
    if "3t" in full and "extraordinary" in full:
        return "Pending", False

    # ── WI / Diggers Hotline: "Ongoing - Working with Excavator" ──
    # Utility começou a trabalhar mas ainda não terminou — bloqueia escavação.
    if "ongoing" in full and ("excavator" in full or "working with" in full):
        return "Pending", False

    # ── GENERIC BLOCKED PATTERNS ──
    # WI / Diggers Hotline manda "Not Marked - Delay - Delay" quando utility
    # pediu prazo SEM ter marcado ainda. Tem que ficar antes do match de "marked"
    # genérico (que pegaria "not marked" como substring → Clear errado).
    BLOCKED = [
        "no response", "no access", "unmarked", "unmark", "not marked",
        "marking delay", "incorrect address", "unclear instruction",
        "ongoing job", "scheduled marking", "late ticket",
    ]
    for b in BLOCKED:
        if b in full:
            return "Pending", False

    # ── GENERIC CLEAR PATTERNS ──
    CLEARED = [
        "1: marked", "1 - marked", "1 marked", "1b", "1c",
        "marked with exception", "2e",
        "2: clear", "2 - clear", "2 clear",
        "4: clear", "4 - clear", "4 clear", "4: private", "4 private line",
        "5: no conflict", "5 - no conflict", "no conflict", "no facilit",
        "5a", "5b", "6c", "joint meet complete",
    ]
    for c in CLEARED:
        if c in full:
            return "Clear", False

    # Guard: "marked" sozinho como substring pode pegar "not marked"/"unmarked"/
    # "remarked". O BLOCKED acima já filtra esses; mantém guard defensivo aqui também.
    if "marked" in full and "not marked" not in full and "unmarked" not in full and "remarked" not in full:
        return "Clear", False
    if "clear" in full and "unclear" not in full:
        return "Clear", False
    if s == "clear":
        return "Clear", False

    # ── AMBIGUOUS: Positive Response / Current sem código reconhecido ──
    if "positive response" in s or s == "current":
        if "3u" in r or "not service provider" in r:
            return "Clear", False
        if "3h" in r or "privately owned" in r:
            return "Clear", False
        for b in BLOCKED:
            if b in r:
                return "Pending", False
        if "marked" in r or "clear" in r or "no conflict" in r or "no facilit" in r:
            return "Clear", False
        return "Pending", True  # ⚠ UNRECOGNIZED: Positive Response/Current sem padrão

    # ── TAGS DE UTILITY (IL/JULIE) ──
    # Algumas utilities respondem só com a categoria delas, em colchetes ou cru.
    # Não é uma resposta real — significa que ainda não posicionaram, mas não é
    # um padrão "novo" que mereça alerta de unrecognized.
    rt_stripped = (response_text or "").strip()
    if re.match(r"^\[[A-Z]{3,5}\]$", rt_stripped):
        return "Pending", False  # tag tipo [COMM], [GAS], [ELEC], [FIBR], [WATR]
    if rt_stripped.upper() in {"ELECTRIC", "GAS", "WATER", "COMM", "TELECOM", "FIBER", "MCIU01"}:
        return "Pending", False  # label cru de tipo de utility

    # ── LATE RESPONSE ──
    # Resposta tardia mas válida; já se sabe que está pendente.
    if "late final" in full or "late response" in full:
        return "Pending", False

    return "Pending", True  # ⚠ UNRECOGNIZED: nenhum padrão reconhecido


def is_in_renewal_grace(ticket, ref_date=None):
    """Determina se um ticket renovado ainda está em período de carência.

    Args:
        ticket: dict com campos old_ticket2, expire_old
        ref_date: data de referência (default: hoje). Injetável para testes.

    Returns: (in_grace, old_ticket_num)
        in_grace: True se renovado E dentro do período de carência
        old_ticket_num: número do ticket antigo (str) ou "" se não aplicável
    """
    old_chain = (ticket.get("old_ticket2") or "").strip()
    if not old_chain:
        return False, ""
    old_num = old_chain.split(" → ")[0].strip()
    old_expire_str = (ticket.get("expire_old") or "").strip()
    if not old_expire_str or old_expire_str == "—":
        return False, old_num
    try:
        exp_str = old_expire_str.split("Time:")[0].strip()
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                exp_dt = datetime.strptime(exp_str, fmt)
                break
            except ValueError:
                continue
        else:
            return False, old_num
    except Exception:
        return False, old_num
    today = ref_date or datetime.now().date()
    if isinstance(today, datetime):
        today = today.date()
    return exp_dt.date() >= today, old_num


def _get_relo_merged_responses(tnum, old_ticket_num, deduped_responses, state):
    """Merge respostas do Relo-No-Show (novo) com herdadas do Standard (antigo).

    WI-only: quando um Relo-No-Show é criado, ele contém APENAS as utilities que
    não responderam ao Standard. As que já responderam ficam com risco no portal
    e não aparecem no Relo. Pra avaliar auto-clear, precisamos combinar:
      - Relo: respostas das utilities que faltaram (tem prioridade)
      - Standard: respostas das utilities que já tinham respondido (herdadas)

    Retorna lista combinada de responses, ou None se não aplicável.
    """
    if state != "WI":
        return None
    if not old_ticket_num:
        return None

    new_utils = {r["utility"] for r in deduped_responses}
    if not new_utils:
        return None

    try:
        old_resps = sb_get(
            "ticket_811_responses",
            f"&ticket_num=eq.{old_ticket_num}&state=eq.{state}"
            "&select=utility_name,status,response_text"
        )
    except Exception as e:
        log.debug(f"[{state}] {tnum}: Relo merge — erro ao buscar respostas do antigo: {e}")
        return None

    if not old_resps:
        return None  # antigo sem respostas → fallback normal

    old_utils = {r["utility_name"] for r in old_resps}
    if len(new_utils) >= len(old_utils):
        return None  # não é Relo parcial (mesmo tamanho ou maior) → lógica normal

    # Merge: herda do antigo, sobrescreve com o novo
    merged = {}
    for r in old_resps:
        merged[r["utility_name"]] = {
            "utility": r["utility_name"],
            "status": r["status"],
            "response": r.get("response_text", ""),
            "_inherited": True
        }
    for r in deduped_responses:
        merged[r["utility"]] = r  # Relo sobrescreve

    inherited_names = [k for k, v in merged.items() if v.get("_inherited")]
    log.info(
        f"[{state}] {tnum}: 🔀 RELO MERGE — {len(new_utils)} no Relo + "
        f"{len(inherited_names)} herdadas do Standard ({old_ticket_num}) "
        f"= {len(merged)} total"
    )

    return list(merged.values())


def _is_valid_utility_name(name):
    """Valida que uma string é nome real de utility (não lixo de UI do portal).

    Rejeita:
      - Vazio, muito curto (<3 chars)
      - Códigos de ID puros (ex: "NI0005", "ID8000") — letras+números sem espaço
      - Botões/filtros da UI do portal Indiana: "All (6)", "Current (3)",
        "Event", "Positive Response", "No Response", "Show all", etc.
      - Cabeçalhos de tabela: "Status", "Date", "Service Area", "Response"

    Exemplos válidos: "DUKE ENERGY", "COMCAST NORTH", "IN AMERICAN WATER"
    Exemplos inválidos: "All (6)", "Current", "NI0005", "ID2227", "Event"
    """
    if not name:
        return False
    n = str(name).strip()
    if len(n) < 3:
        return False
    # Prefixo ID ex: "ID2227"
    if n.startswith("ID") and re.match(r"^ID\d+$", n):
        return False
    # Código puro letras+números COM dígitos: "NI0005", "ID8000" → rejeita
    # Siglas puras só-letras (MCI, AEP, TECO) são utilities reais → aceita
    if re.match(r"^[A-Z0-9]{2,10}$", n) and re.search(r"[0-9]", n):
        return False
    # Lixo da UI do portal: "All (6)", "Current (3)", "Show all (9)"
    if re.match(r"^(All|Current|Show\s+all|Show|Hide|Filter|Previous|Next|Page)\s*\(?\s*\d*\s*\)?\s*$", n, re.IGNORECASE):
        return False
    # Labels de eventos/status puros
    ui_labels = {
        "event", "events", "status", "date", "service area", "response", "responses",
        "positive response", "no response", "entry method", "comments", "comment",
        "web service", "utility", "utilities", "utility name", "service",
        "current", "history", "ticket", "actions", "action",
    }
    if n.lower() in ui_labels:
        return False
    return True


def needs_private_locator(response_text):
    """Detecta se a resposta indica necessidade de private locator (3H)."""
    r = (response_text or "").strip().lower()
    return "3h" in r or "privately owned" in r or "private facility owner" in r


def needs_watch_and_protect(response_text):
    """Detecta se a resposta indica Watch and Protect (código 60 — IL/JULIE).

    W&P = utility tem instalação crítica e exige representante presente durante escavação.
    """
    r = (response_text or "").strip().lower()
    return "watch and protect" in r


# ── │ SECTION: LOCATION_TEXT │ EXTRAIR LOCATION TEXT DO BODY (IN + FL) ────────
def extract_location_text(body, state=""):
    """Extrai campo de localização do body do ticket."""
    if not body:
        return ""

    body_norm = re.sub(r'Locat\s*:', 'Locat:', body, flags=re.IGNORECASE)

    loc_idx = -1
    loc_offset = 0
    for loc_label in ["Locat:", "Location:"]:
        idx = body_norm.find(loc_label)
        if idx >= 0:
            loc_idx = idx
            loc_offset = len(loc_label)
            break

    if loc_idx < 0:
        return ""

    raw = body_norm[loc_idx + loc_offset:loc_idx + 2000]

    stop_pat = (
        r"(.*?)"
        r"(?:"
        r"\n\s*\n"
        r"|\nGrids|\nBoundary|\nService"
        r"|\*\*\*"
        r"|\nBoring|\nRemarks"
        r"|\nWork\s*[Dd]ate|\nDue\s*[Dd]ate|\nWork\s*[Tt]ype"
        r"|\nDone\s*[Ff]or|\nCompany|\nCategory"
        r"|\nHrs\s*notc|\nWhite-lined|\nUg/Oh|\nDepth"
        r"|\n\s*:\s*\n"
        r"|\nType\s*[Oo]f\s*[Ww]ork"
        r"|\nDuration|\nDig\s*[Ss]ite|\nStart\s*[Dd]ate"
        r"|\nExpir|\nTicket\s*#|\nFunction"
        r"|\n[A-Z][A-Za-z ]{2,25}:"
        r")"
    )
    m = re.search(stop_pat, raw, re.DOTALL)
    if m:
        return " ".join(m.group(1).split()).strip()
    return " ".join(raw[:800].split()).strip()


# ── │ SECTION: EXPIRE_PARSE │ EXTRAIR DATA DE VENCIMENTO DO BODY ──────────────
def extract_expire_date(body, ticket_num="", debug=False):
    """Extrai a data de vencimento REAL do ticket 811.

    IMPORTANTE: há 3 datas diferentes que podem aparecer no body:
      - Expires / Expiration Date / Ticket Expires / Exp Date → data real (correta)
      - Due Date / Due                                         → deadline pras utilities
      - Work Date / Start / Legal Date                         → data de início

    Formatos conhecidos:
      - Indiana 811:   "Ticket Expires: 05/13/2026 11:59 PM"
      - Sunshine FL:   "Due Date : 04/15/26 Time: 23:59ET  Exp Date : 05/13/26 Time: 23:59ET"
      - JULIE IL:      expire não está no body texto (retorna "")

    Sempre retorna normalizado como "MM/DD/YYYY" ou "" se não achar.
    """
    def _extract(pattern, text):
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else ""

    # PRIORIDADE: do mais específico/confiável pro mais genérico.
    # "Ticket Expires" antes de "Expires" pra evitar match parcial.
    # "Exp Date" é Sunshine FL — ANTES de genéricos pra não deixar passar.
    # O "?:\s{2,}|\t" no pattern Exp Date evita capturar o "Due Date" que vem na mesma linha.
    # NÃO incluir "Due Date", "Work Date", "Legal Date" — são datas DIFERENTES.
    patterns = [
        (r"Ticket\s+Expires?\s*(?:on)?\s*:\s*([^\n]+)",             "Ticket Expires"),
        (r"Expiration\s+Date\s*:\s*([^\n]+)",                        "Expiration Date"),
        (r"Expiration\s*:\s*([^\n]+)",                                "Expiration"),
        (r"(?<!\w)Expires?\s+on\s*:\s*([^\n]+)",                      "Expires on"),
        (r"(?<!\w)Expires\s*:\s*([^\n]+)",                            "Expires"),
        # Sunshine FL: "Exp Date : 05/13/26 Time: 23:59ET"
        # Usa lookahead pra parar em outro label ou fim de linha.
        (r"(?<!\w)Exp\s+Date\s*:\s*(.+?)(?=\s{2,}\w+\s*Date|\s{2,}Hrs|\s{2,}Work|\n|$)",  "Exp Date"),
        (r"(?<!\w)Expire\s*:\s*([^\n]+)",                             "Expire"),
    ]

    expire_raw = ""
    matched_label = ""
    for pat, label in patterns:
        expire_raw = _extract(pat, body)
        if expire_raw:
            matched_label = label
            break

    if not expire_raw:
        if debug:
            log.warning(f"[ExpireParse] {ticket_num}: nenhum pattern de Expires bateu no body")
        return ""

    if debug:
        log.info(f"[ExpireParse] {ticket_num}: '{matched_label}' → raw='{expire_raw[:80]}'")

    # LIMPEZAS:
    # 1. Remove " at " que alguns portais injetam: "05/13/2026 at 11:59 PM" → "05/13/2026 11:59 PM"
    cleaned = re.sub(r"\s+at\s+", " ", expire_raw, flags=re.IGNORECASE)
    # 2. Remove "Time:" literal do formato Sunshine FL: "05/13/26 Time: 23:59ET" → "05/13/26 23:59ET"
    cleaned = re.sub(r"\s*Time\s*:\s*", " ", cleaned, flags=re.IGNORECASE)
    # 3. Remove sufixos de timezone colados ou separados: "23:59ET" → "23:59", "23:59 EST" → "23:59"
    cleaned = re.sub(r"\s*(ET|EST|EDT|CT|CST|CDT|PT|PST|PDT|MT|MST|MDT|UTC|GMT)\b", "", cleaned, flags=re.IGNORECASE)
    # 4. Normaliza espaços múltiplos
    cleaned = re.sub(r"\s+", " ", cleaned).strip()

    # Pega as primeiras ~5 palavras (suficiente pra data + hora + AM/PM)
    expire_clean = " ".join(cleaned.split()[:5])

    # Tenta formatos do MAIS ESPECÍFICO pro MAIS GENÉRICO
    formats = [
        "%m/%d/%Y %I:%M:%S %p",   # 05/13/2026 11:59:00 PM
        "%m/%d/%Y %I:%M %p",       # 05/13/2026 11:59 PM
        "%m/%d/%Y %H:%M:%S",       # 05/13/2026 23:59:00
        "%m/%d/%Y %H:%M",          # 05/13/2026 23:59
        "%m/%d/%y %H:%M",          # 05/13/26 23:59       (Sunshine FL)
        "%m/%d/%y %I:%M %p",       # 05/13/26 11:59 PM
        "%m/%d/%y %I:%M:%S %p",    # 05/13/26 11:59:00 PM
        "%B %d, %Y %I:%M %p",      # May 13, 2026 11:59 PM
        "%b %d, %Y %I:%M %p",      # May 13, 2026 11:59 PM (abbrev)
        "%B %d, %Y",                # May 13, 2026
        "%b %d, %Y",                # May 13, 2026
        "%m/%d/%Y",                 # 05/13/2026
        "%m-%d-%Y",                 # 05-13-2026
        "%Y-%m-%d",                 # 2026-05-13
        "%m/%d/%y",                 # 05/13/26
    ]
    # Tenta parsear com cada formato pegando prefixo do tamanho apropriado.
    for fmt in formats:
        for attempt_len in [len(fmt) + 5, len(fmt) + 2, len(expire_clean)]:
            try:
                parsed = datetime.strptime(expire_clean[:attempt_len].strip(), fmt)
                return parsed.strftime("%m/%d/%Y")
            except ValueError:
                continue

    # Último recurso: extrai qualquer padrão MM/DD/YYYY ou MM/DD/YY
    m = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", expire_clean)
    if m:
        for fmt in ["%m/%d/%Y", "%m/%d/%y"]:
            try:
                return datetime.strptime(m.group(1), fmt).strftime("%m/%d/%Y")
            except ValueError:
                continue

    if debug:
        log.warning(f"[ExpireParse] {ticket_num}: não parseou '{expire_clean}' com nenhum formato")
    return ""  # Nunca retornar string parcial/crua — melhor vazio


def normalize_expire(s):
    """Normaliza qualquer valor de expire pra 'MM/DD/YYYY' ou ''.

    Aceita:
      - 'MM/DD/YYYY'               → retorna como está
      - 'MM/DD/YY Time: HH:MM'     → legado poluído, limpa
      - 'MM/DD/YY HH:MMET'         → formato Sunshine bruto
      - '05/13/26'                 → converte pra '05/13/2026'
      - '' / None / '—'            → retorna ''

    Usa as mesmas regras do extract_expire_date nas partes de limpeza,
    mas opera sobre um valor já extraído (não sobre body cru).
    Garante que banco/app sempre tenham MM/DD/YYYY consistente.
    """
    if not s:
        return ""
    s = str(s).strip()
    if s in ("—", "-", "N/A", "null", "None"):
        return ""
    # Aplica mesmas limpezas do extract_expire_date
    cleaned = re.sub(r"\s+at\s+", " ", s, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*Time\s*:\s*", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*(ET|EST|EDT|CT|CST|CDT|PT|PST|PDT|MT|MST|MDT|UTC|GMT)\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    prefix = " ".join(cleaned.split()[:5])
    formats = [
        "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
        "%m/%d/%y %H:%M", "%m/%d/%y %I:%M %p", "%m/%d/%y %I:%M:%S %p",
        "%B %d, %Y %I:%M %p", "%b %d, %Y %I:%M %p",
        "%B %d, %Y", "%b %d, %Y",
        "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y",
    ]
    for fmt in formats:
        for attempt_len in [len(fmt) + 5, len(fmt) + 2, len(prefix)]:
            try:
                return datetime.strptime(prefix[:attempt_len].strip(), fmt).strftime("%m/%d/%Y")
            except ValueError:
                continue
    m = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", prefix)
    if m:
        for fmt in ["%m/%d/%Y", "%m/%d/%y"]:
            try:
                return datetime.strptime(m.group(1), fmt).strftime("%m/%d/%Y")
            except ValueError:
                continue
    return ""


def _is_polluted_expire(s):
    """Detecta formato antigo/poluído que deveria ser re-scrapado."""
    if not s:
        return False
    s = str(s)
    if "Time:" in s or re.search(r"\b(ET|EST|EDT)\b", s):
        return True
    # Se não é MM/DD/YYYY limpo, considera poluído
    if not re.fullmatch(r"\d{2}/\d{2}/\d{4}", s.strip()):
        return True
    return False


# ── │ SECTION: CANCEL_DETECT │ DETECTAR CANCELAMENTO ──────────────────────────
def is_ticket_canceled(body):
    if not body:
        return False
    first_line = body.strip().split("\n")[0].strip().upper()
    body_upper = body.upper()
    return (
        first_line.startswith("CANCEL")
        or "FUNCTION: CANCEL" in body_upper
        or "FUNCTION:\nCANCEL" in body_upper
        or ("CANCELED" in body_upper and "REPLACED BY TICKET NUMBER" in body_upper)
        or "DUE TO TICKET BEING CANCELED" in body_upper
    )


# ── │ SECTION: COORDS_ADJUST │ AJUSTE DE POSIÇÃO PELO TEXTO 811 ──────────────
def adjust_coords_by_location(lat, lon, location_text, tipo_out=None):
    if not lat or not lon or not location_text:
        return lat, lon, tipo_out or "Main line"

    text = location_text.upper()
    OFFSET = 0.00012

    is_service = any(w in text for w in ["LATERAL", " REAR ", "REAR OF", "FRONT OF", " FRONT "])
    if is_service:
        tipo = "Service"
        if "NORTH" in text or " N " in text:
            return round(lat + OFFSET, 6), lon, tipo
        if "SOUTH" in text or " S " in text:
            return round(lat - OFFSET, 6), lon, tipo
        if "EAST" in text or " E " in text:
            return lat, round(lon + OFFSET, 6), tipo
        if "WEST" in text or " W " in text:
            return lat, round(lon - OFFSET, 6), tipo
        return round(lat + OFFSET * 0.5, 6), lon, tipo

    if "R/O/W" in text or "ROW TO ROW" in text or "R.O.W" in text:
        return lat, lon, "Main line"

    tipo = "Main line"
    if any(p in text for p in ["NORTH SIDE", "N SIDE", "NORTHSIDE", "ON NORTH", "ALONG NORTH"]):
        return round(lat + OFFSET, 6), lon, tipo
    if any(p in text for p in ["SOUTH SIDE", "S SIDE", "SOUTHSIDE", "ON SOUTH", "ALONG SOUTH"]):
        return round(lat - OFFSET, 6), lon, tipo
    if any(p in text for p in ["EAST SIDE", "E SIDE", "EASTSIDE", "ON EAST", "ALONG EAST"]):
        return lat, round(lon + OFFSET, 6), tipo
    if any(p in text for p in ["WEST SIDE", "W SIDE", "WESTSIDE", "ON WEST", "ALONG WEST"]):
        return lat, round(lon - OFFSET, 6), tipo
    return lat, lon, tipo


# ── │ SECTION: GEOCODING │ GEOCODING ──────────────────────────────────────────
async def geocode_address(street, city, state_code):
    import aiohttp
    STATE_BOUNDS = {
        "IN": {"lat": (37.77, 41.76), "lon": (-88.10, -84.79)},
        "FL": {"lat": (24.39, 31.00), "lon": (-87.63, -79.97)},
        "WI": {"lat": (42.49, 47.08), "lon": (-92.89, -86.25)},
        "IL": {"lat": (36.97, 42.51), "lon": (-91.51, -87.02)},
    }
    query = f"{street}, {city}, {state_code}, USA"
    try:
        # Fix bug #14: Nominatim usage policy exige MAX 1 req/segundo por IP.
        # 0.3s violava o TOS e podia levar a ban do IP em produção. 1.1s dá margem.
        # Ver: https://operations.osmfoundation.org/policies/nominatim/
        await asyncio.sleep(1.1)
        async with aiohttp.ClientSession() as session:
            async with session.get(
                "https://nominatim.openstreetmap.org/search",
                params={"q": query, "format": "json", "limit": 5, "countrycodes": "us"},
                headers={"User-Agent": "OneDrill-811-Sync/2.0"},
                timeout=aiohttp.ClientTimeout(total=10)
            ) as r:
                results = await r.json()

        bounds = STATE_BOUNDS.get(state_code, {})
        lat_range = bounds.get("lat", (-90, 90))
        lon_range = bounds.get("lon", (-180, 180))

        for item in results:
            lat, lon = float(item["lat"]), float(item["lon"])
            if lat_range[0] <= lat <= lat_range[1] and lon_range[0] <= lon <= lon_range[1]:
                return lat, lon
        if city:
            return await geocode_address(street, "", state_code)
    except Exception as e:
        log.debug(f"Geocoding error for '{query}': {e}")
    return None, None


# ── │ SECTION: COUNTY_RESOLVER │ COUNTY RESOLVER (Fase 1 filtro por county) ──
# Resolve o county de um ticket a partir de (location, state, lat, lon).
# Estratégias em cascata: 1) lookup direto 2) fuzzy 3) Nominatim reverse (fallback).

_COUNTIES_DB = None  # carregado lazy

def _load_counties_db():
    """Carrega counties_data.json uma vez. Se arquivo não existe, retorna {} e loga warning."""
    global _COUNTIES_DB
    if _COUNTIES_DB is not None:
        return _COUNTIES_DB
    path = os.path.join(BASE_DIR, "counties_data.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            _COUNTIES_DB = _json.load(f)
        total = sum(len(v) for v in _COUNTIES_DB.values())
        log.info(f"[county] Base carregada: {total} cidades ({', '.join(f'{k}={len(v)}' for k,v in _COUNTIES_DB.items())})")
    except FileNotFoundError:
        log.warning(f"[county] Arquivo counties_data.json não encontrado em {path} — só fallback Nominatim será usado")
        _COUNTIES_DB = {}
    except Exception as e:
        log.error(f"[county] Erro carregando counties_data.json: {e}")
        _COUNTIES_DB = {}
    return _COUNTIES_DB


def _normalize_city_name(name):
    """Normaliza nome de cidade pra lookup: lowercase, remove pontuação, trim, expande abreviações comuns."""
    if not name:
        return ""
    n = name.lower().strip()
    # Insere espaço antes de letras quando grudado em pontuação (ex: "st.petersburg" → "st. petersburg")
    n = re.sub(r"\.(\w)", r". \1", n)
    # Remove pontuação comum
    n = re.sub(r"[.,;'\"]", "", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def _city_variants(name):
    """Gera variantes de busca pra uma cidade: com/sem 'saint', 'fort', com/sem pontuação."""
    base = _normalize_city_name(name)
    variants = {base}
    # st ↔ saint
    if base.startswith("st "):
        variants.add("saint " + base[3:])
    if base.startswith("saint "):
        variants.add("st " + base[6:])
    # ft ↔ fort
    if base.startswith("ft "):
        variants.add("fort " + base[3:])
    if base.startswith("fort "):
        variants.add("ft " + base[5:])
    return variants


async def _reverse_geocode_county(lat, lon):
    """Fallback: Nominatim reverse geocoding pra pegar county.
    Respeita o rate limit 1.1s (fix #14) — compartilha com geocode_address.
    """
    import aiohttp
    try:
        await asyncio.sleep(1.1)
        async with aiohttp.ClientSession() as session:
            async with session.get(
                "https://nominatim.openstreetmap.org/reverse",
                params={"lat": lat, "lon": lon, "format": "json", "zoom": 10, "addressdetails": 1},
                headers={"User-Agent": "OneDrill-811-Sync/2.0"},
                timeout=aiohttp.ClientTimeout(total=10)
            ) as r:
                data = await r.json()
        addr = data.get("address", {}) or {}
        # Nominatim retorna "county" na maioria dos casos; às vezes "Pinellas County" vem com "County" no fim
        county = addr.get("county") or ""
        if county:
            # Remove sufixo " County" se existir — salvamos só o nome
            county = re.sub(r"\s+County\s*$", "", county, flags=re.IGNORECASE).strip()
        return county or None
    except Exception as e:
        log.debug(f"[county] Reverse geocode error lat={lat},lon={lon}: {e}")
        return None


def _extract_city_candidates(location):
    """Extrai candidatos a 'cidade' do campo location, cobrindo os formatos reais vistos:

    - "St. Petersburg, FL"          → ["St. Petersburg"]
    - "Orlando - Tangelo Park"      → ["Orlando", "Tangelo Park"]          (City - Neighborhood)
    - "Vigo - Terre Haute"          → ["Vigo", "Terre Haute"]              (County - City)
    - "ILLINOIS - ELK GROVE"        → ["ILLINOIS", "ELK GROVE"]            (STATE - City)
    - "Pinelas - St. Petersburg"    → ["Pinelas", "St. Petersburg"]        (County typo - City)
    - "St. Petesburg"               → ["St. Petesburg"]                    (City com typo)

    Retorna lista em ordem de tentativa, sem duplicatas.
    """
    if not location:
        return []
    candidates = []
    loc = location.strip()
    # Primeiro segmento antes de vírgula (tira "FL" / "IN" etc do final)
    first_comma = loc.split(",")[0].strip()
    # Se tem hífen, adiciona AMBOS os lados como candidatos
    if " - " in first_comma:
        parts = [p.strip() for p in first_comma.split(" - ") if p.strip()]
        for p in parts:
            if p and p not in candidates:
                candidates.append(p)
    else:
        # Sem hífen, o primeiro segmento inteiro é o candidato
        if first_comma and first_comma not in candidates:
            candidates.append(first_comma)
    return candidates


def _fuzzy_city_match(city_norm, state_db, max_distance=2):
    """Busca cidade na base tolerando pequenas variações (typos).
    Usa distância de Levenshtein simples. max_distance=2 pega casos como:
      'pinelas' → 'pinellas' (dist 1)
      'petesburg' → 'petersburg' (dist 1)
      'jaksonvile' → 'jacksonville' (dist 2)

    Retorna o county se achou match bom, "" caso contrário.
    """
    if not city_norm or not state_db:
        return ""

    def _lev(a, b):
        # Levenshtein iterativo (O(m*n) tempo, O(min(m,n)) memória)
        if len(a) < len(b):
            a, b = b, a
        if not b:
            return len(a)
        # Early exit: se diferença de tamanho já passou o threshold, skip
        if len(a) - len(b) > max_distance:
            return max_distance + 1
        prev_row = list(range(len(b) + 1))
        for i, ca in enumerate(a, 1):
            curr_row = [i]
            for j, cb in enumerate(b, 1):
                ins = curr_row[j - 1] + 1
                dele = prev_row[j] + 1
                sub = prev_row[j - 1] + (ca != cb)
                curr_row.append(min(ins, dele, sub))
            prev_row = curr_row
        return prev_row[-1]

    # Só tenta fuzzy em nomes razoáveis (evita ruído com strings curtas)
    if len(city_norm) < 5:
        return ""

    best_match = None
    best_dist = max_distance + 1
    for city_in_db, county in state_db.items():
        # Só compara com nomes de tamanho similar (otimização)
        if abs(len(city_in_db) - len(city_norm)) > max_distance:
            continue
        d = _lev(city_norm, city_in_db)
        if d < best_dist:
            best_dist = d
            best_match = (city_in_db, county)
            if d == 0:
                break

    return best_match[1] if best_match else ""


async def resolve_county(location, state, lat=None, lon=None):
    """Resolve county de um ticket. Estratégias em cascata:
       1) Lookup exato (cobre 'St. Petersburg, FL', 'Tampa, Hillsborough')
       2) Múltiplos candidatos do location (cobre 'Orlando - Tangelo Park', 'Vigo - Terre Haute')
       3) Variantes (st↔saint, ft↔fort)
       4) Fuzzy match (cobre typos: 'Pinelas' → Pinellas, 'Petesburg' → Petersburg)
       5) Nominatim reverse geocoding se tem lat/lon

    Retorna string (ex: "Pinellas") ou "" se não conseguiu resolver.
    """
    if not state:
        return ""
    state = state.upper()
    db = _load_counties_db()
    state_db = db.get(state, {})

    # Extrai todos os candidatos a cidade do location
    candidates = _extract_city_candidates(location or "")

    # Estratégia 1 + 2 + 3: testa cada candidato com todas as variantes
    for candidate in candidates:
        for variant in _city_variants(candidate):
            if variant in state_db:
                return state_db[variant]

    # Estratégia 4: fuzzy match — tenta typos nos candidatos
    for candidate in candidates:
        candidate_norm = _normalize_city_name(candidate)
        fuzzy_county = _fuzzy_city_match(candidate_norm, state_db)
        if fuzzy_county:
            log.info(f"[county] fuzzy match: {candidate!r} → {fuzzy_county} (FL)")
            return fuzzy_county

    # Estratégia 5: fallback via reverse geocoding
    if lat is not None and lon is not None:
        try:
            lat_f = float(lat)
            lon_f = float(lon)
            county = await _reverse_geocode_county(lat_f, lon_f)
            if county:
                log.info(f"[county] {location!r}, {state}: resolvido via Nominatim → {county}")
                return county
        except (TypeError, ValueError):
            pass

    log.debug(f"[county] Não resolvido: location={location!r}, state={state}, lat={lat}, lon={lon}")
    return ""


# ── │ SECTION: AUTO_LOGIN │ AUTO-LOGIN ────────────────────────────────────────


# Helpers de humanizacao (anti-detect)
async def _human_wait(min_s=0.5, max_s=2.0):
    import random as _rnd
    await asyncio.sleep(_rnd.uniform(min_s, max_s))


async def _human_type_into(page, selector, text):
    import random as _rnd
    try:
        el = page.locator(selector).first
        await el.click()
        await _human_wait(0.2, 0.5)
        for ch in text:
            await page.keyboard.type(ch, delay=_rnd.randint(60, 160))
        return True
    except Exception as e:
        log.debug(f"[type] erro em {selector}: {e}")
        return False


async def auto_login_silent(state):
    log.info(f"[{state}] Tentando auto-login silent (anti-detect)...")

    if state not in PORTALS:
        log.error(f"[{state}] estado nao suportado")
        return False

    user = PORTALS[state]["user"]()
    passwd = PORTALS[state]["pass"]()
    if not user or not passwd:
        log.error(f"[{state}] Credenciais ausentes no .env")
        return False

    perfil = _profile_path(state)

    stealth_fn = None
    try:
        from playwright_stealth import stealth_async as _stealth
        stealth_fn = _stealth
    except ImportError:
        log.warning("[stealth] playwright-stealth nao instalado - mais detectavel")

    try:
        async with async_playwright() as p:
            ctx = await p.chromium.launch_persistent_context(
                perfil,
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-blink-features=AutomationControlled",
                    "--disable-features=IsolateOrigins,site-per-process",
                ],
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/131.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1920, "height": 1080},
                locale="en-US",
                timezone_id="America/New_York",
            )
            page = ctx.pages[0] if ctx.pages else await ctx.new_page()

            if stealth_fn:
                try:
                    await stealth_fn(page)
                except Exception:
                    pass

            page.set_default_timeout(TIMEOUT_PAGE)
            await page.goto(PORTALS[state]["url"], wait_until="domcontentloaded")
            await wait_stable(page)
            await _human_wait(1.5, 3.5)

            url_now = page.url.lower()
            if "login" not in url_now and "signin" not in url_now and "auth" not in url_now:
                log.info(f"[{state}] Ja logado via cookie - URL: {page.url}")
                await ctx.close()
                return True

            email_sels = [
                'input[type="email"]',
                'input[name="email"]',
                'input[name="username"]',
                'input[id*="email"]',
                'input[id*="user"]',
                "#email",
                "#username",
            ]
            pass_sels = [
                'input[type="password"]',
                'input[name="password"]',
                'input[id*="password"]',
                "#password",
            ]

            email_input = None
            for sel in email_sels:
                if await page.locator(sel).count():
                    email_input = sel
                    break
            if not email_input:
                log.error(f"[{state}] Campo email nao encontrado")
                await ctx.close()
                return False

            pass_input = None
            for sel in pass_sels:
                if await page.locator(sel).count():
                    pass_input = sel
                    break
            if not pass_input:
                log.error(f"[{state}] Campo senha nao encontrado")
                await ctx.close()
                return False

            log.debug(f"[{state}] Campos: email={email_input}, pass={pass_input}")

            if not await _human_type_into(page, email_input, user):
                await ctx.close()
                return False
            await _human_wait(0.5, 1.2)

            if not await _human_type_into(page, pass_input, passwd):
                await ctx.close()
                return False
            await _human_wait(0.8, 1.8)

            submit_clicked = False
            for sel in [
                'button[type="submit"]',
                'input[type="submit"]',
                'button:has-text("Login")',
                'button:has-text("Sign In")',
                'button:has-text("Submit")',
                'button:has-text("Log In")',
            ]:
                if await page.locator(sel).count():
                    try:
                        await page.locator(sel).first.click()
                        submit_clicked = True
                        break
                    except Exception:
                        continue
            if not submit_clicked:
                try:
                    await page.locator(pass_input).first.press("Enter")
                    submit_clicked = True
                except Exception:
                    pass

            if not submit_clicked:
                log.error(f"[{state}] Nao consegui submeter")
                await ctx.close()
                return False

            await _human_wait(2, 4)
            try:
                await page.wait_for_load_state("networkidle", timeout=30000)
            except Exception:
                pass
            await wait_stable(page)

            if "login" in page.url.lower() or "signin" in page.url.lower():
                log.error(f"[{state}] Login falhou - URL: {page.url}")
                await ctx.close()
                return False

            try:
                await page.goto(PORTALS[state]["dashboard"], wait_until="domcontentloaded")
                await wait_stable(page)
            except Exception:
                pass

            if "login" in page.url.lower():
                log.error(f"[{state}] Dashboard redireciona pra login")
                await ctx.close()
                return False

            log.info(f"[{state}] AUTO-LOGIN SILENT OK - URL: {page.url}")
            await ctx.close()
            return True

    except Exception as e:
        log.error(f"[{state}] Erro fatal em auto_login_silent: {e}", exc_info=True)
        return False



async def auto_login(state):
    log.warning(f"[{state}] Sessão expirada — abrindo janela para renovação rápida...")
    perfil = _profile_path(state)
    try:
        async with async_playwright() as p:
            ctx2 = await p.chromium.launch_persistent_context(
                perfil, headless=False, args=["--no-sandbox", "--start-maximized"], viewport=None
            )
            page2 = ctx2.pages[0] if ctx2.pages else await ctx2.new_page()
            page2.set_default_timeout(60000)
            await page2.goto(PORTALS[state]["url"], wait_until="domcontentloaded")
            await wait_stable(page2)
            log.info(f"[{state}] Aguardando login manual (max 3 minutos)...")

            for i in range(60):
                await page2.wait_for_timeout(3000)
                url_now = page2.url.lower()
                login_keywords = ["login", "google", "cognito", "accounts", "signin"]
                if not any(kw in url_now for kw in login_keywords) and "811" in url_now:
                    log.info(f"[{state}] Login detectado! URL: {page2.url}")
                    try:
                        await page2.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
                        await wait_stable(page2)
                        await page2.goto(PORTALS[state]["dashboard"], wait_until="domcontentloaded")
                        await wait_stable(page2)
                        log.info(f"[{state}] Cookies persistidos — URL final: {page2.url}")
                    except Exception as e:
                        log.debug(f"[{state}] Navegação pós-login: {e}")
                    await page2.wait_for_timeout(4000)
                    await asyncio.sleep(3)
                    try:
                        await ctx2.close()
                    except Exception:
                        pass
                    await asyncio.sleep(2)
                    return True
                if i % 10 == 0:
                    log.info(f"[{state}] Aguardando login... ({i*3}s)")

            await ctx2.close()
            log.error(f"[{state}] Timeout — login não completado")
            return False
    except Exception as e:
        log.error(f"[{state}] Erro na renovação: {e}")
        return False


# ── │ SECTION: PORTAL_NAV │ HELPERS: NAVEGAÇÃO DO PORTAL ──────────────────────
async def goto_dashboard(page, state):
    """Navega ao dashboard e seleciona filtros corretos."""
    portal_home = PORTALS[state]["home"]
    portal_dashboard = PORTALS[state]["dashboard"]

    await page.goto(portal_home, wait_until="domcontentloaded")
    await wait_stable(page)

    for sel in ['text="Go to Ticket Dashboard"', 'a:has-text("Ticket Dashboard")']:
        if await page.locator(sel).count():
            await click_and_wait(page, page.locator(sel).first, "nav")
            break

    if "dashboard" not in page.url.lower() and "tickets" not in page.url.lower():
        await page.goto(portal_dashboard, wait_until="domcontentloaded")
        await wait_nav(page)

    if "login" in page.url.lower():
        log.warning(f"[{state}] Dashboard redirecionou para login — tentando novamente...")
        await page.goto(portal_home, wait_until="domcontentloaded")
        await wait_nav(page)
        if "login" in page.url.lower():
            return False
        await page.goto(portal_dashboard, wait_until="domcontentloaded")
        await wait_nav(page)
        if "login" in page.url.lower():
            return False

    if state == "IN":
        for sel in ['label:has-text("My Office Tickets")', 'span:has-text("My Office Tickets")', 'text="My Office Tickets"']:
            el = page.locator(sel).first
            if await el.count():
                await click_and_wait(page, el, "stable")
                break
        for sel in ['label:has-text("Show Completed Tickets")', 'text="Show Completed Tickets"', 'input[type="checkbox"]:near(:text("Completed"))']:
            el = page.locator(sel).first
            if await el.count():
                try:
                    cb = page.locator('input[type="checkbox"]').first
                    if not await cb.is_checked():
                        await click_and_wait(page, el, "stable")
                except Exception:
                    await click_and_wait(page, el, "stable")
                break
    elif state == "FL":
        selected = False
        for sel in ['label:has-text("My Office Tickets")', 'span:has-text("My Office Tickets")', 'text="My Office Tickets"']:
            el = page.locator(sel).first
            if await el.count():
                await click_and_wait(page, el, "stable")
                selected = True
                break
        if not selected:
            for sel in ['label:has-text("My Company Tickets")', 'span:has-text("My Company Tickets")', 'text="My Company Tickets"']:
                el = page.locator(sel).first
                if await el.count():
                    await click_and_wait(page, el, "stable")
                    break

    return True


async def set_items_per_page(page, count=100):
    """Muda Items/Page pra 100 no dashboard."""
    try:
        sel = page.locator('.num-per-page-selector mat-select, mat-select:near(:text("Items / Page"))')
        if await sel.count():
            await sel.first.click()
            await page.wait_for_timeout(300)
            opt = page.locator(f'mat-option:has-text("{count}")')
            if await opt.count():
                await opt.first.click()
                await wait_stable(page)
                log.info(f"Items/Page alterado para {count}")
                return True
    except Exception as e:
        log.debug(f"Não conseguiu mudar Items/Page: {e}")
    return False


async def select_office(page, state):
    """Re-seleciona My Office/Company Tickets após voltar ao dashboard."""
    if state == "IN":
        for sel in ['label:has-text("My Office Tickets")', 'text="My Office Tickets"']:
            if await page.locator(sel).count():
                await click_and_wait(page, page.locator(sel).first, "stable")
                break
    elif state == "FL":
        selected = False
        for sel in ['label:has-text("My Office Tickets")', 'text="My Office Tickets"']:
            if await page.locator(sel).count():
                await click_and_wait(page, page.locator(sel).first, "stable")
                selected = True
                break
        if not selected:
            for sel in ['label:has-text("My Company Tickets")', 'text="My Company Tickets"']:
                if await page.locator(sel).count():
                    await click_and_wait(page, page.locator(sel).first, "stable")
                    break


async def filter_ticket(page, tnum):
    """Digita número do ticket no filtro e espera resultados."""
    lbl = page.locator('mat-label:has-text("Filter by Ticket Number")')
    if await lbl.count():
        await lbl.first.click()
    else:
        await page.locator('#mat-input-0').click(force=True)
    await page.wait_for_timeout(80)
    await page.keyboard.press("Control+a")
    await page.keyboard.type(tnum, delay=8)
    # Espera o ticket aparecer na lista em vez de networkidle genérico (mais rápido)
    try:
        await page.get_by_text(tnum, exact=True).first.wait_for(timeout=4000)
    except Exception:
        await page.wait_for_timeout(800)


async def back_to_dashboard(page, state):
    """Volta ao dashboard rapidamente."""
    await page.goto(PORTALS[state]["dashboard"], wait_until="domcontentloaded")
    await wait_stable(page)
    await select_office(page, state)


async def ensure_login(page, ctx, p, state):
    """Verifica login e renova se necessário. Retorna (page, ctx) atualizados."""
    perfil = _profile_path(state)
    if "login" in page.url.lower():
        await ctx.close()
        await asyncio.sleep(1)
        # Fix 2026-05-15: tenta auto_login_silent primeiro (full auto)
        ok = await auto_login_silent(state)
        if not ok:
            log.warning(f"[{state}] auto_login_silent falhou, tentando manual...")
            # FL/IN: avisa o Eric por email antes de abrir janela (o anti-bot trava o silent
            # nesses portais, então a janela manual sempre vai aparecer aqui).
            if state in ("FL", "IN"):
                send_session_expired_alert(state)
            ok = await auto_login(state)
        if not ok:
            return None, None
        await asyncio.sleep(1)
        ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)
        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)
        if "login" in page.url.lower():
            log.error(f"[{state}] Login falhou após renovação — cookies não persistiram")
            await ctx.close()
            return None, None
        log.info(f"[{state}] Sessão renovada com sucesso: {page.url}")
        # Login deu certo: zera a flag de "expirado" pra reabilitar o alerta na próxima expiração.
        if state in ("FL", "IN"):
            _clear_session_flag(state)
    return page, ctx


async def fast_back(page, state):
    """Volta ao dashboard usando go_back (mais rápido que goto completo)."""
    try:
        await page.go_back(wait_until="domcontentloaded")
        await page.wait_for_timeout(300)
        url = page.url.lower()
        if "dashboard" in url or ("tickets" in url and "ticket/" not in url):
            return
    except Exception:
        pass
    await back_to_dashboard(page, state)


# ── │ SECTION: SCRAPE │ SCRAPE (PARALELO) ────────────────────────────────────
async def scrape(state, ticket_numbers, tickets_data=None):
    results = {}
    perfil = _profile_path(state)

    # Tickets que precisam de notes (Text tab)
    needs_notes = set()
    if tickets_data:
        for t in tickets_data:
            notes = (t.get("notes") or "").strip()
            if not notes or notes == "[811 Location]":
                needs_notes.add(t["ticket"])

    # Divide em chunks pras abas paralelas
    chunks = [[] for _ in range(NUM_TABS)]
    for i, tnum in enumerate(ticket_numbers):
        chunks[i % NUM_TABS].append(tnum)
    chunks = [c for c in chunks if c]

    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)

        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)

        # Verifica login
        if "login" in page.url.lower():
            await ctx.close()
            await asyncio.sleep(1)
            ok = await auto_login_silent(state)

            if not ok:

                log.warning(f"[{state}] auto_login_silent falhou, tentando manual...")

                ok = await auto_login(state)
            if not ok:
                return results
            await asyncio.sleep(1)
            ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
            page = ctx.pages[0] if ctx.pages else await ctx.new_page()
            page.set_default_timeout(TIMEOUT_PAGE)
            await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
            await wait_stable(page)
            if "login" in page.url.lower():
                log.error(f"[{state}] Scrape: login falhou após renovação")
                await ctx.close()
                return results

        log.info(f"[{state}] Logado! Abrindo {len(chunks)} abas paralelas...")

        async def process_chunk(chunk, tab_id):
            """Processa um grupo de tickets em uma aba independente."""
            tab_results = {}
            if not chunk:
                return tab_results

            pg = await ctx.new_page()
            pg.set_default_timeout(60000)

            ok = await goto_dashboard(pg, state)
            if not ok:
                log.error(f"[{state}][T{tab_id}] Dashboard inacessível — abortando aba")
                try:
                    await pg.close()
                except Exception:
                    pass
                return tab_results

            log.info(f"[{state}][T{tab_id}] Dashboard OK — {len(chunk)} tickets")

            for idx, tnum in enumerate(chunk):
                log.info(f"[{state}][T{tab_id}] ({idx+1}/{len(chunk)}) Ticket {tnum}")
                result = {"location_text": "", "responses": [], "expire_date": ""}
                MAX_RETRIES = 2

                for _attempt in range(MAX_RETRIES):
                    try:
                        await filter_ticket(pg, tnum)

                        if not await pg.get_by_text(tnum, exact=True).count():
                            log.warning(f"[{state}] {tnum}: não encontrado")
                            tab_results[tnum] = result
                            break

                        await pg.get_by_text(tnum, exact=True).first.click()
                        await wait_stable(pg)

                        # Text tab — SEMPRE lê body pra extrair expire_date;
                        # location_text só se precisa de notes
                        tt = pg.locator('[role="tab"]:has-text("Text")').first
                        if await tt.count():
                            await click_and_wait(pg, tt, "tab")
                        body = await pg.locator("body").inner_text()
                        result["expire_date"] = extract_expire_date(body)
                        if tnum in needs_notes or not tickets_data:
                            result["location_text"] = extract_location_text(body, state=state)

                        # Responses tab
                        rt = pg.locator('[role="tab"]:has-text("Responses")').first
                        if await rt.count():
                            await rt.click()
                            try:
                                await pg.wait_for_selector(
                                    'text="Loading ticket responses..."',
                                    state="hidden", timeout=6000
                                )
                            except Exception:
                                pass
                            await wait_tab_content(pg)

                            # Vários tickets abrem modais Angular ("Marking delay",
                            # "Excavation Date" etc.) que INTERCEPTAM o clique no filtro
                            # "All (N)" e estouram o timeout de 60s — foi o que congelou os
                            # tickets de St. Pete desde 08/06. Alguns aparecem COM ATRASO,
                            # então dispensa em loop (removendo qualquer overlay restante) e
                            # clica "All" com retry curto.
                            filter_clicked = False
                            all_links = pg.locator('text=/^All \\(/')
                            if await all_links.count():
                                for _ftry in range(4):
                                    for _dlg in ("Marking delay", "Excavation Date", "Date of Excavation"):
                                        if await pg.locator(f'h1:has-text("{_dlg}")').count():
                                            await _dismiss_dialog(pg, _dlg, label=tnum)
                                    if await pg.locator('.cdk-overlay-backdrop, mat-dialog-container').count():
                                        try:
                                            await pg.evaluate("document.querySelectorAll('.cdk-overlay-container mat-dialog-container, .cdk-overlay-pane, .cdk-overlay-backdrop').forEach(el => el.remove())")
                                        except Exception:
                                            pass
                                        await pg.wait_for_timeout(250)
                                    try:
                                        await all_links.first.click(timeout=12000)
                                        await wait_stable(pg)
                                        filter_clicked = True
                                        break
                                    except Exception:
                                        await pg.wait_for_timeout(400)

                            if not filter_clicked:
                                cur_links = pg.locator('text=/^Current Only/')
                                if await cur_links.count():
                                    await click_and_wait(pg, cur_links.first, "filter")
                                    filter_clicked = True

                            if not filter_clicked:
                                for sel in ['button:has-text("All")', 'a:has-text("All")']:
                                    if await pg.locator(sel).count():
                                        await click_and_wait(pg, pg.locator(sel).first, "filter")
                                        break

                            await wait_stable(pg)
                            body_text = await pg.locator("body").inner_text()

                            # Salva texto bruto pra debug.
                            # - DEBUG_MODE=1: salva os 2 primeiros de cada aba (amostragem)
                            # - ONEDRILL_DEBUG_TICKET=XXX: força salvar só esse ticket específico
                            should_debug = False
                            if DEBUG_TICKET and str(tnum).strip() == DEBUG_TICKET:
                                should_debug = True
                            elif DEBUG_MODE and idx < 2 and tab_id == 0:
                                should_debug = True
                            if should_debug:
                                with open(f"debug_responses_{state}_{tnum}.txt", "w", encoding="utf-8") as df:
                                    df.write(body_text)
                                log.info(f"[{state}] {tnum}: DEBUG salvo em debug_responses_{state}_{tnum}.txt")

                            # Parse responses (resilient to whitespace/order changes)
                            resp_idx = -1
                            resp_patterns = [
                                r"Status\s+Date\s+Service\s+Area\s+Response",
                                r"Service\s+Area",
                                r"Utility\s+Name",
                            ]
                            for rp in resp_patterns:
                                rm = re.search(rp, body_text, re.IGNORECASE)
                                if rm:
                                    resp_idx = rm.start()
                                    break
                            if resp_idx >= 0:
                                resp_section = body_text[resp_idx:resp_idx + 50000]
                                linhas = resp_section.split("\n")
                                i = 0
                                while i < len(linhas):
                                    linha = linhas[i].strip()
                                    if linha in ["Status", "Date", "Service Area", "Response", "Entry Method", "Comments", ""]:
                                        i += 1
                                        continue

                                    if linha == "Event":
                                        i += 1
                                        while (i < len(linhas)
                                               and linhas[i].strip() not in ["Current", "No Response", "Positive Response", "Event"]
                                               and not linhas[i].strip().startswith("No Response")):
                                            i += 1
                                        continue

                                    if linha.startswith("No Response"):
                                        status_raw = "No Response"
                                        i += 1
                                        utility = linhas[i].strip() if i < len(linhas) else ""
                                        i += 1
                                        if i < len(linhas) and re.match(r"^[A-Z0-9]{2,10}$", linhas[i].strip()):
                                            i += 1
                                        if i < len(linhas) and linhas[i].strip().startswith("ID"):
                                            i += 1
                                        if _is_valid_utility_name(utility):
                                            _cls_status, _cls_unrec = classify(status_raw, "")
                                            result["responses"].append({
                                                "utility": utility, "status_raw": status_raw,
                                                "status": _cls_status,
                                                "response": "", "comment": "",
                                                "_unrecognized": _cls_unrec
                                            })
                                        continue

                                    if linha in ["Current", "Positive Response"]:
                                        status_raw = linha
                                        i += 1
                                        data_line = linhas[i].strip() if i < len(linhas) else ""
                                        i += 1
                                        utility = linhas[i].strip() if i < len(linhas) else ""
                                        i += 1
                                        if i < len(linhas) and re.match(r"^[A-Z0-9]{2,10}$", linhas[i].strip()):
                                            i += 1
                                        if i < len(linhas) and linhas[i].strip().startswith("ID"):
                                            i += 1
                                        response = linhas[i].strip() if i < len(linhas) else ""
                                        i += 1
                                        comment = ""
                                        while i < len(linhas):
                                            l = linhas[i].strip()
                                            if l in ["Web Service", ""] or l.startswith("Entered via"):
                                                i += 1
                                                continue
                                            if l in ["Current", "Positive Response", "Event", "No Response"] or l.startswith("No Response"):
                                                break
                                            comment = l
                                            i += 1
                                            break
                                        responded_date = None
                                        if data_line:
                                            dl = data_line.strip()[:24]
                                            for fmt in [
                                                "%m/%d/%Y %I:%M %p",   # 03/24/2026 08:50 AM
                                                "%m/%d/%Y %I:%M:%S %p",# 03/24/2026 08:50:00 AM
                                                "%m/%d/%Y %H:%M",      # 03/24/2026 08:50
                                                "%m/%d/%Y %H:%M:%S",   # 03/24/2026 08:50:00
                                                "%m/%d/%Y",            # 03/24/2026
                                                "%m/%d/%y %I:%M %p",   # 03/24/26 08:50 AM
                                                "%m/%d/%y %H:%M",      # 03/24/26 08:50
                                                "%m/%d/%y",            # 03/24/26
                                            ]:
                                                try:
                                                    responded_date = datetime.strptime(dl[:len(fmt)+2].strip(), fmt).replace(tzinfo=None).isoformat()
                                                    break
                                                except ValueError:
                                                    continue
                                        if _is_valid_utility_name(utility):
                                            _cls_status, _cls_unrec = classify(status_raw, response + " " + comment)
                                            result["responses"].append({
                                                "utility": utility, "status_raw": status_raw,
                                                "status": _cls_status,
                                                "response": response, "comment": comment,
                                                "responded_date": responded_date,
                                                "_unrecognized": _cls_unrec
                                            })
                                        continue
                                    i += 1

                            log.info(f"[{state}] {tnum}: {len(result['responses'])} utilities")

                        await fast_back(pg, state)
                        break  # Sucesso — sai do retry loop

                    except Exception as e:
                        if _attempt < MAX_RETRIES - 1 and "Timeout" in str(e):
                            log.warning(f"[{state}] {tnum}: Timeout (tentativa {_attempt+1}/{MAX_RETRIES}) — retry...")
                            try:
                                await back_to_dashboard(pg, state)
                            except Exception:
                                pass
                            continue
                        log.error(f"[{state}] {tnum}: ERRO -> {e}")
                        try:
                            await back_to_dashboard(pg, state)
                        except Exception:
                            pass

                tab_results[tnum] = result

            try:
                await pg.close()
            except Exception:
                pass
            log.info(f"[{state}][T{tab_id}] Concluído — {len(chunk)} tickets processados")
            return tab_results

        # Roda todas as abas em paralelo
        try:
            chunk_results = await asyncio.gather(*[process_chunk(c, i) for i, c in enumerate(chunks)])
            for cr in chunk_results:
                results.update(cr)
        finally:
            try:
                await ctx.close()
            except Exception:
                pass
    return results


# ── │ SECTION: SYNC_SUMMARY │ SYNC SUMMARY (estruturado) ─────────────────────
class SyncSummary:
    """Resumo estruturado de uma execução de sync."""

    def __init__(self):
        self.cleared = 0
        self.reverted = 0
        self.canceled = 0
        self.private_locator = 0
        self.watch_protect = 0
        self.pending = 0
        self.confirmed_clear = 0
        self.backfilled = 0
        self.locked_skipped = 0
        self.responses_saved = 0
        self.unrecognized = 0
        self.unrecognized_list = []  # [{ticket, utility, raw_text, state}]
        self.expire_updated = 0

    def __str__(self):
        parts = []
        if self.cleared:
            parts.append(f"{self.cleared} auto-clear")
        if self.reverted:
            parts.append(f"{self.reverted} revertidos")
        if self.canceled:
            parts.append(f"{self.canceled} auto-cancel (DHL)")
        if self.private_locator:
            parts.append(f"{self.private_locator} private-locator")
        if self.watch_protect:
            parts.append(f"👁 {self.watch_protect} watch-protect")
        if self.pending:
            parts.append(f"{self.pending} pendentes")
        if self.confirmed_clear:
            parts.append(f"{self.confirmed_clear} clear confirmados")
        if self.backfilled:
            parts.append(f"{self.backfilled} backfill")
        if self.locked_skipped:
            parts.append(f"{self.locked_skipped} locked")
        if self.unrecognized:
            parts.append(f"⚠ {self.unrecognized} não-reconhecidas")
        if self.expire_updated:
            parts.append(f"📅 {self.expire_updated} datas atualizadas")
        parts.append(f"{self.responses_saved} respostas salvas")
        return " | ".join(parts)


# ── │ SECTION: BATCH_PATCH │ BATCH PATCH HELPER ───────────────────────────────

def _get_latest_response_date(responses, ticket_num=""):
    """Extrai a data da ÚLTIMA utility que respondeu.

    Itera as respostas e encontra o MAX de responded_date.
    Retorna (datetime, is_fallback):
      - is_fallback=False  → data real extraída das respostas
      - is_fallback=True   → nenhuma data válida, usado datetime.now()
    """
    DATE_FMTS = [
        "%Y-%m-%dT%H:%M:%S",      # ISO
        "%Y-%m-%dT%H:%M",          # ISO short
        "%m/%d/%Y %I:%M:%S %p",    # 03/24/2026 08:50:00 AM
        "%m/%d/%Y %I:%M %p",       # 03/24/2026 08:50 AM
        "%m/%d/%Y %H:%M:%S",       # 03/24/2026 08:50:00
        "%m/%d/%Y %H:%M",          # 03/24/2026 08:50
        "%m/%d/%Y",                # 03/24/2026
        "%m/%d/%y",                # 03/24/26
    ]
    latest = None
    unparsed = []
    for resp in responses:
        rd = resp.get("responded_date")
        if not rd:
            continue
        rd_str = str(rd).strip().replace("Z", "")
        # Fix bug #20 Python: pré-normaliza datas com mês/dia sem leading zero (ex: "3/5/26" → "03/05/26").
        # Python strptime exige 2 dígitos no %m e %d — sem isso, falha silenciosamente e cai no
        # fromisoformat, que tbm falha. Tickets com data não parseada usavam datetime.now() como
        # fallback, resultando em responded_at impreciso e "gap" incorreto em análises temporais.
        rd_norm = re.sub(r'\b(\d)/(\d)/', r'0\1/0\2/', rd_str)       # 3/5/26 → 03/05/26
        rd_norm = re.sub(r'\b(\d)/(\d\d)/', r'0\1/\2/', rd_norm)     # 3/05/26 → 03/05/26
        rd_norm = re.sub(r'\b(\d\d)/(\d)/', r'\1/0\2/', rd_norm)     # 03/5/26 → 03/05/26
        parsed = None
        for fmt in DATE_FMTS:
            try:
                parsed = datetime.strptime(rd_norm[:len(fmt)+4].strip(), fmt)
                break
            except ValueError:
                continue
        if not parsed:
            try:
                parsed = datetime.fromisoformat(rd_str)
            except Exception:
                unparsed.append(rd_str)
                continue
        if parsed and (latest is None or parsed > latest):
            latest = parsed
    if latest:
        return latest, False
    if unparsed:
        log.warning(f"[DateParse] {ticket_num}: datas não parseadas: {unparsed[:5]} — usando datetime.now() como fallback")
    elif responses:
        log.warning(f"[DateParse] {ticket_num}: nenhuma resposta tem responded_date — usando datetime.now() como fallback")
    return datetime.now(), True


def sb_batch_patch(table, patches, id_field="id"):
    """Aplica múltiplos patches em batch (fix bug #5).

    Antes: loop chamava sb_patch() individualmente (N HTTP requests).
    Agora: 1 request POST com Prefer: resolution=merge-duplicates + on_conflict=<id_field>.
    Upsert pelo PK é seguro porque o PK está sempre preenchido nos patches → Supabase
    encontra a row existente e faz MERGE, não tenta INSERT vazio.

    Em caso de falha do batch (ex: RLS, NOT NULL surpresa, rede), faz fallback
    pro loop item-por-item antigo pra garantir que ao menos alguns patches passem.
    """
    if not patches:
        return

    # Valida que todos têm o id_field — sem isso, nem pode tentar batch
    valid = [p for p in patches if p.get(id_field) is not None]
    invalid_count = len(patches) - len(valid)
    if invalid_count:
        log.warning(f"[BatchPatch] {invalid_count}/{len(patches)} itens sem {id_field}, pulando")
    if not valid:
        return

    # ── Tentativa 1: BATCH REAL (1 request pra tudo) ──
    try:
        # Divide em lotes grandes pra evitar payload gigante (limite Supabase ~10MB/req)
        for i in range(0, len(valid), BATCH_SIZE):
            chunk = valid[i:i + BATCH_SIZE]
            sb_upsert(table, chunk, on_conflict=id_field)
            if i + BATCH_SIZE < len(valid):
                time.sleep(0.3)
        log.info(f"[BatchPatch] ✅ {len(valid)} patches aplicados em batch em {table}")
        return
    except Exception as e:
        log.warning(f"[BatchPatch] Batch falhou ({e}), tentando fallback 1-por-1...")

    # ── Fallback: loop antigo (seguro mas lento) ──
    success = 0
    for item in valid:
        item_copy = dict(item)
        item_id = item_copy.pop(id_field, None)
        try:
            sb_patch(table, item_id, item_copy)
            success += 1
        except Exception as e:
            log.error(f"[BatchPatch] Erro fallback em {id_field}={item_id}: {e}")
    log.info(f"[BatchPatch] {success}/{len(valid)} patches aplicados em {table} (fallback)")


# ── │ SECTION: UNRECOGNIZED │ RESPOSTAS NÃO RECONHECIDAS (salva no Supabase + alerta por email) 
def save_unrecognized_responses(items):
    """Salva respostas não reconhecidas na tabela unrecognized_responses.

    Tabela deve ser criada no Supabase (ver create_unrecognized_table.sql).
    Se a tabela não existir, apenas loga o aviso.
    """
    if not items:
        return
    now_iso = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    records = []
    for item in items:
        records.append({
            "ticket_num": item["ticket_num"],
            "state": item.get("state", ""),
            "utility_name": item.get("utility_name", ""),
            "status_raw": item.get("status_raw", ""),
            "raw_text": (item.get("raw_text") or "")[:500],
            "detected_at": now_iso,
            "resolved": False,
        })
    try:
        h = {**SB_H, "Prefer": "return=minimal"}
        _sb_request(requests.post, f"{SB_URL}/rest/v1/unrecognized_responses",
                    headers=h, json=records, timeout=15)
        log.info(f"[Unrecognized] {len(records)} respostas não reconhecidas salvas no Supabase")
    except Exception as e:
        # Se a tabela não existir, loga mas não quebra o sync
        log.warning(f"[Unrecognized] Erro ao salvar (tabela existe?): {e}")
        log.warning(f"[Unrecognized] Rode o script SQL para criar a tabela: create_unrecognized_table.sql")


def send_unrecognized_alert(state, items):
    """Envia email com resumo de respostas não reconhecidas."""
    import smtplib
    from email.mime.text import MIMEText

    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_PASS")
    alert_to = os.getenv("ALERT_EMAIL")

    if not all([gmail_user, gmail_pass, alert_to]):
        log.info("[Unrecognized] Email não configurado — alerta pulado")
        return

    subject = f"[OneDrill] ⚠ {len(items)} resposta(s) 811 não reconhecida(s) — {state}"
    body = f"OneDrill 811 Sync detectou {len(items)} resposta(s) que não batem com nenhum padrão conhecido.\n"
    body += f"Estado: {state}\n"
    body += f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
    body += "\n" + "=" * 60 + "\n\n"

    for i, item in enumerate(items, 1):
        body += f"{i}. Ticket: {item['ticket_num']}\n"
        body += f"   Utility: {item.get('utility_name', '—')}\n"
        body += f"   Status raw: {item.get('status_raw', '—')}\n"
        body += f"   Texto: {item.get('raw_text', '—')}\n\n"

    body += "=" * 60 + "\n"
    body += "\nAção necessária: verifique esses tickets manualmente no portal 811.\n"
    body += "Se o padrão se repetir, adicione-o na função classify() do 811_sync.py.\n"

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = gmail_user
    msg["To"] = alert_to

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(gmail_user, gmail_pass)
            s.send_message(msg)
        log.info(f"[Unrecognized] Alerta enviado para {alert_to}")
    except Exception as e:
        log.warning(f"[Unrecognized] Erro ao enviar email: {e}")


# Marcador no history pra não reenviar o alerta do mesmo ticket travado+renovado
LOCKED_RENEWED_MARKER = "[ALERTA] travado+renovado"


def send_locked_renewed_alert(state, items):
    """Email avisando tickets TRAVADOS (status_locked) que foram renovados —
    sinal de que o operador esqueceu de destravar antes de renovar.
    Retorna True se enviou (pra só então marcar no history e não reenviar)."""
    import smtplib
    from email.mime.text import MIMEText

    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_PASS")
    alert_to = os.getenv("ALERT_EMAIL")

    if not all([gmail_user, gmail_pass, alert_to]):
        log.info("[LockedRenewed] Email não configurado — alerta pulado")
        return False

    subject = f"[OneDrill] {len(items)} ticket(s) travado(s) e renovado(s) — {state}"
    body = "Estes tickets estão com status TRAVADO (clear manual 🔒) E foram renovados.\n"
    body += "O normal é DESTRAVAR antes de renovar — confira se esqueceu de algum.\n"
    body += f"Estado: {state}\n"
    body += f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
    body += "\n" + "=" * 60 + "\n\n"

    for i, t in enumerate(items, 1):
        body += f"{i}. Ticket: {t.get('ticket')}\n"
        body += f"   Status: {t.get('status')} (travado)\n"
        body += f"   Renovou: {t.get('old_ticket2')}\n"
        body += f"   Carência (expire_old): {t.get('expire_old') or '—'}\n\n"

    body += "=" * 60 + "\n"
    body += "\nAção: abra o ticket no OneDrill e destrave (🔓) se a renovação exige novas liberações.\n"

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = gmail_user
    msg["To"] = alert_to

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(gmail_user, gmail_pass)
            s.send_message(msg)
        log.info(f"[LockedRenewed] Alerta enviado para {alert_to} ({len(items)} ticket(s))")
        return True
    except Exception as e:
        log.warning(f"[LockedRenewed] Erro ao enviar email: {e}")
        return False


def check_locked_renewed_alert(state):
    """Detecta tickets TRAVADOS (status_locked) renovados nesse estado e, pros que ainda
    não foram alertados, manda 1 email e marca no history (dedup — 1 alerta por ticket)."""
    try:
        rows = sb_get(
            "tickets",
            f"&state=eq.{state}&status_locked=eq.true"
            "&select=id,ticket,status,old_ticket2,expire_old,history",
        )
    except Exception as e:
        log.warning(f"[LockedRenewed] Erro ao buscar travados ({state}): {e}")
        return

    pend = []
    for t in rows:
        if (t.get("status") or "") in ("Closed", "Cancel"):
            continue  # terminal — trabalho concluído, não precisa de ação
        if not (t.get("old_ticket2") or "").strip():
            continue  # não é renovado
        hist = t.get("history") or []
        if any(LOCKED_RENEWED_MARKER in (h.get("action") or "") for h in hist):
            continue  # já alertado antes
        pend.append(t)

    if not pend:
        return

    if not send_locked_renewed_alert(state, pend):
        return  # email desligado/falhou — não marca, tenta de novo no próximo sync

    ts = int(datetime.now().timestamp() * 1000)
    label = datetime.now().strftime("%m/%d/%Y")
    for t in pend:
        hist = t.get("history") or []
        hist.append({"ts": ts, "action": f"{LOCKED_RENEWED_MARKER} — alertado por email em {label}", "color": "#b45309"})
        try:
            sb_patch("tickets", t["id"], {"history": hist})
        except Exception as e:
            log.warning(f"[LockedRenewed] Erro ao marcar {t.get('ticket')}: {e}")


# ── │ SECTION: SESSION_ALERT │ Alerta de sessão expirada (FL/IN) ──────────────
# Dedup: 1 email por estado a cada 24h, via flag local. Reseta quando login OK.

_SESSION_FLAG_TTL = 24 * 60 * 60  # segundos


def _session_flag_path(state):
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, f".session_expired_{state.upper()}.flag")


def _clear_session_flag(state):
    """Remove flag após login bem-sucedido. Idempotente."""
    try:
        p = _session_flag_path(state)
        if os.path.exists(p):
            os.remove(p)
            log.info(f"[SessionExpired] {state}: flag limpa (sessão renovada)")
    except Exception as e:
        log.debug(f"[SessionExpired] {state}: erro ao limpar flag: {e}")


def _touch_session_flag(state):
    try:
        with open(_session_flag_path(state), "w") as f:
            f.write(datetime.now().isoformat())
    except Exception as e:
        log.warning(f"[SessionExpired] {state}: erro ao criar flag: {e}")


def send_session_expired_alert(state):
    """Email avisando que a sessão do portal {state} expirou — precisa login manual.
    Dedup: não reenvia nas próximas 24h por estado. Retorna True se mandou."""
    import smtplib
    from email.mime.text import MIMEText

    flag = _session_flag_path(state)
    if os.path.exists(flag):
        try:
            age = datetime.now().timestamp() - os.path.getmtime(flag)
            if age < _SESSION_FLAG_TTL:
                log.info(f"[SessionExpired] {state}: já alertado há {int(age/3600)}h — skip (dedup)")
                return False
        except Exception:
            pass

    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_PASS")
    alert_to = os.getenv("ALERT_EMAIL")
    if not all([gmail_user, gmail_pass, alert_to]):
        log.info(f"[SessionExpired] {state}: email não configurado — pulado")
        return False

    portal_url = (PORTALS.get(state) or {}).get("url", "")
    subject = f"[OneDrill] Sessão {state} expirada — precisa de login manual"
    body = (
        f"A sessão do portal 811 ({state}) expirou. O sync não consegue continuar sem que você logue manualmente.\n\n"
        f"O que fazer:\n"
        f"  1) Rode keepalive.bat (abre janela já preparada pra login no portal {state}).\n"
        f"     OU abra o portal direto: {portal_url}\n"
        f"  2) Faça login com a conta OneDrill.\n"
        f"  3) Rode rodar_{state}.bat normalmente — o sync pega a sessão renovada automaticamente.\n\n"
        f"Estado: {state}\n"
        f"Data:   {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
        f"\nEste alerta não se repete nas próximas 24h pra esse estado.\n"
    )

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = gmail_user
    msg["To"] = alert_to

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(gmail_user, gmail_pass)
            s.send_message(msg)
        log.info(f"[SessionExpired] {state}: alerta enviado para {alert_to}")
        _touch_session_flag(state)
        return True
    except Exception as e:
        log.warning(f"[SessionExpired] {state}: erro ao enviar email: {e}")
        return False


# ── Alerta de tickets travando no scrape (respostas congeladas em silencio) ──
_SCRAPE_FAIL_PATH = os.path.join(BASE_DIR, "_scrape_fail_streak.json")
_SCRAPE_FAIL_THRESHOLD = 3  # runs consecutivos sem ler NENHUMA resposta = alerta


def send_scrape_failure_alert(state, chronic):
    """Email listando tickets que falham o scrape ha varios runs seguidos."""
    import smtplib
    from email.mime.text import MIMEText
    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_PASS")
    alert_to = os.getenv("ALERT_EMAIL")
    if not all([gmail_user, gmail_pass, alert_to]):
        log.info(f"[ScrapeFail] {state}: email nao configurado — pulado")
        return False
    lines = "\n".join(f"  {tn} — {n} runs seguidos sem resposta" for tn, n in chronic)
    subject = f"[OneDrill] {state}: {len(chronic)} ticket(s) travando no scrape 811"
    body = (
        f"Estes tickets estao ativos (Open/Damage) mas o scrape do portal 811 ({state}) NAO le "
        f"nenhuma resposta ha {_SCRAPE_FAIL_THRESHOLD}+ runs seguidos — as respostas ficam "
        f"CONGELADAS ate resolver:\n\n"
        f"{lines}\n\n"
        f"O que checar: abrir o ticket no portal 811 na mao; ver se ha modal/estado travando o "
        f"scrape (ex.: 'Excavation Date' / 'Marking delay'); conferir 811_sync.log.\n\n"
        f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
    )
    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = gmail_user
    msg["To"] = alert_to
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(gmail_user, gmail_pass)
            s.send_message(msg)
        log.info(f"[ScrapeFail] {state}: alerta enviado ({len(chronic)} tickets) -> {alert_to}")
        return True
    except Exception as e:
        log.warning(f"[ScrapeFail] {state}: erro ao enviar email: {e}")
        return False


def check_scrape_failures(state, tickets_to_scrape, results):
    """Ticket que FOI scrapeado mas voltou SEM nenhuma resposta = falha do scrape.
    Conta runs consecutivos por ticket num JSON e alerta (1x por mudanca/dia) quando
    chega em _SCRAPE_FAIL_THRESHOLD. Chamador envolve em try/except — nunca quebra o sync."""
    scraped = {str(t.get("ticket")).strip() for t in (tickets_to_scrape or []) if t.get("ticket")}
    results = results or {}
    failed = {tn for tn in scraped if tn in results and not (results.get(tn) or {}).get("responses")}
    try:
        with open(_SCRAPE_FAIL_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = {}
    st = data.get(state, {})
    for tn in scraped:
        if tn in failed:
            st[tn] = int(st.get(tn, 0)) + 1
        else:
            st.pop(tn, None)  # leu ok agora -> zera o streak
    data[state] = st
    chronic = sorted([(tn, st[tn]) for tn in st if st[tn] >= _SCRAPE_FAIL_THRESHOLD])
    cur_set = [tn for tn, _ in chronic]
    alerted = data.get("_alerted", {})
    prev = alerted.get(state, {})
    today = datetime.now().strftime("%Y-%m-%d")
    if chronic and (cur_set != prev.get("set") or prev.get("date") != today):
        send_scrape_failure_alert(state, chronic)
        alerted[state] = {"set": cur_set, "date": today}
    elif not cur_set:
        alerted.pop(state, None)
    data["_alerted"] = alerted
    try:
        with open(_SCRAPE_FAIL_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception as e:
        log.warning(f"[ScrapeFail] erro salvando streak: {e}")
    if chronic:
        log.info(f"[{state}] ScrapeFail cronicos (>= {_SCRAPE_FAIL_THRESHOLD} runs): {cur_set}")


# ── │ SECTION: SAVE │ SAVE TO SUPABASE ────────────────────────────────────────
def save_to_supabase(state, results, tickets, grace_old_map=None):
    grace_old_map = grace_old_map or {}
    ticket_map = {t["ticket"]: t for t in tickets}
    summary = SyncSummary()
    all_records = []
    ticket_patches = []  # Acumula patches pra batch no final
    unrecognized_list = []  # Respostas que caíram no fallback do classify()

    for tnum, data in results.items():
        t = ticket_map.get(tnum)

        # Ticket antigo em carência: salva respostas mas pula status update
        if not t and tnum in grace_old_map:
            tid = grace_old_map[tnum]
            latest_by_utility = {}
            for resp in data["responses"]:
                key = resp["utility"]
                if key in latest_by_utility:
                    existing = latest_by_utility[key]
                    ex_is_nr = (existing.get("status_raw") or "").lower().startswith("no response")
                    new_is_nr = (resp.get("status_raw") or "").lower().startswith("no response")
                    if ex_is_nr and not new_is_nr:
                        pass
                    elif not ex_is_nr and new_is_nr:
                        continue
                    else:
                        existing_date = existing.get("responded_date") or ""
                        new_date = resp.get("responded_date") or ""
                        if existing_date and new_date and new_date < existing_date:
                            continue
                latest_by_utility[key] = resp
            deduped = list(latest_by_utility.values())
            now_iso = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
            for resp in deduped:
                parsed_date = resp.get("responded_date")
                all_records.append({
                    "ticket_id": tid, "ticket_num": tnum, "state": state,
                    "utility_name": resp["utility"], "status": resp["status"],
                    "response_text": resp.get("response") or resp.get("comment") or resp.get("status_raw", ""),
                    "synced_at": now_iso,
                    "responded_at": parsed_date if parsed_date else now_iso,
                })
            statuses = [r["status"] for r in deduped]
            pending_names = [r["utility"] for r in deduped if r["status"] == "Pending"]
            log.info(f"[{state}] {tnum} (antigo/carência): {len(deduped)} respostas"
                     + (f" — pendentes: {pending_names}" if pending_names else " — TUDO CLEAR"))
            continue

        if not t:
            continue
        tid = t["id"]
        patch = {"id": tid}  # Acumula campos pra este ticket
        needs_patch = False

        if data["location_text"] and not (t.get("notes") or "").strip():
            patch["notes"] = f"[811 Location] {data['location_text']}"
            needs_patch = True

        # ── ATUALIZAR DATA DE VENCIMENTO (renovações) ──
        scraped_expire = normalize_expire(data.get("expire_date") or "")
        current_expire = (t.get("expire") or "").strip()
        current_normalized = normalize_expire(current_expire)
        current_is_polluted = _is_polluted_expire(current_expire)

        if scraped_expire and scraped_expire != current_normalized:
            # Novo valor válido e diferente → atualiza
            patch["expire"] = scraped_expire
            needs_patch = True
            log.info(f"[{state}] {tnum}: 📅 EXPIRE ATUALIZADO: {current_expire} → {scraped_expire}")
            summary.expire_updated += 1
        elif current_is_polluted and current_normalized and not scraped_expire:
            # Scraper não extraiu, mas podemos limpar o formato poluído do banco
            patch["expire"] = current_normalized
            needs_patch = True
            log.warning(f"[{state}] {tnum}: 🧹 EXPIRE NORMALIZADO: {current_expire} → {current_normalized}")
            summary.expire_updated += 1
        elif current_is_polluted and not current_normalized:
            log.warning(f"[{state}] {tnum}: ⚠ expire poluído '{current_expire}' não normalizável — scraper tb falhou")

        # Dedup responses por utility.
        # Regras:
        #   1. "No Response" / "RE-MARK NOT NEEDED" SEMPRE perde pra resposta real
        #   2. Resposta real NUNCA é sobrescrita por "No Response" ou "RE-MARK NOT NEEDED"
        #   3. "RE-MARK NOT NEEDED" ganha de "No Response" (é ack, melhor que silêncio)
        #   4. Entre duas respostas reais: usa responded_date (mais recente ganha)
        #   5. Se datas não comparáveis: mantém a última (ordem do portal = cronológica)
        #
        # "RE-MARK NOT NEEDED" (Code 21) = extensão sem remarcação. Não é resposta
        # real — significa "minha resposta anterior continua valendo". O portal JULIE
        # retorna respostas de TODAS as revisões (00X…03X), então a revisão anterior
        # com a resposta real está no mesmo dataset. Basta não deixar RE-MARK sobrescrever.
        def _is_non_real(r):
            raw = (r.get("status_raw") or "").lower()
            resp_text = (r.get("response") or "").lower()
            combined = raw + " " + resp_text
            if raw.startswith("no response"):
                return 3  # lowest priority
            if "re-mark not needed" in combined or "remark not needed" in combined or raw == "21":
                return 2  # ack de extensão — melhor que NR, pior que real
            if "late final response" in combined or "late final" in combined or raw == "999":
                return 1  # aviso de atraso — NÃO é resposta real, MARKED/CLEAR real sempre ganha
            return 0      # real response

        latest_by_utility = {}
        for resp in data["responses"]:
            key = resp["utility"]
            if key in latest_by_utility:
                existing = latest_by_utility[key]
                ex_nr = _is_non_real(existing)
                new_nr = _is_non_real(resp)

                # Resposta real (0) SEMPRE ganha de não-real (1, 2)
                if ex_nr and not new_nr:
                    log.debug(f"  [Dedup] {tnum}/{key}: {existing.get('status_raw','')[:30]} → {resp['status']} ({resp.get('status_raw','')[:30]})")
                elif not ex_nr and new_nr:
                    log.debug(f"  [Dedup] {tnum}/{key}: mantém {existing['status']} ({existing.get('status_raw','')[:30]}), ignora {resp.get('status_raw','')[:30]}")
                    continue  # Mantém resposta real
                elif ex_nr and new_nr:
                    # Ambas não-reais: menor valor (maior prioridade) ganha
                    if new_nr > ex_nr:
                        log.debug(f"  [Dedup] {tnum}/{key}: mantém {existing.get('status_raw','')[:30]}, ignora {resp.get('status_raw','')[:30]}")
                        continue  # Ex é melhor (ex: REMARK > NR)
                else:
                    # Ambas reais. Dois casos:
                    #   (a) Datas diferentes → a utility atualizou status no tempo. Mais recente
                    #       ganha (Bug 2026-06-04: antes "pior ganha" descartava MARKED novo em
                    #       favor de Pending antigo. Ex: COMED LOCATE NOT COMPLETE 06/01 → MARKED
                    #       06/02 era descartado, ticket ficava preso em Pending falso.).
                    #   (b) Mesma data (ou sem data) → utility com múltiplos tipos (ELECTRIC + GAS)
                    #       respondeu na mesma revisão. Pior status ganha (segurança: falso-Clear
                    #       é perigoso; equipe pode cavar sem liberação).
                    existing_date = existing.get("responded_date") or ""
                    new_date = resp.get("responded_date") or ""
                    if existing_date and new_date and existing_date != new_date:
                        # (a) Datas diferentes — mais recente ganha
                        if new_date < existing_date:
                            log.debug(f"  [Dedup] {tnum}/{key}: mantém {existing['status']} (data {existing_date} > {new_date})")
                            continue
                        log.info(f"  [Dedup] {tnum}/{key}: {existing['status']}→{resp['status']} (atualização mais recente {existing_date}→{new_date})")
                    else:
                        # (b) Mesma data ou sem data — pior status ganha (multi-tipo)
                        STATUS_PRIORITY = {"Damage": 0, "Pending": 1, "Clear": 2}
                        ex_prio = STATUS_PRIORITY.get(existing.get("status", ""), 1)
                        new_prio = STATUS_PRIORITY.get(resp.get("status", ""), 1)
                        if ex_prio < new_prio:
                            log.info(f"  [Dedup] {tnum}/{key}: mantém {existing['status']} (pior status ganha — ignora {resp['status']})")
                            continue
                        if new_prio < ex_prio:
                            log.info(f"  [Dedup] {tnum}/{key}: {existing['status']}→{resp['status']} (pior status ganha — utility multi-tipo?)")
            latest_by_utility[key] = resp
        deduped_responses = list(latest_by_utility.values())

        # Aplica overrides locais (Frontier Terre Haute, etc)
        try:
            _apply_local_overrides(t, deduped_responses)
        except Exception as _ovr_e:
            log.warning(f"[Override] erro: {_ovr_e}")

        # ── RE-MARK NOT NEEDED: extensão de prazo — skip total ──
        # Quando TODAS as respostas são "RE-MARK NOT NEEDED" (Code 21), é só
        # extensão de prazo sem remarcação. Não são respostas reais.
        # → NÃO salvar no banco (mantém respostas anteriores intactas)
        # → NÃO rodar auto-clear/revert (status fica como estava)
        # O expire já foi atualizado acima (linha ~2445). Seguir com anterior.
        if deduped_responses:
            def _is_remark(r):
                raw = (r.get("status_raw") or "").lower()
                return ("re-mark not needed" in raw or "remark not needed" in raw
                        or raw == "21")
            _all_remark_skip = all(_is_remark(r) for r in deduped_responses)
            if _all_remark_skip:
                log.info(f"[{state}] {tnum}: 📋 Extensão (RE-MARK NOT NEEDED × {len(deduped_responses)}) "
                         f"— mantendo respostas e status anteriores (expire={patch.get('expire', t.get('expire', '?'))})")
                if needs_patch:
                    ticket_patches.append(patch)
                continue

        # Coleta records para bulk upsert
        now_iso = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
        for resp in deduped_responses:
            parsed_date = resp.get("responded_date")
            record = {
                "ticket_id": tid, "ticket_num": tnum, "state": state,
                "utility_name": resp["utility"], "status": resp["status"],
                "response_text": resp.get("response") or resp.get("comment") or resp.get("status_raw", ""),
                "synced_at": now_iso,
                # Sempre inclui responded_at (evita NOT NULL constraint).
                # Se tem data real parseada, usa ela. Senão, fallback pra synced_at.
                # fix_clear_dates diferencia comparando responded_at vs synced_at.
                # Fix bug #15: simplificado — era `parsed_date if parsed_date else (now_iso if X else now_iso)`,
                # com os dois lados do ternário interno iguais (código morto, provável sobra de refactor).
                "responded_at": parsed_date if parsed_date else now_iso,
            }
            all_records.append(record)

        if deduped_responses:
            # ── WI Relo-No-Show: merge com utilities herdadas do Standard ──
            _old_chain = (t.get("old_ticket2") or "").strip()
            _old_num = _old_chain.split(" → ")[0].strip() if _old_chain else ""
            relo_merged = _get_relo_merged_responses(tnum, _old_num, deduped_responses, state)
            eval_responses = relo_merged if relo_merged is not None else deduped_responses

            statuses = [r["status"] for r in eval_responses]
            none_pending = not any(s == "Pending" for s in statuses)
            all_responded = all(s in ("Clear", "Pending", "Cancel") for s in statuses)
            # WI: todas as utilities (excluindo "Not Participating") foram "Closed by DHL"
            # → ticket invalidado (Wis. Stat. §182.0175) → vira Cancel.
            # Not Participating não conta — utility disse "não tenho rede ali", é fora do escopo.
            def _is_not_part(r):
                txt = ((r.get("response", "") or "") + " "
                       + (r.get("comment", "") or "") + " "
                       + (r.get("status_raw", "") or "")).lower()
                return "not participating" in txt or "not service provider" in txt
            relevant_statuses = [r["status"] for r in eval_responses if not _is_not_part(r)]
            all_cancel = bool(relevant_statuses) and all(s == "Cancel" for s in relevant_statuses)
            ticket_locked = t.get("status_locked", False)

            for resp in deduped_responses:
                log.info(f"  [{state}] {tnum} | {resp['utility']}: {resp['status']} ({resp.get('response', '')[:60]})")
            if relo_merged is not None:
                inherited = [r for r in relo_merged if r.get("_inherited")]
                for r in inherited:
                    log.info(f"  [{state}] {tnum} | {r['utility']}: {r['status']} (herdado do Standard)")

            # ── SEGURANÇA: cruzar com banco para detectar Pending que o scrape perdeu ──
            # Se o scrape retornou tudo Clear, verifica se o banco tem Pending
            # que NÃO apareceu neste scrape (ex: portal truncou a lista de utilities).
            # Se encontrar, bloqueia o auto-clear — dados incompletos.
            if none_pending and not ticket_locked:
                scraped_utils = {r["utility"] for r in deduped_responses}
                try:
                    db_pending = sb_get("ticket_811_responses",
                                       f"&ticket_num=eq.{_qv(tnum)}&status=eq.Pending&select=utility_name")
                    missed_pending = {r["utility_name"] for r in db_pending} - scraped_utils
                    if missed_pending:
                        log.warning(f"[{state}] {tnum}: ⚠ SEGURANÇA — {len(missed_pending)} utility(s) Pending no banco ausente(s) no scrape: {list(missed_pending)}")
                        log.warning(f"[{state}] {tnum}: Bloqueando auto-clear (scrape incompleto: {len(deduped_responses)} capturadas, banco tem Pending que faltou)")
                        none_pending = False
                except Exception as e:
                    log.warning(f"[{state}] {tnum}: Erro ao verificar Pending no banco (conservador: bloqueando auto-clear): {e}")
                    none_pending = False

            # ── DETECTAR RESPOSTAS NÃO RECONHECIDAS ──
            for resp in deduped_responses:
                if resp.get("_unrecognized"):
                    raw = (resp.get("response") or resp.get("comment") or resp.get("status_raw", ""))[:200]
                    log.warning(f"[{state}] {tnum}: ⚠ RESPOSTA NÃO RECONHECIDA — {resp['utility']}: '{raw}'")
                    summary.unrecognized += 1
                    summary.unrecognized_list.append({
                        "ticket_num": tnum,
                        "state": state,
                        "utility_name": resp["utility"],
                        "status_raw": resp.get("status_raw", ""),
                        "raw_text": raw,
                    })

            # ── DETECTAR 3H: PRIVATE LOCATOR NECESSÁRIO ──
            private_utils = []
            for resp in deduped_responses:
                resp_text = (resp.get("response", "") + " " + resp.get("comment", "") + " " + resp.get("status_raw", "")).strip()
                if needs_private_locator(resp_text):
                    private_utils.append(resp["utility"])
            if private_utils:
                current_pending = (t.get("pending") or "").strip()
                pvt_tag = "🔒 PRIVATE LOCATOR: " + ", ".join(private_utils)
                if "PRIVATE LOCATOR" not in current_pending:
                    new_pending = (current_pending + "\n" + pvt_tag).strip() if current_pending else pvt_tag
                    patch["pending"] = new_pending
                    needs_patch = True
                    log.warning(f"[{state}] {tnum}: ⚠ PRIVATE LOCATOR necessário — {', '.join(private_utils)}")
                    summary.private_locator += 1

            # ── DETECTAR WATCH AND PROTECT (W&P — código 60, IL) ──
            wp_utils = []
            for resp in deduped_responses:
                resp_text = (resp.get("response", "") + " " + resp.get("comment", "") + " " + resp.get("status_raw", "")).strip()
                if needs_watch_and_protect(resp_text):
                    wp_utils.append(resp["utility"])
            if wp_utils:
                current_pending = (patch.get("pending") or t.get("pending") or "").strip()
                wp_tag = "⚠️ WATCH & PROTECT: " + ", ".join(wp_utils)
                if "WATCH & PROTECT" not in current_pending:
                    new_pending = (current_pending + "\n" + wp_tag).strip() if current_pending else wp_tag
                    patch["pending"] = new_pending
                    needs_patch = True
                    log.warning(f"[{state}] {tnum}: ⚠ WATCH & PROTECT — representante obrigatório: {', '.join(wp_utils)}")
                    summary.watch_protect += 1

            # ── STATUS LOCKED: NUNCA alterar ──
            if ticket_locked:
                log.info(f"[{state}] {tnum}: 🔒 STATUS TRAVADO (manual) — nenhuma alteração automática")
                summary.locked_skipped += 1
                if needs_patch:
                    ticket_patches.append(patch)
                continue

            # ── RENOVAÇÃO: Período de graça ──
            old_status = (t.get("status_old") or "").strip()
            in_grace, old_ticket_num = is_in_renewal_grace(t)

            if in_grace:
                # Primeiro: checa se o ticket NOVO tem Pending. Se sim, NÃO protege —
                # processa normalmente pra disparar a reversão Clear→Open.
                # Regra rigorosa: graça do antigo não pode mascarar pendências reais do novo.
                new_has_pending = any(r["status"] == "Pending" for r in deduped_responses)
                if new_has_pending:
                    log.info(f"[{state}] {tnum}: 🔄 RENOVAÇÃO em graça, MAS ticket novo tem pendências — processando normalmente (graça não protege falso-Clear)")
                    # Cai fora do branch de graça, segue pro auto-clear/revert abaixo
                else:
                    real_old_clear = False
                    if old_ticket_num:
                        try:
                            old_resps = sb_get("ticket_811_responses", f"&ticket_num=eq.{_qv(old_ticket_num)}&select=status,response_text")
                            if old_resps and len(old_resps) > 0:
                                # Cancel (Closed by DHL) = utility NUNCA respondeu → NÃO é Clear.
                                # Not Participating = utility sem rede na área → seguro, ignorar.
                                # TODAS as utilities relevantes devem estar em status liberador.
                                released = {"Clear", "Private", "Marked", "Unmarked"}
                                def _is_not_part(r):
                                    return "not participating" in ((r.get("response_text") or "") + " " + (r.get("status") or "")).lower()
                                relevant = [r for r in old_resps if not _is_not_part(r)]
                                all_relevant_clear = len(relevant) > 0 and all(r.get("status") in released for r in relevant)
                                if all_relevant_clear:
                                    real_old_clear = True
                                    if old_status != "Clear":
                                        log.info(f"[{state}] {tnum}: 🔄 RENOVAÇÃO — status_old={old_status or 'Open'} mas utilities REAIS do antigo ({old_ticket_num}) estão todas Clear")
                                else:
                                    not_clear = [r.get("status") for r in relevant if r.get("status") not in released]
                                    log.info(f"[{state}] {tnum}: 🔄 RENOVAÇÃO — antigo NÃO é Clear real ({len(not_clear)} utilities não-liberadas: {not_clear[:5]})")
                        except Exception as e:
                            log.debug(f"[{state}] {tnum}: Erro ao checar utilities do antigo: {e}")

                    if old_status == "Clear" or real_old_clear:
                        log.info(f"[{state}] {tnum}: 🔄 RENOVAÇÃO (graça até {t.get('expire_old', '')}) — antigo {'Clear (verificado)' if real_old_clear else 'Clear'}, mantém")
                        if needs_patch:
                            ticket_patches.append(patch)
                        continue
                    elif old_status:
                        log.info(f"[{state}] {tnum}: 🔄 RENOVAÇÃO (graça até {t.get('expire_old', '')}) — antigo era {old_status}, processando normalmente")
                    else:
                        log.warning(f"[{state}] {tnum}: 🔄 RENOVAÇÃO (graça até {t.get('expire_old', '')}) — status do antigo DESCONHECIDO, protegendo por precaução")
                        if needs_patch:
                            ticket_patches.append(patch)
                        continue

            # ── AUTO-RESOLVE: ticket inteiro com todas utilities em Cancel ──
            # WI: "Closed by DHL após 10 working days" — utilities não respondeu no prazo legal,
            # mas o sistema (Diggers) admin-fechou, liberando o ticket pra trabalhar com
            # cautela. Vira CLEAR com lembrete "verificar marcações em campo".
            # Outros estados (FL/IN/IL): all_cancel = ticket realmente cancelado → Cancel.
            if all_cancel and t.get("status") not in ("Cancel", "Clear"):
                now_ts = int(datetime.now().timestamp() * 1000)
                now_label = datetime.now().strftime('%m/%d/%Y')
                hist = t.get("history") or []
                if state == "WI":
                    note = f"[AUTO 811] Clear em {now_label} — todas utilities Closed by DHL (10 working days). ⚠ VERIFICAR MARCAÇÕES EM CAMPO antes de escavar."
                    hist.append({"ts": now_ts, "action": note, "color": "#f59e0b"})
                    new_notes = append_auto_note(patch.get("notes") or t.get("notes"), note)
                    patch.update({"status": "Clear", "notes": new_notes, "history": hist})
                    needs_patch = True
                    log.info(f"[{state}] {tnum}: AUTO-CLEAR (DHL closed) — todas utilities Closed by DHL, marcado Clear com lembrete de verificação em campo")
                    summary.cleared += 1
                else:
                    cancel_note = f"[AUTO 811] Cancelado em {now_label} — ticket invalidado (Closed by DHL após 10 working days)"
                    hist.append({"ts": now_ts, "action": cancel_note, "color": "#6d28d9"})
                    new_notes = append_auto_note(patch.get("notes") or t.get("notes"), cancel_note)
                    patch.update({"status": "Cancel", "notes": new_notes, "history": hist})
                    needs_patch = True
                    log.info(f"[{state}] {tnum}: AUTO-CANCEL — todas utilities Closed by DHL (ticket invalidado)")
                    summary.canceled += 1

            # ── WI: MIX CLEAR + CANCEL → TICKET CLEAR ──
            # Cenário WI: algumas utilities responderam (Marked/Clear) e outras
            # foram "Closed by DHL" (Cancel). Closed by DHL NÃO é pendência real
            # (prazo expirou, utility não vai responder mais). Ticket está liberado.
            # Só pra WI — outros estados não usam essa lógica.
            if state == "WI" and not all_cancel and none_pending and t.get("status") == "Open":
                wi_has_cancel = any(s == "Cancel" for s in relevant_statuses)
                wi_has_clear = any(s == "Clear" for s in relevant_statuses)
                wi_all_resolved = bool(relevant_statuses) and all(
                    s in ("Clear", "Cancel") for s in relevant_statuses
                )
                if wi_all_resolved and wi_has_clear and wi_has_cancel:
                    clear_dt, is_fallback = _get_latest_response_date(deduped_responses, ticket_num=tnum)
                    clear_ts = int(datetime.now().timestamp() * 1000)
                    clear_label = clear_dt.strftime('%m/%d/%Y')
                    hist = t.get("history") or []
                    n_cancel = sum(1 for s in relevant_statuses if s == "Cancel")
                    clear_note = f"[AUTO 811] Clear em {clear_label} — {n_cancel} utility(s) Closed by DHL (ignoradas)"
                    hist.append({"ts": clear_ts, "action": clear_note, "color": "#16a34a"})
                    new_notes = append_auto_note(patch.get("notes") or t.get("notes"), clear_note)
                    patch.update({"status": "Clear", "notes": new_notes, "history": hist})
                    needs_patch = True
                    log.info(f"[{state}] {tnum}: AUTO-CLEAR (WI mix) — {len(relevant_statuses)-n_cancel} Clear + {n_cancel} Closed by DHL")
                    summary.cleared += 1

            if none_pending and all_responded and t.get("status") == "Open":
                # AUTO-CLEAR (Fix 2026-05-14): clear_ts = now (quando ticket mudou pra Clear)
                # clear_label = data real da ultima resposta (info no historico)
                clear_dt, is_fallback = _get_latest_response_date(deduped_responses, ticket_num=tnum)
                clear_ts = int(datetime.now().timestamp() * 1000)
                clear_label = clear_dt.strftime('%m/%d/%Y')
                hist = t.get("history") or []
                clear_note = f"[AUTO 811] Clear em {clear_label}"
                hist.append({"ts": clear_ts, "action": clear_note, "color": "#16a34a"})
                new_notes = append_auto_note(patch.get("notes") or t.get("notes"), clear_note)
                patch.update({"status": "Clear", "notes": new_notes, "history": hist})
                needs_patch = True
                log.info(f"[{state}] {tnum}: AUTO-CLEAR (data{'⚠ FALLBACK sync-time' if is_fallback else ' real'}: {clear_label})")
                summary.cleared += 1

            elif none_pending and all_responded and t.get("status") == "Clear":
                # Backfill: se não tem evento de clear no histórico, adiciona
                hist = t.get("history") or []
                has_clear_evt = any(
                    "auto 811" in (h.get("action", "")).lower() and "revertido" not in (h.get("action", "")).lower()
                    for h in hist
                )
                has_clear_evt = has_clear_evt or any("→ clear" in (h.get("action", "")).lower() for h in hist)

                if not has_clear_evt:
                    # Tenta usar data real das respostas; fallback pra notas; fallback pra now
                    backfill_dt, is_fallback = _get_latest_response_date(deduped_responses, ticket_num=tnum)
                    # Se fallback (sem datas reais), tenta extrair das notas
                    if is_fallback:
                        notes_text = t.get("notes") or ""
                        date_match = re.search(r'\[AUTO 811\] Clear em (\d{1,2}/\d{1,2}/\d{4})', notes_text)
                        if date_match:
                            try:
                                backfill_dt = datetime.strptime(date_match.group(1), "%m/%d/%Y")
                            except Exception:
                                pass
                    backfill_ts = int(backfill_dt.timestamp() * 1000)
                    backfill_label = backfill_dt.strftime('%m/%d/%Y')

                    hist.append({"ts": backfill_ts, "action": f"[AUTO 811] Clear em {backfill_label}", "color": "#16a34a"})
                    patch["history"] = hist
                    needs_patch = True
                    log.info(f"[{state}] {tnum}: Clear confirmado + BACKFILL histórico ({backfill_label})")
                    summary.backfilled += 1
                else:
                    log.info(f"[{state}] {tnum}: Clear confirmado")
                    summary.confirmed_clear += 1

            elif not none_pending and t.get("status") == "Open":
                pending_utils = [r["utility"] for r in deduped_responses if r["status"] == "Pending"]
                log.info(f"[{state}] {tnum}: Pendente — aguardando: {pending_utils}")
                summary.pending += 1

            elif not none_pending and t.get("status") == "Clear":
                # REVERTENDO Clear→Open
                pending_utils = [r["utility"] for r in deduped_responses if r["status"] == "Pending"]
                now_ts = int(datetime.now().timestamp() * 1000)
                hist = t.get("history") or []
                revert_note = f"[AUTO 811] Revertido Clear→Open em {datetime.now().strftime('%m/%d/%Y')} — pendente: {', '.join(pending_utils)}"
                hist.append({"ts": now_ts, "action": f"[AUTO 811] Revertido Clear→Open — {', '.join(pending_utils)}", "color": "#dc2626"})
                new_notes = append_auto_note(patch.get("notes") or t.get("notes"), revert_note)
                log.warning(f"[{state}] {tnum}: REVERTENDO Clear→Open — pendente: {pending_utils}")
                patch.update({"status": "Open", "notes": new_notes, "history": hist})
                needs_patch = True
                summary.reverted += 1

        if needs_patch:
            ticket_patches.append(patch)

    # ── Batch ticket patches (substitui N+1 sb_patch individuais) ──
    if ticket_patches:
        log.info(f"[{state}] Aplicando {len(ticket_patches)} ticket patches em batch...")
        sb_batch_patch("tickets", ticket_patches, id_field="id")

    # Bulk upsert responses
    if all_records:
        try:
            for i in range(0, len(all_records), BATCH_SIZE):
                batch = all_records[i:i + BATCH_SIZE]
                sb_upsert("ticket_811_responses", batch)
                if i + BATCH_SIZE < len(all_records):
                    time.sleep(0.5)
            summary.responses_saved = len(all_records)
            log.info(f"[{state}] Bulk upsert: {summary.responses_saved} registros salvos")
        except Exception as e:
            log.error(f"[{state}] Erro no bulk upsert: {e} — tentando individual...")
            for rec in all_records:
                try:
                    sb_upsert("ticket_811_responses", rec)
                    summary.responses_saved += 1
                except Exception as e2:
                    log.error(f"Erro individual {rec.get('ticket_num')}/{rec.get('utility_name')}: {e2}")

    # ── Salvar respostas não reconhecidas no Supabase ──
    if summary.unrecognized_list:
        save_unrecognized_responses(summary.unrecognized_list)
        send_unrecognized_alert(state, summary.unrecognized_list)

    # ── Alerta: tickets travados que foram renovados (esqueceu de destravar antes de renovar) ──
    check_locked_renewed_alert(state)

    return summary


# ── │ SECTION: IMPORT │ IMPORTAR TICKETS NOVOS ────────────────────────────────

async def _scrape_bodies_parallel(state, ticket_numbers):
    """Scrape body (Text tab) de múltiplos tickets em paralelo.

    Abre NUM_TABS abas, cada uma processa um chunk. Retorna dict {tnum: body_text}.
    Usado pelo import_new_tickets pra coletar bodies sem bloquear 1 por 1.
    """
    results = {}
    if not ticket_numbers:
        return results

    perfil = _profile_path(state)
    n_tabs = min(NUM_TABS, len(ticket_numbers))
    chunks = [[] for _ in range(n_tabs)]
    for i, tnum in enumerate(ticket_numbers):
        chunks[i % n_tabs].append(tnum)
    chunks = [c for c in chunks if c]

    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)

        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)

        if "login" in page.url.lower():
            await ctx.close()
            await asyncio.sleep(1)
            ok = await auto_login_silent(state)
            if not ok:
                log.warning(f"[{state}] auto_login_silent falhou, tentando manual...")
                ok = await auto_login(state)
            if not ok:
                return results
            await asyncio.sleep(1)
            ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
            page = ctx.pages[0] if ctx.pages else await ctx.new_page()
            page.set_default_timeout(TIMEOUT_PAGE)
            await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
            await wait_stable(page)
            if "login" in page.url.lower():
                log.error(f"[{state}] Import bodies: login falhou após renovação")
                await ctx.close()
                return results

        log.info(f"[{state}] Import bodies: {len(ticket_numbers)} tickets em {len(chunks)} abas paralelas")

        async def _body_chunk(chunk, tab_id):
            tab_results = {}
            if not chunk:
                return tab_results
            pg = await ctx.new_page()
            pg.set_default_timeout(60000)

            ok = await goto_dashboard(pg, state)
            if not ok:
                log.error(f"[{state}][T{tab_id}] Dashboard inacessível — abortando aba")
                try:
                    await pg.close()
                except Exception:
                    pass
                return tab_results

            for idx, tnum in enumerate(chunk):
                log.info(f"[{state}][T{tab_id}] Import ({idx+1}/{len(chunk)}) {tnum}")
                for _attempt in range(2):
                    try:
                        await filter_ticket(pg, tnum)
                        if not await pg.get_by_text(tnum, exact=True).count():
                            log.warning(f"[{state}] {tnum}: não encontrado no portal")
                            break
                        await pg.get_by_text(tnum, exact=True).first.click()
                        await wait_stable(pg)

                        tt = pg.locator('[role="tab"]:has-text("Text")').first
                        if await tt.count():
                            await click_and_wait(pg, tt, "tab")

                        body = await pg.locator("body").inner_text()
                        tab_results[tnum] = body
                        await fast_back(pg, state)
                        break
                    except Exception as e:
                        if _attempt == 0 and "Timeout" in str(e):
                            log.warning(f"[{state}][T{tab_id}] {tnum}: Timeout — retry...")
                            try:
                                await back_to_dashboard(pg, state)
                            except Exception:
                                pass
                            continue
                        log.error(f"[{state}][T{tab_id}] {tnum}: ERRO → {e}")
                        try:
                            await back_to_dashboard(pg, state)
                        except Exception:
                            pass

            try:
                await pg.close()
            except Exception:
                pass
            log.info(f"[{state}][T{tab_id}] Import chunk: {len(tab_results)}/{len(chunk)} bodies coletados")
            return tab_results

        try:
            chunk_results = await asyncio.gather(*[_body_chunk(c, i) for i, c in enumerate(chunks)])
            for cr in chunk_results:
                results.update(cr)
        finally:
            try:
                await ctx.close()
            except Exception:
                pass
    return results


async def import_new_tickets(state, triggered_by="manual"):
    """Importa tickets novos do portal 811.

    Fluxo otimizado em 5 fases:
      1. Coleta números de tickets novos (1 aba, serial — lê labels no dashboard)
      2. Scrape bodies em paralelo (NUM_TABS abas simultâneas)
      3. Parse fields de cada body (Python puro, sem browser)
      4. Geocoding em batch (Nominatim, 1.1s entre chamadas)
      5. Batch upsert no Supabase
    """
    log.info(f"[{state}] === Importando tickets novos ===")
    existing = sb_get("tickets", f"&state=eq.{state}")
    existing_nums = {t["ticket"] for t in existing}
    projects = sb_get("projects")
    perfil = _profile_path(state)

    canceled_set = get_canceled_set(state)
    if canceled_set:
        log.info(f"[{state}] Cache: {len(canceled_set)} tickets cancelados conhecidos — serão pulados")

    # ── FASE 1: Coletar números de tickets novos (1 aba, serial) ──
    all_new_nums = []
    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)

        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)

        page, ctx = await ensure_login(page, ctx, p, state)
        if not page:
            return 0

        log.info(f"[{state}] Logado para importação")
        ok = await goto_dashboard(page, state)
        if not ok:
            log.error(f"[{state}] Não conseguiu acessar dashboard — abortando importação")
            try:
                await ctx.close()
            except Exception:
                pass
            return 0

        await set_items_per_page(page, 100)

        page_num = 1
        empty_pages = 0
        last_page_tickets = set()

        while page_num <= MAX_IMPORT_PAGES:
            if page_num == MAX_IMPORT_PAGES:
                log.warning(f"[{state}] Limite de {MAX_IMPORT_PAGES} páginas atingido — pode haver mais tickets não importados")

            log.info(f"[{state}] Lendo página {page_num}...")
            ticket_labels = page.locator("label.column-fixed")
            count = await ticket_labels.count()
            page_nums = []
            all_on_page = set()
            collected_set = set(all_new_nums)

            for i in range(count):
                txt = (await ticket_labels.nth(i).inner_text()).strip()
                if re.match(r"^\d{8,12}$", txt):
                    all_on_page.add(txt)
                    if txt in canceled_set:
                        continue
                    if txt not in existing_nums and txt not in collected_set:
                        page_nums.append(txt)
                        collected_set.add(txt)

            # Detecta loop de paginação
            if all_on_page and all_on_page == last_page_tickets:
                log.info(f"[{state}] Página {page_num} idêntica a anterior — fim da paginação")
                break
            last_page_tickets = all_on_page

            log.info(f"[{state}] Pág {page_num}: {len(page_nums)} tickets novos (de {len(all_on_page)} na página)")
            if len(page_nums) == 0:
                empty_pages += 1
            else:
                empty_pages = 0

            all_new_nums.extend(page_nums)

            next_btn = page.get_by_text("Next").first
            if await next_btn.count() and await next_btn.is_enabled():
                await click_and_wait(page, next_btn, "nav")
                await select_office(page, state)
                page_num += 1
            else:
                break

        try:
            await ctx.close()
        except Exception:
            pass

    if not all_new_nums:
        log.info(f"[{state}] Nenhum ticket novo encontrado")
        return 0

    log.info(f"[{state}] FASE 1 OK: {len(all_new_nums)} tickets novos — abrindo scrape paralelo")

    # ── FASE 2: Scrape bodies em paralelo (NUM_TABS abas) ──
    bodies = await _scrape_bodies_parallel(state, all_new_nums)
    log.info(f"[{state}] FASE 2 OK: {len(bodies)}/{len(all_new_nums)} bodies coletados")

    # ── FASE 3: Parse fields de cada body (Python puro, sem browser) ──
    parsed_tickets = []
    for tnum in all_new_nums:
        body = bodies.get(tnum)
        if not body:
            log.warning(f"[{state}] {tnum}: body não obtido — pulando")
            continue

        if is_ticket_canceled(body):
            replaced_match = re.search(r"REPLACED\s+BY\s+TICKET\s+(?:NUMBER:?\s*)(\d+)", body, re.IGNORECASE)
            replaced_by = replaced_match.group(1) if replaced_match else "N/A"
            log.warning(f"[{state}] {tnum}: CANCELADO (substituído por {replaced_by}) — pulando")
            add_to_canceled_cache(state, tnum)
            continue

        def extract(pattern, text, default=""):
            m = re.search(pattern, text, re.IGNORECASE)
            return m.group(1).strip() if m else default

        tnum_ext = extract(r"Ticket\s*:\s*(\d+)", body) or tnum
        state_code = extract(r"State:\s*(\w{2})", body, state)
        city = extract(r"Cityname:\s*([^\n]+)", body)
        if not city:
            city = extract(r"GeoPlace:\s*([^\n]+)", body)
        township = extract(r"Twp:\s*([^\n]+)", body)
        street_name = extract(r"Street\s*:\s*([^\n]+)", body)
        street_number = extract(r"Address\s*:\s*(\d+[^\n]*)", body)
        if street_number and street_name and not street_name[0].isdigit():
            street = f"{street_number.strip()} {street_name.strip()}"
        else:
            street = street_name
        work_type = extract(r"(?:Work Type|Type of Work|Work type)\s*:\s*([^\n]+)", body)
        job_id = extract(r"Job\s*(?:ID|#|Number)?\s*:\s*([^\n]+)", body)

        done_for_raw = extract(r"(?:Done\s*for|Work\s*(?:being\s*)?done\s*for)\s*:\s*([^\n]+)", body)
        client_811 = ""
        prime_811 = ""
        if done_for_raw:
            parts = [p.strip() for p in done_for_raw.replace("\\", "/").split("/") if p.strip()]
            if len(parts) >= 2:
                prime_811 = parts[0]
                client_811 = parts[1]
            elif len(parts) == 1:
                client_811 = parts[0]
            log.info(f"[{state}] {tnum_ext}: Done for → client={client_811}, prime={prime_811}")

        expire_str = normalize_expire(extract_expire_date(body))
        if expire_str:
            log.info(f"[{state}] {tnum_ext}: Expire → {expire_str}")
        location = extract_location_text(body, state=state)

        # Match com projeto — APENAS por Job#.
        project_id = None
        if job_id:
            for proj in projects:
                if job_id.strip() in (proj.get("name", "") + proj.get("description", "")):
                    project_id = proj["id"]
                    break

        # Boundary coords (instant, sem geocoding)
        geo_lat, geo_lon = None, None
        needs_geocode = False
        boundary_match = re.search(
            r"Boundary:\s*n\s*([\d.]+)\s+s\s*([\d.]+)\s+w\s*([-\d.]+)\s+e\s*([-\d.]+)",
            body, re.IGNORECASE
        )
        if boundary_match:
            n, s_val, w, e = (
                float(boundary_match.group(1)), float(boundary_match.group(2)),
                float(boundary_match.group(3)), float(boundary_match.group(4))
            )
            geo_lat = round((n + s_val) / 2, 6)
            geo_lon = round((w + e) / 2, 6)
        elif street and city:
            needs_geocode = True

        # Verificar se substitui ticket anterior
        old_ticket_num = ""
        old_expire_str = ""
        old_status_str = ""
        inherited_path = None
        old_tkt_match = re.search(r"Old\s*Tkt\s*:\s*(\d+)", body, re.IGNORECASE)
        if old_tkt_match:
            old_ticket_num = old_tkt_match.group(1)
        if not old_ticket_num:
            rep_match = re.search(r"Replaces?\s+(?:Ticket\s*)?(?:Number:?\s*)?(\d+)", body, re.IGNORECASE)
            if rep_match:
                old_ticket_num = rep_match.group(1)

        if old_ticket_num:
            log.info(f"[{state}] {tnum_ext}: Substitui ticket anterior {old_ticket_num}")
            try:
                old_tickets = sb_get("tickets", f"&ticket=eq.{_qv(old_ticket_num)}&state=eq.{_qv(state_code)}")
                if old_tickets:
                    ot = old_tickets[0]
                    if ot.get("field_path"):
                        inherited_path = ot["field_path"]
                        log.info(f"[{state}] {tnum_ext}: Trajeto herdado do ticket {old_ticket_num} ({len(inherited_path)} pts)")
                    old_expire_str = normalize_expire(ot.get("expire") or "")
                    old_status_str = (ot.get("status") or "").strip()
                    if old_expire_str or old_status_str:
                        log.info(f"[{state}] {tnum_ext}: Graça capturada — old_status={old_status_str!r}, old_expire={old_expire_str!r}")
            except Exception as e:
                log.debug(f"Erro ao buscar ticket anterior {old_ticket_num}: {e}")

        parsed_tickets.append({
            "tnum_ext": tnum_ext, "state_code": state_code,
            "city": city, "township": township, "street": street,
            "work_type": work_type, "job_id": job_id,
            "client_811": client_811, "prime_811": prime_811,
            "expire_str": expire_str, "location": location,
            "project_id": project_id,
            "geo_lat": geo_lat, "geo_lon": geo_lon, "needs_geocode": needs_geocode,
            "old_ticket_num": old_ticket_num, "old_expire_str": old_expire_str,
            "old_status_str": old_status_str, "inherited_path": inherited_path,
        })

    log.info(f"[{state}] FASE 3 OK: {len(parsed_tickets)} tickets parseados")

    # ── FASE 4: Geocoding em batch (sem browser, só Nominatim) ──
    geocode_count = sum(1 for t in parsed_tickets if t["needs_geocode"])
    if geocode_count:
        log.info(f"[{state}] Geocodando {geocode_count} endereços...")
        done = 0
        for t in parsed_tickets:
            if t["needs_geocode"]:
                t["geo_lat"], t["geo_lon"] = await geocode_address(t["street"], t["city"], t["state_code"])
                done += 1
                if done % 10 == 0:
                    log.info(f"[{state}] Geocoding {done}/{geocode_count}...")

    # ── FASE 5: Build ticket data e batch upsert ──
    new_tickets = []
    for t in parsed_tickets:
        work_type_final = t["work_type"] or "Main line"
        if t["geo_lat"] and t["geo_lon"]:
            t["geo_lat"], t["geo_lon"], work_type_final = adjust_coords_by_location(
                t["geo_lat"], t["geo_lon"], t["location"], work_type_final
            )

        ticket_county = ""
        try:
            loc_for_county = f"{t['city']}, {t['township']}".strip(", ")
            ticket_county = await resolve_county(loc_for_county, t["state_code"], t["geo_lat"], t["geo_lon"])
        except Exception as e:
            log.debug(f"[{state}] {t['tnum_ext']}: erro resolvendo county: {e}")

        history_entries = [
            {"ts": int(datetime.now().timestamp() * 1000), "action": f"Importado 811 - {t['city']}, {t['state_code']}", "color": "#10a574"}
        ]
        if t["inherited_path"]:
            history_entries.append(
                {"ts": int(datetime.now().timestamp() * 1000), "action": f"Trajeto herdado do ticket {t['old_ticket_num']}", "color": "#6d28d9"}
            )
        if t["old_ticket_num"] and (t["old_expire_str"] or t["old_status_str"]):
            history_entries.append(
                {"ts": int(datetime.now().timestamp() * 1000),
                 "action": f"[RENOVAÇÃO] {t['old_ticket_num']} → {t['tnum_ext']} (graça até {t['old_expire_str'] or 'N/A'}, status antigo: {t['old_status_str'] or 'N/A'})",
                 "color": "#7c3aed"}
            )

        ticket_data = {
            "ticket": t["tnum_ext"], "company": "One Drill", "state": t["state_code"],
            "location": f"{t['city']}, {t['township']}".strip(", "), "address": t["street"],
            "status": "Open", "expire": t["expire_str"], "footage": 0,
            "client": t["client_811"], "prime": t["prime_811"], "tipo": work_type_final,
            "job": t["job_id"] or "", "notes": f"[811 Location] {t['location']}" if t["location"] else "",
            "project_id": t["project_id"], "pending": "", "old_ticket2": t["old_ticket_num"],
            "status_old": t["old_status_str"], "expire_old": t["old_expire_str"], "field_path": t["inherited_path"],
            "geocoded_lat": t["geo_lat"], "geocoded_lon": t["geo_lon"],
            "county": ticket_county,
            "history": history_entries, "attachments": [],
        }
        new_tickets.append(ticket_data)
        log.info(f"[{state}] Preparado: {t['tnum_ext']}  {t['city']}, {t['township']}{' [' + ticket_county + ' Co]' if ticket_county else ''}")

    # Batch upsert
    inserted = 0
    to_insert = [td for td in new_tickets if td['ticket'] not in existing_nums]
    for td in new_tickets:
        if td['ticket'] in existing_nums:
            log.warning(f"[{state}] {td['ticket']}: DUPLICATA — já existe no sistema, pulando")

    if to_insert:
        try:
            for i in range(0, len(to_insert), BATCH_SIZE):
                chunk = to_insert[i:i + BATCH_SIZE]
                sb_upsert("tickets", chunk, on_conflict="ticket")
                inserted += len(chunk)
                for td in chunk:
                    existing_nums.add(td['ticket'])
                if i + BATCH_SIZE < len(to_insert):
                    time.sleep(0.3)
            log.info(f"[{state}] ✅ Batch upsert: {inserted} tickets inseridos em {((len(to_insert)-1)//BATCH_SIZE)+1} request(s)")
        except Exception as e:
            log.warning(f"[{state}] Batch falhou ({e}), tentando 1-por-1 como fallback...")
            inserted = 0
            for td in to_insert:
                try:
                    if td['ticket'] not in existing_nums:
                        sb_insert("tickets", td)
                        existing_nums.add(td['ticket'])
                    inserted += 1
                    log.info(f"[{state}] [OK fallback] {td['ticket']}")
                except Exception as e2:
                    log.error(f"[{state}] Erro inserindo {td['ticket']}: {e2}")

    log.info(f"[{state}] === Importação: {inserted} tickets novos ===")
    return inserted


# ── │ SECTION: RESCRAPE │ REESCRAPER: ATUALIZA NOTES + EXPIRE ─────────────────
async def rescrape_notes(state, force=False):
    perfil = _profile_path(state)
    tickets = sb_get("tickets", f"&state=eq.{state}&status=in.(Open,Damage,Clear)&order=ticket")
    if not tickets:
        log.info(f"[{state}] Nenhum ticket ativo para re-scrape")
        return

    if force:
        to_fix = tickets
    else:
        to_fix = [
            t for t in tickets
            if not (t.get("notes") or "").strip()
            or (t.get("notes") or "").strip() == "[811 Location]"
            or not (t.get("expire") or "").strip()
        ]

    log.info(f"[{state}] Re-scrape{'(FORCE)' if force else ''}: {len(to_fix)} tickets (de {len(tickets)} ativos)")
    if not to_fix:
        log.info(f"[{state}] Nada a fazer")
        return

    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)

        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)
        page, ctx = await ensure_login(page, ctx, p, state)
        if not page:
            return
        log.info(f"[{state}] Logado para re-scrape")
        ok = await goto_dashboard(page, state)
        if not ok:
            log.error(f"[{state}] Re-scrape: não conseguiu acessar dashboard — abortando")
            await ctx.close()
            return

        fixed = 0
        for idx, t in enumerate(to_fix):
            tnum = t["ticket"]
            tid = t["id"]
            log.info(f"[{state}] Re-scrape ({idx+1}/{len(to_fix)}) {tnum}")
            try:
                await filter_ticket(page, tnum)
                if not await page.get_by_text(tnum, exact=True).count():
                    log.warning(f"[{state}] {tnum}: não encontrado")
                    continue
                await page.get_by_text(tnum, exact=True).first.click()
                await wait_stable(page)

                tt = page.locator('[role="tab"]:has-text("Text")').first
                if await tt.count():
                    await click_and_wait(page, tt, "tab")

                body = await page.locator("body").inner_text()
                location = extract_location_text(body, state=state)
                patch_data = {}

                if location and (not (t.get("notes") or "").strip() or (t.get("notes") or "").strip() == "[811 Location]"):
                    patch_data["notes"] = f"[811 Location] {location}"

                current_exp = (t.get("expire") or "").strip()
                if not current_exp or _is_polluted_expire(current_exp):
                    expire_str = normalize_expire(extract_expire_date(body))
                    if expire_str and expire_str != current_exp:
                        patch_data["expire"] = expire_str

                if location and t.get("geocoded_lat") and t.get("geocoded_lon"):
                    new_lat, new_lon, new_tipo = adjust_coords_by_location(
                        t["geocoded_lat"], t["geocoded_lon"], location, t.get("tipo")
                    )
                    if new_lat != t["geocoded_lat"] or new_lon != t["geocoded_lon"] or new_tipo != t.get("tipo"):
                        patch_data.update({"geocoded_lat": new_lat, "geocoded_lon": new_lon, "tipo": new_tipo})

                if patch_data:
                    sb_patch("tickets", tid, patch_data)
                    fixed += 1
                await back_to_dashboard(page, state)

            except Exception as e:
                log.error(f"[{state}] Re-scrape {tnum}: ERRO → {e}")
                try:
                    await back_to_dashboard(page, state)
                except Exception:
                    pass

        await ctx.close()
    log.info(f"[{state}] === Re-scrape concluído: {fixed}/{len(to_fix)} atualizados ===")


# ── │ SECTION: CLEANUP │ LIMPAR TICKETS CANCELADOS ────────────────────────────
async def cleanup_canceled(state):
    perfil = _profile_path(state)
    canceled = 0

    # Fase 1: REMOVIDA — tickets com status=Cancel ficam no banco como histórico.
    # Política: cancelado no 811 → marca como Cancel, NUNCA deleta.
    # O filter_tickets_for_sync já os exclui do scrape loop, então custo zero.
    # Preserva auditoria, footage, respostas 811 e histórico no OneDrill.

    # Fase 2: Verificar tickets ativos no portal
    tickets = sb_get("tickets", f"&state=eq.{state}&status=in.(Open,Damage,Clear)&order=ticket")
    if not tickets:
        log.info(f"[{state}] Cleanup: nenhum ticket ativo para verificar")
        log.info(f"[{state}] === Cleanup concluído: {canceled} removidos ===")
        return canceled

    log.info(f"[{state}] Cleanup F2: verificando {len(tickets)} tickets no portal...")

    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)

        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)
        page, ctx = await ensure_login(page, ctx, p, state)
        if not page:
            return canceled
        log.info(f"[{state}] Cleanup: logado")
        ok = await goto_dashboard(page, state)
        if not ok:
            log.error(f"[{state}] Cleanup: não conseguiu acessar dashboard — abortando")
            await ctx.close()
            return canceled

        for idx, t in enumerate(tickets):
            tnum = t["ticket"]
            tid = t["id"]
            try:
                await filter_ticket(page, tnum)
                if not await page.get_by_text(tnum, exact=True).count():
                    continue
                await page.get_by_text(tnum, exact=True).first.click()
                await wait_stable(page)

                body = await page.locator("body").inner_text()
                tt = page.locator('[role="tab"]:has-text("Text")').first
                if await tt.count():
                    await click_and_wait(page, tt, "tab")
                    body = await page.locator("body").inner_text()

                if is_ticket_canceled(body):
                    replaced_match = re.search(r"REPLACED\s+BY\s+TICKET\s+(?:NUMBER:?\s*)(\d+)", body, re.IGNORECASE)
                    replaced_by = replaced_match.group(1) if replaced_match else "N/A"
                    try:
                        # NÃO deleta — marca como Cancel pra preservar histórico no OneDrill.
                        # Mantém ticket, respostas 811, footage e histórico acessíveis no app.
                        # add_to_canceled_cache garante que import_new_tickets não re-importa.
                        hist = t.get("history") or []
                        hist.append({
                            "ts": int(datetime.now().timestamp() * 1000),
                            "action": f"[AUTO 811] CANCELADO no portal — substituído por {replaced_by}",
                            "color": "#6d28d9"
                        })
                        sb_patch("tickets", tid, {
                            "status": "Cancel",
                            "history": hist
                        })
                        canceled += 1
                        add_to_canceled_cache(state, tnum)
                        log.warning(f"[{state}] Cleanup: {tnum} marcado como Cancel (substituído por {replaced_by})")
                    except Exception as e:
                        log.error(f"[{state}] Cleanup: erro marcando {tnum}: {e}")

                await back_to_dashboard(page, state)

            except Exception as e:
                log.error(f"[{state}] Cleanup {tnum}: ERRO → {e}")
                try:
                    await back_to_dashboard(page, state)
                except Exception:
                    pass

        await ctx.close()

    log.info(f"[{state}] === Cleanup concluído: {canceled} tickets cancelados removidos ===")
    return canceled


# ── │ SECTION: FILTER_SYNC │ CLEAR TICKET CACHE ───────────────────────────────
def filter_tickets_for_sync(tickets, state):
    """Filtra tickets que realmente precisam de scraping.

    - Open / Damage: SEMPRE verificar
    - Clear SEM expire: SEMPRE verificar (provavelmente ticket renovado
      que teve expire zerado — precisa buscar data nova no portal)
    - Clear RENOVADO (old_ticket2 preenchido): SEMPRE verificar (pode ter
      pendências novas que o cache não detectou porque o registro de Pending
      tem state/ticket_num quebrado em algum canto)
    - Clear: Pular se TODAS utilities já responderam E último sync < CLEAR_CACHE_HOURS
    """
    must_check = [t for t in tickets if t.get("status") in ("Open", "Damage")]
    clear_tickets = [t for t in tickets if t.get("status") == "Clear"]

    # Clear sem expire → ticket renovado esperando atualização pelo scraper.
    # Move imediatamente pra lista de verificação (fura qualquer cache).
    renewal_pending = [t for t in clear_tickets if not (t.get("expire") or "").strip()]
    if renewal_pending:
        nums = ", ".join(t["ticket"] for t in renewal_pending[:5])
        extra = "" if len(renewal_pending) <= 5 else f" (+{len(renewal_pending)-5} mais)"
        log.info(f"[{state}] {len(renewal_pending)} ticket(s) Clear sem expire — "
                 f"forçando verificação: {nums}{extra}")
        must_check.extend(renewal_pending)
    clear_tickets = [t for t in clear_tickets if (t.get("expire") or "").strip()]

    # Clear + renovado → força verificação. O ciclo de resposta do ticket
    # novo pode mudar rápido quando o antigo termina a graça, e o cache
    # de 24h é longo demais pra esse cenário.
    renewed_clear = [t for t in clear_tickets if (t.get("old_ticket2") or "").strip()]
    if renewed_clear:
        nums = ", ".join(t["ticket"] for t in renewed_clear[:5])
        extra = "" if len(renewed_clear) <= 5 else f" (+{len(renewed_clear)-5} mais)"
        log.info(f"[{state}] {len(renewed_clear)} ticket(s) Clear renovados — "
                 f"forçando verificação: {nums}{extra}")
        must_check.extend(renewed_clear)
    clear_tickets = [t for t in clear_tickets if not (t.get("old_ticket2") or "").strip()]

    if not clear_tickets:
        return must_check, 0

    clear_nums = [t["ticket"] for t in clear_tickets]
    try:
        pending_responses = sb_get(
            "ticket_811_responses",
            f"&state=eq.{state}&status=eq.Pending"
            f"&ticket_num=in.({','.join(clear_nums)})"
            "&select=ticket_num"
        )
        tickets_with_pending = {str(r["ticket_num"]) for r in pending_responses}
    except Exception as e:
        log.warning(f"[{state}] Cache: erro ao buscar pendentes, verificando todos: {e}")
        return tickets, 0

    try:
        last_sync = sb_get(
            "sync_811_log",
            f"&state=eq.{state}&status=eq.success&order=finished_at.desc&limit=1"
        )
        if last_sync and last_sync[0].get("finished_at"):
            last_sync_time = datetime.fromisoformat(last_sync[0]["finished_at"])
            if last_sync_time.tzinfo is not None:
                last_sync_time = last_sync_time.replace(tzinfo=None)
        else:
            last_sync_time = None
    except Exception:
        last_sync_time = None

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    cache_cutoff = now - timedelta(hours=CLEAR_CACHE_HOURS)

    skipped = 0
    for t in clear_tickets:
        tnum = t["ticket"]

        if tnum in tickets_with_pending:
            must_check.append(t)
            continue

        if not last_sync_time or last_sync_time < cache_cutoff:
            must_check.append(t)
            continue

        skipped += 1

    return must_check, skipped


# ── │ SECTION: SYNC │ SYNC STATE ──────────────────────────────────────────────
async def sync_state(state, triggered_by="manual"):
    log.info(f"{'='*55}")
    log.info(f"  OneDrill 811 Sync  {state}  [{triggered_by}]")
    log.info(f"{'='*55}")
    lid = log_start(state, triggered_by)
    checked = 0
    summary = SyncSummary()
    try:
        if not PORTALS[state]["user"]() or not PORTALS[state]["pass"]():
            raise ValueError(f"Credenciais faltando para {state}")

        all_tickets = sb_get("tickets", f"&state=eq.{state}&status=in.(Open,Damage,Clear)&order=ticket")
        if not all_tickets:
            log.info(f"[{state}] Nenhum ticket ativo")
            log_finish(lid, 0, 0)
            return

        tickets_to_scrape, skipped = filter_tickets_for_sync(all_tickets, state)
        if skipped > 0:
            log.info(f"[{state}] Cache: {skipped} tickets Clear pulados (sem pendências, verificados recentemente)")

        checked = len(tickets_to_scrape)
        if not tickets_to_scrape:
            log.info(f"[{state}] Nenhum ticket precisa verificação agora ({len(all_tickets)} ativos, {skipped} em cache)")
            log_finish(lid, 0, 0)
            return

        nums = [t["ticket"] for t in tickets_to_scrape]

        # ── Inclui tickets ANTIGOS de renovações em carência (respostas podem ter atualizado) ──
        grace_old_map = {}  # old_ticket_num → new_ticket_id
        nums_set = set(nums)
        for t in all_tickets:
            in_grace, old_num = is_in_renewal_grace(t)
            if not in_grace or not old_num or old_num in nums_set:
                continue
            nums.append(old_num)
            nums_set.add(old_num)
            grace_old_map[old_num] = t["id"]
            log.info(f"[{state}] Incluindo ticket antigo {old_num} (carência de {t['ticket']})")

        log.info(f"[{state}] {checked} tickets para verificar (de {len(all_tickets)} ativos)"
                 + (f" + {len(grace_old_map)} antigos em carência" if grace_old_map else ""))
        results = await scrape(state, nums, tickets_data=tickets_to_scrape)

        summary = save_to_supabase(state, results, all_tickets, grace_old_map=grace_old_map)
        log_finish(lid, checked, summary.responses_saved)
        try:
            check_scrape_failures(state, tickets_to_scrape, results)
        except Exception as _e:
            log.warning(f"[{state}] check_scrape_failures falhou: {_e}")
        log.info(f"[{state}] CONCLUÍDO  {checked} verificados, {skipped} em cache | {summary}")

    except Exception as e:
        log.error(f"[{state}] FALHOU: {e}")
        log_finish(lid, checked, summary.responses_saved, "error", str(e))


async def validate_sessions():
    """Valida sessões FL e IN sequencialmente antes do sync paralelo.
    Se alguma sessão expirou, abre janela de login uma de cada vez."""
    for state in ["FL", "IN"]:
        perfil = _profile_path(state)
        log.info(f"[{state}] Validando sessão...")
        try:
            async with async_playwright() as p:
                ctx = await p.chromium.launch_persistent_context(
                    perfil, headless=True, args=["--no-sandbox"]
                )
                page = ctx.pages[0] if ctx.pages else await ctx.new_page()
                page.set_default_timeout(15000)
                await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
                await page.wait_for_timeout(3000)
                needs_login = "login" in page.url.lower()
                await ctx.close()

            if needs_login:
                log.warning(f"[{state}] Sessão expirada — abrindo login manual...")
                ok = await auto_login_silent(state)

                if not ok:

                    log.warning(f"[{state}] auto_login_silent falhou, tentando manual...")

                    ok = await auto_login(state)
                if ok:
                    log.info(f"[{state}] ✅ Sessão renovada")
                else:
                    log.error(f"[{state}] ❌ Login falhou — sync do {state} pode falhar")
            else:
                log.info(f"[{state}] ✅ Sessão válida")
        except Exception as e:
            log.error(f"[{state}] Erro ao validar sessão: {e}")


async def sync_all(triggered_by="manual"):
    await validate_sessions()
    await asyncio.gather(
        sync_state("IN", triggered_by),
        sync_state("FL", triggered_by)
    )


async def sync_and_import(state, triggered_by="manual"):
    await import_new_tickets(state, triggered_by)
    await sync_state(state, triggered_by)


async def sync_and_import_all(triggered_by="manual"):
    """Roda IN e FL em paralelo."""
    await asyncio.gather(
        sync_and_import("IN", triggered_by),
        sync_and_import("FL", triggered_by)
    )


async def cleanup_all():
    """Cleanup IN e FL em paralelo."""
    await asyncio.gather(
        cleanup_canceled("IN"),
        cleanup_canceled("FL")
    )


def cleanup_wi_dhl_clears():
    """Cleanup retroativo: tickets WI Clear cujas TODAS as responses são "Closed by DHL".

    Razão: a regra antiga do classify() mapeava "Closed by DHL" como Clear, o que era
    incorreto. Wis. Stat. §182.0175 invalida o ticket após 10 working days sem positive
    response. A nova regra (2026-05) classifica como Cancel. Esta função aplica
    retroativamente: encontra tickets WI Clear cujas responses são todas "Closed by DHL"
    e os reverte pra Cancel (com histórico explicativo).

    Idempotente — pode rodar várias vezes, só afeta tickets que ainda não foram revertidos.
    """
    log.info("[CLEANUP-WI-DHL] Iniciando cleanup retroativo...")

    clear_tickets = sb_get("tickets", "&state=eq.WI&status=eq.Clear&select=id,ticket,status,history,notes")
    if not clear_tickets:
        log.info("[CLEANUP-WI-DHL] Nenhum ticket WI Clear encontrado")
        return

    log.info(f"[CLEANUP-WI-DHL] Analisando {len(clear_tickets)} tickets Clear...")

    affected = []  # [(ticket_dict, [response_dicts]), ...]

    for t in clear_tickets:
        tnum = t["ticket"]
        resps = sb_get(
            "ticket_811_responses",
            f"&ticket_num=eq.{_qv(tnum)}&select=id,status,response_text,utility_name"
        )
        if not resps:
            continue

        def _is_dhl_closed(r):
            txt = (r.get("response_text") or "").lower()
            return "closed by dhl" in txt or "closed by diggers" in txt

        if all(_is_dhl_closed(r) for r in resps):
            affected.append((t, resps))

    if not affected:
        log.info("[CLEANUP-WI-DHL] Nenhum ticket se encaixa (todas responses Closed by DHL)")
        return

    log.info(f"[CLEANUP-WI-DHL] {len(affected)} ticket(s) Clear→Cancel:")
    for t, _ in affected:
        log.info(f"   - {t['ticket']}")

    ticket_patches = []
    response_patches = []
    cancel_ts = int(datetime.now().timestamp() * 1000)
    cancel_label = datetime.now().strftime('%m/%d/%Y')
    cancel_note = f"[AUTO 811] Cancelado em {cancel_label} — ticket invalidado (Closed by DHL após 10 working days)"

    for t, resps in affected:
        hist = t.get("history") or []
        hist.append({"ts": cancel_ts, "action": cancel_note, "color": "#6d28d9"})
        new_notes = append_auto_note(t.get("notes"), cancel_note)
        ticket_patches.append({
            "id": t["id"],
            "status": "Cancel",
            "notes": new_notes,
            "history": hist,
        })
        for r in resps:
            response_patches.append({"id": r["id"], "status": "Cancel"})

    if ticket_patches:
        log.info(f"[CLEANUP-WI-DHL] Aplicando {len(ticket_patches)} ticket patches...")
        sb_batch_patch("tickets", ticket_patches, id_field="id")

    if response_patches:
        log.info(f"[CLEANUP-WI-DHL] Aplicando {len(response_patches)} response patches...")
        sb_batch_patch("ticket_811_responses", response_patches, id_field="id")

    log.info(f"[CLEANUP-WI-DHL] ✅ Concluído: {len(affected)} tickets revertidos pra Cancel")


def reclassify_wi_responses():
    """Re-aplica classify() em todas as responses WI do banco e re-checa auto-cancel.

    Útil quando a regra do classify muda e queremos atualizar dados existentes sem
    re-scrapear o portal. Lê cada response WI, chama classify(status_raw, response_text),
    patcheia se status mudou. Depois pra cada ticket afetado, re-roda a lógica
    de all_cancel (excluindo "Not Participating" que é fora do escopo).

    Idempotente — pode rodar várias vezes.
    """
    log.info("[RECLASSIFY-WI] Iniciando reclassify retroativo...")

    resps = sb_get(
        "ticket_811_responses",
        "&state=eq.WI&select=id,ticket_num,utility_name,status,response_text,status_raw"
    )
    if not resps:
        log.info("[RECLASSIFY-WI] Nenhuma response WI encontrada")
        return

    log.info(f"[RECLASSIFY-WI] {len(resps)} responses WI carregadas")

    response_patches = []
    changed_tickets = set()

    for r in resps:
        resp_text = r.get("response_text") or ""
        status_raw = r.get("status_raw") or ""
        new_status, _ = classify(status_raw, resp_text)
        current = r.get("status")
        if new_status != current:
            response_patches.append({"id": r["id"], "status": new_status})
            changed_tickets.add(r["ticket_num"])
            log.info(f"   {r['ticket_num']}/{r['utility_name']}: {current} → {new_status}")

    if not response_patches:
        log.info("[RECLASSIFY-WI] Nenhuma response precisa atualizar")
    else:
        log.info(f"[RECLASSIFY-WI] Aplicando {len(response_patches)} response patches...")
        sb_batch_patch("ticket_811_responses", response_patches, id_field="id")
        log.info(f"[RECLASSIFY-WI] ✅ {len(response_patches)} responses atualizadas")

    if not changed_tickets:
        log.info("[RECLASSIFY-WI] Concluído (sem mudanças)")
        return

    log.info(f"[RECLASSIFY-WI] Re-checando auto-cancel em {len(changed_tickets)} ticket(s)...")

    def _is_not_part(r):
        txt = ((r.get("response_text") or "") + " " + (r.get("status_raw") or "")).lower()
        return "not participating" in txt or "not service provider" in txt

    ticket_patches = []
    cancel_ts = int(datetime.now().timestamp() * 1000)
    cancel_label = datetime.now().strftime('%m/%d/%Y')
    cancel_note = f"[AUTO 811] Cancelado em {cancel_label} — ticket invalidado (Closed by DHL após 10 working days)"

    for tnum in sorted(changed_tickets):
        t_rows = sb_get(
            "tickets",
            f"&state=eq.WI&ticket=eq.{_qv(tnum)}&select=id,ticket,status,history,notes,status_locked"
        )
        if not t_rows:
            continue
        t = t_rows[0]
        if t.get("status") == "Cancel":
            continue
        if t.get("status_locked", False):
            log.info(f"   [RECLASSIFY-WI] {tnum}: 🔒 status_locked — pulando")
            continue

        t_resps = sb_get(
            "ticket_811_responses",
            f"&ticket_num=eq.{_qv(tnum)}&select=status,response_text,status_raw"
        )
        if not t_resps:
            continue

        relevant = [r for r in t_resps if not _is_not_part(r)]
        if not relevant:
            continue

        if all(r.get("status") == "Cancel" for r in relevant):
            hist = t.get("history") or []
            hist.append({"ts": cancel_ts, "action": cancel_note, "color": "#6d28d9"})
            new_notes = append_auto_note(t.get("notes"), cancel_note)
            ticket_patches.append({
                "id": t["id"],
                "status": "Cancel",
                "notes": new_notes,
                "history": hist,
            })
            log.info(f"   [RECLASSIFY-WI] {tnum}: AUTO-CANCEL ({len(relevant)} relevantes, todas Cancel)")

    if ticket_patches:
        log.info(f"[RECLASSIFY-WI] Aplicando {len(ticket_patches)} ticket patches...")
        sb_batch_patch("tickets", ticket_patches, id_field="id")
        log.info(f"[RECLASSIFY-WI] ✅ {len(ticket_patches)} tickets → Cancel")
    else:
        log.info("[RECLASSIFY-WI] Nenhum ticket precisa virar Cancel")

    log.info("[RECLASSIFY-WI] Concluído")


# ── │ SECTION: JULIE │ JULIE (Illinois 811) — SCRAPE PÚBLICO ──────────────────

async def scrape_julie_ticket(page, tnum, retry=True):
    """Scrape um ticket no JULIE — retorna dict no formato padrão.

    Navega para a página de busca antes de cada pesquisa (evita stale state).
    Se não encontrar na primeira tentativa, recarrega e tenta mais uma vez.
    """
    result = {"location_text": "", "responses": [], "expire_date": ""}

    # Navega para página de busca (garante estado limpo)
    await page.goto(JULIE_URL, wait_until="domcontentloaded")
    await page.wait_for_timeout(1500)

    # Limpa input e digita ticket
    inp = page.locator('input[type="text"], input[type="search"]').first
    if not await inp.count():
        log.warning(f"[IL] {tnum}: input de busca não encontrado — recarregando")
        await page.goto(JULIE_URL, wait_until="domcontentloaded")
        await page.wait_for_timeout(2000)
        inp = page.locator('input[type="text"], input[type="search"]').first
        if not await inp.count():
            log.error(f"[IL] {tnum}: input de busca não encontrado após reload")
            return result

    await inp.click()
    await inp.fill("")
    await page.wait_for_timeout(200)
    await inp.fill(tnum)

    # Clica busca (ícone de lupa)
    btn = page.locator('button:near(input):visible').first
    try:
        await btn.click()
    except Exception:
        # Fallback: Enter
        await inp.press("Enter")
    await page.wait_for_timeout(4000)
    await wait_stable(page)

    body = await page.locator("body").inner_text()
    result["expire_date"] = extract_expire_date(body)

    # Verifica se o ticket foi encontrado — checa se a tabela "Ticket Details" existe
    # NÃO usar "no matching records" no body inteiro porque esse texto aparece
    # na tabela Responses vazia quando o ticket existe mas nenhuma utility respondeu ainda
    has_ticket_details = await page.locator('text="Ticket Details"').count() > 0
    if not has_ticket_details:
        # Fallback: checa se tem alguma tabela com dados do ticket
        has_any_table = await page.locator('table').count() > 0
        if not has_any_table or ("no matching records" in body.lower() and "ticket details" not in body.lower()):
            # Retry: recarrega e tenta mais uma vez
            if retry:
                log.debug(f"[IL] {tnum}: não encontrado na 1ª tentativa — retry")
                return await scrape_julie_ticket(page, tnum, retry=False)
            log.info(f"[IL] {tnum}: não encontrado no JULIE")
            return result

    # Parse tabelas via JS
    data = await page.evaluate("""() => {
        const tables = document.querySelectorAll('table');
        const result = {details: [], pending: [], responses: []};

        tables.forEach((table) => {
            const rows = table.querySelectorAll('tr');
            const headers = [];
            const data = [];
            rows.forEach((row, ri) => {
                const cells = row.querySelectorAll('th, td');
                const vals = Array.from(cells).map(c => c.innerText.trim());
                if (ri === 0) {
                    headers.push(...vals);
                } else if (vals.some(v => v.length > 0)) {
                    data.push(vals);
                }
            });

            const hdr = headers.join('|').toLowerCase();
            if (hdr.includes('ticket') && hdr.includes('revision') && hdr.includes('address')) {
                result.details = data;
            } else if (hdr.includes('member code') && hdr.includes('facility type') && !hdr.includes('response')) {
                result.pending = data;
            } else if (hdr.includes('response') && hdr.includes('member name') && hdr.includes('responded')) {
                result.responses = data;
            }
        });
        return result;
    }""")

    # Location text: Address + Street + Cross Street
    if data.get("details"):
        row = data["details"][0]
        # Cols: Ticket(0), Revision(1), Address(2), Street(3), Cross Street(4), Company(5), Locate By(6), Attachments(7)
        if len(row) >= 5:
            parts = [p for p in [row[2], row[3], row[4]] if p and p.strip() and p.strip() != "-"]
            result["location_text"] = " / ".join(parts) if parts else ""

    # Pending → No Response
    for row in data.get("pending", []):
        # Cols: Member Code(0), Facility Type(1), Member Name(2), Due(3), Comments(4), Attachments(5)
        if len(row) >= 3:
            member_name = (row[2] or "").strip()
            facility_type = (row[1] or "").strip()
            if member_name and member_name != "-" and "no matching" not in member_name.lower():
                result["responses"].append({
                    "utility": member_name,
                    "status_raw": "No Response",
                    "status": "Pending",
                    "response": "",
                    "comment": f"[{facility_type}]" if facility_type else "",
                })

    # Responses → parse respostas reais
    for row in data.get("responses", []):
        # Cols: Revision(0), Member Code(1), Facility Type(2), Member Name(3), Response(4), Description(5), Responded(6), Comments(7), URL(8), Attachments(9)
        if len(row) >= 7:
            member_name = (row[3] or "").strip()
            response_code = (row[4] or "").strip()
            description = (row[5] or "").strip()
            responded_str = (row[6] or "").strip()
            comment = (row[7] if len(row) > 7 else "").strip()

            if member_name and member_name != "-" and "no matching" not in member_name.lower():
                # ── IGNORAR ENTRADAS DE AGENDAMENTO (IL/JULIE) ──
                # Não são respostas reais de status — são negociações de schedule
                desc_upper = (description or "").upper()
                code_upper = (response_code or "").upper()
                IL_SCHEDULE_PATTERNS = [
                    "DECLINED CODE 50",
                    "ACCEPTED CODE 50",
                    "LOCATOR AND EXCAVATOR AGREED",
                    "DOCUMENTED ALTERNATE MARKING SCHEDULE",
                    "ALTERNATE DATE REQUESTED",
                ]
                is_schedule = any(pat in desc_upper or pat in code_upper for pat in IL_SCHEDULE_PATTERNS)
                if is_schedule:
                    log.debug(f"[IL] {tnum}: IGNORANDO entrada de agendamento — {member_name}: {description[:80]}")
                    continue

                # Parse responded date
                responded_date = None
                if responded_str and responded_str != "-":
                    for fmt in [
                        "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p",
                        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
                        "%m/%d/%Y",
                    ]:
                        try:
                            responded_date = datetime.strptime(
                                responded_str[:len(fmt)+4].strip(), fmt
                            ).replace(tzinfo=None).isoformat()
                            break
                        except ValueError:
                            continue

                status, _cls_unrec = classify(response_code, description + " " + comment)
                # status_raw = texto descritivo (MARKED, RE-MARK NOT NEEDED, etc)
                # NÃO usar response_code numérico (21, 20, 999) — a dedup _is_non_real()
                # e o all-RE-MARK skip dependem de texto legível em status_raw.
                raw_text = description if description and description != "-" else response_code
                result["responses"].append({
                    "utility": member_name,
                    "status_raw": raw_text,
                    "status": status,
                    "response": description,
                    "comment": comment,
                    "responded_date": responded_date,
                    "_unrecognized": _cls_unrec
                })

    log.info(f"[IL] {tnum}: {len(result['responses'])} utilities")
    return result


async def scrape_il(ticket_numbers, tickets_data=None):
    """Scrape batch de tickets IL no JULIE (público, sem login).

    Usa NUM_TABS abas paralelas para acelerar o processo.
    """
    results = {}
    if not ticket_numbers:
        return results

    # Divide em chunks para abas paralelas
    chunks = [[] for _ in range(NUM_TABS)]
    for i, tnum in enumerate(ticket_numbers):
        chunks[i % NUM_TABS].append(tnum)
    chunks = [c for c in chunks if c]

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox"])

        async def process_chunk(chunk, tab_id):
            """Processa um grupo de tickets em uma aba independente."""
            tab_results = {}
            page = await browser.new_page()
            page.set_default_timeout(30000)

            for idx, tnum in enumerate(chunk):
                log.info(f"[IL][T{tab_id}] ({idx+1}/{len(chunk)}) Ticket {tnum}")
                try:
                    result = await scrape_julie_ticket(page, tnum)
                    tab_results[tnum] = result
                except Exception as e:
                    log.error(f"[IL][T{tab_id}] {tnum}: ERRO → {e}")
                    tab_results[tnum] = {"location_text": "", "responses": [], "expire_date": ""}

            try:
                await page.close()
            except Exception:
                pass
            return tab_results

        log.info(f"[IL] Abrindo {len(chunks)} abas paralelas...")
        chunk_results = await asyncio.gather(*[process_chunk(c, i) for i, c in enumerate(chunks)])
        for cr in chunk_results:
            results.update(cr)

        await browser.close()

    return results


async def sync_il(triggered_by="manual"):
    """Sync completo Illinois — JULIE público (sem login)."""
    log.info(f"{'='*55}")
    log.info(f"  OneDrill 811 Sync  IL (JULIE)  [{triggered_by}]")
    log.info(f"{'='*55}")
    lid = log_start("IL", triggered_by)
    checked = 0
    summary = SyncSummary()
    try:
        all_tickets = sb_get("tickets", "&state=eq.IL&status=in.(Open,Damage,Clear)&order=ticket")
        if not all_tickets:
            log.info("[IL] Nenhum ticket ativo")
            log_finish(lid, 0, 0)
            return

        tickets_to_scrape, skipped = filter_tickets_for_sync(all_tickets, "IL")
        if skipped > 0:
            log.info(f"[IL] Cache: {skipped} tickets Clear pulados")

        checked = len(tickets_to_scrape)
        if not tickets_to_scrape:
            log.info(f"[IL] Nenhum ticket precisa verificação ({len(all_tickets)} ativos, {skipped} em cache)")
            log_finish(lid, 0, 0)
            return

        nums = [t["ticket"] for t in tickets_to_scrape]
        log.info(f"[IL] {checked} tickets para verificar (de {len(all_tickets)} ativos)")
        results = await scrape_il(nums, tickets_data=tickets_to_scrape)

        summary = save_to_supabase("IL", results, all_tickets)
        log_finish(lid, checked, summary.responses_saved)
        log.info(f"[IL] CONCLUÍDO  {checked} verificados, {skipped} em cache | {summary}")

    except Exception as e:
        log.error(f"[IL] FALHOU: {e}")
        log_finish(lid, checked, summary.responses_saved, "error", str(e))


# ── │ SECTION: IL_IMPORT │ IMPORT IL (JULIE Ticket Entry — portal cliente) ────
#
# Em IL os tickets são tipicamente inseridos manualmente. Esse módulo automatiza
# a importação via portal cliente PALINKASSE, espelhando o fluxo de import_wi.
#
# Fluxo:
#   1. Login em julie_ticketentry.html
#   2. Menu top → Search → Ticket Search
#   3. Dropdown "Search for County" → COOK → Search
#   4. Parse grid → filter Company ONEDRILL → dedupe revisão mais recente
#   5. Compara com Supabase via campo `expire`:
#        - Não existe → INSERT (abre detalhe pra extrair work_type/remarks)
#        - Existe com expire diferente → UPDATE expire + history (extensão)
#        - Existe com expire igual → skip


def _il_split_ticket_revision(full):
    """'>A261140377-03X' → ('A261140377', '03X'). Remove marcador '>' se presente."""
    s = (full or "").strip().lstrip(">").strip()
    if "-" not in s:
        return s, ""
    base, rev = s.split("-", 1)
    return base.strip(), rev.strip()


def _il_parse_completed_ts(completed_str):
    """Parse 'MM/DD/YYYY HH:MM AM' do Completed pra comparação."""
    s = (completed_str or "").strip()
    for fmt in ["%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y"]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return datetime.min


def _il_extract_expire(s):
    """Extrai 'MM/DD/YYYY' da string Expires do grid (ex: '07/06/2026 11:59 PM')."""
    if not s:
        return ""
    m = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", s)
    if not m:
        return ""
    try:
        return datetime.strptime(m.group(1), "%m/%d/%Y").strftime("%m/%d/%Y")
    except ValueError:
        return ""


async def _il_login(page):
    """Login no JULIE Ticket Entry (newtin.julie1call.com).

    Returns: True se login OK, False caso contrário.
    """
    if not IL_USER or not IL_PASS:
        log.error("[IL] Credenciais IL_USER/IL_PASS não definidas no .env")
        return False

    log.info(f"[IL] Login: navegando para {JULIE_TICKETENTRY_URL}")
    await page.goto(JULIE_TICKETENTRY_URL, wait_until="domcontentloaded")
    await page.wait_for_timeout(2000)
    await wait_stable(page)

    body = await page.locator("body").inner_text()
    if "logged in as" in body.lower():
        log.info(f"[IL] Já logado")
        return True

    user_input = None
    for sel in ['input[name="Account"]', 'input[name="account"]', 'input[type="text"]:visible']:
        loc = page.locator(sel).first
        try:
            if await loc.count():
                user_input = loc
                break
        except Exception:
            continue

    pass_input = page.locator('input[type="password"]').first
    if not user_input or not await pass_input.count():
        log.error("[IL] Login: campo Account ou Password não encontrado")
        await page.screenshot(path=os.path.join(BASE_DIR, "debug_il_login_fail.png"))
        return False

    await user_input.click()
    await user_input.fill(IL_USER)
    await pass_input.click()
    await pass_input.fill(IL_PASS)

    submitted = False
    for sel in ['button:has-text("Submit")', 'input[type="submit"]', 'button[type="submit"]']:
        loc = page.locator(sel).first
        try:
            if await loc.count():
                await loc.click()
                submitted = True
                break
        except Exception:
            continue
    if not submitted:
        await pass_input.press("Enter")

    await page.wait_for_timeout(3000)
    await wait_stable(page)

    body = await page.locator("body").inner_text()
    if "logged in as" in body.lower():
        log.info(f"[IL] ✅ Login OK como {IL_USER}")
        return True

    log.error(f"[IL] Login falhou — verifique IL_USER/IL_PASS no .env")
    await page.screenshot(path=os.path.join(BASE_DIR, "debug_il_login_fail.png"))
    return False


async def _il_open_search_screen(page):
    """Click no botão Search do menu top (não confundir com Search do painel do mapa).

    O menu top tem: Inquire | New | Recent | Test | Search | Log out.
    O painel inferior do mapa tem: Home Search Places LatLong Grids Layers.
    Distinguir pelo contexto (left-of Log out / right-of Test).
    """
    # Tenta primeiro o botão específico do menu top (id fixo)
    btn_direct = page.locator('#btnTicketSearch')
    try:
        if await btn_direct.count() and await btn_direct.is_visible():
            await btn_direct.click()
            log.info("[IL] Search menu: click direto em #btnTicketSearch")
            await page.wait_for_timeout(4000)
            await wait_stable(page)
            body_parts = []
            try:
                body_parts.append(await page.locator("body").inner_text())
            except Exception:
                pass
            for fr in page.frames:
                if fr == page.main_frame:
                    continue
                try:
                    body_parts.append(await fr.locator("body").inner_text())
                except Exception:
                    continue
            body_all = "\n".join(body_parts).lower()
            if ("search for street" in body_all or "search for place" in body_all
                    or "search for county" in body_all or "tickets for county" in body_all):
                return True
            log.info("[IL] #btnTicketSearch clicou mas search form não apareceu, tentando scan")
    except Exception:
        pass

    # Fallback: Enumera TODOS os elementos com texto/value "Search" visíveis no DOM,
    # com bounding box. Filtra o do TOP (menor Y) — esse é o do menu Inquire|New|Recent|Test|Search|Log out.
    candidates = await page.evaluate("""() => {
        const out = [];
        const tags = ['button', 'input', 'a', 'span', 'div'];
        tags.forEach(tag => {
            document.querySelectorAll(tag).forEach(el => {
                const txt = (el.innerText || el.value || '').trim();
                if (txt !== 'Search') return;
                const r = el.getBoundingClientRect();
                if (r.width === 0 || r.height === 0) return;
                out.push({
                    tag: el.tagName,
                    id: el.id || '',
                    cls: el.className || '',
                    x: Math.round(r.x), y: Math.round(r.y),
                    w: Math.round(r.width), h: Math.round(r.height),
                });
            });
        });
        return out;
    }""")
    log.info(f"[IL] Search candidates: {candidates}")
    if not candidates:
        log.error("[IL] Nenhum elemento 'Search' visível encontrado")
        await page.screenshot(path=os.path.join(BASE_DIR, "debug_il_search_menu_fail.png"))
        return False

    # Escolhe o de MENOR Y (top menu fica no topo da página)
    top_search = min(candidates, key=lambda c: c["y"])
    log.info(f"[IL] Search menu: alvo top → {top_search}")

    # Click via coordenadas no centro do elemento — bypassa qualquer ID/handler weird
    cx = top_search["x"] + top_search["w"] // 2
    cy = top_search["y"] + top_search["h"] // 2
    try:
        await page.mouse.click(cx, cy)
        log.info(f"[IL] Search menu: click em ({cx},{cy})")
    except Exception as e:
        log.error(f"[IL] Click em coordenadas falhou: {e}")
        await page.screenshot(path=os.path.join(BASE_DIR, "debug_il_search_menu_fail.png"))
        return False
    await page.wait_for_timeout(4000)
    await wait_stable(page)

    # Coleta texto de TODOS os frames (Ticket Search pode estar em iframe)
    body_parts = []
    try:
        body_parts.append(await page.locator("body").inner_text())
    except Exception:
        pass
    for fr in page.frames:
        if fr == page.main_frame:
            continue
        try:
            body_parts.append(await fr.locator("body").inner_text())
        except Exception:
            continue
    body_all = "\n".join(body_parts).lower()

    has_search_form = ("search for street" in body_all or "search for place" in body_all
                       or "search for county" in body_all or "tickets for county" in body_all)
    if not has_search_form:
        log.error(f"[IL] Tela Ticket Search não apareceu — frames={len(page.frames)}")
        await page.screenshot(path=os.path.join(BASE_DIR, "debug_il_search_screen_fail.png"))
        try:
            with open(os.path.join(BASE_DIR, "debug_il_after_search_click.html"), "w", encoding="utf-8") as f:
                f.write(await page.content())
        except Exception:
            pass
        return False
    return True


async def _il_search_county(page, county=IL_SEARCH_COUNTY):
    """Seleciona County e clica Search. Espera grid carregar."""
    log.info(f"[IL] Search por County={county}")

    county_select = None
    count_selects = await page.locator('select').count()
    log.info(f"[IL] {count_selects} selects no DOM")
    for i in range(count_selects):
        sel = page.locator('select').nth(i)
        try:
            # Pula selects invisíveis ou disabled (ex: selCounty do Digsite Information)
            if not await sel.is_visible():
                continue
            if not await sel.is_enabled():
                continue
            options = await sel.locator('option').all_text_contents()
            opts_upper = [o.strip().upper() for o in options]
            # Distingue dropdown County do Place: Place tem ABINGDON CIT, County não.
            if county.upper() in opts_upper and "ABINGDON CIT" not in opts_upper:
                try:
                    sel_id = await sel.get_attribute("id") or ""
                except Exception:
                    sel_id = ""
                log.info(f"[IL] County select encontrado: idx={i}, id='{sel_id}', opts={len(options)}")
                county_select = sel
                break
        except Exception:
            continue

    if not county_select:
        log.error("[IL] Dropdown County (Ticket Search) não encontrado")
        await page.screenshot(path=os.path.join(BASE_DIR, "debug_il_county_dropdown_fail.png"))
        return False

    await county_select.select_option(label=county)
    await page.wait_for_timeout(500)

    # Pega bounding box do select pra achar o botão Search adjacente (mesma linha Y).
    try:
        sel_box = await county_select.bounding_box()
    except Exception:
        sel_box = None

    # Enumera todos os botões/inputs "Search" visíveis + enabled na página
    search_btns = await page.evaluate("""() => {
        const out = [];
        const isVisible = (el) => {
            const s = window.getComputedStyle(el);
            if (s.display === 'none' || s.visibility === 'hidden') return false;
            const r = el.getBoundingClientRect();
            return r.width > 0 && r.height > 0;
        };
        document.querySelectorAll('button, input[type="button"], input[type="submit"]').forEach(el => {
            if (el.disabled) return;
            const txt = ((el.innerText || el.value) || '').trim();
            if (txt !== 'Search') return;
            if (!isVisible(el)) return;
            const r = el.getBoundingClientRect();
            out.push({tag: el.tagName, id: el.id || '', x: Math.round(r.x), y: Math.round(r.y), w: Math.round(r.width), h: Math.round(r.height)});
        });
        return out;
    }""")
    log.info(f"[IL] Search buttons disponíveis: {search_btns}")

    # Escolhe o Search mais próximo do select County (mesma Y, X > sel_box right)
    # Tenta múltiplas estratégias de click: force_click via id, JS click, dispatch, coords.
    clicked = False
    target = None
    if sel_box and search_btns:
        sy = sel_box["y"]
        sx_right = sel_box["x"] + sel_box["width"]
        adjacent = [b for b in search_btns if abs(b["y"] - sy) < 30 and b["x"] >= sx_right - 5]
        if adjacent:
            target = min(adjacent, key=lambda b: b["x"])
            log.info(f"[IL] Search County alvo → {target}")

    if target:
        target_id = target.get("id", "")
        if target_id:
            loc = page.locator(f"#{target_id}").first
            for strategy in ("force_click", "js_click", "dispatch"):
                try:
                    if strategy == "force_click":
                        await loc.click(force=True, timeout=5000)
                    elif strategy == "js_click":
                        await loc.evaluate("el => el.click()")
                    else:
                        await loc.evaluate("""el => {
                            ['mousedown','mouseup','click'].forEach(t => {
                                el.dispatchEvent(new MouseEvent(t, {bubbles:true, cancelable:true, view:window}));
                            });
                        }""")
                    log.info(f"[IL] Search County: clicou via #{target_id} ({strategy})")
                    clicked = True
                    break
                except Exception as e:
                    log.debug(f"[IL] {strategy} em #{target_id} falhou: {e}")
                    continue
        if not clicked:
            cx = target["x"] + target["w"] // 2
            cy = target["y"] + target["h"] // 2
            try:
                await page.mouse.click(cx, cy)
                log.info(f"[IL] Search County: click em coords ({cx},{cy})")
                clicked = True
            except Exception as e:
                log.warning(f"[IL] Click coords falhou: {e}")

    if not clicked:
        log.warning("[IL] Botão Search County não encontrado — aguardando trigger automático")

    # JULIE pode demorar pra carregar grid com county inteiro
    await page.wait_for_timeout(8000)
    await wait_stable(page)
    return True


async def _il_parse_grid(page):
    """Lê a tabela 'Tickets' e retorna lista de dicts."""
    data = await page.evaluate("""() => {
        const tables = document.querySelectorAll('table');
        for (const t of tables) {
            const firstRow = t.querySelectorAll('tr')[0];
            if (!firstRow) continue;
            const headers = Array.from(firstRow.querySelectorAll('th, td'))
                .map(c => (c.innerText || '').trim().toLowerCase());
            const hdrStr = headers.join('|');
            if (hdrStr.includes('ticket') && hdrStr.includes('completed') && hdrStr.includes('expires')) {
                const rows = Array.from(t.querySelectorAll('tr')).slice(1)
                    .map(r => Array.from(r.querySelectorAll('td, th'))
                        .map(c => (c.innerText || '').trim()));
                return {headers, rows};
            }
        }
        return null;
    }""")

    if not data:
        log.warning("[IL] Grid: tabela não encontrada")
        return []

    headers = data["headers"]

    def _idx(name):
        for i, h in enumerate(headers):
            if name in h:
                return i
        return -1

    cols = {
        "ticket": _idx("ticket"),
        "completed": _idx("completed"),
        "expires": _idx("expires"),
        "county": _idx("county"),
        "place": _idx("place"),
        "address": _idx("address"),
        "street": _idx("street"),
        "cross": _idx("cross"),
        "caller": _idx("caller"),
        "company": _idx("company"),
    }

    def _cell(r, key):
        i = cols.get(key, -1)
        return r[i] if 0 <= i < len(r) else ""

    rows = []
    for r in data["rows"]:
        if not r or cols["ticket"] < 0:
            continue
        full = _cell(r, "ticket")
        base, rev = _il_split_ticket_revision(full)
        if not base:
            continue
        rows.append({
            "ticket_full": full.lstrip(">").strip(),
            "ticket_base": base,
            "revision": rev,
            "completed": _cell(r, "completed"),
            "expires": _cell(r, "expires"),
            "county": _cell(r, "county"),
            "place": _cell(r, "place"),
            "address_num": _cell(r, "address"),
            "street": _cell(r, "street"),
            "cross_street": _cell(r, "cross"),
            "caller": _cell(r, "caller"),
            "company": _cell(r, "company"),
        })

    log.info(f"[IL] Grid: {len(rows)} linhas")
    return rows


def _filter_il_onedrill(rows):
    """Mantém só Company ONEDRILL/ONE DRILL (case insensitive, ignora espaços)."""
    out = []
    for r in rows:
        norm = (r.get("company") or "").upper().replace(" ", "")
        if norm == "ONEDRILL":
            out.append(r)
    return out


def _dedupe_latest_revision(rows):
    """Pra cada ticket_base, mantém row com Completed mais recente."""
    by_base = {}
    for r in rows:
        base = r["ticket_base"]
        ts = _il_parse_completed_ts(r["completed"])
        if base not in by_base or ts > by_base[base]["_ts"]:
            by_base[base] = {**r, "_ts": ts}
    return [{k: v for k, v in r.items() if k != "_ts"} for r in by_base.values()]


async def _il_open_detail(page, row):
    """Click na linha do ticket pra abrir o detalhe na tela main.

    ⚠ Seletor ainda não validado contra HTML real — pode precisar ajuste após
    o primeiro run (debug_screenshot + il_detail_*.html ficam no BASE_DIR).
    """
    tnum_full = row["ticket_full"]
    tnum_base = row["ticket_base"]
    log.info(f"[IL] Abrindo detalhe: {tnum_full}")

    clicked = False
    for sel in [
        f'tr:has(td:text-is("{tnum_full}"))',
        f'tr:has(td:has-text("{tnum_base}"))',
        f'td:text-is("{tnum_full}")',
        f'a:has-text("{tnum_full}")',
    ]:
        loc = page.locator(sel).first
        try:
            if await loc.count():
                await loc.dblclick()
                clicked = True
                break
        except Exception:
            continue

    if not clicked:
        log.warning(f"[IL] Detail: linha {tnum_full} não encontrada")
        return False

    await page.wait_for_timeout(2500)
    await wait_stable(page)
    return True


async def _il_parse_detail(page):
    """Extrai dados do form de detalhe (Excavator/Digsite/Work/Members).

    ⚠ Defensivo — pode precisar ajuste depois do print do detalhe real.
    """
    try:
        data = await page.evaluate("""() => {
            const byLabel = (label) => {
                const lab = label.toLowerCase();
                const elements = document.querySelectorAll('td, th, label, span, div');
                for (const el of elements) {
                    const txt = (el.innerText || '').trim().toLowerCase();
                    if (txt === lab || txt === lab + ':') {
                        let target = el.nextElementSibling;
                        while (target) {
                            const inp = target.querySelector?.('input, textarea, select');
                            if (inp) return (inp.value || inp.innerText || '').trim();
                            if (target.tagName === 'TD' && (target.innerText || '').trim()) {
                                return target.innerText.trim();
                            }
                            target = target.nextElementSibling;
                        }
                        const parent = el.parentElement;
                        if (parent) {
                            const inp = parent.querySelector('input, textarea, select');
                            if (inp && inp !== el) return (inp.value || inp.innerText || '').trim();
                        }
                    }
                }
                return '';
            };

            const members = [];
            const all = document.querySelectorAll('*');
            for (const el of all) {
                const txt = (el.innerText || '').trim();
                if (txt.toLowerCase() === 'members' && el.children.length < 5) {
                    let container = el.parentElement || el;
                    for (let i = 0; i < 4 && container; i++) {
                        const rows = container.querySelectorAll('tr, li');
                        if (rows.length >= 2) {
                            rows.forEach(r => {
                                const t = (r.innerText || '').trim();
                                if (t && t.toLowerCase() !== 'members') members.push(t);
                            });
                            break;
                        }
                        container = container.parentElement;
                    }
                    break;
                }
            }

            return {
                work_type: byLabel('Work Type'),
                extent: byLabel('Extent'),
                premark: byLabel('Premark'),
                done_for: byLabel('Done For'),
                begin_date: byLabel('Begin Date'),
                dig_by_date: byLabel('Dig By Date'),
                expires_date: byLabel('Expires Date'),
                remarks: byLabel('Remarks'),
                company: byLabel('Company'),
                caller: byLabel('Caller'),
                county: byLabel('County'),
                place: byLabel('Place'),
                addr_street: byLabel('Addr/Street'),
                cross_st: byLabel('Cross St'),
                members_raw: members,
            };
        }""")
        return data or {}
    except Exception as e:
        log.warning(f"[IL] Parse detail falhou: {e}")
        return {}


def _build_il_notes(row, detail):
    """Monta string de notes pra ticket novo IL."""
    parts = []
    addr = (row.get("address_num") or "").strip()
    street = (row.get("street") or "").strip()
    cross = (row.get("cross_street") or "").strip()
    place = (row.get("place") or "").strip()
    if addr or street:
        parts.append(f"{addr} {street}".strip())
    if cross:
        parts.append(f"Cross: {cross}")
    if place:
        parts.append(place)
    if detail.get("remarks"):
        parts.append(f"[Remarks] {detail['remarks']}")
    return "\n".join(parts)


def _find_renewal_candidate(row, existing_tickets):
    """Detecta se ticket_base novo é renovação de um antigo (heurística address+place+timing)."""
    new_addr = f"{row.get('address_num', '')} {row.get('street', '')}".strip().upper()
    new_place = (row.get("place") or "").strip().upper()
    if not new_addr or not new_place:
        return None

    today = datetime.now().date()
    for t in existing_tickets:
        if (t.get("state") or "").upper() != "IL":
            continue
        if t.get("ticket") == row["ticket_base"]:
            continue
        addr = (t.get("address") or "").strip().upper()
        loc = (t.get("location") or "").strip().upper()
        if addr != new_addr or new_place not in loc:
            continue
        try:
            exp_str = normalize_expire(t.get("expire") or "")
            if not exp_str:
                continue
            exp_dt = datetime.strptime(exp_str, "%m/%d/%Y").date()
            if 0 <= (today - exp_dt).days <= 7:
                return t
        except Exception:
            continue
    return None


async def import_il(triggered_by="manual"):
    """Importa novos tickets IL via JULIE Ticket Entry (portal cliente).

    Returns: total de tickets inseridos + atualizados.
    """
    if not IL_USER or not IL_PASS:
        log.error("[IL] import_il: IL_USER/IL_PASS não definidos no .env")
        return 0

    log.info(f"{'='*55}")
    log.info(f"  OneDrill 811  IMPORT IL (JULIE Ticket Entry)  [{triggered_by}]")
    log.info(f"{'='*55}")

    existing = sb_get("tickets", "&state=eq.IL")
    existing_by_base = {t["ticket"]: t for t in existing}
    log.info(f"[IL] {len(existing_by_base)} tickets IL no Supabase")

    inserted = 0
    updated = 0
    skipped = 0
    renewals = 0

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox"])
        page = await browser.new_page()
        page.set_default_timeout(60000)

        try:
            if not await _il_login(page):
                return 0
            if not await _il_open_search_screen(page):
                return 0
            if not await _il_search_county(page, IL_SEARCH_COUNTY):
                return 0

            rows = await _il_parse_grid(page)
            rows = _filter_il_onedrill(rows)
            rows = _dedupe_latest_revision(rows)
            log.info(f"[IL] {len(rows)} tickets ONEDRILL após filter+dedupe")

            debug_count = 0
            for row in rows:
                base = row["ticket_base"]
                expire_grid = _il_extract_expire(row["expires"])
                existing_t = existing_by_base.get(base)

                if existing_t:
                    expire_saved = normalize_expire(existing_t.get("expire") or "")
                    if expire_saved == expire_grid:
                        skipped += 1
                        continue
                    hist = existing_t.get("history") or []
                    hist.append({
                        "ts": int(datetime.now().timestamp() * 1000),
                        "action": f"[AUTO IL] Revisão {row['revision']}: expire {expire_saved or '?'} → {expire_grid}",
                        "color": "#7c3aed",
                    })
                    try:
                        sb_patch("tickets", existing_t["id"], {
                            "expire": expire_grid,
                            "history": hist,
                        })
                        updated += 1
                        log.info(f"[IL] {base}: expire {expire_saved}→{expire_grid} (rev {row['revision']})")
                    except Exception as e:
                        log.error(f"[IL] {base}: erro patch — {e}")
                    continue

                # Ticket novo: insere só com dados da grid (sync_il pega utilities depois)
                # Não abre detalhe — evita ter que voltar pra grid (frágil) e o
                # sync_il via portal público já preenche utilities/respostas.
                detail = {}

                renewal_old = _find_renewal_candidate(row, existing)
                old_ticket_num = ""
                old_status_str = ""
                old_expire_str = ""
                if renewal_old:
                    old_ticket_num = renewal_old["ticket"]
                    old_status_str = (renewal_old.get("status") or "").strip()
                    old_expire_str = normalize_expire(renewal_old.get("expire") or "")
                    renewals += 1
                    log.info(f"[IL] {base}: renovação de {old_ticket_num} ({old_status_str}, exp {old_expire_str})")

                location_str = f"ILLINOIS - {(row.get('place') or '').strip()}"
                address_full = f"{(row.get('address_num') or '').strip()} {(row.get('street') or '').strip()}".strip()
                work_type = (detail.get("work_type") or "").strip() or "Service"

                history_entries = [{
                    "ts": int(datetime.now().timestamp() * 1000),
                    "action": f"[AUTO IL] Importado do JULIE — revisão {row['revision']}",
                    "color": "#10a574",
                }]
                if old_ticket_num:
                    history_entries.append({
                        "ts": int(datetime.now().timestamp() * 1000),
                        "action": f"[RENOVAÇÃO] {old_ticket_num} → {base} (graça até {old_expire_str or 'N/A'})",
                        "color": "#7c3aed",
                    })

                ticket_data = {
                    "ticket": base,
                    "company": "One Drill",
                    "state": "IL",
                    "location": location_str,
                    "address": address_full,
                    "status": "Open",
                    "expire": expire_grid,
                    "footage": 0,
                    "client": "",
                    "prime": "",
                    "tipo": work_type,
                    "job": "",
                    "notes": _build_il_notes(row, detail),
                    "pending": "",
                    "old_ticket2": old_ticket_num,
                    "status_old": old_status_str,
                    "expire_old": old_expire_str,
                    "county": (row.get("county") or "").strip(),
                    "history": history_entries,
                    "attachments": [],
                }
                try:
                    sb_insert("tickets", ticket_data)
                    existing_by_base[base] = ticket_data
                    existing.append(ticket_data)
                    inserted += 1
                    log.info(f"[IL] ✅ {base}: novo ticket inserido ({row.get('place', '')})")
                except Exception as e:
                    log.error(f"[IL] {base}: erro insert — {e}")

                # Não precisa voltar pra grid — não saímos dela (sem _il_open_detail).

        finally:
            await browser.close()

    log.info(f"[IL] === Import IL concluído: {inserted} novos, {updated} atualizados, "
             f"{renewals} renovações, {skipped} sem mudança ===")
    return inserted + updated


async def sync_and_import_il(triggered_by="manual"):
    """Import + Sync IL completo (equivalente a sync_and_import_wi)."""
    imported = await import_il(triggered_by)
    await sync_il(triggered_by)
    return imported


# ── │ SECTION: PDF_HELPERS │ Helpers pra geração de PDFs ─────────────────────

def _sanitize_folder(name):
    """Remove caracteres inválidos pra nome de pasta."""
    if not name or not name.strip():
        return "(Sem Nome)"
    return re.sub(r'[/\\:*?"<>|]', '-', name.strip()).rstrip('.')


def _build_renewal_groups(all_tickets):
    """Agrupa tickets por cadeia de renovação via old_ticket2.

    Retorna dict: ticket_id → {
        'folder': 'NUM1 - NUM2 - NUM3' ou None (avulso),
        'members': [ticket_numbers sorted]
    }
    """
    by_num = {}
    for t in all_tickets:
        num = (t.get('ticket') or '').strip()
        if num:
            by_num[num] = t

    # parent map: new_ticket → oldest predecessor
    parent = {}
    for t in all_tickets:
        num = (t.get('ticket') or '').strip()
        old_raw = (t.get('old_ticket2') or '').strip()
        if num and old_raw:
            # "A → B" format — pega o primeiro (mais antigo)
            parts = [p.strip() for p in old_raw.split('→') if p.strip()]
            if parts:
                parent[num] = parts[0]

    def find_root(num, visited=None):
        if visited is None:
            visited = set()
        if num in visited:
            return num
        visited.add(num)
        return find_root(parent[num], visited) if num in parent else num

    # Agrupar por raiz
    groups = {}
    for t in all_tickets:
        num = (t.get('ticket') or '').strip()
        if not num:
            continue
        root = find_root(num)
        groups.setdefault(root, set()).add(num)

    # Resultado
    result = {}
    for root, members in groups.items():
        sorted_members = sorted(members)
        folder = ' - '.join(sorted_members) if len(sorted_members) > 1 else None
        for num in sorted_members:
            if num in by_num:
                result[by_num[num]['id']] = {
                    'folder': folder,
                    'members': sorted_members
                }
    return result


def _pdf_disk_map(state):
    """Mapa {numero_do_arquivo (sem .pdf) -> path} de todos os PDFs em disco pra um estado.
    Usado pra validar se um ticket JÁ TEM PDF salvo (por qualquer nº da cadeia de renovação),
    em vez de confiar só no anexo 'ticket_pdf' do banco (que pode estar dessincronizado)."""
    d = {}
    for p in glob.glob(os.path.join(BASE_DIR, "pdfs", state, "**", "*.pdf"), recursive=True):
        d[os.path.splitext(os.path.basename(p))[0]] = p
    return d


def _ticket_has_pdf_on_disk(t, disk_map, min_bytes):
    """True se existe arquivo PDF (>min_bytes) pro nº atual OU algum nº da cadeia de renovação."""
    nums = [str(t.get("ticket") or "").strip()]
    nums += [x.strip() for x in (t.get("old_ticket2") or "").split(" → ") if x.strip()]
    for n in nums:
        p = disk_map.get(n)
        if p:
            try:
                if os.path.getsize(p) > min_bytes:
                    return True
            except OSError:
                pass
    return False


def _compute_pdf_paths(t, projects_map, renewal_groups, base_dir):
    """Computa paths do PDF principal e duplicata Damage.

    Retorna dict com: pdf_path, damage_path (ou None), pdf_filename, query_tnum, used_old
    """
    tid = t['id']
    tnum = (t.get('ticket') or '').strip()
    state = (t.get('state') or '?').upper()
    prime = _sanitize_folder(t.get('prime') or '(SEM PRIME)')

    proj_id = t.get('project_id')
    proj_name = '(Sem Projeto)'
    if proj_id and proj_id in projects_map:
        proj_name = projects_map[proj_id].get('name') or '(Sem Projeto)'
    proj_name = _sanitize_folder(proj_name)

    group = renewal_groups.get(tid, {})
    renewal_folder = group.get('folder')

    query_tnum, used_old = _pdf_query_number(t)
    pdf_filename = f"{query_tnum}.pdf" if used_old else f"{tnum}.pdf"

    if renewal_folder:
        pdf_dir = os.path.join(base_dir, "pdfs", state, prime, proj_name, _sanitize_folder(renewal_folder))
    else:
        pdf_dir = os.path.join(base_dir, "pdfs", state, prime, proj_name)
    pdf_path = os.path.join(pdf_dir, pdf_filename)

    damage_path = None
    if t.get('status') == 'Damage':
        if renewal_folder:
            dmg_dir = os.path.join(base_dir, "Damage", state, prime, proj_name, _sanitize_folder(renewal_folder))
        else:
            dmg_dir = os.path.join(base_dir, "Damage", state, prime, proj_name)
        damage_path = os.path.join(dmg_dir, pdf_filename)

    return {
        'pdf_path': pdf_path,
        'damage_path': damage_path,
        'pdf_filename': pdf_filename,
        'query_tnum': query_tnum,
        'used_old': used_old
    }


async def _il_pdf_go_back(page):
    """Volta pro grid de Ticket Search após ver Full Ticket de um ticket.

    Clica Exit até sair de todas as modais/views e voltar pro grid.
    Se o grid sumir, re-abre Search → County.
    """
    # Clica Exit repetidamente (Full Ticket Exit, depois Inquire Exit)
    for attempt in range(4):
        found_exit = False
        for sel in ['input[value="Exit"]', 'button:has-text("Exit")', 'a:has-text("Exit")']:
            loc = page.locator(sel).first
            try:
                if await loc.count() and await loc.is_visible():
                    await loc.click()
                    found_exit = True
                    await page.wait_for_timeout(1500)
                    break
            except Exception:
                continue
        if not found_exit:
            break
        await wait_stable(page)
        # Checa se já voltou pro grid
        try:
            body = await page.locator("body").inner_text()
            if "tickets for county" in body.lower() or "search for county" in body.lower():
                return True
        except Exception:
            pass

    # Grid não apareceu — tenta re-abrir
    try:
        body = await page.locator("body").inner_text()
        if "tickets for county" in body.lower():
            return True
    except Exception:
        pass

    log.info("[IL] PDF: grid sumiu, re-abrindo Search County...")
    try:
        if await _il_open_search_screen(page):
            if await _il_search_county(page, IL_SEARCH_COUNTY):
                return True
    except Exception as e:
        log.warning(f"[IL] PDF: re-search falhou: {e}")

    # Fallback: navega direto pro JULIE TE (session cookies persistem)
    log.info("[IL] PDF: fallback — navegando direto pra JULIE TE...")
    try:
        await page.goto(JULIE_TICKETENTRY_URL, wait_until="domcontentloaded")
        await page.wait_for_timeout(5000)
        await wait_stable(page)
        if await _il_open_search_screen(page):
            if await _il_search_county(page, IL_SEARCH_COUNTY):
                return True
    except Exception as e:
        log.warning(f"[IL] PDF: fallback JULIE TE falhou: {e}")
    return False


async def save_ticket_pdfs_il(force=False):
    """Salva PDF de tickets IL via JULIE Ticket Entry (headless).

    Fluxo por ticket:
      1. Login JULIE Ticket Entry → menu Search → County COOK → grid
      2. Dblclick ticket no grid → Inquire view
      3. Click "Full Ticket" → modal Full Ticket display
      4. Click "Print (large)" → abre janela nova (about:blank com texto completo)
      5. page.pdf() na janela de print → salva PDF
      6. Fecha popup → Exit Full Ticket → Exit Inquire → volta pro grid

    Headless — não precisa de pyautogui. Pode usar o PC normalmente.
    Estrutura: pdfs/IL/{PRIME}/{PROJECT}/{ticket}.pdf (ou subpasta de renovação).
    Damage duplicado em Damage/IL/{PRIME}/{PROJECT}/{ticket}.pdf.
    """
    if not IL_USER or not IL_PASS:
        log.error("[IL] PDF: IL_USER/IL_PASS não definidos no .env")
        return

    all_tickets = sb_get("tickets", "&state=eq.IL&status=in.(Clear,Damage,Completed)&order=ticket")
    if not all_tickets:
        log.info("[IL] PDF: nenhum ticket Clear/Damage/Completed")
        return

    # Busca projetos pra resolver nome pelo project_id
    projects = sb_get("projects", "&select=id,name") or []
    projects_map = {p['id']: p for p in projects}
    renewal_groups = _build_renewal_groups(all_tickets)

    if not force:
        # Valida se o ticket JÁ TEM arquivo PDF salvo (nº atual OU cadeia), não só o anexo do banco.
        _disk = _pdf_disk_map("IL")
        all_tickets = [t for t in all_tickets if not _ticket_has_pdf_on_disk(t, _disk, 5000)]

    if not all_tickets:
        log.info("[IL] PDF: todos os tickets Clear/Damage/Completed já têm PDF")
        return

    # Mapa ticket_base → dados Supabase pra match com grid
    need_pdf = {}
    for t in all_tickets:
        tnum = (t.get("ticket") or "").strip()
        if tnum:
            need_pdf[tnum] = t

    log.info("=" * 55)
    log.info(f"  SAVE-PDF IL (JULIE Ticket Entry): {len(need_pdf)} tickets")
    log.info("=" * 55)

    saved = 0
    errors = 0

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox"])
        page = await browser.new_page()
        page.set_default_timeout(60000)

        try:
            # ── Login ────────────────────────────────────────────────
            if not await _il_login(page):
                return

            # ── Search County → grid ─────────────────────────────────
            if not await _il_open_search_screen(page):
                return
            if not await _il_search_county(page, IL_SEARCH_COUNTY):
                return

            # ── Parse grid e match com nossos tickets ────────────────
            grid_rows = await _il_parse_grid(page)
            grid_ours = _filter_il_onedrill(grid_rows)
            grid_latest = _dedupe_latest_revision(grid_ours)
            log.info(f"[IL] PDF: {len(grid_latest)} tickets ONEDRILL no grid, "
                     f"{len(need_pdf)} precisam PDF")

            processed = 0
            for row in grid_latest:
                base = row["ticket_base"]
                if base not in need_pdf:
                    continue

                t = need_pdf[base]
                tnum = base
                tid = t.get("id", "")
                processed += 1

                # Computa paths via helper centralizado
                paths = _compute_pdf_paths(t, projects_map, renewal_groups, BASE_DIR)
                query_tnum = paths['query_tnum']
                used_old = paths['used_old']
                pdf_filename = paths['pdf_filename']
                full_path = os.path.abspath(paths['pdf_path'])

                if used_old:
                    log.info(f"  {tnum}: RENOVADO em grace — PDF nomeado como {query_tnum}")

                os.makedirs(os.path.dirname(full_path), exist_ok=True)

                if not force and os.path.exists(full_path):
                    sz = os.path.getsize(full_path)
                    if sz > 5000:
                        log.info(f"  {tnum}: PDF já existe ({round(sz/1024)}KB), pulando")
                        continue

                try:
                    log.info(f"  ({processed}/{len(need_pdf)}) {tnum}...")

                    # ── 1. Dblclick no ticket no grid → Inquire view ──
                    if not await _il_open_detail(page, row):
                        log.warning(f"  {tnum}: não encontrou no grid")
                        errors += 1
                        continue

                    # ── 2. Click "Full Ticket" ────────────────────────
                    ft_clicked = False
                    for sel in [
                        'input[value="Full Ticket"]',
                        'button:has-text("Full Ticket")',
                        'a:has-text("Full Ticket")',
                        'text="Full Ticket"',
                        'input[value*="Full"]',
                    ]:
                        loc = page.locator(sel).first
                        try:
                            if await loc.count():
                                try:
                                    await loc.click(timeout=5000)
                                except Exception:
                                    await loc.evaluate("el => el.click()")
                                ft_clicked = True
                                break
                        except Exception:
                            continue

                    if not ft_clicked:
                        log.warning(f"  {tnum}: botão 'Full Ticket' não encontrado")
                        await page.screenshot(
                            path=os.path.join(BASE_DIR, f"debug_il_pdf_no_ft_{tnum}.png"))
                        errors += 1
                        await _il_pdf_go_back(page)
                        continue

                    await page.wait_for_timeout(2500)
                    await wait_stable(page)

                    # ── 3. Click "Print (large)" → captura popup ──────
                    print_page = None
                    try:
                        async with page.expect_popup(timeout=10000) as popup_info:
                            pl_clicked = False
                            for sel in [
                                'input[value="Print (large)"]',
                                'button:has-text("Print (large)")',
                                'input[value*="Print"][value*="large" i]',
                            ]:
                                loc = page.locator(sel).first
                                try:
                                    if await loc.count():
                                        await loc.click()
                                        pl_clicked = True
                                        break
                                except Exception:
                                    continue

                            if not pl_clicked:
                                # Fallback: procura qualquer elemento com texto "Print" + "large"
                                loc = page.locator('text=/Print.*large/i').first
                                if await loc.count():
                                    await loc.click()
                                    pl_clicked = True
                                else:
                                    raise Exception("botão 'Print (large)' não encontrado")

                        print_page = await popup_info.value
                    except Exception as e:
                        log.warning(f"  {tnum}: popup Print (large) falhou: {e}")
                        await page.screenshot(
                            path=os.path.join(BASE_DIR, f"debug_il_pdf_no_print_{tnum}.png"))
                        errors += 1
                        await _il_pdf_go_back(page)
                        continue

                    # ── 4. Gera PDF da janela de print (about:blank) ──
                    await print_page.wait_for_load_state("domcontentloaded")
                    await print_page.wait_for_timeout(2000)

                    await print_page.emulate_media(media="screen")
                    await print_page.pdf(path=full_path, format="Letter",
                                         print_background=True)

                    # Fecha janela de print
                    try:
                        await print_page.close()
                    except Exception:
                        pass

                    # Fallback screenshot se PDF pequeno
                    pdf_ok = (os.path.exists(full_path)
                              and os.path.getsize(full_path) > 3000)
                    if not pdf_ok:
                        log.info(f"  {tnum}: page.pdf() insuficiente — fallback screenshot")
                        # Re-abre Print (large) pra tirar screenshot
                        try:
                            async with page.expect_popup(timeout=8000) as popup2:
                                for sel in ['input[value="Print (large)"]',
                                            'button:has-text("Print (large)")']:
                                    loc = page.locator(sel).first
                                    try:
                                        if await loc.count():
                                            await loc.click()
                                            break
                                    except Exception:
                                        continue
                            p2 = await popup2.value
                            await p2.wait_for_load_state("domcontentloaded")
                            await p2.wait_for_timeout(1500)
                            temp_png = full_path + ".tmp.png"
                            await p2.screenshot(path=temp_png, full_page=True)
                            await p2.close()
                            if os.path.exists(temp_png) and os.path.getsize(temp_png) > 1000:
                                from PIL import Image as _PILImage
                                img = _PILImage.open(temp_png)
                                if img.mode in ("RGBA", "P"):
                                    img = img.convert("RGB")
                                img.save(full_path, "PDF", resolution=100.0)
                                log.info(f"  {tnum}: screenshot→PDF OK")
                            try:
                                os.remove(temp_png)
                            except Exception:
                                pass
                        except Exception as ss_err:
                            log.warning(f"  {tnum}: screenshot fallback falhou: {ss_err}")

                    # ── 5. Valida + registra ──────────────────────────
                    if os.path.exists(full_path) and os.path.getsize(full_path) > 3000:
                        file_size = os.path.getsize(full_path)
                        log.info(f"  ✅ {tnum}: PDF salvo ({round(file_size/1024)}KB)")

                        # Duplica pra pasta Damage se aplicável
                        if paths['damage_path']:
                            dmg_full = os.path.abspath(paths['damage_path'])
                            os.makedirs(os.path.dirname(dmg_full), exist_ok=True)
                            shutil.copy2(full_path, dmg_full)
                            log.info(f"  {tnum}: cópia Damage salva")

                        attachments = t.get("attachments") or []
                        attachments = [a for a in attachments if a.get("type") != "ticket_pdf"]
                        att = {
                            "name": pdf_filename,
                            "type": "ticket_pdf",
                            "saved_at": datetime.now().isoformat(),
                            "size_kb": round(file_size / 1024, 1)
                        }
                        if used_old:
                            att["old_ticket"] = query_tnum
                            att["new_ticket"] = tnum
                        attachments.append(att)
                        hist = t.get("history") or []
                        action_txt = (f"📄 PDF salvo ({round(file_size/1024)}KB)"
                                      + (f" — usado # antigo {query_tnum}" if used_old else ""))
                        hist.append({
                            "ts": int(datetime.now().timestamp() * 1000),
                            "action": action_txt,
                            "color": "#7c3aed"
                        })
                        sb_patch("tickets", tid, {"attachments": attachments, "history": hist})
                        saved += 1
                    else:
                        sz = os.path.getsize(full_path) if os.path.exists(full_path) else 0
                        log.warning(f"  ⚠ {tnum}: PDF não salvo ou pequeno ({sz}B)")
                        errors += 1

                    # ── 6. Volta pro grid ─────────────────────────────
                    await _il_pdf_go_back(page)

                except Exception as e:
                    log.error(f"  ❌ {tnum}: {e}")
                    errors += 1
                    try:
                        await _il_pdf_go_back(page)
                    except Exception:
                        pass

        finally:
            await browser.close()

    log.info("=" * 55)
    log.info(f"  SAVE-PDF IL CONCLUÍDO: {saved} salvos, {errors} erros")
    log.info("=" * 55)


# ── │ SECTION: WI │ WISCONSIN (Diggers Hotline) — SCRAPE PÚBLICO ──────────────

async def _wait_for_diggers_element(page, selector, timeout_s=15):
    """Espera por um elemento (via seletor CSS) em qualquer frame da página.

    Retorna (frame, locator) ou (None, None) se timeout.
    Usado pra encontrar o input ou tabela de resultado em qualquer frame ativo.
    """
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        for frame in page.frames:
            try:
                loc = frame.locator(selector).first
                if await loc.count() > 0:
                    return frame, loc
            except Exception:
                continue
        await page.wait_for_timeout(500)
    return None, None


async def scrape_diggers_ticket(page, tnum, retry=True, debug_dump=False):
    """Scrape um ticket no Diggers Hotline (Wisconsin) — formato padrão.

    O portal é uma SPA ExtJS 4.1. Estrutura confirmada:
      - Carregamento inicial: main page (header + toolbar) + iframe MpWelcome
      - Após click em #findTicketsButton-btnEl: ExtJS REMOVE o iframe e renderiza
        a view "Find" DIRETO no main frame (não em iframe novo)
      - Input: <input name="ticket-number"> dentro do main DOM
      - Botão: <button> com texto "Search" (CSS faz lowercase visualmente)
      - Resultado: injetado em <div class="x-gc-mp-pnl-ticket-find-result"> no main

    Status WI conhecidos:
      - "No Response"                        → Pending (utility não respondeu)
      - "Not Participating"                  → Clear   (não atende a área)
      - "Ongoing - Working with Excavator"   → Pending (em andamento)
      - "Marked"/"Cleared"/"Completed"       → Clear   (resposta normal)
      - "Damage"/"Damaged"                   → Damage

    debug_dump: salva screenshot + HTML quando algo falha.
    """
    result = {"location_text": "", "responses": [], "expire_date": ""}

    # ── 1. Navega pro portal ──────────────────────────────────────────────────
    await page.goto(DIGGERS_URL, wait_until="domcontentloaded")
    await page.wait_for_timeout(2500)  # ExtJS demora pra inicializar
    await wait_stable(page)

    # ── 2. Espera ExtJS renderizar o botão Find Tickets ──────────────────────
    try:
        await page.wait_for_selector('#findTicketsButton-btnEl', timeout=15000, state="visible")
    except Exception:
        if debug_dump:
            try:
                base = BASE_DIR
                ss = os.path.join(base, f"debug_wi_{tnum}_no_button.png")
                await page.screenshot(path=ss, full_page=True)
                log.error(f"[WI] {tnum}: botão Find Tickets não apareceu — debug: {ss}")
            except Exception:
                pass
        if retry:
            log.warning(f"[WI] {tnum}: botão Find Tickets não apareceu — retry")
            return await scrape_diggers_ticket(page, tnum, retry=False, debug_dump=True)
        return result

    # ── 3. Clica no botão Find Tickets ───────────────────────────────────────
    try:
        await page.locator('#findTicketsButton-btnEl').click()
    except Exception:
        try:
            await page.evaluate("""() => {
                const btn = document.querySelector('#findTicketsButton-btnEl');
                if (btn) btn.click();
            }""")
        except Exception as e:
            log.error(f"[WI] {tnum}: erro clicando Find Tickets: {e}")
            return result

    await page.wait_for_timeout(1500)

    # ── 4. Encontra o input pelo seletor estável (name="ticket-number") ──────
    # ExtJS gera IDs dinâmicos, mas o atributo name é estável.
    target_frame, inp = await _wait_for_diggers_element(
        page, 'input[name="ticket-number"]', timeout_s=15
    )

    if not target_frame or not inp:
        if debug_dump:
            try:
                base = BASE_DIR
                ss_path = os.path.join(base, f"debug_wi_{tnum}_no_input.png")
                html_path = os.path.join(base, f"debug_wi_{tnum}_no_input.html")
                await page.screenshot(path=ss_path, full_page=True)
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(await page.content())
                frame_urls = [f.url for f in page.frames]
                log.error(f"[WI] {tnum}: input ticket-number não encontrado.")
                log.error(f"  Debug: {ss_path}")
                log.error(f"  HTML:  {html_path}")
                log.error(f"  Frames: {frame_urls}")
            except Exception as e:
                log.error(f"[WI] {tnum}: erro salvando debug: {e}")
        if retry:
            log.warning(f"[WI] {tnum}: input não apareceu — retry")
            return await scrape_diggers_ticket(page, tnum, retry=False, debug_dump=True)
        return result

    log.debug(f"[WI] {tnum}: input encontrado em {target_frame.url[:100]}")

    # ── 5. Digita o ticket no input ──────────────────────────────────────────
    try:
        await inp.click()
        await inp.fill("")
        await page.wait_for_timeout(200)
        await inp.fill(tnum)
    except Exception as e:
        log.error(f"[WI] {tnum}: erro digitando ticket: {e}")
        return result

    # ── 6. Clica no botão "Search" ───────────────────────────────────────────
    # Texto real é "Search" (S maiúsculo); CSS aplica text-transform pra lowercase.
    # Playwright :has-text() é case-sensitive, então procuramos "Search".
    btn = target_frame.locator('button:has-text("Search")').first
    try:
        if await btn.count():
            await btn.click()
        else:
            await inp.press("Enter")
    except Exception:
        try:
            await inp.press("Enter")
        except Exception:
            log.error(f"[WI] {tnum}: erro clicando Search")
            return result

    # ── 7. Espera resultado carregar ─────────────────────────────────────────
    await page.wait_for_timeout(4000)
    await wait_stable(page)

    # Resultado é injetado no painel de resultado do main frame (ou em iframe novo).
    # Procura "Positive Response" OU "TICKET #" pra confirmar que carregou.
    result_frame, _ = await _wait_for_diggers_element(
        page, 'text=/Positive Response/i', timeout_s=10
    )
    if not result_frame:
        result_frame, _ = await _wait_for_diggers_element(
            page, 'text=/TICKET #/i', timeout_s=5
        )
    if not result_frame:
        result_frame = target_frame

    try:
        body = await result_frame.locator("body").inner_text()
    except Exception:
        body = await page.locator("body").inner_text()

    # Verifica se ticket foi encontrado
    body_lower = body.lower()
    has_ticket = (
        tnum in body
        and ("ticket #" in body_lower or "positive response" in body_lower or "diggers hotline" in body_lower)
    )
    if not has_ticket:
        if retry:
            log.debug(f"[WI] {tnum}: não encontrado na 1ª tentativa — retry")
            return await scrape_diggers_ticket(page, tnum, retry=False)
        log.info(f"[WI] {tnum}: não encontrado no Diggers")
        return result

    # ── 8. Extrai location_text (Address / Place / County) ───────────────────
    location_parts = []
    addr_match = re.search(r"\bAddress\s*:\s*([^\n]+)", body)
    place_match = re.search(r"\bPlace\s*:\s*([^\n]+)", body)
    county_match = re.search(r"\bCounty\s*:\s*([^\n]+)", body)
    if addr_match:
        location_parts.append(addr_match.group(1).strip())
    if place_match:
        location_parts.append(place_match.group(1).strip())
    if county_match:
        location_parts.append(county_match.group(1).strip() + " County")
    if location_parts:
        result["location_text"] = " / ".join(location_parts)

    # ── 9. Parse tabela Positive Response via JS dentro do frame ────────────
    js_extract = """() => {
        const out = {responses: []};
        const tables = document.querySelectorAll('table');
        for (const table of tables) {
            const txt = (table.innerText || '').toLowerCase();
            if (!txt.includes('status') || !txt.includes('name')) continue;
            if (!txt.includes('facilities') && !txt.includes('phone') && !txt.includes('code')) continue;

            const rows = Array.from(table.querySelectorAll('tr'));
            let headers = [];
            let headerIdx = -1;
            for (let ri = 0; ri < rows.length; ri++) {
                const cells = rows[ri].querySelectorAll('th, td');
                const vals = Array.from(cells).map(c => (c.innerText || '').trim());
                const lower = vals.map(v => v.toLowerCase());
                if (lower.some(v => v === 'status') && lower.some(v => v === 'name')) {
                    headers = lower;
                    headerIdx = ri;
                    break;
                }
            }
            if (headerIdx < 0) continue;

            const idx = (key) => headers.findIndex(h => h.includes(key));
            const iStatus = idx('status'),
                  iCode   = idx('code'),
                  iName   = idx('name'),
                  iFac    = idx('facilities'),
                  iPhone  = idx('phone');

            for (let ri = headerIdx + 1; ri < rows.length; ri++) {
                const cells = rows[ri].querySelectorAll('td');
                const vals = Array.from(cells).map(c => (c.innerText || '').trim());
                if (vals.length < 3) continue;
                if (vals.every(v => !v)) continue;

                const get = (i) => (i >= 0 && i < vals.length) ? vals[i] : '';
                const status = get(iStatus);
                const code   = get(iCode);
                let name     = get(iName);
                let facilities = get(iFac);
                const phone  = get(iPhone);

                if (!name) continue;

                // CRÍTICO: a célula "name" pode conter múltiplas linhas (nome + bullets de eventos).
                // Pegamos só a 1ª linha não-vazia — o nome real da utility.
                // Mesmo tratamento pra status e facilities (que podem ter eventos).
                name       = name.split('\\n').map(s => s.trim()).filter(Boolean)[0] || '';
                facilities = facilities.split('\\n').map(s => s.trim()).filter(Boolean)[0] || '';

                // Procura TODOS os eventos "Mmm DD, YYYY H:MM AM/PM || comentário" e pega o MAIS RECENTE.
                // Múltiplos eventos podem aparecer na mesma célula (ex: TIME WARNER CABLE com 3 eventos).
                const fullText = vals.join('\\n');
                const eventRegex = /([A-Za-z]{3,9}\\s+\\d{1,2},\\s+\\d{4}\\s+\\d{1,2}:\\d{2}\\s*[AP]M)\\s*\\|\\|\\s*([^\\n|]+)/g;
                const matches = [...fullText.matchAll(eventRegex)];
                let respondedDate = null;
                let comment = '';
                if (matches.length > 0) {
                    let bestIdx = 0;
                    let bestDate = new Date(matches[0][1]);
                    for (let mi = 1; mi < matches.length; mi++) {
                        const d = new Date(matches[mi][1]);
                        if (!isNaN(d) && (isNaN(bestDate) || d > bestDate)) {
                            bestDate = d;
                            bestIdx = mi;
                        }
                    }
                    respondedDate = matches[bestIdx][1].trim();
                    comment = matches[bestIdx][2].trim();
                }

                out.responses.push({status, code, name, facilities, phone, comment, respondedDate});
            }
        }
        return out;
    }"""

    try:
        data = await result_frame.evaluate(js_extract)
    except Exception:
        data = await page.evaluate(js_extract)

    # ── 10. Processa respostas ──────────────────────────────────────────────
    for row in data.get("responses", []):
        name = (row.get("name") or "").strip()
        code = (row.get("code") or "").strip()
        status_raw = (row.get("status") or "").strip()
        facilities = (row.get("facilities") or "").strip()
        comment = (row.get("comment") or "").strip()
        rd_str = row.get("respondedDate")

        if not name:
            continue

        if name.lower() in ("name", "status", "code", "facilities", "phone"):
            continue

        # Remove sufixo de código do nome (3 padrões observados no Diggers):
        #   1) " - CODE"   → "TIME WARNER CABLE - TWC30"
        #   2) " (CODE)"   → "WISCONSIN DOT - ITS EQUIPMENT (ITS02)"
        #   3) " CODE"     → "RACINE WATER UTILITY RWU01" / "WINDSTREAM NRL02"
        if code:
            for pat in [
                re.compile(r'\s*-\s*' + re.escape(code) + r'\s*$'),
                re.compile(r'\s*\(' + re.escape(code) + r'\)\s*$'),
                re.compile(r'\s+' + re.escape(code) + r'\s*$'),
            ]:
                if pat.search(name):
                    name = pat.sub('', name).strip(' -')
                    break

        # Catch-all fallback: remove "(CODE)" ou " CODE" no fim, mesmo sem code definido.
        # CODE = 2-5 letras maiúsculas + 1-4 dígitos (típico Diggers: RWU01, TWC30, ITS02, NRL02, DOT02).
        name = re.sub(r"\s*\([A-Z]{2,5}\d{1,4}\)\s*$", "", name).strip()
        name = re.sub(r"\s+[A-Z]{2,5}\d{1,4}\s*$", "", name).strip(' -')

        if not _is_valid_utility_name(name):
            log.debug(f"[WI] {tnum}: nome inválido descartado: {name!r}")
            continue

        # Parse responded_date (formato Diggers: "May 04, 2026 9:21 AM")
        responded_date = None
        if rd_str:
            for fmt in [
                "%b %d, %Y %I:%M %p",
                "%B %d, %Y %I:%M %p",
                "%b %d, %Y %I:%M:%S %p",
                "%B %d, %Y %I:%M:%S %p",
            ]:
                try:
                    responded_date = datetime.strptime(rd_str.strip(), fmt).replace(tzinfo=None).isoformat()
                    break
                except ValueError:
                    continue

        # DISTINÇÃO: "Not Participating" (coluna Facilities) vs "Closed by DHL" genérico
        # Se Facilities diz "Not Participating" → utility realmente não participa → Clear
        # Se NÃO diz → prazo expirou sem resposta → Cancel (closed by DHL)
        fac_low = facilities.lower()
        is_not_part = "not participating" in fac_low or "not service provider" in fac_low
        if is_not_part:
            # Força classify a ver "not participating" no texto combinado
            cls_status, cls_unrec = classify(status_raw, "not participating " + comment + " " + facilities)
            response_text = "Not Participating"
        else:
            cls_status, cls_unrec = classify(status_raw, comment + " " + facilities)
            response_text = comment if comment else status_raw

        result["responses"].append({
            "utility": name,
            "status_raw": status_raw,
            "status": cls_status,
            "response": response_text,
            "comment": comment,
            "responded_date": responded_date,
            "_unrecognized": cls_unrec,
        })

    log.info(f"[WI] {tnum}: {len(result['responses'])} utilities")
    return result


async def scrape_wi(ticket_numbers, tickets_data=None):
    """Scrape batch de tickets WI no Diggers Hotline (público, sem login).

    Por enquanto roda SERIAL (1 aba) — portais legados costumam ter problemas
    de concorrência. Se ficar lento, dá pra aumentar pra NUM_TABS depois.
    """
    results = {}
    if not ticket_numbers:
        return results

    WI_NUM_TABS = 1  # serial por enquanto — sobe depois se funcionar

    chunks = [[] for _ in range(WI_NUM_TABS)]
    for i, tnum in enumerate(ticket_numbers):
        chunks[i % WI_NUM_TABS].append(tnum)
    chunks = [c for c in chunks if c]

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox"])

        async def process_chunk(chunk, tab_id):
            tab_results = {}
            page = await browser.new_page()
            page.set_default_timeout(30000)

            for idx, tnum in enumerate(chunk):
                log.info(f"[WI][T{tab_id}] ({idx+1}/{len(chunk)}) Ticket {tnum}")
                try:
                    debug = (idx < 2)
                    result = await scrape_diggers_ticket(page, tnum, debug_dump=debug)
                    tab_results[tnum] = result
                except Exception as e:
                    log.error(f"[WI][T{tab_id}] {tnum}: ERRO → {e}")
                    tab_results[tnum] = {"location_text": "", "responses": [], "expire_date": ""}

            try:
                await page.close()
            except Exception:
                pass
            return tab_results

        log.info(f"[WI] Abrindo {len(chunks)} aba(s) — modo {'serial' if WI_NUM_TABS == 1 else 'paralelo'}")
        chunk_results = await asyncio.gather(*[process_chunk(c, i) for i, c in enumerate(chunks)])
        for cr in chunk_results:
            results.update(cr)

        await browser.close()

    return results


async def sync_wi(triggered_by="manual"):
    """Sync completo Wisconsin — Diggers Hotline público (sem login)."""
    log.info(f"{'='*55}")
    log.info(f"  OneDrill 811 Sync  WI (Diggers Hotline)  [{triggered_by}]")
    log.info(f"{'='*55}")
    lid = log_start("WI", triggered_by)
    checked = 0
    summary = SyncSummary()
    try:
        all_tickets = sb_get("tickets", "&state=eq.WI&status=in.(Open,Damage,Clear)&order=ticket")
        if not all_tickets:
            log.info("[WI] Nenhum ticket ativo")
            log_finish(lid, 0, 0)
            return

        tickets_to_scrape, skipped = filter_tickets_for_sync(all_tickets, "WI")
        if skipped > 0:
            log.info(f"[WI] Cache: {skipped} tickets Clear pulados")

        checked = len(tickets_to_scrape)
        if not tickets_to_scrape:
            log.info(f"[WI] Nenhum ticket precisa verificação ({len(all_tickets)} ativos, {skipped} em cache)")
            log_finish(lid, 0, 0)
            return

        nums = [t["ticket"] for t in tickets_to_scrape]
        log.info(f"[WI] {checked} tickets para verificar (de {len(all_tickets)} ativos)")
        results = await scrape_wi(nums, tickets_data=tickets_to_scrape)

        summary = save_to_supabase("WI", results, all_tickets)
        log_finish(lid, checked, summary.responses_saved)
        log.info(f"[WI] CONCLUÍDO  {checked} verificados, {skipped} em cache | {summary}")

    except Exception as e:
        log.error(f"[WI] FALHOU: {e}")
        log_finish(lid, checked, summary.responses_saved, "error", str(e))


# ── │ SECTION: WI_IMPORT │ WISCONSIN — IMPORTAR TICKETS VIA PORTAL LOGADO ────
#
# O portal PÚBLICO (DIGGERS_URL) serve pra consultar respostas de tickets
# já conhecidos. Pro IMPORT de tickets novos, é preciso logar no portal
# cliente (DIGGERS_CLIENT_URL) e fazer "Excavator Search" por range de datas.
#
# Fluxo:
#   1. Login → https://geocall.diggershotline.com/geocall/client/login
#   2. Clicar "+" (novo search) → "Excavator search"
#   3. Trocar datas (from_date → to_date) e pesquisar
#   4. Lista de tickets retornada → conferir com Supabase
#   5. Clicar em cada ticket novo → scrape detalhes (endereço, expire, etc.)
#   6. Upsert no Supabase
#   7. Salvar data da última pesquisa pra próxima execução

WI_LAST_SEARCH_FILE = os.path.join(BASE_DIR, "wi_last_search.json")


def _get_wi_last_search_date():
    """Retorna última data de pesquisa WI (MM/DD/YYYY) ou '' se nunca rodou."""
    try:
        with open(WI_LAST_SEARCH_FILE, "r") as f:
            data = _json.load(f)
            return data.get("last_date", "")
    except (FileNotFoundError, _json.JSONDecodeError):
        return ""


def _set_wi_last_search_date(date_str):
    """Salva a data de última pesquisa WI."""
    with open(WI_LAST_SEARCH_FILE, "w", encoding="utf-8") as f:
        _json.dump({"last_date": date_str, "updated_at": datetime.now().isoformat()}, f)
    log.info(f"[WI] Última pesquisa salva: {date_str}")


async def _wi_debug_screenshot(page, label, tnum=""):
    """Salva screenshot de debug do portal WI (só se DEBUG_MODE ou primeiros tickets)."""
    try:
        tag = f"_{tnum}" if tnum else ""
        ss = os.path.join(BASE_DIR, f"debug_wi_import_{label}{tag}.png")
        await page.screenshot(path=ss, full_page=True)
        log.debug(f"[WI] Debug screenshot: {ss}")
        return ss
    except Exception as e:
        log.debug(f"[WI] Erro salvando screenshot: {e}")
        return ""


async def _login_diggers(page):
    """Login no portal cliente do Diggers Hotline (ExtJS 4.1).

    Tenta múltiplos seletores porque ExtJS gera IDs dinâmicos.
    Salva screenshot de debug se falhar.

    Returns: True se login OK, False caso contrário.
    """
    if not WI_USER or not WI_PASS:
        log.error("[WI] Credenciais WI_USER/WI_PASS não definidas no .env")
        return False

    log.info(f"[WI] Navegando para login: {DIGGERS_CLIENT_URL}")
    await page.goto(DIGGERS_CLIENT_URL, wait_until="domcontentloaded")
    await page.wait_for_timeout(4000)
    await wait_stable(page)

    # Se já logado (redirecionou pra home), pula login
    if "login" not in page.url.lower():
        log.info(f"[WI] Já logado — URL: {page.url[:80]}")
        return True

    # ── Encontrar campos de login ──
    # ExtJS: IDs dinâmicos, mas name/type são estáveis
    user_input = None
    for sel in [
        'input[name="j_username"]',
        'input[name="username"]',
        'input[name="user"]',
        'input[name="loginId"]',
        'input[placeholder*="user" i]',
        'input[placeholder*="User" i]',
        'input[placeholder*="email" i]',
    ]:
        loc = page.locator(sel).first
        try:
            if await loc.count():
                user_input = loc
                log.debug(f"[WI] Login: username field → {sel}")
                break
        except Exception:
            continue

    # Se não achou pelo name, tenta o 1º input text visível
    if not user_input:
        text_inputs = page.locator('input[type="text"]:visible')
        if await text_inputs.count():
            user_input = text_inputs.first
            log.debug("[WI] Login: username field → primeiro input[type=text] visível")

    # Também checa iframes (ExtJS pode carregar login em iframe)
    if not user_input:
        for frame in page.frames:
            if frame == page.main_frame:
                continue
            for sel in ['input[name="j_username"]', 'input[name="username"]', 'input[type="text"]']:
                try:
                    loc = frame.locator(sel).first
                    if await loc.count():
                        user_input = loc
                        log.debug(f"[WI] Login: username em iframe {frame.url[:60]}")
                        break
                except Exception:
                    continue
            if user_input:
                break

    pass_input = None
    for sel in [
        'input[name="j_password"]',
        'input[name="password"]',
        'input[type="password"]',
    ]:
        loc = page.locator(sel).first
        try:
            if await loc.count():
                pass_input = loc
                log.debug(f"[WI] Login: password field → {sel}")
                break
        except Exception:
            continue

    if not pass_input:
        for frame in page.frames:
            if frame == page.main_frame:
                continue
            loc = frame.locator('input[type="password"]').first
            try:
                if await loc.count():
                    pass_input = loc
                    break
            except Exception:
                continue

    if not user_input or not pass_input:
        ss = await _wi_debug_screenshot(page, "login_no_fields")
        log.error(f"[WI] Login: campos username/password não encontrados — debug: {ss}")
        # Dump HTML pra análise
        try:
            html_path = os.path.join(BASE_DIR, "debug_wi_login_page.html")
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(await page.content())
            log.error(f"[WI] HTML salvo: {html_path}")
        except Exception:
            pass
        return False

    # ── Preencher e submeter ──
    try:
        await user_input.click()
        await user_input.fill("")
        await page.wait_for_timeout(150)
        await user_input.fill(WI_USER)
        await page.wait_for_timeout(200)

        await pass_input.click()
        await pass_input.fill("")
        await page.wait_for_timeout(150)
        await pass_input.fill(WI_PASS)
        await page.wait_for_timeout(200)
    except Exception as e:
        log.error(f"[WI] Login: erro preenchendo campos: {e}")
        return False

    # ── Clicar botão de login ──
    clicked = False
    for sel in [
        'button:has-text("Log In")',
        'button:has-text("Login")',
        'button:has-text("Sign In")',
        'input[type="submit"]',
        'button[type="submit"]',
        '#loginButton',
        'a:has-text("Log In")',
        '.x-btn:has-text("Log")',
    ]:
        try:
            loc = page.locator(sel).first
            if await loc.count():
                await loc.click()
                clicked = True
                log.debug(f"[WI] Login: clicou → {sel}")
                break
        except Exception:
            continue

    if not clicked:
        # Fallback: Enter no password
        try:
            await pass_input.press("Enter")
            log.debug("[WI] Login: fallback Enter no password")
        except Exception:
            pass

    await page.wait_for_timeout(5000)
    await wait_stable(page)

    # ── Verifica sucesso ──
    # Se URL ainda tem "login", pode ter falhado
    url_lower = page.url.lower()
    if "login" in url_lower and "client" in url_lower:
        # Checa se há mensagem de erro visível
        try:
            body_text = await page.locator("body").inner_text()
            if any(kw in body_text.lower() for kw in ["invalid", "incorrect", "failed", "error", "wrong"]):
                ss = await _wi_debug_screenshot(page, "login_failed")
                log.error(f"[WI] Login falhou (credenciais inválidas?) — debug: {ss}")
                return False
        except Exception:
            pass
        # Pode ser que o login redirecionou mas URL não mudou — espera mais
        await page.wait_for_timeout(3000)
        if "login" in page.url.lower():
            ss = await _wi_debug_screenshot(page, "login_stuck")
            log.error(f"[WI] Login: ainda na página de login após 8s — debug: {ss}")
            return False

    log.info(f"[WI] Login OK — URL: {page.url[:100]}")
    return True


async def _search_diggers_excavator(page, from_date, to_date, debug=True):
    """Pesquisa tickets no Diggers Hotline por range de data (Excavator Search).

    from_date/to_date: formato "MM/DD/YYYY"
    Retorna: list of dicts [{ticket, ticket_type, ...}, ...] com tickets encontrados.

    Fluxo confirmado no portal (ExtJS 4.1 — sidebar navigation):
      1. Clicar "Excavator Search" no sidebar esquerdo (abaixo de "Search")
      2. Preencher date range nos campos do form
      3. Clicar botão Search
      4. Ler grid de resultados (cada row = 1 ticket)
      5. Paginar se necessário
    """
    found_tickets = []

    if debug:
        await _wi_debug_screenshot(page, "before_search")

    # ── 1. Clicar "Excavator Search" no sidebar esquerdo ──
    # Layout do sidebar confirmado:
    #   Search (header)
    #     ├── Ticket
    #     ├── Excavator Search    ← este
    #     └── Excavator By Number
    exc_clicked = False

    # O sidebar tem "Search" com um botão [+] pra expandir os sub-itens.
    # Primeiro: expandir "Search" clicando no [+] ou no texto "Search"
    # Depois: clicar em "Excavator Search" que aparece embaixo.

    # Passo 1: Expandir "Search" no sidebar
    expanded = False
    # Tenta clicar no [+] icon ao lado de "Search"
    for sel in [
        # ExtJS tree: o expand icon é um <img> ou <span> com classe de expand
        'img[class*="plus"]:near(:text("Search"))',
        'span[class*="plus"]:near(:text("Search"))',
        # Ou o proprio texto "Search" pode expandir ao clicar
    ]:
        try:
            loc = page.locator(sel).first
            if await loc.count() and await loc.is_visible():
                await loc.click()
                expanded = True
                log.debug(f"[WI] Search: expandiu via {sel}")
                break
        except Exception:
            continue

    if not expanded:
        # Tenta JS: encontra elemento "Search" no sidebar e clica pra expandir
        try:
            await page.evaluate("""() => {
                // Procura todos os elementos com texto "Search"
                const walker = document.createTreeWalker(
                    document.body, NodeFilter.SHOW_TEXT, null, false
                );
                while (walker.nextNode()) {
                    const text = walker.currentNode.textContent.trim();
                    if (text === 'Search') {
                        const el = walker.currentNode.parentElement;
                        if (!el) continue;
                        // Procura o [+] icon no mesmo container ou irmao
                        const parent = el.closest('div, tr, li, span') || el.parentElement;
                        if (!parent) continue;
                        // Clica no [+] icon se existir
                        const plus = parent.querySelector('img[class*="plus"], img[class*="expand"], .x-tree-ec-icon, .x-tool-expand');
                        if (plus) {
                            plus.click();
                            return 'plus';
                        }
                        // Senao, clica no proprio texto
                        el.click();
                        return 'text';
                    }
                }
                return false;
            }""")
            expanded = True
            log.debug("[WI] Search: expandiu via JS")
        except Exception as e:
            log.warning(f"[WI] Search: erro expandindo sidebar: {e}")

    if not expanded:
        # Fallback direto: tenta clicar em "Search" pelo texto
        for sel in [
            'text="Search"',
            'span:has-text("Search")',
            'a:has-text("Search")',
            'div:text-is("Search")',
        ]:
            try:
                loc = page.locator(sel).first
                if await loc.count() and await loc.is_visible():
                    await loc.click()
                    expanded = True
                    log.debug(f"[WI] Search: clicou texto Search -> {sel}")
                    break
            except Exception:
                continue

    await page.wait_for_timeout(2000)

    if debug:
        await _wi_debug_screenshot(page, "after_expand_search")

    # Passo 2: Clicar em "Excavator Search" (agora visivel apos expandir)
    for sel in [
        'a:has-text("Excavator Search")',
        'span:has-text("Excavator Search")',
        'div:has-text("Excavator Search")',
        '.x-tree-node-text:has-text("Excavator Search")',
        'text="Excavator Search"',
    ]:
        try:
            loc = page.locator(sel).first
            if await loc.count() and await loc.is_visible():
                await loc.click()
                exc_clicked = True
                log.debug(f"[WI] Search: clicou 'Excavator Search' -> {sel}")
                break
        except Exception:
            continue

    # JS fallback pra "Excavator Search"
    if not exc_clicked:
        try:
            result = await page.evaluate("""() => {
                const walker = document.createTreeWalker(
                    document.body, NodeFilter.SHOW_TEXT, null, false
                );
                while (walker.nextNode()) {
                    const text = walker.currentNode.textContent.trim();
                    if (text === 'Excavator Search') {
                        const el = walker.currentNode.parentElement;
                        if (el) { el.click(); return true; }
                    }
                }
                return false;
            }""")
            if result:
                exc_clicked = True
                log.debug("[WI] Search: 'Excavator Search' via JS walker")
        except Exception:
            pass

    if not exc_clicked:
        ss = await _wi_debug_screenshot(page, "no_excavator_search")
        log.error(f"[WI] Search: link 'Excavator Search' nao encontrado — debug: {ss}")
        return found_tickets

    await page.wait_for_timeout(3000)
    await wait_stable(page)

    if debug:
        await _wi_debug_screenshot(page, "search_form_loaded")

    # ── 2. Preencher date range ──
    # O form de Excavator Search tem campos de data (from/to)
    # ExtJS date fields: input com name contendo "date" ou "Date"
    date_filled = False

    # Estratégia A: procurar por name/class de date fields
    date_inputs = []
    for sel in [
        'input[name*="date" i]',
        'input[name*="Date"]',
        'input[name*="from" i]',
        'input[name*="start" i]',
        'input[name*="begin" i]',
        'input.x-form-date-field',
    ]:
        try:
            locs = page.locator(sel)
            count = await locs.count()
            for i in range(count):
                el = locs.nth(i)
                if await el.is_visible():
                    date_inputs.append(el)
        except Exception:
            continue

    # Estratégia B: inputs que já contêm data ou que parecem date fields
    if len(date_inputs) < 2:
        try:
            all_inputs = page.locator('input[type="text"]:visible, input:not([type]):visible')
            count = await all_inputs.count()
            for i in range(count):
                el = all_inputs.nth(i)
                try:
                    name = await el.get_attribute("name") or ""
                    value = await el.input_value() or ""
                    cls = await el.get_attribute("class") or ""
                    if ("date" in name.lower() or "date" in cls.lower()
                            or re.match(r"\d{1,2}/\d{1,2}/\d{2,4}", value)):
                        if el not in date_inputs:
                            date_inputs.append(el)
                except Exception:
                    continue
        except Exception:
            pass

    # Dedup (pode ter pego o mesmo campo 2x)
    if len(date_inputs) > 2:
        # Pega só os 2 primeiros únicos
        unique = []
        seen_ids = set()
        for el in date_inputs:
            try:
                el_id = await el.get_attribute("id") or ""
                if el_id and el_id in seen_ids:
                    continue
                seen_ids.add(el_id)
                unique.append(el)
            except Exception:
                unique.append(el)
        date_inputs = unique[:2]

    if len(date_inputs) >= 2:
        try:
            # From date
            await date_inputs[0].click(click_count=3)
            await page.wait_for_timeout(100)
            await date_inputs[0].fill(from_date)
            await date_inputs[0].press("Tab")
            await page.wait_for_timeout(500)

            # To date
            await date_inputs[1].click(click_count=3)
            await page.wait_for_timeout(100)
            await date_inputs[1].fill(to_date)
            await date_inputs[1].press("Tab")
            await page.wait_for_timeout(500)

            date_filled = True
            log.info(f"[WI] Search: datas preenchidas: {from_date} -> {to_date}")
        except Exception as e:
            log.warning(f"[WI] Search: erro preenchendo datas: {e}")

    if not date_filled:
        # JS fallback: encontra campos de data pelo valor ou name e seta
        try:
            await page.evaluate(f"""() => {{
                const inputs = document.querySelectorAll('input');
                const dateFields = [];
                for (const inp of inputs) {{
                    const n = (inp.name || '').toLowerCase();
                    const v = inp.value || '';
                    const cls = (inp.className || '').toLowerCase();
                    if (n.includes('date') || cls.includes('date')
                        || /\\d{{1,2}}\\/\\d{{1,2}}\\/\\d{{2,4}}/.test(v)) {{
                        dateFields.push(inp);
                    }}
                }}
                if (dateFields.length >= 2) {{
                    const set = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value'
                    ).set;
                    set.call(dateFields[0], '{from_date}');
                    dateFields[0].dispatchEvent(new Event('input', {{ bubbles: true }}));
                    dateFields[0].dispatchEvent(new Event('change', {{ bubbles: true }}));
                    set.call(dateFields[1], '{to_date}');
                    dateFields[1].dispatchEvent(new Event('input', {{ bubbles: true }}));
                    dateFields[1].dispatchEvent(new Event('change', {{ bubbles: true }}));
                    return true;
                }}
                return false;
            }}""")
            date_filled = True
            log.info(f"[WI] Search: datas via JS: {from_date} -> {to_date}")
        except Exception as e:
            ss = await _wi_debug_screenshot(page, "date_fill_failed")
            log.error(f"[WI] Search: datas nao preenchidas — debug: {ss}, erro: {e}")
            return found_tickets

    await page.wait_for_timeout(500)

    # ── 3. Clicar OK/Search (dialog usa "Ok", nao "Search") ──
    search_clicked = False
    for sel in [
        'button:has-text("Ok")',
        'a:has-text("Ok")',
        '.x-btn:has-text("Ok")',
        'button:has-text("OK")',
        'button:has-text("Search")',
        'a:has-text("Search")',
        '.x-btn:has-text("Search")',
        'input[value="Ok" i]',
        'input[value="Search" i]',
        'button[type="submit"]',
    ]:
        try:
            locs = page.locator(sel)
            count = await locs.count()
            for i in range(count):
                el = locs.nth(i)
                if await el.is_visible():
                    await el.click()
                    search_clicked = True
                    log.debug(f"[WI] Search: clicou Search -> {sel}")
                    break
            if search_clicked:
                break
        except Exception:
            continue

    if not search_clicked:
        ss = await _wi_debug_screenshot(page, "no_search_button")
        log.error(f"[WI] Search: botao 'Search' nao encontrado — debug: {ss}")
        return found_tickets

    # ── 4. Esperar resultados ──
    log.info("[WI] Search: aguardando resultados...")
    await page.wait_for_timeout(8000)
    await wait_stable(page)

    if debug:
        await _wi_debug_screenshot(page, "search_results")

    # ── 5. Ler grid de resultados ──
    found_tickets = await _parse_wi_search_results(page)

    if not found_tickets:
        log.info("[WI] Search: nenhum resultado encontrado")
        try:
            body = await page.locator("body").inner_text()
            if "no results" in body.lower() or "no records" in body.lower() or "0 ticket" in body.lower():
                log.info("[WI] Search: portal confirma 0 resultados")
        except Exception:
            pass
    else:
        log.info(f"[WI] Search: {len(found_tickets)} tickets encontrados")

    # ── 6. Paginacao (só se achou resultados) ──
    MAX_SEARCH_PAGES = 200          # safety net
    page_num = 1
    seen_tickets = {t["ticket"] for t in found_tickets}

    while found_tickets and page_num < MAX_SEARCH_PAGES:
        # Detecta botão Next — ExtJS usa classes CSS (não attr disabled)
        next_clicked = False
        for sel in [
            'button:has-text("Next")',
            '.x-tbar-page-next',
            'a:has-text("Next")',
            'button[data-qtip="Next Page"]',
        ]:
            try:
                loc = page.locator(sel).first
                if await loc.count() and await loc.is_visible():
                    # ExtJS disabled check: classes "x-btn-disabled" / "x-item-disabled"
                    cls = await loc.get_attribute("class") or ""
                    if "disabled" in cls.lower():
                        continue  # botão existe mas está desabilitado
                    # Checa parent tbm (ExtJS coloca disabled no wrapper <a> do botão)
                    try:
                        parent_cls = await loc.locator("..").get_attribute("class") or ""
                        if "disabled" in parent_cls.lower():
                            continue
                    except Exception:
                        pass
                    await loc.click()
                    next_clicked = True
                    break
            except Exception:
                continue

        if not next_clicked:
            break

        page_num += 1
        log.info(f"[WI] Search: pagina {page_num}...")
        await page.wait_for_timeout(4000)
        await wait_stable(page)

        page_results = await _parse_wi_search_results(page)
        if not page_results:
            break

        # Dedup: se TODOS os tickets desta página já foram vistos → parou de paginar de verdade
        new_on_page = [t for t in page_results if t["ticket"] not in seen_tickets]
        if not new_on_page:
            log.info(f"[WI] Search: pagina {page_num} retornou só duplicatas — fim da paginação")
            break

        for t in new_on_page:
            seen_tickets.add(t["ticket"])
            found_tickets.append(t)

    log.info(f"[WI] Search total: {len(found_tickets)} tickets em {page_num} pagina(s)")
    return found_tickets


async def _parse_wi_search_results(page):
    """Extrai dados ricos da grid de resultados do Excavator Search.

    Colunas confirmadas (screenshot):
      Number | Date Time Ticket | Start Date Time | Company Name |
      Company Phone | Caller Name | County | Place (Municipality) | Address | Street

    Retorna: list of dicts com campos parseados por ticket.
    """
    results = []

    # ── Estratégia 1: ExtJS 3.4 Store API + DOM (.x-grid3-row) ──
    # Portal usa ExtJS 3.4.0: sem ComponentQuery, sem StoreManager.
    # Grid DOM: .x-grid3-row (cada row é uma <table> separada dentro de um <div>).
    js_extjs = r"""() => {
        const out = {rows: [], colMap: {}, allFields: [], method: '', debug: ''};
        try {
            if (typeof Ext === 'undefined') { out.debug = 'no-Ext'; return out; }
            try { out.debug += 'v=' + (Ext.version || '?') + ' '; } catch(e) {}

            // ── Store API: Ext.StoreMgr (ExtJS 3.x) ──
            var bestStore = null, bestCount = 0;
            if (Ext.StoreMgr && Ext.StoreMgr.each) {
                try {
                    var stores = [];
                    Ext.StoreMgr.each(function(s) { stores.push(s); });
                    out.debug += 'stores=' + stores.length + ' ';
                    for (var si = 0; si < stores.length; si++) {
                        try {
                            var c = stores[si].getCount();
                            if (c > bestCount) { bestCount = c; bestStore = stores[si]; }
                        } catch(e) {}
                    }
                } catch(e) { out.debug += 'sm-err=' + e.message + ' '; }
            }

            // Fallback: Ext.ComponentMgr.all (Ext 3.x)
            if (!bestStore && Ext.ComponentMgr && Ext.ComponentMgr.all) {
                try {
                    var items = Ext.ComponentMgr.all.items || [];
                    out.debug += 'comps=' + items.length + ' ';
                    for (var ci = 0; ci < items.length; ci++) {
                        try {
                            var comp = items[ci];
                            if (comp.getStore) {
                                var store = comp.getStore();
                                if (store && store.getCount) {
                                    var cnt = store.getCount();
                                    if (cnt > bestCount) { bestCount = cnt; bestStore = store; }
                                }
                            }
                        } catch(e) {}
                    }
                } catch(e) { out.debug += 'cm-err=' + e.message + ' '; }
            }

            // Se achou store com dados, extrai registros
            if (bestStore && bestCount > 0) {
                out.debug += 'bestStore=' + bestCount + ' ';
                // Column mapping via grid.colModel (ExtJS 3.x)
                try {
                    if (Ext.ComponentMgr && Ext.ComponentMgr.all) {
                        var cmpItems = Ext.ComponentMgr.all.items || [];
                        for (var gi = 0; gi < cmpItems.length; gi++) {
                            try {
                                var g = cmpItems[gi];
                                if (g.getStore && g.getStore() === bestStore && g.colModel) {
                                    var cm = g.colModel;
                                    var cc = cm.getColumnCount ? cm.getColumnCount() : (cm.config ? cm.config.length : 0);
                                    for (var i = 0; i < cc; i++) {
                                        try {
                                            var hdr = cm.getColumnHeader ? cm.getColumnHeader(i) : (cm.config[i].header || '');
                                            var di = cm.getDataIndex ? cm.getDataIndex(i) : (cm.config[i].dataIndex || '');
                                            if (hdr && di) out.colMap[hdr] = di;
                                        } catch(e) {}
                                    }
                                    break;
                                }
                            } catch(e) {}
                        }
                    }
                } catch(e) {}
                out.debug += 'cols=' + Object.keys(out.colMap).length + ' ';

                // Extrair registros do store
                try {
                    var records = bestStore.getRange();
                    if (records.length > 0) out.allFields = Object.keys(records[0].data);
                    for (var ri = 0; ri < records.length; ri++) {
                        var data = {};
                        var rdata = records[ri].data;
                        for (var key in rdata) {
                            if (rdata.hasOwnProperty(key)) {
                                var v = rdata[key];
                                data[key] = (v != null && v !== undefined) ? String(v) : '';
                            }
                        }
                        out.rows.push(data);
                    }
                    out.method = 'extjs3-store';
                    out.debug += 'extracted=' + out.rows.length + ' ';
                } catch(e) { out.debug += 'extract-err=' + e.message + ' '; }
                return out;
            }
            out.debug += 'no-store-data ';
        } catch(e) { out.debug += 'api-err=' + e.message + ' '; }

        // ── Estratégia 2: DOM parsing ExtJS 3.x (.x-grid3-row) ──
        try {
            var gridRows = document.querySelectorAll('.x-grid3-row');
            out.debug += 'grid3rows=' + gridRows.length + ' ';

            // Headers: .x-grid3-hd-inner
            var hdSels = ['.x-grid3-hd-inner', '.x-grid3-hd span'];
            for (var hi = 0; hi < hdSels.length; hi++) {
                var hds = document.querySelectorAll(hdSels[hi]);
                if (hds.length >= 5) {
                    for (var hj = 0; hj < hds.length; hj++) {
                        var txt = (hds[hj].innerText || '').trim();
                        if (txt) out.colMap[txt] = String(hj);
                    }
                    out.debug += 'hdr=' + hdSels[hi] + '(' + hds.length + ') ';
                    break;
                }
            }

            // Cada .x-grid3-row tem .x-grid3-cell-inner com os valores
            for (var ri = 0; ri < gridRows.length; ri++) {
                var cellDivs = gridRows[ri].querySelectorAll('.x-grid3-cell-inner');
                var cells = [];
                if (cellDivs.length >= 3) {
                    for (var ci = 0; ci < cellDivs.length; ci++) {
                        cells.push((cellDivs[ci].innerText || '').trim());
                    }
                } else {
                    // Fallback: td dentro da table do row
                    var tds = gridRows[ri].querySelectorAll('td');
                    for (var ti = 0; ti < tds.length; ti++) {
                        cells.push((tds[ti].innerText || '').trim());
                    }
                }
                if (cells.length < 3) continue;

                var hasTicket = false;
                for (var vi = 0; vi < cells.length; vi++) {
                    if (/^\d{8,12}$/.test(cells[vi])) { hasTicket = true; break; }
                }
                if (!hasTicket) continue;

                var rowData = {};
                for (var di = 0; di < cells.length; di++) {
                    rowData[String(di)] = cells[di];
                }
                out.rows.push(rowData);
            }

            if (out.rows.length > 0) {
                out.method = 'extjs3-dom';
                out.debug += 'dom-extracted=' + out.rows.length + ' ';
            }
        } catch(e) { out.debug += 'dom-err=' + e.message + ' '; }

        return out;
    }"""

    extjs_data = {}
    try:
        extjs_data = await page.evaluate(js_extjs)
    except Exception as e:
        log.debug(f"[WI] ExtJS Store API: erro: {e}")

    method = extjs_data.get("method", "")
    debug_info = extjs_data.get("debug", "")
    col_map = extjs_data.get("colMap", {})
    all_fields = extjs_data.get("allFields", [])
    store_rows = extjs_data.get("rows", [])

    log.debug(f"[WI] Grid parse: method={method} debug={debug_info}")
    if col_map:
        log.debug(f"[WI] Grid col map: {col_map}")
    if all_fields:
        log.debug(f"[WI] Store fields: {all_fields}")
    if store_rows and store_rows[0]:
        log.debug(f"[WI] Store sample row: { {k: v for k, v in list(store_rows[0].items())[:6]} }")
    if not store_rows:
        log.debug(f"[WI] Store rows: 0 — ExtJS API retornou vazio")

    # Parse registros do ExtJS Store
    if store_rows and method:
        def _find_field(keywords):
            """Acha field name por keywords — tenta header da coluna e depois nome do campo."""
            for text, field in col_map.items():
                if any(k in text.lower() for k in keywords):
                    return field
            for f in all_fields:
                fl = f.lower()
                if any(k in fl for k in keywords):
                    return f
            return None

        f_number   = _find_field(["number", "ticketnumber", "ticket_number", "tktnum"])
        f_start    = _find_field(["start date", "start_date", "startdate"])
        f_county   = _find_field(["county"])
        f_place    = _find_field(["place", "municipality"])
        f_address  = _find_field(["address"])
        f_street   = _find_field(["street"])
        f_company  = _find_field(["company name", "company_name", "companyname"])
        f_caller   = _find_field(["caller"])
        f_type     = _find_field(["ticket_type", "tickettype", "tkttype", "type_name"])
        f_marking  = _find_field(["marking", "markinginstructions"])
        f_remarks  = _find_field(["remarks", "remark"])
        f_intersec = _find_field(["intersection", "workintersection"])

        log.debug(f"[WI] Field map: number={f_number} start={f_start} county={f_county} "
                  f"place={f_place} addr={f_address} street={f_street} type={f_type}")

        for row_data in store_rows:
            ticket = ""
            if f_number:
                val = row_data.get(f_number, "").strip()
                if re.match(r"^\d{6,14}$", val):
                    ticket = val
            if not ticket:
                for key, val in row_data.items():
                    v = val.strip()
                    if re.match(r"^\d{8,12}$", v):
                        ticket = v
                        break
            if not ticket:
                continue

            item = {
                "ticket": ticket,
                "start_date": row_data.get(f_start, "") if f_start else "",
                "county":     row_data.get(f_county, "") if f_county else "",
                "city":       row_data.get(f_place, "") if f_place else "",
                "address_num": row_data.get(f_address, "") if f_address else "",
                "street":     row_data.get(f_street, "") if f_street else "",
                "caller":     row_data.get(f_caller, "") if f_caller else "",
                "cells":      list(row_data.values()),
            }
            if f_type:
                item["ticket_type"] = row_data.get(f_type, "")
            if f_marking:
                item["marking_instructions"] = row_data.get(f_marking, "")
            if f_remarks:
                item["remarks"] = row_data.get(f_remarks, "")
            if f_intersec:
                item["intersection"] = row_data.get(f_intersec, "")
            results.append(item)

        # Dedup e retorna se ExtJS Store deu certo
        seen = set()
        deduped = []
        for r in results:
            if r["ticket"] not in seen:
                seen.add(r["ticket"])
                deduped.append(r)
        log.debug(f"[WI] Grid parse via {method}: {len(deduped)} tickets extraídos")
        return deduped

    # ── Estratégia 2 (fallback): DOM parsing ──
    log.debug("[WI] ExtJS Store API sem resultados — fallback DOM parsing")
    js_extract = r"""() => {
        const out = {headers: [], rows: [], debug: ''};

        // ── Headers: tenta seletores do mais específico ao mais genérico ──
        const headerSels = [
            '.x-grid3-hd-inner',
            '.x-column-header-text-inner',
            '.x-column-header-text',
            '.x-column-header .x-column-header-inner',
            '.x-grid-hd-text',
            'th',
        ];
        for (const sel of headerSels) {
            const cells = document.querySelectorAll(sel);
            if (cells.length >= 5) {
                out.headers = Array.from(cells).map(h => (h.innerText || '').trim()).filter(Boolean);
                out.debug += `hdr=${sel}(${cells.length}) `;
                break;
            }
        }

        // ── Rows: ExtJS 3.x primeiro, depois 4.x ──
        // ExtJS 3.x: cada .x-grid3-row é um div com table interna
        const grid3Rows = document.querySelectorAll('.x-grid3-row');
        if (grid3Rows.length >= 1) {
            // ExtJS 3.x path: extrai cells de cada .x-grid3-row
            for (const gRow of grid3Rows) {
                let cells = gRow.querySelectorAll('.x-grid3-cell-inner');
                if (cells.length < 3) cells = gRow.querySelectorAll('td');
                const vals = Array.from(cells).map(c => (c.innerText || '').trim());
                if (vals.length < 3) continue;
                let ticket = '';
                for (const v of vals) {
                    if (/^\d{8,12}$/.test(v)) { ticket = v; break; }
                }
                if (ticket) out.rows.push(vals);
            }
            out.debug += 'grid3=' + grid3Rows.length + '/' + out.rows.length + ' ';
            if (out.rows.length > 0) return out;
        }

        const rowSels = [
            '.x-grid-table .x-grid-row',
            '.x-grid-table tr',
            '.x-grid-view table tr',
            '.x-grid-body table tr',
            'table.x-grid-table tr',
        ];
        let rows = [];
        let winSel = '';
        for (const sel of rowSels) {
            try {
                const found = document.querySelectorAll(sel);
                if (found.length > rows.length) { rows = found; winSel = sel; }
            } catch(e) {}
        }

        // Fallback AMPLO: percorre TODAS as tables, acha a que tem mais rows com ticket numbers
        if (rows.length < 3) {
            const tables = document.querySelectorAll('table');
            let bestTable = null, bestCount = 0;
            for (const table of tables) {
                const trs = table.querySelectorAll('tr');
                let ticketCount = 0;
                for (const tr of trs) {
                    const tds = tr.querySelectorAll('td, .x-grid-cell-inner');
                    for (const td of tds) {
                        if (/^\d{8,12}$/.test((td.innerText || '').trim())) { ticketCount++; break; }
                    }
                }
                if (ticketCount > bestCount) { bestCount = ticketCount; bestTable = table; }
            }
            if (bestTable && bestCount > rows.length) {
                rows = bestTable.querySelectorAll('tr');
                winSel = 'fallback-best-table';
            }
        }

        out.debug += `rows=${winSel}(${rows.length})`;

        // ── Extrai cells de cada row ──
        for (const row of rows) {
            let cells = row.querySelectorAll('.x-grid3-cell-inner');
            if (cells.length < 3) cells = row.querySelectorAll('.x-grid-cell-inner');
            if (cells.length < 3) cells = row.querySelectorAll('td > div');
            if (cells.length < 3) cells = row.querySelectorAll('td');
            const vals = Array.from(cells).map(c => (c.innerText || '').trim());
            if (vals.length < 3) continue;

            let ticket = '';
            for (const v of vals) {
                if (/^\d{8,12}$/.test(v)) { ticket = v; break; }
            }
            if (ticket) out.rows.push(vals);
        }
        return out;
    }"""

    try:
        data = await page.evaluate(js_extract)
    except Exception as e:
        log.warning(f"[WI] Parse search results: erro JS: {e}")
        data = {"headers": [], "rows": []}

    headers = data.get("headers", [])
    rows = data.get("rows", [])
    debug_info = data.get("debug", "")

    log.debug(f"[WI] Grid parse: {debug_info}, extracted_rows={len(rows)}")
    if headers:
        log.debug(f"[WI] Grid headers ({len(headers)}): {headers[:12]}")
    if rows and rows[0]:
        log.debug(f"[WI] Grid sample row ({len(rows[0])} cells): {rows[0][:5]}")

    # Mapeia índice de headers conhecidos
    def find_col(keywords):
        for i, h in enumerate(headers):
            hl = h.lower()
            if any(k in hl for k in keywords):
                return i
        return -1

    col_number = find_col(["number"])
    col_start = find_col(["start date", "start_date"])
    col_county = find_col(["county"])
    col_place = find_col(["place", "municipality"])
    col_address = find_col(["address"])
    col_street = find_col(["street"])
    col_company = find_col(["company name"])
    col_caller = find_col(["caller"])
    col_date_ticket = find_col(["date time ticket", "date_time_ticket"])

    # Se não achou headers, usa posições fixas do layout confirmado:
    #   0=Number, 1=DateTimeTicket, 2=StartDateTime, 3=CompanyName,
    #   4=CompanyPhone, 5=CallerName, 6=County, 7=Place, 8=Address, 9=Street
    if col_number < 0:
        col_number = 0
    if col_start < 0:
        col_start = 2
    if col_county < 0:
        col_county = 6
    if col_place < 0:
        col_place = 7
    if col_address < 0:
        col_address = 8
    if col_street < 0:
        col_street = 9

    def get_cell(row, idx):
        if 0 <= idx < len(row):
            return row[idx].strip()
        return ""

    for row in rows:
        ticket = ""
        for v in row:
            if re.match(r"^\d{8,12}$", v):
                ticket = v
                break
        if not ticket:
            continue

        results.append({
            "ticket": ticket,
            "start_date": get_cell(row, col_start),
            "county": get_cell(row, col_county),
            "city": get_cell(row, col_place),
            "address_num": get_cell(row, col_address),
            "street": get_cell(row, col_street),
            "caller": get_cell(row, col_caller if col_caller >= 0 else 5),
            "cells": row,
        })

    # Dedup
    seen = set()
    deduped = []
    for r in results:
        if r["ticket"] not in seen:
            seen.add(r["ticket"])
            deduped.append(r)

    return deduped


async def _scrape_diggers_ticket_detail(page, tnum, debug=False):
    """Scrape detalhes de um ticket WI a partir do portal logado.

    Layout confirmado do formulario (ExtJS 4.1):
      - Tab "Ticket - Excavator Se..." com ticket# no header
      - Seção "Excavator" (esquerda): caller, company, Working For
      - Seção "Work Location" (centro): Ticket/Work, Start Date, Address,
        St/County/Place, Distance/Direction, Marking Instructions, Remarks
      - Seção "Member" (direita): lista de utilities com codigo e telefone
      - Mapa (centro-inferior)

    Retorna dict com campos parseados + ticket_type pra filtrar Relo-No-Show.
    """
    detail = {
        "ticket_type": "",  # "Standard", "Relo-No-Show", etc. — pra filtrar
        "address": "", "city": "", "county": "", "township": "",
        "expire": "", "location_text": "", "work_type": "",
        "client": "", "prime": "", "job_id": "",
        "old_ticket": "", "responses": [],
        "geo_lat": None, "geo_lon": None,
        "footage": 0, "marking_instructions": "", "remarks": "",
        "duration_days": 0, "start_date": "",
    }

    # ── 1. Clicar no ticket na grid de resultados ──
    ticket_clicked = False

    # Tenta seletores diretos do ticket number (pode ser link, td, span)
    for sel in [
        f'td:has-text("{tnum}")',
        f'.x-grid-cell:has-text("{tnum}")',
        f'a:has-text("{tnum}")',
        f'span:has-text("{tnum}")',
        f'div:has-text("{tnum}")',
    ]:
        try:
            loc = page.locator(sel).first
            if await loc.count():
                await loc.click()
                ticket_clicked = True
                log.debug(f"[WI] Detail {tnum}: clicou -> {sel}")
                break
        except Exception:
            continue

    # Fallback: double-click (ExtJS grids abrem em double-click)
    if not ticket_clicked:
        for sel in [f'td:has-text("{tnum}")', f'.x-grid-cell:has-text("{tnum}")']:
            try:
                loc = page.locator(sel).first
                if await loc.count():
                    await loc.dblclick()
                    ticket_clicked = True
                    log.debug(f"[WI] Detail {tnum}: dblclick -> {sel}")
                    break
            except Exception:
                continue

    # JS fallback
    if not ticket_clicked:
        try:
            result = await page.evaluate(f"""() => {{
                const walker = document.createTreeWalker(
                    document.body, NodeFilter.SHOW_TEXT, null, false
                );
                while (walker.nextNode()) {{
                    if (walker.currentNode.textContent.trim() === '{tnum}') {{
                        const el = walker.currentNode.parentElement;
                        if (el) {{ el.click(); return true; }}
                    }}
                }}
                return false;
            }}""")
            if result:
                ticket_clicked = True
        except Exception:
            pass

    if not ticket_clicked:
        log.warning(f"[WI] Detail {tnum}: nao encontrou na grid")
        return detail

    await page.wait_for_timeout(5000)
    await wait_stable(page)

    if debug:
        await _wi_debug_screenshot(page, "ticket_detail", tnum)

    # ── 2. Extrair dados via JavaScript (lê form inputs + texto) ──
    # O formulario do Diggers usa inputs (muitos readonly) com labels em <td>
    js_extract_fields = """() => {
        const out = {};

        // ── Extrair TODOS os input values com seus names ──
        const inputs = document.querySelectorAll('input, select, textarea');
        for (const inp of inputs) {
            const name = inp.name || '';
            const id = inp.id || '';
            const val = inp.value || '';
            if (val) {
                if (name) out['input_' + name] = val;
                if (id) out['input_id_' + id] = val;
            }
        }

        // ── Extrair pares label:valor do body text ──
        // O portal renderiza como "Label:" seguido de valor na mesma linha ou na proxima
        const body = document.body.innerText || '';
        const lines = body.split('\\n').map(l => l.trim()).filter(Boolean);

        // Procura patterns especificos do formulario confirmado
        for (let i = 0; i < lines.length; i++) {
            const line = lines[i];

            // Ticket/Work (tipo do ticket)
            if (/^Ticket\\/Work/i.test(line)) {
                // O valor pode estar na mesma linha ou na proxima
                const rest = line.replace(/^Ticket\\/Work\\s*:?\\s*/i, '').trim();
                if (rest) {
                    out.ticket_work = rest;
                } else if (i + 1 < lines.length) {
                    out.ticket_work = lines[i + 1];
                }
            }

            // Start Date/Time
            if (/^Start Date/i.test(line)) {
                const rest = line.replace(/^Start Date\\/Time\\s*:?\\s*/i, '').trim();
                if (rest) out.start_date = rest;
                else if (i + 1 < lines.length) out.start_date = lines[i + 1];
            }

            // St/County/Place
            if (/^St\\/County\\/Place/i.test(line)) {
                const rest = line.replace(/^St\\/County\\/Place\\s*:?\\s*/i, '').trim();
                if (rest) out.st_county_place = rest;
                else if (i + 1 < lines.length) out.st_county_place = lines[i + 1];
            }

            // Working For
            if (/^Working\\s*(For)?/i.test(line)) {
                const rest = line.replace(/^Working\\s*For\\s*:?\\s*/i, '').trim();
                if (rest) out.working_for = rest;
                else if (i + 1 < lines.length) out.working_for = lines[i + 1];
            }

            // Duration
            if (/^Duration/i.test(line)) {
                const rest = line.replace(/^Duration\\s*:?\\s*/i, '').trim();
                if (rest) out.duration = rest;
                else if (i + 1 < lines.length) out.duration = lines[i + 1];
            }

            // Distance/Direction (footage)
            if (/^Distance/i.test(line)) {
                const rest = line.replace(/^Distance\\/Direction\\s*:?\\s*/i, '').trim();
                if (rest) out.distance = rest;
                else if (i + 1 < lines.length) out.distance = lines[i + 1];
            }

            // Intersection 1
            if (/^Intersection 1/i.test(line)) {
                const rest = line.replace(/^Intersection 1\\s*:?\\s*/i, '').trim();
                if (rest) out.intersection = rest;
                else if (i + 1 < lines.length) out.intersection = lines[i + 1];
            }

            // Marking Instructions
            if (/^Marking/i.test(line)) {
                const rest = line.replace(/^Marking\\s*Instructions?\\s*:?\\s*/i, '').trim();
                if (rest) out.marking = rest;
                else if (i + 1 < lines.length) out.marking = lines[i + 1];
            }

            // Remarks
            if (/^Remarks/i.test(line)) {
                const rest = line.replace(/^Remarks\\s*:?\\s*/i, '').trim();
                if (rest) out.remarks = rest;
                else if (i + 1 < lines.length) {
                    // Remarks pode ter multiplas linhas
                    let rem = [];
                    for (let j = i + 1; j < Math.min(i + 5, lines.length); j++) {
                        if (/^(Refresh|Layers|Links|X:|Pan)/.test(lines[j])) break;
                        if (/^\\w+\\s*:/.test(lines[j]) && !/^(CONTACT|RELO|BORE)/i.test(lines[j])) break;
                        rem.push(lines[j]);
                    }
                    out.remarks = rem.join(' ');
                }
            }

            // Address (pode ter multiplas partes)
            if (/^Address\\s*:/i.test(line) && !out.address_line) {
                const rest = line.replace(/^Address\\s*:?\\s*/i, '').trim();
                if (rest) out.address_line = rest;
                else if (i + 1 < lines.length) out.address_line = lines[i + 1];
            }
        }

        // ── Extrair Member list (utilities no painel direito) ──
        // Formato: "UTILITY NAME CODE\\nCODE\\n(xxx) xxx-xxxx"
        // Procura padroes de utilities com codigo e telefone
        const memberSection = body.indexOf('Member');
        if (memberSection >= 0) {
            out.member_text = body.substring(memberSection, memberSection + 2000);
        }

        // Full body pra fallback
        out._body = body.substring(0, 5000);

        return out;
    }"""

    fields = {}
    try:
        fields = await page.evaluate(js_extract_fields)
    except Exception as e:
        log.warning(f"[WI] Detail {tnum}: erro JS extraindo campos: {e}")
        # Tenta em iframes
        for frame in page.frames:
            if frame == page.main_frame:
                continue
            try:
                fields = await frame.evaluate(js_extract_fields)
                if fields.get("ticket_work") or fields.get("st_county_place"):
                    break
            except Exception:
                continue

    body = fields.get("_body", "")

    # ── 3. Parse campos extraidos ──

    # Ticket type (CRITICO: filtrar Relo-No-Show)
    ticket_work = fields.get("ticket_work", "")
    if ticket_work:
        # O campo mostra "Relo-No-Show  ELECTRIC/TELEPHONE/CABLE INSTALLATION"
        # ou "Standard  ELECTRIC/TELEPHONE/CABLE INSTALLATION"
        tw_parts = re.split(r'\s{2,}|\t', ticket_work, maxsplit=1)
        detail["ticket_type"] = tw_parts[0].strip()
        if len(tw_parts) > 1:
            detail["work_type"] = tw_parts[1].strip()
        else:
            detail["work_type"] = ticket_work
        log.info(f"[WI] {tnum}: ticket_type = {detail['ticket_type']!r}")

    # St/County/Place: "WI  RACINE  MOUNT PLEASANT VILLAGE"
    scp = fields.get("st_county_place", "")
    if scp:
        scp_parts = re.split(r'\s{2,}|\t', scp)
        scp_parts = [p.strip() for p in scp_parts if p.strip()]
        if len(scp_parts) >= 2:
            detail["county"] = scp_parts[1]  # RACINE
        if len(scp_parts) >= 3:
            detail["city"] = scp_parts[2]    # MOUNT PLEASANT VILLAGE

    # Address: O portal tem campos separados (num / street / type / suffix)
    # Na screenshot: [vazio] | EMERALD | DR
    # Tambem pode ter fields de input com name contendo "address", "street"
    addr_parts = []
    for key in sorted(fields.keys()):
        if key.startswith("input_") and any(w in key.lower() for w in ["street", "address", "stname"]):
            val = fields[key].strip()
            if val and val not in addr_parts and not val.isdigit():
                addr_parts.append(val)
    if addr_parts:
        detail["address"] = " ".join(addr_parts)
    elif fields.get("address_line"):
        detail["address"] = fields["address_line"]

    # Se address veio vazio mas tem intersection, usa intersection
    intersection = fields.get("intersection", "")
    if not detail["address"] and intersection:
        detail["address"] = intersection

    # Location text
    location_parts = []
    if detail["address"]:
        location_parts.append(detail["address"])
    if detail["city"]:
        location_parts.append(detail["city"])
    if detail["county"]:
        location_parts.append(detail["county"] + " County")
    detail["location_text"] = " / ".join(location_parts) if location_parts else ""

    # Working For: "Five Stars / AT&T" → prime="Five Stars", client="AT&T"
    working_for = fields.get("working_for", "")
    if working_for:
        parts = [p.strip() for p in working_for.split("/") if p.strip()]
        if len(parts) >= 2:
            detail["prime"] = parts[0]
            detail["client"] = parts[1]
        elif len(parts) == 1:
            detail["client"] = parts[0]

    # Start Date e Duration → calcular Expire
    start_date_raw = fields.get("start_date", "")
    duration_raw = fields.get("duration", "")
    detail["start_date"] = start_date_raw

    if start_date_raw:
        # Parse "05/20/2026 9:00 AM" ou "05/20/2026"
        start_dt = None
        for fmt in ["%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M", "%m/%d/%Y"]:
            try:
                clean = re.split(r'\s{2,}|\t', start_date_raw)[0].strip()
                start_dt = datetime.strptime(clean, fmt)
                break
            except ValueError:
                continue

        if start_dt and duration_raw:
            # Parse "30 DAYS" → 30
            dur_match = re.search(r"(\d+)", duration_raw)
            if dur_match:
                dur_days = int(dur_match.group(1))
                detail["duration_days"] = dur_days
                expire_dt = start_dt + timedelta(days=dur_days)
                detail["expire"] = expire_dt.strftime("%m/%d/%Y")
                log.debug(f"[WI] {tnum}: expire calculado = {detail['expire']} "
                          f"(start={clean} + {dur_days}d)")

    # Se expire nao calculou, tenta extrair do body texto
    if not detail["expire"] and body:
        detail["expire"] = normalize_expire(extract_expire_date(body, tnum))

    # Distance/Direction → footage
    distance = fields.get("distance", "")
    if distance:
        ft_match = re.search(r"(\d+)\s*(?:FT|FEET|ft)", distance, re.IGNORECASE)
        if ft_match:
            detail["footage"] = int(ft_match.group(1))

    # Marking Instructions + Remarks → notes
    detail["marking_instructions"] = fields.get("marking", "")
    detail["remarks"] = fields.get("remarks", "")

    # ── 4. Parse utility responses ──
    # No portal logado, a Positive Response table pode aparecer como no portal publico
    # OU a lista "Member" no painel direito (sem status de resposta — so lista as utilities)
    # Tentamos a tabela de Positive Response primeiro

    js_extract_responses = """() => {
        const out = {responses: []};
        const tables = document.querySelectorAll('table');
        for (const table of tables) {
            const txt = (table.innerText || '').toLowerCase();
            if (!txt.includes('status') || !txt.includes('name')) continue;
            if (!txt.includes('facilities') && !txt.includes('phone') && !txt.includes('code')) continue;

            const rows = Array.from(table.querySelectorAll('tr'));
            let headers = [];
            let headerIdx = -1;
            for (let ri = 0; ri < rows.length; ri++) {
                const cells = rows[ri].querySelectorAll('th, td');
                const vals = Array.from(cells).map(c => (c.innerText || '').trim());
                const lower = vals.map(v => v.toLowerCase());
                if (lower.some(v => v === 'status') && lower.some(v => v === 'name')) {
                    headers = lower;
                    headerIdx = ri;
                    break;
                }
            }
            if (headerIdx < 0) continue;

            const idx = (key) => headers.findIndex(h => h.includes(key));
            const iStatus = idx('status'),
                  iCode   = idx('code'),
                  iName   = idx('name'),
                  iFac    = idx('facilities'),
                  iPhone  = idx('phone');

            for (let ri = headerIdx + 1; ri < rows.length; ri++) {
                const cells = rows[ri].querySelectorAll('td');
                const vals = Array.from(cells).map(c => (c.innerText || '').trim());
                if (vals.length < 3) continue;
                if (vals.every(v => !v)) continue;

                const get = (i) => (i >= 0 && i < vals.length) ? vals[i] : '';
                const status = get(iStatus);
                const code   = get(iCode);
                let name     = get(iName);
                let facilities = get(iFac);
                const phone  = get(iPhone);

                if (!name) continue;

                name       = name.split('\\n').map(s => s.trim()).filter(Boolean)[0] || '';
                facilities = facilities.split('\\n').map(s => s.trim()).filter(Boolean)[0] || '';

                const fullText = vals.join('\\n');
                const eventRegex = /([A-Za-z]{3,9}\\s+\\d{1,2},\\s+\\d{4}\\s+\\d{1,2}:\\d{2}\\s*[AP]M)\\s*\\|\\|\\s*([^\\n|]+)/g;
                const matches = [...fullText.matchAll(eventRegex)];
                let respondedDate = null;
                let comment = '';
                if (matches.length > 0) {
                    let bestIdx = 0;
                    let bestDate = new Date(matches[0][1]);
                    for (let mi = 1; mi < matches.length; mi++) {
                        const d = new Date(matches[mi][1]);
                        if (!isNaN(d) && (isNaN(bestDate) || d > bestDate)) {
                            bestDate = d;
                            bestIdx = mi;
                        }
                    }
                    respondedDate = matches[bestIdx][1].trim();
                    comment = matches[bestIdx][2].trim();
                }

                out.responses.push({status, code, name, facilities, phone, comment, respondedDate});
            }
        }
        return out;
    }"""

    resp_data = {"responses": []}
    try:
        resp_data = await page.evaluate(js_extract_responses)
    except Exception:
        pass

    # Se nao achou tabela de Positive Response no portal logado,
    # as respostas serao obtidas depois pelo sync_wi (portal publico)
    for row in resp_data.get("responses", []):
        name = (row.get("name") or "").strip()
        code = (row.get("code") or "").strip()
        status_raw = (row.get("status") or "").strip()
        facilities = (row.get("facilities") or "").strip()
        comment = (row.get("comment") or "").strip()
        rd_str = row.get("respondedDate")

        if not name:
            continue
        if name.lower() in ("name", "status", "code", "facilities", "phone"):
            continue

        # Remove sufixo de codigo do nome (padroes Diggers)
        if code:
            for pat in [
                re.compile(r'\s*-\s*' + re.escape(code) + r'\s*$'),
                re.compile(r'\s*\(' + re.escape(code) + r'\)\s*$'),
                re.compile(r'\s+' + re.escape(code) + r'\s*$'),
            ]:
                if pat.search(name):
                    name = pat.sub('', name).strip(' -')
                    break
        name = re.sub(r"\s*\([A-Z]{2,5}\d{1,4}\)\s*$", "", name).strip()
        name = re.sub(r"\s+[A-Z]{2,5}\d{1,4}\s*$", "", name).strip(' -')

        if not _is_valid_utility_name(name):
            continue

        responded_date = None
        if rd_str:
            for fmt in ["%b %d, %Y %I:%M %p", "%B %d, %Y %I:%M %p",
                        "%b %d, %Y %I:%M:%S %p", "%B %d, %Y %I:%M:%S %p"]:
                try:
                    responded_date = datetime.strptime(rd_str.strip(), fmt).replace(tzinfo=None).isoformat()
                    break
                except ValueError:
                    continue

        fac_low = facilities.lower()
        is_not_part = "not participating" in fac_low or "not service provider" in fac_low
        if is_not_part:
            cls_status, cls_unrec = classify(status_raw, "not participating " + comment + " " + facilities)
            response_text = "Not Participating"
        else:
            cls_status, cls_unrec = classify(status_raw, comment + " " + facilities)
            response_text = comment if comment else status_raw

        detail["responses"].append({
            "utility": name,
            "status_raw": status_raw,
            "status": cls_status,
            "response": response_text,
            "comment": comment,
            "responded_date": responded_date,
            "_unrecognized": cls_unrec,
        })

    log.info(f"[WI] Detail {tnum}: type={detail['ticket_type']!r} | "
             f"{detail['city'] or '?'}, {detail['county'] or '?'} | "
             f"expire={detail['expire'] or 'N/A'} | {len(detail['responses'])} utilities")

    return detail


async def import_wi(triggered_by="manual"):
    """Importa tickets novos do WI via Diggers Hotline portal (com login).

    Fluxo em 6 fases:
      1. Login no portal cliente
      2. Excavator Search por range de data
      3. Filtra tickets novos (não existem no Supabase)
      4. Scrape detalhes de cada ticket novo
      5. Geocoding em batch
      6. Batch upsert no Supabase + sync de respostas
    """
    if not WI_USER or not WI_PASS:
        log.error("[WI] import_wi: WI_USER/WI_PASS não definidos no .env — abortando")
        return 0

    log.info(f"{'='*55}")
    log.info(f"  OneDrill 811  IMPORT WI (Diggers Hotline)  [{triggered_by}]")
    log.info(f"{'='*55}")

    # ── Determinar range de datas ──
    last_date = _get_wi_last_search_date()
    today_str = datetime.now().strftime("%m/%d/%Y")

    if last_date:
        from_date = last_date
        log.info(f"[WI] Pesquisando de {from_date} até {today_str} (última pesquisa)")
    else:
        from_date = f"01/01/{datetime.now().year}"
        log.info(f"[WI] Primeira execução — pesquisando de {from_date} até {today_str}")

    # ── Tickets existentes no Supabase ──
    existing = sb_get("tickets", "&state=eq.WI")
    existing_nums = {t["ticket"] for t in existing}
    projects = sb_get("projects")
    log.info(f"[WI] {len(existing_nums)} tickets WI já no Supabase")

    canceled_set = get_canceled_set("WI")
    if canceled_set:
        log.info(f"[WI] Cache: {len(canceled_set)} tickets cancelados conhecidos")

    perfil = _profile_path("WI")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox"])
        page = await browser.new_page()
        page.set_default_timeout(60000)

        # ── FASE 1: Login ──
        ok = await _login_diggers(page)
        if not ok:
            log.error("[WI] Import: login falhou — abortando")
            await browser.close()
            return 0

        # ── FASE 2: Excavator Search ──
        search_results = await _search_diggers_excavator(page, from_date, today_str, debug=True)

        if not search_results:
            log.info("[WI] Import: nenhum ticket encontrado na pesquisa")
            _set_wi_last_search_date(today_str)
            await browser.close()
            return 0

        # ── FASE 3: Filtrar tickets novos (importa TODOS os tipos, incluindo Relo-No-Show) ──
        new_tickets_to_scrape = []
        for item in search_results:
            tnum = item["ticket"]
            if tnum in existing_nums:
                continue
            if tnum in canceled_set:
                continue
            new_tickets_to_scrape.append(tnum)

        log.info(f"[WI] Search retornou {len(search_results)} tickets, "
                 f"{len(new_tickets_to_scrape)} novos para importar")

        if not new_tickets_to_scrape:
            log.info("[WI] Import: nenhum ticket novo para importar")
            _set_wi_last_search_date(today_str)
            await browser.close()
            return 0

        # ── FASE 4: Dados da grid + respostas via portal público ──
        # A grid do Excavator Search já tem: county, city, address, street, start_date.
        # O portal público dá: utility responses.
        # Expire: calculado de start_date + 30 dias (duração padrão WI).
        # Não tenta clicar tickets na grid (ExtJS não abre detail ao clicar row).

        # Mapeia dados da grid por ticket number
        grid_data_map = {}
        for item in search_results:
            grid_data_map[item["ticket"]] = item

        await browser.close()  # Fecha browser do portal logado

        # ── Scrape respostas em batch via portal público ──
        log.info(f"[WI] Buscando respostas de {len(new_tickets_to_scrape)} tickets no portal publico...")
        pub_results = await scrape_wi(new_tickets_to_scrape)

        parsed_tickets = []
        for tnum in new_tickets_to_scrape:
            grid = grid_data_map.get(tnum, {})
            pub = pub_results.get(tnum, {})

            county = grid.get("county", "")
            city = grid.get("city", "")
            # Address = numero da casa + rua  (grid cols separados)
            addr_num = grid.get("address_num", "")
            street = grid.get("street", "")
            address = f"{addr_num} {street}".strip() if (addr_num or street) else ""

            # Start Date → expire (+30 dias padrão WI)
            # Store retorna JS Date string: "Wed May 20 2026 09:00:00 GMT-0400 (...)"
            start_raw = grid.get("start_date", "")
            expire = ""
            if start_raw:
                sdt = None
                # Primeiro: JS Date string (do ExtJS Store)
                js_clean = re.sub(r"\s*GMT[+-]\d{4}.*$", "", start_raw).strip()
                for fmt in ["%a %b %d %Y %H:%M:%S", "%a %b %d %Y",
                            "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p",
                            "%m/%d/%Y %H:%M", "%m/%d/%Y"]:
                    try:
                        sdt = datetime.strptime(js_clean, fmt)
                        break
                    except ValueError:
                        continue
                if not sdt:
                    # Fallback: tenta o raw original (pode ser formato portal)
                    for fmt in ["%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p",
                                "%m/%d/%Y %H:%M", "%m/%d/%Y"]:
                        try:
                            clean = re.split(r"\s{2,}|\t", start_raw)[0].strip()
                            sdt = datetime.strptime(clean, fmt)
                            break
                        except ValueError:
                            continue
                if sdt:
                    expire = (sdt + timedelta(days=30)).strftime("%m/%d/%Y")

            # Location text (da grid OU do pub)
            location_text = pub.get("location_text", "")
            if not location_text:
                parts = [p for p in [address, city, f"{county} County" if county else ""] if p]
                location_text = " / ".join(parts)

            # Responses do portal público
            responses = pub.get("responses", [])

            # Campos extras do ExtJS Store
            marking = grid.get("marking_instructions", "")
            remarks = grid.get("remarks", "")
            intersection = grid.get("intersection", "")
            if not address and intersection:
                address = intersection

            parsed_tickets.append({
                "tnum": tnum,
                "detail": {
                    "ticket_type": grid.get("ticket_type", "Standard"),
                    "address": address, "city": city, "county": county,
                    "township": "",  # WI não tem township separado; city = place
                    "expire": expire, "location_text": location_text,
                    "work_type": "Main line",
                    "client": "", "prime": "", "job_id": "",
                    "old_ticket": "", "responses": responses,
                    "geo_lat": None, "geo_lon": None,
                    "footage": 0,
                    "marking_instructions": marking if isinstance(marking, str) else "",
                    "remarks": remarks if isinstance(remarks, str) else "",
                    "duration_days": 30, "start_date": start_raw,
                },
            })

    log.info(f"[WI] FASE 4 OK: {len(parsed_tickets)} tickets com detalhes")

    if not parsed_tickets:
        _set_wi_last_search_date(today_str)
        return 0

    # ── FASE 5: Geocoding + Build ticket data ──
    new_ticket_data = []
    geocode_count = 0

    for item in parsed_tickets:
        tnum = item["tnum"]
        d = item["detail"]

        city = d.get("city", "")
        township = d.get("township", "")
        street = d.get("address", "")
        county = d.get("county", "")

        geo_lat = d.get("geo_lat")
        geo_lon = d.get("geo_lon")

        # Geocode se não tiver boundary coords
        if not geo_lat and street and city:
            geo_lat, geo_lon = await geocode_address(street, city, "WI")
            geocode_count += 1

        # Adjust coords by location
        work_type = d.get("work_type") or "Main line"
        if geo_lat and geo_lon:
            location_for_adjust = d.get("location_text") or ""
            geo_lat, geo_lon, work_type = adjust_coords_by_location(
                geo_lat, geo_lon, location_for_adjust, work_type
            )

        # Resolve county se não veio do portal
        if not county and (city or township):
            try:
                county = await resolve_county(
                    f"{city}, {township}".strip(", "), "WI", geo_lat, geo_lon
                )
            except Exception:
                pass

        # Match com projeto por Job#
        project_id = None
        job_id = d.get("job_id", "")
        if job_id:
            for proj in projects:
                if job_id.strip() in (proj.get("name", "") + proj.get("description", "")):
                    project_id = proj["id"]
                    break

        # Verificar ticket antigo (renovação)
        old_ticket_num = d.get("old_ticket", "")
        old_expire_str = ""
        old_status_str = ""
        inherited_path = None

        if old_ticket_num:
            log.info(f"[WI] {tnum}: Substitui ticket anterior {old_ticket_num}")
            try:
                old_tickets = sb_get("tickets", f"&ticket=eq.{_qv(old_ticket_num)}&state=eq.WI")
                if old_tickets:
                    ot = old_tickets[0]
                    if ot.get("field_path"):
                        inherited_path = ot["field_path"]
                    old_expire_str = normalize_expire(ot.get("expire") or "")
                    old_status_str = (ot.get("status") or "").strip()
            except Exception as e:
                log.debug(f"[WI] Erro buscando ticket antigo {old_ticket_num}: {e}")

        # Build history
        history_entries = [
            {"ts": int(datetime.now().timestamp() * 1000),
             "action": f"Importado 811 - {city or '?'}, WI",
             "color": "#10a574"}
        ]
        if inherited_path:
            history_entries.append(
                {"ts": int(datetime.now().timestamp() * 1000),
                 "action": f"Trajeto herdado do ticket {old_ticket_num}",
                 "color": "#6d28d9"}
            )
        if old_ticket_num and (old_expire_str or old_status_str):
            history_entries.append(
                {"ts": int(datetime.now().timestamp() * 1000),
                 "action": f"[RENOVAÇÃO] {old_ticket_num} → {tnum} (graça até {old_expire_str or 'N/A'}, status antigo: {old_status_str or 'N/A'})",
                 "color": "#7c3aed"}
            )

        location_str = f"{city}, {township}".strip(", ")

        # Notes: combina location + marking instructions + remarks
        notes_parts = []
        if d.get("location_text"):
            notes_parts.append(f"[811 Location] {d['location_text']}")
        if d.get("marking_instructions"):
            notes_parts.append(f"[Marking] {d['marking_instructions']}")
        if d.get("remarks"):
            notes_parts.append(f"[Remarks] {d['remarks']}")
        notes = "\n".join(notes_parts)

        # Footage do campo Distance/Direction
        ticket_footage = d.get("footage", 0) or 0

        ticket_data = {
            "ticket": tnum, "company": "One Drill", "state": "WI",
            "location": location_str, "address": street,
            "status": "Open", "expire": d.get("expire", ""),
            "footage": ticket_footage,
            "client": d.get("client", ""), "prime": d.get("prime", ""),
            "tipo": work_type,
            "job": job_id, "notes": notes,
            "project_id": project_id, "pending": "",
            "old_ticket2": old_ticket_num,
            "status_old": old_status_str, "expire_old": old_expire_str,
            "field_path": inherited_path,
            "geocoded_lat": geo_lat, "geocoded_lon": geo_lon,
            "county": county,
            "history": history_entries, "attachments": [],
        }
        new_ticket_data.append(ticket_data)
        log.info(f"[WI] Preparado: {tnum}  {location_str}"
                 f"{' [' + county + ' Co]' if county else ''}")

    if geocode_count:
        log.info(f"[WI] Geocoding: {geocode_count} endereços processados")

    # ── FASE 6: Batch upsert ──
    inserted = 0
    to_insert = [td for td in new_ticket_data if td['ticket'] not in existing_nums]

    if to_insert:
        try:
            for i in range(0, len(to_insert), BATCH_SIZE):
                chunk = to_insert[i:i + BATCH_SIZE]
                sb_upsert("tickets", chunk, on_conflict="ticket")
                inserted += len(chunk)
                for td in chunk:
                    existing_nums.add(td['ticket'])
                if i + BATCH_SIZE < len(to_insert):
                    time.sleep(0.3)
            log.info(f"[WI] ✅ Batch upsert: {inserted} tickets inseridos")
        except Exception as e:
            log.warning(f"[WI] Batch falhou ({e}), tentando 1-por-1...")
            inserted = 0
            for td in to_insert:
                try:
                    if td['ticket'] not in existing_nums:
                        sb_insert("tickets", td)
                        existing_nums.add(td['ticket'])
                    inserted += 1
                except Exception as e2:
                    log.error(f"[WI] Erro inserindo {td['ticket']}: {e2}")

    # ── Salvar respostas coletadas durante o import ──
    all_responses = []
    now_iso = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    for item in parsed_tickets:
        tnum = item["tnum"]
        detail = item["detail"]
        if not detail["responses"]:
            continue
        # Buscar ticket_id do Supabase
        try:
            t_list = sb_get("tickets", f"&ticket=eq.{_qv(tnum)}&state=eq.WI&select=id")
            if not t_list:
                continue
            tid = t_list[0]["id"]
        except Exception:
            continue

        for resp in detail["responses"]:
            all_responses.append({
                "ticket_id": tid, "ticket_num": tnum, "state": "WI",
                "utility_name": resp["utility"], "status": resp["status"],
                "response_text": resp.get("response") or resp.get("status_raw", ""),
                "synced_at": now_iso,
                "responded_at": resp.get("responded_date") or now_iso,
            })

    if all_responses:
        resp_saved = 0
        RESP_CHUNK = 10
        for i in range(0, len(all_responses), RESP_CHUNK):
            chunk = all_responses[i:i + RESP_CHUNK]
            try:
                sb_upsert("ticket_811_responses", chunk)
                resp_saved += len(chunk)
            except Exception:
                # Fallback 1-por-1
                for single in chunk:
                    try:
                        sb_upsert("ticket_811_responses", [single])
                        resp_saved += 1
                    except Exception as e2:
                        log.debug(f"[WI] Resp skip: {single.get('ticket_num')} {single.get('utility_name')}: {e2}")
            if i + RESP_CHUNK < len(all_responses):
                time.sleep(0.2)
        log.info(f"[WI] {resp_saved}/{len(all_responses)} respostas salvas do import")

    # ── Salvar data da última pesquisa ──
    _set_wi_last_search_date(today_str)

    log.info(f"[WI] === Import WI concluído: {inserted} tickets novos ===")
    return inserted


async def sync_and_import_wi(triggered_by="manual"):
    """Import + Sync WI completo (equivalente ao sync_and_import de IN/FL)."""
    imported = await import_wi(triggered_by)
    await sync_wi(triggered_by)
    return imported


# ── │ SECTION: CONTACTS_FL │ SCRAPE CONTATOS DE UTILITIES (FL - Sunshine 811)
def get_already_processed_tickets(state="FL"):
    """Retorna set de ticket_ref que já têm contatos salvos."""
    try:
        encoded_state = urllib.parse.quote(state, safe='')
        data = sb_get("utility_contacts", f"&state=eq.{encoded_state}&select=ticket_ref&ticket_ref=not.is.null")
        refs = {row["ticket_ref"] for row in data if row.get("ticket_ref")}
        log.info(f"[{state}] [SKIP-CHECK] {len(refs)} tickets já têm contatos salvos")
        return refs
    except Exception as e:
        log.warning(f"[{state}] Não foi possível checar tickets processados: {e}")
        return set()


def sb_upsert_contact(data):
    """Upsert contato — on_conflict utility_name+contact_name+state."""
    h = {**SB_H, "Prefer": "resolution=merge-duplicates,return=minimal"}
    allowed_fields = {"utility_name", "state", "phone_main", "phone_alt", "phone_emergency", "ticket_ref", "contact_name", "notes"}
    clean_data = {k: v for k, v in data.items() if k in allowed_fields and v}
    r = _sb_request(
        requests.post,
        f"{SB_URL}/rest/v1/utility_contacts?on_conflict=utility_name,contact_name,state",
        headers=h, json=clean_data, timeout=20
    )
    return r


def parse_contact_table(body, ticket_num, state):
    """Parse contatos de utilities do body da página Find Ticket (FL) ou Responses (IN).

    Inclui validação: descarta blocos com confiança baixa (utility name duvidoso, sem telefone).
    """
    contacts = []
    lines_list = body.split("\n")

    start_idx = -1
    for i, line in enumerate(lines_list):
        if "Positive Response" in line or ("Service Area" in line and "Contact" in line):
            start_idx = i + 1
            break

    if start_idx < 0:
        return contacts

    # Extrai blocos (cada bloco começa com No ou Yes)
    blocks = []
    current_block = []
    for i in range(start_idx, len(lines_list)):
        line = lines_list[i].strip()
        if not line:
            continue
        if line in ("No", "Yes") and current_block:
            blocks.append(current_block)
            current_block = [line]
        else:
            current_block.append(line)
    if current_block:
        blocks.append(current_block)

    phone_pat = re.compile(r'\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}')

    # Palavras que NÃO são nomes de utility (falsos positivos comuns)
    INVALID_UTILITY_NAMES = {
        "POSITIVE RESPONSE", "NO RESPONSE", "SERVICE AREA", "CONTACT",
        "DATE", "STATUS", "RESPONSE", "ENTRY METHOD", "COMMENTS",
        "TICKET", "SEARCH", "HOME", "DASHBOARD", "FILTER",
    }

    for block in blocks:
        if len(block) < 5:
            continue

        util_name = ""
        svc_code = ""
        util_type = ""

        # Encontra utility name com validação
        for j, bline in enumerate(block):
            if bline in ("No", "Yes"):
                continue
            # Validação: >5 chars, all caps, não é telefone, não é data, não é falso positivo
            if (len(bline) > 5
                    and bline == bline.upper()
                    and not phone_pat.search(bline)
                    and not bline.startswith("Date")
                    and bline.strip() not in INVALID_UTILITY_NAMES
                    and not re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}", bline)):  # Não é data
                util_name = bline
                if j + 1 < len(block):
                    next_l = block[j + 1].strip()
                    if len(next_l) <= 10 and next_l == next_l.upper():
                        svc_code = next_l
                    if j + 2 < len(block):
                        type_l = block[j + 2].strip()
                        if (type_l == type_l.upper()
                                and not phone_pat.search(type_l)
                                and not type_l.startswith("Date")
                                and len(type_l) > 2):
                            util_type = type_l
                break

        if not util_name:
            continue

        # Encontra telefones com nome anterior
        phone_entries = []
        for j, bline in enumerate(block):
            phones = phone_pat.findall(bline)
            if phones:
                phone = phones[0].strip()
                name = ""
                if j > 0:
                    prev = block[j - 1].strip()
                    if not phone_pat.search(prev) and prev not in ("No", "Yes") and not prev.startswith("Date"):
                        name = prev
                name_in_line = bline.split(phone)[0].strip().rstrip("(").strip()
                if name_in_line and len(name_in_line) > 2:
                    name = name_in_line
                phone_entries.append({"name": name, "phone": phone})

        # Encontra respondent
        respondent = ""
        for bline in block:
            resp_match = re.search(r'Respondent[:\s]*(.+)', bline, re.IGNORECASE)
            if resp_match:
                respondent = resp_match.group(1).strip()
                if "(" in respondent:
                    respondent = respondent.split("(")[0].strip()
                break

        # Validação: bloco sem telefone E sem respondent = provavelmente lixo
        if not phone_entries and not respondent:
            log.debug(f"[{state}] {ticket_num}: bloco '{util_name[:30]}' descartado — sem telefone nem respondent")
            continue

        # Cria registro por PESSOA
        seen_names = set()
        roles = ["Contact", "Alternate", "Emergency"]
        for k, entry in enumerate(phone_entries[:3]):
            name = entry["name"]
            phone = entry["phone"]
            if not name or name == util_name or name == svc_code or name == util_type:
                continue
            name_key = name.upper().strip("* ").strip()
            if name_key in seen_names:
                continue
            seen_names.add(name_key)
            contacts.append({
                "utility_name": util_name,
                "service_area_code": svc_code,
                "contact_name": name.strip("* ").strip(),
                "phone_main": phone,
                "state": state,
                "ticket_ref": ticket_num,
                "notes": (roles[k] if k < 3 else "Other") + " | " + util_type,
            })

        if respondent:
            resp_key = respondent.upper().strip()
            if resp_key not in seen_names:
                contacts.append({
                    "utility_name": util_name,
                    "service_area_code": svc_code,
                    "contact_name": respondent,
                    "phone_main": "",
                    "state": state,
                    "ticket_ref": ticket_num,
                    "notes": "Respondent | " + util_type,
                })

    return contacts


async def scrape_contacts(state="FL", ticket_numbers=None, force=False):
    """Scrape contatos de utilities via Find Ticket (FL + IN)."""
    perfil = _profile_path(state)

    if not ticket_numbers:
        tickets_db = sb_get("tickets", f"&state=eq.{state}&status=in.(Open,Damage,Clear)&order=ticket")
        ticket_numbers = [t["ticket"] for t in tickets_db]

    if not ticket_numbers:
        log.info(f"[{state}] Nenhum ticket para buscar contatos")
        return 0

    log.info(f"[{state}] === Scraping contatos de {len(ticket_numbers)} tickets ===")
    saved = 0

    if not force:
        already_done = get_already_processed_tickets(state)
        ticket_numbers = [t for t in ticket_numbers if t not in already_done]
        log.info(f"[{state}] Após skip: {len(ticket_numbers)} tickets a processar")
    else:
        log.info(f"[{state}] --force ativo: processando todos os {len(ticket_numbers)} tickets")

    if not ticket_numbers:
        log.info(f"[{state}] Todos tickets já processados — nada a fazer")
        return 0

    # Pré-carrega contatos existentes em memória
    existing_contacts_set = set()
    try:
        all_existing = sb_get("utility_contacts", f"&state=eq.{state}&select=utility_name,contact_name")
        for row in all_existing:
            key = (row.get("utility_name", "").strip().upper(), row.get("contact_name", "").strip().upper())
            existing_contacts_set.add(key)
        log.info(f"[{state}] {len(existing_contacts_set)} contatos existentes em memória")
    except Exception as e:
        log.warning(f"[{state}] Não foi possível pré-carregar contatos: {e}")

    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(perfil, headless=False, args=["--no-sandbox"])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)

        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)

        if "login" in page.url.lower():
            await ctx.close()
            await asyncio.sleep(1)
            ok = await auto_login_silent(state)

            if not ok:

                log.warning(f"[{state}] auto_login_silent falhou, tentando manual...")

                ok = await auto_login(state)
            if not ok:
                return 0
            await asyncio.sleep(1)
            ctx = await p.chromium.launch_persistent_context(perfil, headless=False, args=["--no-sandbox"])
            page = ctx.pages[0] if ctx.pages else await ctx.new_page()
            page.set_default_timeout(TIMEOUT_PAGE)

        for idx, tnum in enumerate(ticket_numbers):
            log.info(f"[{state}] Contatos ({idx+1}/{len(ticket_numbers)}) Ticket {tnum}")
            try:
                if state == "FL":
                    find_url = "https://exactix.sunshine811.com/findTicketByNumberAndPhone"
                    await page.goto(find_url, wait_until="domcontentloaded")
                    await page.wait_for_timeout(2500)

                    if "login" in page.url.lower():
                        sl = page.locator('a:has-text("Search here")').first
                        if await sl.count():
                            await sl.click()
                            await page.wait_for_timeout(2500)

                    # Mouse click em cada campo + digitar (simula pessoa real)
                    inputs = page.locator('input:visible')
                    ic = await inputs.count()

                    if ic >= 2:
                        box1 = await inputs.nth(0).bounding_box()
                        if box1:
                            await page.mouse.click(box1['x'] + box1['width']/2, box1['y'] + box1['height']/2)
                            await page.wait_for_timeout(200)
                            await page.keyboard.type(tnum, delay=40)
                            await page.wait_for_timeout(300)

                        # Preenche telefone via JavaScript (bypass Angular mask que adiciona "x")
                        phone_input = inputs.nth(1)
                        await phone_input.evaluate("""el => {
                            // Seta valor nativo via Angular
                            const nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                                window.HTMLInputElement.prototype, 'value'
                            ).set;
                            nativeInputValueSetter.call(el, '(321) 947-3131');
                            el.dispatchEvent(new Event('input', { bubbles: true }));
                            el.dispatchEvent(new Event('change', { bubbles: true }));
                            el.dispatchEvent(new Event('blur', { bubbles: true }));
                        }""")
                        await page.wait_for_timeout(300)

                        log.info(f"[{state}] {tnum}: campos preenchidos via mouse+keyboard")

                    search_btn = page.locator('button:has-text("Search")').first
                    box3 = await search_btn.bounding_box() if await search_btn.count() else None
                    if box3:
                        await page.mouse.click(box3['x'] + box3['width']/2, box3['y'] + box3['height']/2)
                        log.info(f"[{state}] {tnum}: Search clicado")

                    try:
                        await page.wait_for_selector('text="Positive Response", mat-card, table', timeout=8000)
                    except Exception:
                        await page.wait_for_timeout(3000)

                    if idx < 2:
                        ss2 = os.path.join(BASE_DIR, f"debug_after_search_{state}_{tnum}.png")
                        await page.screenshot(path=ss2, full_page=True)
                        log.info(f"[{state}] Screenshots salvos")

                    await page.wait_for_timeout(500)

                elif state == "IN":
                    await page.goto(f"https://811.indiana811.org/tickets/{tnum}", wait_until="domcontentloaded")
                    await wait_stable(page)
                    resp_tab = page.locator('a:has-text("Responses"), [role="tab"]:has-text("Responses")').first
                    if await resp_tab.count():
                        await resp_tab.click()
                        await wait_stable(page, timeout=5000)

                body = await page.locator("body").inner_text()

                if idx < 2:
                    debug_path = os.path.join(BASE_DIR, f"debug_contacts_{state}_{tnum}.txt")
                    with open(debug_path, "w", encoding="utf-8") as f:
                        f.write(f"URL: {page.url}\n\n{body}")
                    log.info(f"[{state}] Debug salvo: {debug_path}")

                contacts = parse_contact_table(body, tnum, state)

                if contacts:
                    for contact in contacts:
                        db_record = {
                            "utility_name": contact.get("utility_name", ""),
                            "contact_name": contact.get("contact_name", ""),
                            "state": state,
                            "phone_main": contact.get("phone_main", ""),
                            "ticket_ref": tnum,
                            "notes": contact.get("notes", ""),
                        }
                        try:
                            cname = (db_record.get("contact_name", "") or "").strip()
                            uname = (db_record.get("utility_name", "") or "").strip()
                            dedup_key = (uname.upper(), cname.upper())
                            if cname and uname and dedup_key in existing_contacts_set:
                                continue
                            sb_upsert_contact(db_record)
                            if cname and uname:
                                existing_contacts_set.add(dedup_key)
                            saved += 1
                            log.info(f"[{state}] SALVO: {uname} / {cname} / {db_record.get('phone_main','')}")
                        except Exception as e:
                            log.error(f"[{state}] Erro salvando contato {db_record.get('utility_name','?')} / {db_record.get('contact_name','?')}: {e}")
                            if saved == 0:
                                log.error(f"[{state}] Data enviado: {db_record}")
                    log.info(f"[{state}] {tnum}: {len(contacts)} contatos extraídos")
                else:
                    log.warning(f"[{state}] {tnum}: nenhum contato encontrado")

            except Exception as e:
                log.error(f"[{state}] Contatos {tnum}: ERRO → {e}")

        try:
            await ctx.close()
        except Exception:
            pass

    log.info(f"[{state}] === Contatos: {saved} salvos/atualizados de {len(ticket_numbers)} tickets ===")
    return saved


# ── │ SECTION: UTILITY_HELPERS │ UTILITY HELPERS ──────────────────────────────
def get_contacts_for_utility(utility_name, state="FL"):
    try:
        data = sb_get("utility_contacts", f"&utility_name=eq.{_qv(utility_name)}&state=eq.{_qv(state)}")
        return data[0] if data else None
    except Exception:
        return None


def get_all_contacts(state="FL"):
    try:
        return sb_get("utility_contacts", f"&state=eq.{state}&order=utility_name")
    except Exception:
        return []


# ── │ SECTION: EXPORT_EXCEL │ EXPORT EXCEL ────────────────────────────────────
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        log.info("Exportando Excel...")
        tickets = sb_get("tickets", "&order=ticket")
        projects = sb_get("projects", "&order=name")
        proj_map = {p["id"]: p["name"] for p in projects}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets"

        headers = [
            "Ticket #", "Projeto", "Cliente", "Prime", "Estado", "Local",
            "Status", "Footage", "Expira", "Tipo", "Endereço", "Job #",
            "Pending", "Empresa",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a6cf0")
            cell.alignment = Alignment(horizontal="center")

        STATUS_COLOR = {"Open": "FFCCCC", "Clear": "CCFFCC", "Damage": "FFE5B4", "Closed": "EEEEEE", "Cancel": "E0D7F7"}
        for row, t in enumerate(tickets, 2):
            vals = [
                t.get("ticket", ""), proj_map.get(t.get("project_id", ""), ""),
                t.get("client", ""), t.get("prime", ""), t.get("state", ""),
                t.get("location", ""), t.get("status", ""), t.get("footage", 0),
                t.get("expire", ""), t.get("tipo", ""), t.get("address", ""),
                t.get("job", ""), t.get("pending", ""), t.get("company", ""),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor=STATUS_COLOR.get(t.get("status", ""), "FFFFFF"))

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        # Aba Projetos
        wp = wb.create_sheet("Projetos")
        ph = ["Nome", "Cliente", "Estado", "Status", "Total Feet", "Tickets"]
        for col, h in enumerate(ph, 1):
            cell = wp.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a6cf0")

        ticket_count = {}
        for t in tickets:
            pid = t.get("project_id", "")
            ticket_count[pid] = ticket_count.get(pid, 0) + 1

        for row, p_data in enumerate(projects, 2):
            for col, val in enumerate([
                p_data.get("name", ""), p_data.get("client", ""), p_data.get("state", ""),
                p_data.get("status", ""), p_data.get("total_feet", 0),
                ticket_count.get(p_data["id"], 0),
            ], 1):
                wp.cell(row=row, column=col, value=val)

        fname = f"OneDrill_Tickets_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(fname)
        log.info(f"Excel salvo: {fname}")
        return fname
    except Exception as e:
        log.error(f"Erro ao exportar Excel: {e}")
        return None


# ── │ SECTION: DEBUG_SCREENSHOT │ DEBUG ───────────────────────────────────────
async def debug_screenshot(state):
    perfil = _profile_path(state)
    if not os.path.exists(perfil):
        print(f"Perfil não encontrado: {perfil}")
        return
    async with playwright_context(state, headless=False) as (p, ctx, page):
        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_nav(page)
        log.info(f"[{state}] URL: {page.url}")
        if "login" in page.url.lower():
            log.warning(f"[{state}] Sessão expirada!")
            return
        await page.screenshot(path=f"debug_2_after_login_{state}.png")
        for sel in ['text="Go to Ticket Dashboard"', 'a:has-text("Ticket Dashboard")']:
            if await page.locator(sel).count():
                await click_and_wait(page, page.locator(sel), "nav")
                break
        await page.screenshot(path=f"debug_3_dashboard_{state}.png")
        with open(f"debug_page_{state}.html", "w", encoding="utf-8") as f:
            f.write(await page.content())
        await page.wait_for_timeout(20000)


# ── │ SECTION: BACKFILL │ BACKFILL HISTORY ────────────────────────────────────
def backfill_history():
    """Adiciona eventos de clear no histórico para tickets Clear que não têm.
    Usa data real da última resposta 811 (não datetime.now).
    """
    log.info("=== BACKFILL: Adicionando eventos de clear no histórico ===")
    all_clear = sb_get("tickets", "&status=eq.Clear&order=ticket")
    if not all_clear:
        log.info("Nenhum ticket Clear encontrado")
        return

    fixed = 0
    for t in all_clear:
        tid = t["id"]
        tnum = t["ticket"]
        state = t.get("state", "")
        hist = t.get("history") or []

        has_clear_evt = any(
            "auto 811" in (h.get("action", "")).lower() and "revertido" not in (h.get("action", "")).lower()
            for h in hist
        )
        has_clear_evt = has_clear_evt or any("→ clear" in (h.get("action", "")).lower() for h in hist)
        if has_clear_evt:
            continue

        # ── Busca data real da última resposta 811 ──
        # Filtra datas corrompidas (responded_at ≈ synced_at = fallback do sync)
        backfill_dt = None
        try:
            resps = sb_get(
                "ticket_811_responses",
                f"&ticket_num=eq.{tnum}&state=eq.{state}"
                "&responded_at=not.is.null&select=responded_at,synced_at"
            )
            real_dates = []
            for r in resps:
                ra = r.get("responded_at")
                sa = r.get("synced_at")
                if not ra:
                    continue
                try:
                    ra_dt = datetime.fromisoformat(str(ra).replace("Z", ""))
                    if sa:
                        sa_dt = datetime.fromisoformat(str(sa).replace("Z", ""))
                        if abs((ra_dt - sa_dt).total_seconds()) / 3600 < 6:
                            continue
                    real_dates.append(ra_dt)
                except Exception:
                    continue
            if real_dates:
                backfill_dt = max(real_dates)
        except Exception as e:
            log.debug(f"  {tnum}: erro ao buscar respostas: {e}")

        # Fallback: tenta extrair das notas
        if not backfill_dt:
            notes_text = t.get("notes") or ""
            date_match = re.search(r'\[AUTO 811\] Clear em (\d{1,2}/\d{1,2}/\d{4})', notes_text)
            if date_match:
                try:
                    backfill_dt = datetime.strptime(date_match.group(1), "%m/%d/%Y")
                except Exception:
                    pass

        # Último fallback: agora
        if not backfill_dt:
            backfill_dt = datetime.now()

        backfill_ts = int(backfill_dt.timestamp() * 1000)
        backfill_label = backfill_dt.strftime('%m/%d/%Y')

        hist.append({"ts": backfill_ts, "action": f"[AUTO 811] Clear em {backfill_label}", "color": "#16a34a"})
        sb_patch("tickets", tid, {"history": hist})
        log.info(f"  {tnum}: BACKFILL → {backfill_label}")
        fixed += 1

    log.info(f"=== BACKFILL CONCLUÍDO: {fixed} tickets corrigidos (de {len(all_clear)} Clear) ===")


def fix_clear_dates():
    """Corrige datas de clear no histórico de tickets que têm a data do sync
    em vez da data real da última resposta 811.

    Para cada ticket Clear:
    1. Busca o evento [AUTO 811] ou → clear no history
    2. Busca a data real da última resposta na tabela ticket_811_responses
    3. Se a data do history for diferente da data real, corrige o timestamp
    """
    log.info("=" * 55)
    log.info("  FIX-DATES: Corrigindo datas de clear no histórico")
    log.info("=" * 55)

    all_clear = sb_get("tickets", "&status=eq.Clear&order=ticket")
    if not all_clear:
        log.info("Nenhum ticket Clear encontrado")
        return

    fixed = 0
    skipped = 0
    no_response = 0

    for t in all_clear:
        tid = t["id"]
        tnum = t["ticket"]
        state = t.get("state", "")
        hist = t.get("history") or []

        # Encontra o evento de clear no history
        clear_evt_idx = None
        for i, h in enumerate(hist):
            a = (h.get("action", "") or "").lower()
            if ("auto 811" in a and "revertido" not in a) or "→ clear" in a or "auto-clear" in a:
                clear_evt_idx = i
                # Pega o ÚLTIMO evento de clear (mais recente)

        if clear_evt_idx is None:
            skipped += 1
            continue

        current_ts = hist[clear_evt_idx].get("ts", 0)

        # Busca data real da última resposta 811.
        # IMPORTANTE: responded_at pode estar corrompido (= synced_at quando parser
        # não capturou a data). Filtra apenas respostas onde responded_at difere
        # de synced_at por mais de 6h (indicando data real do portal).
        real_dt = None
        try:
            resps = sb_get(
                "ticket_811_responses",
                f"&ticket_num=eq.{tnum}&state=eq.{state}"
                "&responded_at=not.is.null&select=responded_at,synced_at"
            )
            real_dates = []
            for r in resps:
                ra = r.get("responded_at")
                if not ra:
                    continue
                try:
                    ra_dt = datetime.fromisoformat(str(ra).replace("Z", ""))
                    # Aceita todas as datas — o parser agora captura datas reais.
                    # Antes filtrava diff<6h vs synced_at, mas isso descartava
                    # respostas legítimas do mesmo dia.
                    real_dates.append(ra_dt)
                except Exception:
                    continue
            if real_dates:
                real_dt = max(real_dates)
        except Exception as e:
            log.debug(f"  {tnum}: erro ao buscar respostas: {e}")

        if not real_dt:
            no_response += 1
            continue

        real_ts = int(real_dt.timestamp() * 1000)
        real_label = real_dt.strftime('%m/%d/%Y')

        # Compara: se a diferença for > 1 hora, corrige
        diff_hours = abs(current_ts - real_ts) / 3600000
        if diff_hours < 1:
            skipped += 1
            continue

        # Corrige o timestamp e o texto do evento
        old_label = datetime.fromtimestamp(current_ts / 1000).strftime('%m/%d/%Y') if current_ts else "?"
        hist[clear_evt_idx]["ts"] = real_ts
        hist[clear_evt_idx]["action"] = f"[AUTO 811] Clear em {real_label}"

        # Também corrige nas notas se existir
        notes = t.get("notes") or ""
        if f"[AUTO 811] Clear em {old_label}" in notes:
            notes = notes.replace(
                f"[AUTO 811] Clear em {old_label}",
                f"[AUTO 811] Clear em {real_label}"
            )

        patch_data = {"history": hist, "updated_at": real_dt.isoformat()}
        if notes != (t.get("notes") or ""):
            patch_data["notes"] = notes

        sb_patch("tickets", tid, patch_data)
        log.info(f"  {tnum}: {old_label} → {real_label} (corrigido, diff={diff_hours:.0f}h)")
        fixed += 1

    log.info(f"{'=' * 55}")
    log.info(f"  FIX-DATES CONCLUÍDO:")
    log.info(f"    {fixed} corrigidos")
    log.info(f"    {skipped} já corretos / sem evento clear")
    log.info(f"    {no_response} sem data de resposta no banco")
    log.info(f"    {len(all_clear)} total Clear")
    log.info(f"{'=' * 55}")






# ── SECTION: LOCAL_OVERRIDES ────────────────────────────────────
# Regras manuais aplicadas DEPOIS do classify do portal 811.
# Match case-insensitive + substring em location + utility.
# Pra adicionar regra: editar local_overrides.json e rodar --apply-overrides 1 vez.

_OVERRIDES_JSON = os.path.join(BASE_DIR, "local_overrides.json")

def _load_overrides():
    if os.path.exists(_OVERRIDES_JSON):
        try:
            with open(_OVERRIDES_JSON, "r", encoding="utf-8") as f:
                data = _json.load(f)
            if isinstance(data, list) and data:
                log.info(f"[Overrides] {len(data)} regra(s) carregada(s) de local_overrides.json")
                return data
        except Exception as e:
            log.warning(f"[Overrides] Erro ao ler local_overrides.json: {e} — usando fallback hardcoded")
    return [
        {
            "name": "Frontier Indiana",
            "state_match": "IN",
            "utility_match": "frontier",
            "force_status": "Clear",
            "reason": "Override local: Frontier sempre Clear em todo o Indiana (validado em campo, Eric 2026-06-05)",
        },
    ]

LOCAL_OVERRIDES = _load_overrides()


def _match_override(ovr, location_lower, state_upper):
    """True se a regra casa com location/state do ticket.
    location_match: substring match (case-insensitive). state_match: igualdade exata.
    Regra DEVE ter pelo menos um dos dois — sem nenhum matcher = invalida (skip).
    Se ambos presentes, AMBOS devem casar (AND)."""
    loc_m = (ovr.get("location_match") or "").lower()
    st_m = (ovr.get("state_match") or "").upper()
    if not loc_m and not st_m:
        return False  # regra invalida — sem matcher
    if loc_m and loc_m not in location_lower:
        return False
    if st_m and st_m != state_upper:
        return False
    return True


def _apply_local_overrides(ticket, deduped_responses):
    """Aplica LOCAL_OVERRIDES no array de responses (in-place).
    Cada regra exige location_match e/ou state_match (pelo menos um).
    state_match exige igualdade EXATA do estado (ex: 'IN' so casa state='IN'),
    blindando contra falsos positivos entre estados."""
    location_lower = (ticket.get("location") or "").lower()
    state_upper = (ticket.get("state") or "").upper()
    altered = 0
    for ovr in LOCAL_OVERRIDES:
        if not _match_override(ovr, location_lower, state_upper):
            continue
        for resp in deduped_responses:
            uname = (resp.get("utility") or "").lower()
            if ovr["utility_match"].lower() not in uname:
                continue
            old_status = resp.get("status")
            if old_status == ovr["force_status"]:
                continue
            log.info(
                f"  [OVERRIDE] {ticket.get('ticket')} [{ticket.get('state')}] - "
                f"{resp.get('utility')}: {old_status} -> {ovr['force_status']} ({ovr['name']})"
            )
            resp["status"] = ovr["force_status"]
            resp["response"] = ovr["reason"]
            resp["comment"] = ovr["reason"]
            resp["status_raw"] = ovr["reason"]
            resp["_override"] = ovr["name"]
            altered += 1
    return altered


def apply_overrides_now(target_state=None):
    """Aplica LOCAL_OVERRIDES retroativamente em tickets existentes.

    Tambem dispara AUTO-CLEAR se todas as responses ficaram Clear
    (sem isso, dashboard 'Cleared Hoje' nao mostra o ticket).

    Inclui RENOVACOES: percorre tanto tickets antigos quanto renovados,
    aplica override em todos os que batem com location_match.
    """
    log.info("=" * 55)
    log.info("  APPLY-OVERRIDES: aplicando overrides retroativamente")
    if target_state:
        log.info(f"  Estado: {target_state}")
    for ovr in LOCAL_OVERRIDES:
        parts = []
        if ovr.get('location_match'):
            parts.append(f"loc='{ovr['location_match']}'")
        if ovr.get('state_match'):
            parts.append(f"state='{ovr['state_match']}'")
        parts.append(f"util='{ovr.get('utility_match','?')}'")
        log.info(f"  - {ovr['name']}: {' + '.join(parts)} = {ovr.get('force_status','?')}")
    log.info("=" * 55)

    query = "&status=in.(Open,Damage,Clear)&order=ticket&select=id,ticket,location,state,status,history,notes,old_ticket2"
    if target_state:
        query += f"&state=eq.{target_state}"

    tickets = sb_get("tickets", query)
    if not tickets:
        log.info("Nenhum ticket")
        return

    # Log dos tickets que VAO ser afetados (antes de mexer)
    candidates = []
    for _t in tickets:
        _loc = (_t.get("location") or "").lower()
        _st = (_t.get("state") or "").upper()
        for _ovr in LOCAL_OVERRIDES:
            if _match_override(_ovr, _loc, _st):
                _renew = _t.get("old_ticket2")
                _ri = f" (renovou {_renew})" if _renew else ""
                candidates.append(f"  [{_t.get('state','?')}] {_t['ticket']} status={_t.get('status')}{_ri} <- {_ovr['name']}")
                break
    if candidates:
        log.info(f"Tickets candidatos a override ({len(candidates)}):")
        for c in candidates:
            log.info(c)
    else:
        log.info("Nenhum ticket bateu nenhum override")
        return

    total_updated = 0
    tickets_with_match = 0
    tickets_auto_cleared = 0

    for t in tickets:
        location = (t.get("location") or "").lower()
        state_upper = (t.get("state") or "").upper()

        matched_overrides = [o for o in LOCAL_OVERRIDES if _match_override(o, location, state_upper)]
        if not matched_overrides:
            continue

        responses = sb_get(
            "ticket_811_responses",
            f"&ticket_num=eq.{t['ticket']}&select=id,utility_name,status,response_text"
        )

        ticket_changed = False

        for ovr in matched_overrides:
            for r in responses:
                uname = (r.get("utility_name") or "").lower()
                if ovr["utility_match"].lower() not in uname:
                    continue
                if r.get("status") == ovr["force_status"]:
                    continue
                try:
                    sb_patch("ticket_811_responses", r["id"], {
                        "status": ovr["force_status"],
                        "response_text": ovr["reason"],
                        "synced_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
                    })
                    log.info(
                        f"  [{t['state']}] {t['ticket']} - {r['utility_name']}: "
                        f"{r['status']} -> {ovr['force_status']} ({ovr['name']})"
                    )
                    r["status"] = ovr["force_status"]
                    total_updated += 1
                    ticket_changed = True
                except Exception as e:
                    log.error(f"  erro ao atualizar response {r['id']}: {e}")

        if ticket_changed:
            tickets_with_match += 1

        # AUTO-CLEAR + BACKFILL HISTORICO
        # Cobre 2 cenarios:
        #   1. Ticket Open/Damage com todas responses Clear -> muda pra Clear + adiciona entry
        #   2. Ticket JA Clear sem entry [AUTO 811] no historico -> so adiciona entry
        #      (pra aparecer em 'Cleared Hoje' do dashboard)
        try:
            n_responses = len(responses)
            statuses_summary = sorted(set((r.get("status") or "?") for r in responses))
            all_clear = n_responses > 0 and all(r.get("status") == "Clear" for r in responses)
            hist = t.get("history") or []
            has_clear_entry = any(
                "[AUTO 811] Clear em" in (h.get("action") or "")
                and "Revertido" not in (h.get("action") or "")
                for h in hist
            )
            log.info(
                f"    [DBG] {t['ticket']}: status={t.get('status')} "
                f"responses={n_responses} statuses={statuses_summary} "
                f"all_clear={all_clear} has_clear_entry={has_clear_entry}"
            )

            # Apenas auto-clear quando Open/Damage com todas responses Clear,
            # ou backfill quando Clear sem entry historico. NUNCA adiciona entry de HOJE
            # se ticket clareou em data ANTERIOR (pois seria mentira no historico).
            if n_responses > 0 and all_clear:
                needs_action = False
                patch_data = {}
                action_type = ""

                if t.get("status") in ("Open", "Damage"):
                    needs_action = True
                    patch_data["status"] = "Clear"
                    action_type = "AUTO-CLEAR"
                elif t.get("status") == "Clear" and not has_clear_entry:
                    needs_action = True
                    action_type = "BACKFILL HISTORICO"

                if needs_action:
                    now_dt = datetime.now()
                    now_ts = int(now_dt.timestamp() * 1000)
                    clear_label = now_dt.strftime("%m/%d/%Y")
                    clear_note = f"[AUTO 811] Clear em {clear_label} (override local)"
                    hist.append({"ts": now_ts, "action": clear_note, "color": "#16a34a"})
                    new_notes = append_auto_note(t.get("notes"), clear_note)
                    patch_data["history"] = hist
                    patch_data["notes"] = new_notes
                    try:
                        sb_patch("tickets", t["id"], patch_data)
                        log.info(f"  [{t['state']}] {t['ticket']}: {action_type}")
                        tickets_auto_cleared += 1
                    except Exception as e:
                        log.error(f"  [{t['ticket']}] erro: {e}")
        except Exception as e:
            log.warning(f"  [{t['ticket']}] erro check auto-clear: {e}")

    log.info("=" * 55)
    log.info(f"  APPLY-OVERRIDES CONCLUIDO:")
    log.info(f"    Tickets candidatos:      {len(candidates)}")
    log.info(f"    Tickets com mudanca:     {tickets_with_match}")
    log.info(f"    Responses atualizadas:   {total_updated}")
    log.info(f"    Tickets auto-clarados:   {tickets_auto_cleared}")
    log.info("=" * 55)




def undo_fake_overrides_today(target_state=None):
    """Desfaz entries falsas '(override local)' adicionadas HOJE em historico.

    Necessario apos rodar apply_overrides_now versao com bug RE-CONFIRMA HOJE,
    que adicionou entries de hoje em tickets que clarearam em outras datas.
    """
    log.info("=" * 55)
    log.info("  UNDO-FAKE-OVERRIDES: removendo entries falsas de hoje")
    log.info("=" * 55)

    today_midnight = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_ms = int(today_midnight.timestamp() * 1000)

    query = "&status=in.(Open,Damage,Clear)&order=ticket&select=id,ticket,state,history"
    if target_state:
        query += f"&state=eq.{target_state}"

    tickets = sb_get("tickets", query)
    if not tickets:
        log.info("Nenhum ticket")
        return

    fixed = 0
    for t in tickets:
        hist = t.get("history") or []
        if not hist:
            continue

        # Filtra entries que NAO sao (override local) com ts >= hoje
        new_hist = []
        removed = 0
        for h in hist:
            action = h.get("action") or ""
            ts = h.get("ts", 0)
            is_override_today = (
                "(override local)" in action
                and ts >= today_ms
            )
            if is_override_today:
                removed += 1
            else:
                new_hist.append(h)

        if removed > 0:
            try:
                sb_patch("tickets", t["id"], {"history": new_hist})
                log.info(f"  [{t.get('state','?')}] {t['ticket']}: removidas {removed} entries falsas")
                fixed += 1
            except Exception as e:
                log.error(f"  [{t['ticket']}] erro: {e}")

    log.info("=" * 55)
    log.info(f"  UNDO concluido: {fixed} tickets ajustados")
    log.info("=" * 55)



def debug_ticket_history(ticket_num):
    """Mostra historico completo de um ticket pra debug.

    Lista todas entries do history com ts formatado, action e cor.
    Util pra entender por que ticket aparece (ou nao) em Cleared Hoje.
    """
    log.info("=" * 55)
    log.info(f"  DEBUG HISTORICO - Ticket: {ticket_num}")
    log.info("=" * 55)

    tickets = sb_get("tickets", f"&ticket=eq.{_qv(ticket_num)}&select=id,ticket,state,status,location,history,old_ticket2")
    if not tickets:
        log.info(f"Ticket {ticket_num} nao achado")
        return

    t = tickets[0]
    log.info(f"Estado: {t.get('state')}")
    log.info(f"Status: {t.get('status')}")
    log.info(f"Location: {t.get('location')}")
    log.info(f"Old ticket: {t.get('old_ticket2') or '(nenhum)'}")
    log.info("")

    hist = t.get("history") or []
    if not hist:
        log.info("(historico vazio)")
        return

    log.info(f"History ({len(hist)} entries):")
    for i, h in enumerate(hist, 1):
        ts = h.get("ts", 0)
        try:
            dt_str = datetime.fromtimestamp(ts/1000).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            dt_str = f"ts={ts}"
        action = h.get("action", "")
        color = h.get("color", "")
        log.info(f"  {i}. [{dt_str}] {action} (color={color})")

    log.info("")
    log.info("=" * 55)
    log.info("  Responses (utilities):")
    log.info("=" * 55)
    responses = sb_get("ticket_811_responses", f"&ticket_num=eq.{_qv(ticket_num)}&select=utility_name,status,response_text,synced_at&order=utility_name")
    for r in responses:
        log.info(f"  {r.get('utility_name'):40} {r.get('status'):8} synced={r.get('synced_at','')[:19]}")
        rt = (r.get('response_text') or '')[:80]
        if rt:
            log.info(f"      response: {rt}")
    log.info("=" * 55)


def apply_today_clears(target_state=None, target_date=None):
    """Adiciona entry '[AUTO 811] Last utility Clear em hoje' no historico de
    tickets que ja estavam Clear mas cujas utilities responderam HOJE.

    Cenario: ticket era Clear (status), mas no portal 811 utilities responderam
    HOJE. O sync detecta as responses mas nao adiciona entry no historico
    porque o status nao mudou. Resultado: dashboard 'Cleared Hoje' nao pega.

    Esse script roda 1x e adiciona a entry retroativamente.
    """
    log.info("=" * 55)
    log.info("  APPLY-TODAY-CLEARS: marcando tickets com atividade hoje")
    if target_state:
        log.info(f"  Estado: {target_state}")
    log.info("=" * 55)

    if target_date:
        try:
            target_dt = datetime.strptime(target_date, "%Y-%m-%d")
        except ValueError:
            log.error(f"Data invalida: {target_date} (use YYYY-MM-DD)")
            return
        log.info(f"  Data alvo: {target_date}")
    else:
        target_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_midnight = target_dt
    today_ms = int(today_midnight.timestamp() * 1000)
    today_key = today_midnight.strftime("%Y-%m-%d")
    entry_ts = int(datetime.now().timestamp() * 1000)
    entry_label = datetime.now().strftime("%m/%d/%Y")

    query = "&status=eq.Clear&order=ticket&select=id,ticket,state,history,notes"
    if target_state:
        query += f"&state=eq.{target_state}"

    tickets = sb_get("tickets", query)
    if not tickets:
        log.info("Nenhum ticket Clear")
        return

    log.info(f"Analisando {len(tickets)} tickets Clear...")

    updated = 0
    skipped_no_today_response = 0
    skipped_already_today = 0

    for t in tickets:
        # Verifica se TEM resposta de hoje (responded_at >= hoje)
        responses = sb_get(
            "ticket_811_responses",
            f"&ticket_num=eq.{t['ticket']}&status=eq.Clear&order=responded_at.desc&limit=1&select=responded_at,utility_name"
        )
        if not responses:
            continue

        latest = responses[0].get("responded_at", "")
        if not latest or latest[:10] != today_key:
            skipped_no_today_response += 1
            continue

        # Tem resposta de hoje! Verifica se historico ja tem entry de hoje
        hist = t.get("history") or []
        has_today_entry = False
        for h in hist:
            ts = h.get("ts", 0)
            action = h.get("action") or ""
            if ts >= today_ms and "[AUTO 811]" in action and "Revertido" not in action:
                has_today_entry = True
                break

        if has_today_entry:
            skipped_already_today += 1
            continue

        utility = responses[0].get("utility_name", "?")
        clear_note = f"[AUTO 811] Last utility Clear em {today_key} ({utility})"
        hist.append({"ts": entry_ts, "action": clear_note, "color": "#16a34a"})

        try:
            sb_patch("tickets", t["id"], {"history": hist})
            log.info(f"  [{t.get('state','?')}] {t['ticket']}: entry adicionada ({utility} respondeu {latest[:10]})")
            updated += 1
        except Exception as e:
            log.error(f"  [{t['ticket']}] erro: {e}")

    log.info("=" * 55)
    log.info(f"  APPLY-TODAY-CLEARS CONCLUIDO:")
    log.info(f"    Tickets atualizados:           {updated}")
    log.info(f"    Sem resposta de hoje:          {skipped_no_today_response}")
    log.info(f"    Ja tinha entry de hoje:        {skipped_already_today}")
    log.info("=" * 55)


# ── SECTION: EMAIL_SCANNER ──────────────────────────────────────
def _decode_email_header(h):
    """Decodifica header de email com multiple encodings."""
    if not h:
        return ""
    try:
        import email
        from email.header import decode_header
        parts = decode_header(h)
        out = ""
        for txt, enc in parts:
            if isinstance(txt, bytes):
                out += txt.decode(enc or "utf-8", errors="replace")
            else:
                out += txt
        return out
    except Exception:
        return str(h)


def _extract_email_body(msg):
    """Extrai body de email (plain ou HTML reduzido)."""
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype in ("text/plain", "text/html"):
                try:
                    payload = part.get_payload(decode=True)
                    if payload:
                        body = payload.decode(part.get_content_charset() or "utf-8", errors="replace")
                        if ctype == "text/html":
                            body = re.sub(r"<[^>]+>", " ", body)
                            body = re.sub(r"&nbsp;", " ", body)
                            body = re.sub(r"\s+", " ", body)
                        return body
                except Exception:
                    pass
    else:
        try:
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode(msg.get_content_charset() or "utf-8", errors="replace")
                if msg.get_content_type() == "text/html":
                    body = re.sub(r"<[^>]+>", " ", body)
                    body = re.sub(r"\s+", " ", body)
                return body
        except Exception:
            return ""
    return ""


def scan_emails_for_responses(commit=False, state_filter=None, days_back=7):
    """Varre Gmail buscando emails de status change das utilities 811.

    Eric reportou que utilities respondem 'Not Participating' no portal
    mas mandam email confirmando status. Esse scanner faz parser desses
    emails e atualiza no banco SE a utility ta como Not Participating.

    Args:
        commit: Se True, atualiza banco. Se False (default), so loga (dry run).
        state_filter: Se passado, filtra so tickets desse estado.
        days_back: Quantos dias pra tras varrer emails (default 7).
    """
    import imaplib
    import email as email_mod

    GMAIL_USER = os.getenv("GMAIL_USER")
    GMAIL_PASS = os.getenv("GMAIL_PASS")

    if not GMAIL_USER or not GMAIL_PASS:
        log.error("[EmailScan] GMAIL_USER/GMAIL_PASS nao definidos no .env")
        return

    log.info("=" * 55)
    log.info(f"  EMAIL SCANNER {'(DRY RUN)' if not commit else '(COMMIT)'}")
    log.info(f"  Buscando emails dos ultimos {days_back} dias")
    if state_filter:
        log.info(f"  Estado: {state_filter}")
    log.info("=" * 55)

    try:
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(GMAIL_USER, GMAIL_PASS)
        imap.select("INBOX")
    except Exception as e:
        log.error(f"[EmailScan] Erro IMAP login: {e}")
        return

    since_date = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")
    status, msg_ids = imap.search(None, f'(SINCE {since_date})')

    if status != "OK" or not msg_ids[0]:
        log.info("[EmailScan] Nenhum email no periodo")
        imap.logout()
        return

    ids = msg_ids[0].split()
    log.info(f"[EmailScan] Analisando {len(ids)} emails...")

    matched = 0
    updated = 0
    skipped_neg = 0
    skipped_no_ticket = 0
    skipped_no_match = 0
    errors = 0

    for msg_id in ids:
        try:
            # PEEK em vez de RFC822: lê o corpo SEM marcar o email como lido (\Seen).
            status, msg_data = imap.fetch(msg_id, "(BODY.PEEK[])")
            if status != "OK" or not msg_data or not msg_data[0]:
                continue

            raw = msg_data[0][1]
            msg = email_mod.message_from_bytes(raw)

            subject = _decode_email_header(msg.get("Subject", ""))
            sender = _decode_email_header(msg.get("From", ""))

            # Filtro grosso por subject
            subj_lower = subject.lower()
            if not any(kw in subj_lower for kw in ["ticket", "status change", "cleared", "positive response"]):
                continue

            body = _extract_email_body(msg)
            if not body:
                continue

            # Parser
            ticket_m = re.search(r"Ticket\s*#?:?\s*(\d{8,})", body)
            if not ticket_m:
                # Tenta no subject
                ticket_m = re.search(r"Ticket\s*#?\s*(\d{8,})", subject)
            member_m = re.search(r"Member\s*Code:?\s*(\w+)", body)
            response_m = re.search(r"Response:?\s*_?(\w+)", body)
            if not response_m:
                # Tenta "Work Performed"
                response_m = re.search(r"Work\s*Performed:?\s*(\w+)", body)

            if not (ticket_m and member_m and response_m):
                continue

            ticket_num = ticket_m.group(1)
            member_code = member_m.group(1).upper()
            response = response_m.group(1).upper()

            matched += 1
            log.info(f"  [{ticket_num}] member={member_code} response={response} from={sender[:40]}")

            if response not in ("CLEARED", "MARKED", "COMPLETE", "CLEAR", "COMPLETED"):
                log.info(f"    -> response nao-positiva, pula")
                skipped_neg += 1
                continue

            tickets = sb_get("tickets", f"&ticket=eq.{_qv(ticket_num)}&select=id,state,ticket")
            if not tickets:
                log.warning(f"    -> ticket {ticket_num} nao achado no banco")
                skipped_no_ticket += 1
                continue

            t = tickets[0]
            if state_filter and t.get("state") != state_filter:
                continue

            responses_at_ticket = sb_get(
                "ticket_811_responses",
                f"&ticket_num=eq.{ticket_num}&select=id,utility_name,status,response_text"
            )

            target = None
            for r in responses_at_ticket:
                uname = (r.get("utility_name") or "").upper()
                rt = (r.get("response_text") or "").lower()
                if member_code in uname:
                    if "not participating" in rt or "not service provider" in rt or r.get("status") == "Pending":
                        target = r
                        break

            if not target:
                log.info(f"    -> nenhuma utility com '{member_code}' marcada como Not Participating/Pending no ticket")
                skipped_no_match += 1
                continue

            log.info(f"    -> MATCH: {target['utility_name']} (era: '{(target.get('response_text') or '')[:50]}')")

            if commit:
                new_response_text = f"Email confirmation ({member_code}): {response}"
                try:
                    sb_patch("ticket_811_responses", target["id"], {
                        "status": "Clear",
                        "response_text": new_response_text,
                        "synced_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
                    })
                    log.info(f"    -> ATUALIZADO no banco")
                    updated += 1
                except Exception as e:
                    log.error(f"    -> ERRO ao atualizar: {e}")
                    errors += 1
            else:
                log.info(f"    -> [DRY RUN] atualizaria pra Clear")
                updated += 1
        except Exception as e:
            log.warning(f"  Erro processando email: {e}")
            errors += 1
            continue

    try:
        imap.logout()
    except Exception:
        pass

    log.info("=" * 55)
    log.info(f"  EMAIL SCANNER CONCLUIDO:")
    log.info(f"    Emails analisados:        {len(ids)}")
    log.info(f"    Parser identificou:       {matched}")
    log.info(f"    Updated{' (DRY)' if not commit else ''}:                  {updated}")
    log.info(f"    Skipped (resp negativa):  {skipped_neg}")
    log.info(f"    Skipped (ticket no banco):{skipped_no_ticket}")
    log.info(f"    Skipped (sem match):      {skipped_no_match}")
    log.info(f"    Erros:                    {errors}")
    log.info("=" * 55)


def scan_emails_debug(days_back=3):
    """Lista TODOS emails do Gmail (subject, remetente, preview) SEM aplicar filtros.
    Diagnostico pra entender o formato dos emails que chegam e ajustar o parser do
    scan_emails_for_responses (que hoje identifica 0 emails)."""
    import imaplib
    import email as email_mod

    GMAIL_USER = os.getenv("GMAIL_USER")
    GMAIL_PASS = os.getenv("GMAIL_PASS")
    if not GMAIL_USER or not GMAIL_PASS:
        log.error("[EmailDebug] GMAIL_USER/GMAIL_PASS nao definidos no .env")
        return

    log.info("=" * 55)
    log.info(f"  EMAIL DEBUG — listando emails dos ultimos {days_back} dias")
    log.info("=" * 55)

    try:
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(GMAIL_USER, GMAIL_PASS)
        imap.select("INBOX")
    except Exception as e:
        log.error(f"[EmailDebug] Erro IMAP login: {e}")
        return

    since_date = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")
    status, msg_ids = imap.search(None, f'(SINCE {since_date})')
    if status != "OK" or not msg_ids[0]:
        log.info("[EmailDebug] Nenhum email no periodo")
        try:
            imap.logout()
        except Exception:
            pass
        return

    ids = msg_ids[0].split()
    log.info(f"[EmailDebug] {len(ids)} emails encontrados\n")

    for i, msg_id in enumerate(ids, 1):
        try:
            # PEEK em vez de RFC822: lê o corpo SEM marcar o email como lido (\Seen).
            status, msg_data = imap.fetch(msg_id, "(BODY.PEEK[])")
            if status != "OK" or not msg_data or not msg_data[0]:
                continue
            raw = msg_data[0][1]
            msg = email_mod.message_from_bytes(raw)
            subject = _decode_email_header(msg.get("Subject", ""))
            sender = _decode_email_header(msg.get("From", ""))
            date_hdr = msg.get("Date", "")
            body = _extract_email_body(msg) or ""
            preview = " ".join(body[:300].split())  # colapsa whitespace pra 1 linha
            log.info(f"[{i:3d}] FROM:    {sender[:80]}")
            log.info(f"      SUBJECT: {subject[:120]}")
            log.info(f"      DATE:    {date_hdr}")
            log.info(f"      BODY:    {preview}")
            log.info("")
        except Exception as e:
            log.warning(f"  Erro processando email #{i}: {e}")

    try:
        imap.logout()
    except Exception:
        pass

    log.info("=" * 55)
    log.info(f"[EmailDebug] LISTAGEM CONCLUIDA — {len(ids)} emails")
    log.info("=" * 55)


# ── │ SECTION: COMPARE_WI_EXCEL │ Compara Excel com Supabase ─────────────────
def compare_wi_excel(xlsx_path):
    """Compara planilha .xlsx com tickets WI no Supabase.

    Detecta colunas pela primeira linha (busca por nomes contendo 'ticket' e 'status').
    Reporta:
      - Tickets na planilha que NÃO estão no banco
      - Tickets no banco WI que NÃO estão na planilha
      - Status divergente em tickets que existem nos dois
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl não instalado. Rode: pip install openpyxl")
        return

    if not os.path.exists(xlsx_path):
        log.error(f"Arquivo não encontrado: {xlsx_path}")
        return

    log.info(f"[CompareWI] Lendo {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log.error("[CompareWI] Planilha vazia")
        return

    # Detecta header (primeira linha que tem 'ticket' em alguma célula)
    header_row_idx = None
    header_idx = {}
    for i, row in enumerate(rows):
        cells = [(str(c) if c is not None else "").strip().lower() for c in row]
        if any("ticket" in c for c in cells):
            header_row_idx = i
            for j, c in enumerate(cells):
                if c:
                    header_idx[c] = j
            break

    if header_row_idx is None:
        log.error("[CompareWI] Cabeçalho não encontrado (esperado 'Ticket #' ou similar)")
        return

    tnum_col = next((v for k, v in header_idx.items() if "ticket" in k), None)
    status_col = next((v for k, v in header_idx.items() if k == "status" or k.endswith("status")), None)
    state_col = next((v for k, v in header_idx.items() if k == "state" or k == "estado"), None)

    if tnum_col is None:
        log.error("[CompareWI] Coluna 'Ticket' não encontrada no header")
        return

    excel_tickets = {}
    for row in rows[header_row_idx + 1:]:
        if not row or tnum_col >= len(row):
            continue
        tnum_raw = row[tnum_col]
        if tnum_raw is None:
            continue
        tnum = str(tnum_raw).strip()
        if not tnum or not tnum.replace(".", "").replace("-", "").isdigit():
            continue
        # Normaliza (planilha pode ter ".0" final pra ints)
        if tnum.endswith(".0"):
            tnum = tnum[:-2]
        status = ""
        if status_col is not None and status_col < len(row) and row[status_col] is not None:
            status = str(row[status_col]).strip()
        excel_tickets[tnum] = {"status": status}

    log.info(f"[CompareWI] Planilha: {len(excel_tickets)} tickets únicos")

    db_tickets_list = sb_get("tickets", "&state=eq.WI&select=ticket,status,location,client,prime,job")
    db_tickets = {}
    for t in db_tickets_list:
        tn = str(t.get("ticket") or "").strip()
        if tn:
            db_tickets[tn] = t

    log.info(f"[CompareWI] Banco WI: {len(db_tickets)} tickets")

    excel_set = set(excel_tickets.keys())
    db_set = set(db_tickets.keys())

    missing_in_db = sorted(excel_set - db_set)
    extra_in_db = sorted(db_set - excel_set)
    in_both = excel_set & db_set

    status_mismatch = []
    for tnum in sorted(in_both):
        excel_status = (excel_tickets[tnum]["status"] or "").lower()
        db_status = (db_tickets[tnum].get("status") or "").lower()
        if excel_status and db_status and excel_status != db_status:
            status_mismatch.append((tnum, excel_tickets[tnum]["status"], db_tickets[tnum].get("status")))

    log.info("=" * 60)
    log.info("  COMPARAÇÃO Excel × Supabase (state=WI)")
    log.info("=" * 60)
    log.info(f"  Planilha (unique):       {len(excel_set)}")
    log.info(f"  Banco WI:                {len(db_set)}")
    log.info(f"  Em AMBOS:                {len(in_both)}")
    log.info(f"  Faltando no BANCO:       {len(missing_in_db)}")
    log.info(f"  Extra no BANCO:          {len(extra_in_db)}")
    log.info(f"  Status divergente:       {len(status_mismatch)}")
    log.info("=" * 60)

    if missing_in_db:
        log.info("")
        log.info("📋 NA PLANILHA, AUSENTE NO BANCO:")
        for tnum in missing_in_db:
            log.info(f"   - {tnum} (status excel: {excel_tickets[tnum]['status'] or '?'})")

    if extra_in_db:
        log.info("")
        log.info("💾 NO BANCO, AUSENTE NA PLANILHA:")
        for tnum in extra_in_db:
            t = db_tickets[tnum]
            loc = t.get("location") or "?"
            client = t.get("client") or ""
            prime = t.get("prime") or ""
            log.info(f"   - {tnum} | {t.get('status','?'):8} | {loc} | client={client} prime={prime}")

    if status_mismatch:
        log.info("")
        log.info("⚠ STATUS DIVERGENTE (planilha ≠ banco):")
        for tnum, ex, db in status_mismatch:
            log.info(f"   - {tnum}: planilha={ex:8} banco={db}")

    log.info("=" * 60)


# ── │ SECTION: EMAIL_SCAN_WI │ EMAIL SCAN WISCONSIN (Outlook 365) ────────────
def list_outlook_folders():
    """Lista pastas IMAP da conta Outlook 365. Útil pra configurar OUTLOOK_FOLDER.

    Usa OUTLOOK_USER e OUTLOOK_PASS do .env. Imprime a hierarquia completa.
    """
    import imaplib

    USER = os.getenv("OUTLOOK_USER")
    PASS = os.getenv("OUTLOOK_PASS")
    HOST = os.getenv("OUTLOOK_HOST", "outlook.office365.com")

    if not USER or not PASS:
        log.error("[OutlookFolders] OUTLOOK_USER/OUTLOOK_PASS não definidos no .env")
        return

    try:
        imap = imaplib.IMAP4_SSL(HOST)
        imap.login(USER, PASS)
        status, folders = imap.list()
        if status != "OK":
            log.error(f"[OutlookFolders] Erro ao listar pastas: {status}")
            imap.logout()
            return
        log.info("=" * 55)
        log.info(f"  PASTAS IMAP de {USER}")
        log.info("=" * 55)
        for f in folders:
            line = f.decode("utf-8", errors="replace") if isinstance(f, bytes) else str(f)
            log.info(f"  {line}")
        log.info("=" * 55)
        imap.logout()
    except Exception as e:
        log.error(f"[OutlookFolders] Erro: {e}")


def scan_emails_wi(commit=False, days_back=14):
    """Scan emails da pasta Winsconsin (Outlook 365) buscando confirmações de utilities.

    Formato esperado (Windstream PRS / KorWeb):
        Subject: Ticket NNNNNNNN for CODE - Status Change
        Body:    Ticket: NNNNNNNN
                 Member Code: CODE
                 Response: _CLEARED (ou _MARKED, _COMPLETE, etc.)

    Quando match: utility WI que tinha status=Cancel (Closed by DHL) ou Pending
    é atualizada pra Clear, com response_text "Email confirmation (CODE): CLEARED em DD/MM/YYYY",
    e o ticket recebe entry no histórico.

    Args:
        commit: Se True, atualiza banco. Se False (dry run), só loga o que faria.
        days_back: Quantos dias pra trás varrer (default 14).
    """
    import imaplib
    import email as email_mod

    USER = os.getenv("OUTLOOK_USER")
    PASS = os.getenv("OUTLOOK_PASS")
    HOST = os.getenv("OUTLOOK_HOST", "outlook.office365.com")
    FOLDER = os.getenv("OUTLOOK_FOLDER", "Inbox/Clientes/Five Stars/Winsconsin")

    if not USER or not PASS:
        log.error("[EmailScanWI] OUTLOOK_USER/OUTLOOK_PASS não definidos no .env")
        return

    log.info("=" * 55)
    log.info(f"  EMAIL SCAN WI {'(DRY RUN)' if not commit else '(COMMIT)'}")
    log.info(f"  Pasta: {FOLDER}")
    log.info(f"  Buscando emails dos últimos {days_back} dias")
    log.info("=" * 55)

    try:
        imap = imaplib.IMAP4_SSL(HOST)
        imap.login(USER, PASS)
    except Exception as e:
        log.error(f"[EmailScanWI] Erro IMAP login: {e}")
        return

    folder_select = f'"{FOLDER}"' if " " in FOLDER else FOLDER
    status, _ = imap.select(folder_select)
    if status != "OK":
        log.error(f"[EmailScanWI] Pasta '{FOLDER}' não encontrada. Use --list-outlook-folders pra ver as pastas reais")
        try:
            imap.logout()
        except Exception:
            pass
        return

    since_date = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")
    status, msg_ids = imap.search(None, f'(SINCE {since_date})')

    if status != "OK" or not msg_ids or not msg_ids[0]:
        log.info("[EmailScanWI] Nenhum email no período")
        imap.logout()
        return

    ids = msg_ids[0].split()
    log.info(f"[EmailScanWI] Analisando {len(ids)} emails...")

    matched = 0
    updated = 0
    skipped_neg = 0
    skipped_no_ticket = 0
    skipped_no_match = 0
    errors = 0
    affected_tickets = set()

    for msg_id in ids:
        try:
            # PEEK em vez de RFC822: lê o corpo SEM marcar o email como lido (\Seen).
            status, msg_data = imap.fetch(msg_id, "(BODY.PEEK[])")
            if status != "OK" or not msg_data or not msg_data[0]:
                continue

            raw = msg_data[0][1]
            msg = email_mod.message_from_bytes(raw)

            subject = _decode_email_header(msg.get("Subject", ""))
            sender = _decode_email_header(msg.get("From", ""))

            # ── Parse subject — formato "Ticket NNNNN for CODE - Status Change" ──
            subj_m = re.search(r"Ticket\s+(\d{8,})\s+for\s+(\w+)", subject, re.IGNORECASE)
            if not subj_m:
                continue
            ticket_num = subj_m.group(1)
            member_code = subj_m.group(2).upper()

            # Body fields (estrutura KorWeb)
            body = _extract_email_body(msg)
            response = ""
            completed_on = ""
            if body:
                resp_m = re.search(r"Response\s*:?\s*_?(\w+)", body)
                if resp_m:
                    response = resp_m.group(1).upper()
                comp_m = re.search(r"Completed\s*on\s*:?\s*([\d/]+\s+[\d:]+\s*[AP]M)", body)
                if comp_m:
                    completed_on = comp_m.group(1).strip()

            if not response:
                # Tenta achar via Work Performed
                if body:
                    wp_m = re.search(r"Work\s*Performed[\s\W]+(\w+)", body)
                    if wp_m:
                        response = wp_m.group(1).upper()

            matched += 1
            log.info(f"  [{ticket_num}] code={member_code} response={response or '?'} completed={completed_on or '?'} from={sender[:40]}")

            if response not in ("CLEARED", "MARKED", "COMPLETE", "CLEAR", "COMPLETED", "NOCONFLICT"):
                log.info("    -> response não-positiva, pula")
                skipped_neg += 1
                continue

            # ── Match ticket no banco — só WI ──
            tickets_db = sb_get("tickets", f"&state=eq.WI&ticket=eq.{_qv(ticket_num)}&select=id,state,ticket,status,history,notes")
            if not tickets_db:
                log.warning(f"    -> ticket WI {ticket_num} não achado no banco")
                skipped_no_ticket += 1
                continue
            t = tickets_db[0]

            # ── Match utility por member_code (sufixo no utility_name) ──
            t_resps = sb_get(
                "ticket_811_responses",
                f"&ticket_num=eq.{_qv(ticket_num)}&select=id,utility_name,status,response_text"
            )
            target = None
            for r in t_resps:
                uname = (r.get("utility_name") or "").upper()
                if member_code in uname:
                    target = r
                    break

            if not target:
                log.info(f"    -> nenhuma utility com code '{member_code}' no ticket {ticket_num}")
                skipped_no_match += 1
                continue

            old_status = target.get("status")
            old_text = (target.get("response_text") or "")[:60]
            log.info(f"    -> MATCH: {target['utility_name']} ({old_status}: '{old_text}')")

            if old_status in ("Clear", "Marked"):
                log.info("    -> já está Clear/Marked, pula")
                continue

            if commit:
                new_text = f"Email confirmation ({member_code}): {response}"
                if completed_on:
                    new_text += f" em {completed_on}"
                try:
                    sb_patch("ticket_811_responses", target["id"], {
                        "status": "Clear",
                        "response_text": new_text,
                        "synced_at": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(),
                    })
                    # Histórico do ticket
                    hist = t.get("history") or []
                    hist_ts = int(datetime.now().timestamp() * 1000)
                    hist_note = f"[AUTO EMAIL WI] {target['utility_name']} respondeu por email — {response}"
                    if completed_on:
                        hist_note += f" em {completed_on}"
                    hist.append({"ts": hist_ts, "action": hist_note, "color": "#0ea5e9"})
                    new_notes = append_auto_note(t.get("notes"), hist_note)
                    sb_patch("tickets", t["id"], {"history": hist, "notes": new_notes})
                    log.info("    -> ATUALIZADO no banco (response + histórico)")
                    updated += 1
                    affected_tickets.add(ticket_num)
                except Exception as e:
                    log.error(f"    -> ERRO ao atualizar: {e}")
                    errors += 1
            else:
                log.info(f"    -> [DRY RUN] atualizaria {old_status} → Clear ({response})")
                updated += 1
                affected_tickets.add(ticket_num)
        except Exception as e:
            log.warning(f"  Erro processando email: {e}")
            errors += 1
            continue

    try:
        imap.logout()
    except Exception:
        pass

    log.info("=" * 55)
    log.info("  EMAIL SCAN WI CONCLUIDO:")
    log.info(f"    Emails analisados:        {len(ids)}")
    log.info(f"    Parser identificou:       {matched}")
    log.info(f"    Updated{' (DRY)' if not commit else ''}:                  {updated}")
    log.info(f"    Tickets afetados:         {len(affected_tickets)}")
    log.info(f"    Skipped (resp negativa):  {skipped_neg}")
    log.info(f"    Skipped (ticket no banco):{skipped_no_ticket}")
    log.info(f"    Skipped (utility no ticket): {skipped_no_match}")
    log.info(f"    Erros:                    {errors}")
    log.info("=" * 55)


# ── │ SECTION: PDF_SAVE │ SAVE PDF (via Print Dialog — simula humano) ────────


def fix_clear_ts(target_state=None):
    """Fix 2026-05-14: corrige timestamp dos [AUTO 811] Clear no historico.

    Tickets gravados pelo codigo antigo ficaram com ts = data da resposta da
    utility (pode ser ontem) em vez do momento do auto-clear (hoje). Resultado:
    aparecem no dia errado no dashboard 'Cleared Hoje'.

    Aproximacao: usa synced_at MAX das responses Clear do ticket como ts.
    """
    log.info("=" * 55)
    log.info("  FIX-CLEAR-TS: Corrigindo timestamp dos [AUTO 811] Clear")
    if target_state:
        log.info(f"  Estado: {target_state}")
    log.info("=" * 55)

    query = "&status=in.(Clear,Damage)&order=ticket"
    if target_state:
        query += f"&state=eq.{target_state}"

    all_tickets = sb_get("tickets", query)
    if not all_tickets:
        log.info("Nenhum ticket encontrado")
        return

    log.info(f"Processando {len(all_tickets)} tickets...")
    fixed = 0
    skipped_no_problem = 0
    skipped_no_response = 0
    skipped_no_history = 0

    for t in all_tickets:
        hist = t.get("history") or []
        if not hist:
            skipped_no_history += 1
            continue

        changed = False

        for entry in hist:
            action = entry.get("action", "")
            ts = entry.get("ts", 0)

            if "[AUTO 811] Clear em" not in action or "Revertido" in action:
                continue

            m = re.search(r"\[AUTO 811\] Clear em (\d{1,2}/\d{1,2}/\d{4})", action)
            if not m:
                continue

            try:
                label_dt = datetime.strptime(m.group(1), "%m/%d/%Y")
                label_ts = int(label_dt.timestamp() * 1000)
            except ValueError:
                continue

            # Se ts ja eh maior que label + 36h, foi corrigido
            if ts > label_ts + 36 * 3600 * 1000:
                skipped_no_problem += 1
                continue

            responses = sb_get(
                "ticket_811_responses",
                f"&ticket_num=eq.{t['ticket']}&status=eq.Clear"
                f"&order=synced_at.desc&limit=1&select=synced_at"
            )

            if not responses or not responses[0].get("synced_at"):
                skipped_no_response += 1
                continue

            try:
                synced_str = responses[0]["synced_at"]
                if synced_str.endswith("Z"):
                    synced_str = synced_str[:-1] + "+00:00"
                synced_dt = datetime.fromisoformat(synced_str)
                new_ts = int(synced_dt.timestamp() * 1000)

                if new_ts > ts:
                    log.info(
                        f"  [{t.get('state','?')}] {t['ticket']}: "
                        f"ts {datetime.fromtimestamp(ts/1000).strftime('%Y-%m-%d')} -> "
                        f"{datetime.fromtimestamp(new_ts/1000).strftime('%Y-%m-%d')}"
                    )
                    entry["ts"] = new_ts
                    changed = True
                    fixed += 1
                else:
                    skipped_no_problem += 1
            except Exception as e:
                log.warning(f"  [{t['ticket']}] erro parse synced_at: {e}")
                skipped_no_response += 1

        if changed:
            try:
                sb_patch("tickets", t["id"], {"history": hist})
            except Exception as e:
                log.error(f"  [{t['ticket']}] erro ao salvar: {e}")

    log.info("=" * 55)
    log.info(f"  FIX-CLEAR-TS CONCLUIDO:")
    log.info(f"    {fixed} corrigidos")
    log.info(f"    {skipped_no_problem} ja corretos")
    log.info(f"    {skipped_no_response} sem synced_at no banco")
    log.info(f"    {skipped_no_history} sem historico")
    log.info("=" * 55)



def _find_chrome_hwnd():
    """Encontra a janela do Chrome/Chromium via Win32 API."""
    import ctypes
    import ctypes.wintypes

    results = []

    @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)
    def callback(hwnd, _):
        if ctypes.windll.user32.IsWindowVisible(hwnd):
            buf = ctypes.create_unicode_buffer(512)
            ctypes.windll.user32.GetWindowTextW(hwnd, buf, 512)
            title = buf.value
            if any(kw in title for kw in ['Chrome', 'Chromium', 'Ticket', 'Sunshine', '811']):
                results.append((hwnd, title))
        return True

    ctypes.windll.user32.EnumWindows(callback, 0)
    if results:
        # Prioriza janela com "Ticket" no título (é a que está com o print dialog)
        for hwnd, title in results:
            if 'Ticket' in title:
                return hwnd, title
        return results[0]
    return None, ""


def _get_window_rect(hwnd):
    """Retorna (left, top, right, bottom) da janela."""
    import ctypes
    import ctypes.wintypes
    rect = ctypes.wintypes.RECT()
    ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
    return rect.left, rect.top, rect.right, rect.bottom


def _bring_to_front(hwnd):
    """Traz janela pro foreground."""
    import ctypes
    ctypes.windll.user32.SetForegroundWindow(hwnd)


def _pdf_query_number(t):
    """Decide qual número usar pra buscar o ticket nos portais durante geração de PDF.

    Pra tickets RENOVADOS ainda em grace period, o ticket NOVO foi enviado mas
    pode não ter respostas das utilities (que responderam no ANTIGO). Pra
    salvar o PDF com evidência legal das respostas Clear, busca pelo número
    ANTIGO durante grace. Fora de grace, usa o NOVO normal.

    Retorna (query_number, is_old). is_old=True significa que usou o antigo.
    """
    tnum_new = (t.get("ticket") or "").strip()
    in_grace, old_num = is_in_renewal_grace(t)
    if in_grace and old_num:
        return old_num, True
    return tnum_new, False


# CDP Page.printToPDF params — replica o Chrome "Save as PDF" dialog.
# Margins 0.5in em todos os lados (Chrome "Save as PDF" default).
_CDP_PDF_PARAMS = {
    "printBackground": True,
    "paperWidth": 8.5,
    "paperHeight": 11,
    "marginTop": 0.5,
    "marginBottom": 0.5,
    "marginLeft": 0.5,
    "marginRight": 0.5,
}

# CSS injetado antes do PDF — compensa diff de font metrics CDP vs Chrome dialog:
#  - break-inside: avoid → impede quebra de página no meio de row da tabela SA
#  - letter-spacing: 0.03em → ajusta wrapping pra igualar paginação do Chrome
#    (CDP headless rende ~3% mais estreito que Chrome dialog nas mesmas fonts)
_PRINT_FIX_CSS = (
    "tr { break-inside: avoid !important; page-break-inside: avoid !important; } "
    "* { letter-spacing: 0.03em !important; }"
)


async def save_ticket_pdfs(state="FL", force=False):
    """Salva PDF de tickets Clear, Damage e Completed via Print Text nativo do portal.

    Headless — pode usar o PC normalmente enquanto roda.
    Estrutura: pdfs/{STATE}/{PRIME}/{PROJECT}/{ticket}.pdf (ou subpasta de renovação).
    Damage duplicado em Damage/{STATE}/{PRIME}/{PROJECT}/{ticket}.pdf.

    Fluxo por ticket (idêntico ao que o usuário faz manualmente):
      1. Dashboard → Filtrar ticket
      2. Menu 3 pontos (more_vert) → bloquear window.print()
      3. Clicar "Print Text" → Angular renderiza print component
      4. CDP Page.printToPDF com margins 0.5in → PDF idêntico ao "Save as PDF" do Chrome
    """

    all_tickets = sb_get("tickets", f"&state=eq.{state}&status=in.(Clear,Damage,Completed)&order=ticket")
    if not all_tickets:
        log.info(f"[{state}] PDF: nenhum ticket Clear/Damage/Completed")
        return

    # Busca projetos pra resolver nome pelo project_id
    projects = sb_get("projects", "&select=id,name") or []
    projects_map = {p['id']: p for p in projects}
    renewal_groups = _build_renewal_groups(all_tickets)

    if not force:
        # Valida se o ticket JÁ TEM arquivo PDF salvo (pelo nº atual OU da cadeia de renovação),
        # não confia só no anexo 'ticket_pdf' do banco. Re-salva só quem realmente não tem o arquivo.
        _disk = _pdf_disk_map(state)
        all_tickets = [t for t in all_tickets if not _ticket_has_pdf_on_disk(t, _disk, 10000)]

    if not all_tickets:
        log.info(f"[{state}] PDF: todos os tickets Clear/Damage/Completed já têm PDF")
        return

    log.info("=" * 55)
    log.info(f"  SAVE-PDF: {len(all_tickets)} tickets Clear/Damage/Completed ({state})")
    log.info(f"  Headless — pode usar o PC normalmente")
    log.info("=" * 55)

    perfil = _profile_path(state)
    saved = 0
    errors = 0

    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(
            perfil, headless=True, args=["--no-sandbox"],
        )
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)

        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)

        if "login" in page.url.lower():
            log.warning(f"[{state}] PDF: sessão expirada — tentando login...")
            await ctx.close()
            await asyncio.sleep(1)
            ok = await auto_login_silent(state)

            if not ok:
                log.error(f"[{state}] PDF: login falhou (auto_login_silent)")
                return
            await asyncio.sleep(1)
            ctx = await p.chromium.launch_persistent_context(
                perfil, headless=True, args=["--no-sandbox"],
            )
            page = ctx.pages[0] if ctx.pages else await ctx.new_page()
            page.set_default_timeout(TIMEOUT_PAGE)

        ok = await goto_dashboard(page, state)
        if not ok:
            log.error(f"[{state}] PDF: dashboard inacessível")
            await ctx.close()
            return

        log.info(f"[{state}] PDF: dashboard OK — processando {len(all_tickets)} tickets")

        for idx, t in enumerate(all_tickets):
            tnum = t["ticket"]
            tid = t["id"]

            # Computa paths via helper centralizado
            paths = _compute_pdf_paths(t, projects_map, renewal_groups, BASE_DIR)
            query_tnum = paths['query_tnum']
            used_old = paths['used_old']
            pdf_filename = paths['pdf_filename']
            full_path = os.path.abspath(paths['pdf_path'])

            if used_old:
                log.info(f"  {tnum}: 🔄 RENOVADO em grace — usando número antigo {query_tnum} pro PDF")

            os.makedirs(os.path.dirname(full_path), exist_ok=True)

            if not force and os.path.exists(full_path):
                sz = os.path.getsize(full_path)
                if sz > 10000:
                    log.info(f"  {tnum}: PDF já existe ({round(sz/1024)}KB), pulando")
                    continue

            try:
                log.info(f"  ({idx+1}/{len(all_tickets)}) {tnum} (busca: {query_tnum})...")

                # 1. Filtrar no dashboard
                await filter_ticket(page, query_tnum)

                # 2. Clicar no menu 3 pontos (more_vert) do ticket
                menu_btn = page.locator('button:has(mat-icon:text("more_vert"))')
                if not await menu_btn.count():
                    menu_btn = page.locator('mat-icon:text("more_vert")')
                if not await menu_btn.count():
                    log.warning(f"  {tnum} (busca {query_tnum}): menu 3 pontos não encontrado")
                    errors += 1
                    continue
                await menu_btn.first.click()
                await page.wait_for_timeout(800)

                # 3. Bloquear window.print() ANTES de clicar Print Text
                #    O portal chama window.print() imediatamente após renderizar o
                #    print component — bloquear impede que o dialog abra e resete a página
                await page.evaluate("""() => {
                    window.__printCalled = false;
                    window.print = function() {
                        window.__printCalled = true;
                        console.log('BLOCKED window.print()');
                    };
                }""")

                # 4. Clicar "Print Text" no menu
                print_text_item = page.locator(
                    '[role="menuitem"]:has-text("Print Text"), '
                    'button:has-text("Print Text")'
                )
                if not await print_text_item.count():
                    log.warning(f"  {tnum}: 'Print Text' não encontrado no menu")
                    await page.keyboard.press("Escape")
                    errors += 1
                    continue
                await print_text_item.first.click()

                # 5. Esperar window.print() ser bloqueado (confirma render do componente)
                rendered = False
                for _i in range(30):
                    await page.wait_for_timeout(500)
                    if await page.evaluate("() => window.__printCalled"):
                        rendered = True
                        break

                if not rendered:
                    log.warning(f"  {tnum}: Print component não renderizou (timeout 15s)")
                    await back_to_dashboard(page, state)
                    errors += 1
                    continue

                # 6. Tempo extra pro conteúdo carregar completamente (SA table, responses)
                await page.wait_for_timeout(3000)

                # 7. Injeta CSS fix antes do PDF:
                #    - break-inside: avoid em tr → não quebra row no meio
                #    - letter-spacing: 0.03em → compensa font metrics CDP vs Chrome
                await page.evaluate(
                    "(css) => { var s = document.createElement('style');"
                    " s.textContent = css; document.head.appendChild(s); }",
                    _PRINT_FIX_CSS
                )
                await page.wait_for_timeout(300)

                # 8. CDP Page.printToPDF — margins 0.5in (Chrome "Save as PDF" defaults).
                #    CDP direto + CSS fix produz ~85KB, mesma paginação do Chrome dialog.
                cdp = await ctx.new_cdp_session(page)
                try:
                    cdp_result = await cdp.send('Page.printToPDF', _CDP_PDF_PARAMS)
                    final_pdf = base64.b64decode(cdp_result.get('data', ''))
                finally:
                    await cdp.detach()

                with open(full_path, "wb") as f:
                    f.write(final_pdf)

                # 9. Verificar se salvou
                if os.path.exists(full_path) and os.path.getsize(full_path) > 10000:
                    file_size = os.path.getsize(full_path)
                    log.info(f"  ✅ {tnum}: PDF Print Text salvo ({round(file_size/1024)}KB)")

                    # Duplica pra pasta Damage se aplicável
                    if paths['damage_path']:
                        dmg_full = os.path.abspath(paths['damage_path'])
                        os.makedirs(os.path.dirname(dmg_full), exist_ok=True)
                        shutil.copy2(full_path, dmg_full)
                        log.info(f"  {tnum}: 📋 cópia Damage salva")

                    attachments = t.get("attachments") or []
                    attachments = [a for a in attachments if a.get("type") != "ticket_pdf"]
                    att = {
                        "name": pdf_filename,
                        "type": "ticket_pdf",
                        "saved_at": datetime.now().isoformat(),
                        "size_kb": round(file_size / 1024, 1)
                    }
                    if used_old:
                        att["old_ticket"] = query_tnum
                        att["new_ticket"] = tnum
                    attachments.append(att)
                    hist = t.get("history") or []
                    action_txt = (f"📄 PDF salvo ({round(file_size/1024)}KB)"
                                  + (f" — usado # antigo {query_tnum}" if used_old else ""))
                    hist.append({
                        "ts": int(datetime.now().timestamp() * 1000),
                        "action": action_txt,
                        "color": "#7c3aed"
                    })
                    sb_patch("tickets", tid, {"attachments": attachments, "history": hist})
                    saved += 1
                else:
                    sz = os.path.getsize(full_path) if os.path.exists(full_path) else 0
                    log.warning(f"  ⚠ {tnum}: PDF não salvo ou pequeno ({sz}B)")
                    errors += 1

                # 10. Volta ao dashboard pro próximo ticket
                await back_to_dashboard(page, state)

            except Exception as e:
                log.error(f"  ❌ {tnum}: {e}")
                errors += 1
                try:
                    await back_to_dashboard(page, state)
                except Exception:
                    pass

        await ctx.close()

    log.info("=" * 55)
    log.info(f"  SAVE-PDF CONCLUÍDO: {saved} salvos, {errors} erros")
    log.info("=" * 55)


async def save_ticket_pdfs_wi(force=False):
    """Salva PDF de tickets WI via page.pdf() — Diggers Hotline é público, headless.

    Fluxo por ticket:
      1. Portal Diggers → Find Tickets → Digita ticket → Search
      2. Espera iframe com resultado (URL: .../client/item/ticket/{id}?pr=true)
      3. Navega direto pra URL do iframe (página limpa com ticket + Positive Response)
      4. page.pdf() gera PDF com tudo (info + members + responses)

    Estrutura: pdfs/WI/{PRIME}/{PROJECT}/{ticket}.pdf (ou subpasta de renovação).
    Damage duplicado em Damage/WI/{PRIME}/{PROJECT}/{ticket}.pdf.
    """
    all_tickets = sb_get("tickets", "&state=eq.WI&status=in.(Clear,Damage,Completed)&order=ticket")
    if not all_tickets:
        log.info("[WI] PDF: nenhum ticket Clear/Damage/Completed")
        return

    # Busca projetos pra resolver nome pelo project_id
    projects = sb_get("projects", "&select=id,name") or []
    projects_map = {p['id']: p for p in projects}
    renewal_groups = _build_renewal_groups(all_tickets)

    if not force:
        # Valida se o ticket JÁ TEM arquivo PDF salvo (nº atual OU cadeia), não só o anexo do banco.
        _disk = _pdf_disk_map("WI")
        all_tickets = [t for t in all_tickets if not _ticket_has_pdf_on_disk(t, _disk, 5000)]

    if not all_tickets:
        log.info("[WI] PDF: todos os tickets Clear/Damage/Completed já têm PDF")
        return

    log.info("=" * 55)
    log.info(f"  SAVE-PDF WI (DIGGERS): {len(all_tickets)} tickets Clear/Damage/Completed")
    log.info("=" * 55)

    saved = 0
    errors = 0

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox"])
        page = await browser.new_page()
        page.set_default_timeout(30000)

        for idx, t in enumerate(all_tickets):
            tnum = t["ticket"]
            tid = t["id"]

            # Computa paths via helper centralizado
            paths = _compute_pdf_paths(t, projects_map, renewal_groups, BASE_DIR)
            query_tnum = paths['query_tnum']
            used_old = paths['used_old']
            pdf_filename = paths['pdf_filename']
            full_path = os.path.abspath(paths['pdf_path'])

            if used_old:
                log.info(f"  {tnum}: 🔄 RENOVADO em grace — usando número antigo {query_tnum} pro PDF")

            os.makedirs(os.path.dirname(full_path), exist_ok=True)

            if not force and os.path.exists(full_path):
                sz = os.path.getsize(full_path)
                if sz > 5000:
                    log.info(f"  {tnum}: PDF já existe ({round(sz/1024)}KB), pulando")
                    continue

            try:
                log.info(f"  ({idx+1}/{len(all_tickets)}) {tnum} (busca: {query_tnum})...")

                # ── 1. Navega pro portal Diggers ─────────────────────────────
                await page.goto(DIGGERS_URL, wait_until="domcontentloaded")
                await page.wait_for_timeout(3000)  # ExtJS demora pra inicializar

                # ── 2. Espera e clica no botão Find Tickets ──────────────────
                try:
                    await page.wait_for_selector('#findTicketsButton-btnEl', timeout=15000, state="visible")
                    await page.locator('#findTicketsButton-btnEl').click()
                except Exception:
                    try:
                        await page.evaluate("""() => {
                            const btn = document.querySelector('#findTicketsButton-btnEl');
                            if (btn) btn.click();
                        }""")
                    except Exception as e:
                        log.error(f"  {tnum}: erro clicando Find Tickets: {e}")
                        errors += 1
                        continue

                await page.wait_for_timeout(2000)

                # ── 3. Encontra input e digita ticket ────────────────────────
                inp = page.locator('input[name="ticket-number"]').first
                if not await inp.count():
                    # Fallback: procurar em frames
                    target_frame, inp = await _wait_for_diggers_element(
                        page, 'input[name="ticket-number"]', timeout_s=15
                    )
                    if not target_frame or not inp:
                        log.warning(f"  {tnum}: input ticket-number não encontrado")
                        errors += 1
                        continue

                await inp.click()
                await inp.fill(query_tnum)
                await inp.press("Enter")

                # ── 4. Espera iframe com resultado do ticket ─────────────────
                ticket_frame_url = None
                for _attempt in range(20):
                    await page.wait_for_timeout(500)
                    for frame in page.frames:
                        if '/client/item/ticket/' in frame.url:
                            ticket_frame_url = frame.url
                            break
                    if ticket_frame_url:
                        break

                if not ticket_frame_url:
                    log.warning(f"  {tnum}: iframe do ticket não apareceu no Diggers")
                    errors += 1
                    continue

                # Garantir que inclui Positive Response (?pr=true)
                if 'pr=true' not in ticket_frame_url:
                    sep = '&' if '?' in ticket_frame_url else '?'
                    ticket_frame_url += f'{sep}pr=true'

                # ── 5. Navega direto pra URL limpa (ticket + responses) ──────
                await page.goto(ticket_frame_url, wait_until="domcontentloaded")
                await page.wait_for_timeout(2000)

                # ── 6. Gera PDF da página completa (ticket + Positive Response)
                await page.emulate_media(media="screen")
                await page.pdf(
                    path=full_path, format="Letter", print_background=True,
                    margin={"top": "0.4in", "bottom": "0.4in",
                            "left": "0.4in", "right": "0.4in"},
                )

                if os.path.exists(full_path) and os.path.getsize(full_path) > 3000:
                    file_size = os.path.getsize(full_path)
                    log.info(f"  ✅ {tnum}: PDF salvo ({round(file_size/1024)}KB)")

                    # Duplica pra pasta Damage se aplicável
                    if paths['damage_path']:
                        dmg_full = os.path.abspath(paths['damage_path'])
                        os.makedirs(os.path.dirname(dmg_full), exist_ok=True)
                        shutil.copy2(full_path, dmg_full)
                        log.info(f"  {tnum}: 📋 cópia Damage salva")

                    attachments = t.get("attachments") or []
                    attachments = [a for a in attachments if a.get("type") != "ticket_pdf"]
                    att = {
                        "name": pdf_filename,
                        "type": "ticket_pdf",
                        "saved_at": datetime.now().isoformat(),
                        "size_kb": round(file_size / 1024, 1)
                    }
                    if used_old:
                        att["old_ticket"] = query_tnum
                        att["new_ticket"] = tnum
                    attachments.append(att)
                    hist = t.get("history") or []
                    action_txt = (f"📄 PDF salvo ({round(file_size/1024)}KB)"
                                  + (f" — usado # antigo {query_tnum}" if used_old else ""))
                    hist.append({
                        "ts": int(datetime.now().timestamp() * 1000),
                        "action": action_txt,
                        "color": "#7c3aed"
                    })
                    sb_patch("tickets", tid, {"attachments": attachments, "history": hist})
                    saved += 1
                else:
                    sz = os.path.getsize(full_path) if os.path.exists(full_path) else 0
                    log.warning(f"  ⚠ {tnum}: PDF não salvo ou pequeno ({sz}B)")
                    errors += 1

            except Exception as e:
                log.error(f"  ❌ {tnum}: {e}")
                errors += 1

        await browser.close()

    log.info("=" * 55)
    log.info(f"  SAVE-PDF WI CONCLUÍDO: {saved} salvos, {errors} erros")
    log.info("=" * 55)


# ── │ SECTION: BACKUP │ BACKUP ────────────────────────────────────────────────
def backup_database(backup_dir=None, keep_days=30):
    """Exporta todas as tabelas do Supabase para JSON datado.
    
    Estrutura: backups/2026-04-08/tickets.json, projects.json, etc.
    Mantém os últimos 'keep_days' dias de backup.
    """
    import json, shutil
    from datetime import timedelta

    today = datetime.now().strftime("%Y-%m-%d")
    base = backup_dir or os.path.join(BASE_DIR, "backups")
    folder = os.path.join(base, today)
    os.makedirs(folder, exist_ok=True)

    tables = [
        ("tickets", "&order=id"),
        ("projects", "&order=id"),
        ("ticket_811_responses", "&order=id"),
        ("utility_contacts", "&order=id"),
        ("sync_811_log", "&order=id"),
        ("app_roles", ""),
    ]

    log.info("=" * 55)
    log.info("  BACKUP: Exportando banco de dados")
    log.info(f"  Destino: {folder}")
    log.info("=" * 55)

    total_rows = 0
    for table_name, query in tables:
        try:
            # Paginate: Supabase returns max 1000 rows per request
            all_rows = []
            offset = 0
            batch_size = 1000
            while True:
                data = sb_get(table_name, f"{query}&limit={batch_size}&offset={offset}")
                if not data:
                    break
                all_rows.extend(data)
                if len(data) < batch_size:
                    break
                offset += batch_size

            filepath = os.path.join(folder, f"{table_name}.json")
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(all_rows, f, ensure_ascii=False, indent=2, default=str)
            
            total_rows += len(all_rows)
            log.info(f"  ✅ {table_name}: {len(all_rows)} registros")
        except Exception as e:
            log.error(f"  ❌ {table_name}: {e}")

    # Gerar resumo
    summary = {
        "date": today,
        "tables": {},
        "total_rows": total_rows,
    }
    for table_name, _ in tables:
        fp = os.path.join(folder, f"{table_name}.json")
        if os.path.exists(fp):
            summary["tables"][table_name] = {
                "rows": len(json.load(open(fp, encoding="utf-8"))),
                "size_kb": round(os.path.getsize(fp) / 1024, 1),
            }
    with open(os.path.join(folder, "_summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    log.info(f"  📊 Total: {total_rows} registros exportados")

    # Cleanup: manter apenas os últimos 'keep_days' dias
    cutoff = (datetime.now() - timedelta(days=keep_days)).strftime("%Y-%m-%d")
    removed = 0
    if os.path.isdir(base):
        for d in sorted(os.listdir(base)):
            dpath = os.path.join(base, d)
            if os.path.isdir(dpath) and d < cutoff and d != today:
                try:
                    shutil.rmtree(dpath)
                    removed += 1
                except Exception:
                    pass
    if removed:
        log.info(f"  🗑️ {removed} backup(s) antigo(s) removido(s) (>{keep_days} dias)")

    log.info("=" * 55)
    log.info(f"  BACKUP CONCLUÍDO: {folder}")
    log.info("=" * 55)
    return folder


# ── │ SECTION: SELF_TESTS │ SELF TESTS (rodar com --selftest) ────────────────
def run_self_tests():
    """Testes internos para classify(), needs_private_locator(), append_auto_note() etc."""
    passed = 0
    failed = 0

    def _assert(label, expected, actual):
        nonlocal passed, failed
        if expected == actual:
            passed += 1
            print(f"  ✅ {label}")
        else:
            failed += 1
            print(f"  ❌ {label}: esperado={expected!r}, recebeu={actual!r}")

    print("\n=== OneDrill 811 Self-Tests ===\n")

    # ── classify() tests ──
    print("classify():")
    # CLEAR codes
    _assert("Clear: 1 Marked", "Clear", classify("Current", "1: Marked - Underground facilities have been marked")[0])
    _assert("Clear: 1B High-Profile", "Clear", classify("Current", "1B: Marked with Exceptions - High-Profile Utility")[0])
    _assert("Clear: 1C Work by Owner", "Clear", classify("Current", "1C: Work Being Done by Facility Owner")[0])
    _assert("Clear: 2 Clear", "Clear", classify("Current", "2: Clear - No underground facilities in the proposed excavation")[0])
    _assert("Clear: 2E Marked Exceptions", "Clear", classify("Current", "2E: Marked with Exceptions - Marked within confines")[0])
    _assert("Pending: 3E Already Performed", "Pending", classify("Current", "3E: Unmarked - Excavation Already Performed or Canceled")[0])
    _assert("Clear: 3U Not service", "Clear", classify("Positive Response", "3U: Unmarked - Not service provider for this location")[0])
    _assert("Clear: 3H Private", "Clear", classify("Current", "3H: Unmarked - Privately owned facilities on property")[0])
    _assert("Clear: 4 Clear No Facilities", "Clear", classify("Current", "4: Clear No Facilities")[0])
    _assert("Clear: 4 Private Line", "Clear", classify("Current", "4: Private Line - not responsibility of Indiana 811")[0])
    _assert("Clear: 5 No Conflict", "Clear", classify("Current", "5: No Conflict - utility is outside of the requested work site")[0])
    _assert("Clear: 5A Documents", "Clear", classify("Current", "5A: Design Notice - Documents Provided")[0])
    _assert("Clear: 5B Design Marked", "Clear", classify("Current", "5B: Design Notice - Marked")[0])
    _assert("Clear: 6C Joint Meet Complete", "Clear", classify("Current", "6C: Joint Meet Complete")[0])
    _assert("Clear: 1A Marked with Exceptions", "Clear", classify("Current", "1A: Marked with Exceptions - Do Not Excavate, High-Profile Utility")[0])
    _assert("Clear: 1A High-Profile variant", "Clear", classify("Current", "1A - Marked with Exceptions - Do Not Excavate, High-Profile Utility - Do not excavate")[0])
    _assert("Clear: status=Clear", "Clear", classify("Clear", "")[0])
    _assert("Clear: no facilit", "Clear", classify("Current", "No facilities in area")[0])
    # PENDING codes
    _assert("Pending: 3A No Access", "Pending", classify("Current", "3A: Unmarked - Could Not Gain Access to Property")[0])
    _assert("Pending: 3B Incorrect Address", "Pending", classify("Current", "3B: Unmarked - Incorrect Address Information")[0])
    _assert("Pending: 3C Marking Delay", "Pending", classify("Current", "3C: Unmarked - Marking Delay - Do not excavate until resolved")[0])
    _assert("Pending: 3D Instructions Unclear", "Pending", classify("Current", "3D: Unmarked - Marking Instructions are Unclear")[0])
    _assert("Pending: 3F Untonable", "Pending", classify("Current", "3F: Unmarked - Line is untonable")[0])
    _assert("Pending: 3G Ongoing", "Pending", classify("Current", "3G: Unmarked - Ongoing - partially marked")[0])
    _assert("Pending: 6A Joint Meet Conflict", "Pending", classify("Current", "6A: Joint Meet Conflict")[0])
    _assert("Pending: 6A Active Facilities FL", "Pending", classify("Current", "6A: Active Facilities are present - DO NOT demolish until the member notifies you the site is clear")[0])
    _assert("Pending: 6B Joint Meet Accepted", "Pending", classify("Current", "6B: Joint Meet Accepted")[0])
    _assert("Pending: 8 Ongoing Job", "Pending", classify("Current", "8: Ongoing Job - locate technician and excavator have established")[0])
    _assert("Pending: No Response", "Pending", classify("No Response", "")[0])
    _assert("Pending: unmarked generic", "Pending", classify("Current", "unmarked")[0])
    _assert("Pending: positive ambiguous", "Pending", classify("Positive Response", "")[0])
    # UNRECOGNIZED flag tests
    _assert("Recognized: 1 Marked → False", False, classify("Current", "1: Marked - Underground facilities have been marked")[1])
    _assert("Recognized: No Response → False", False, classify("No Response", "")[1])
    _assert("Unrecognized: empty Positive Response → True", True, classify("Positive Response", "")[1])
    _assert("Unrecognized: unknown text → True", True, classify("SomeWeirdStatus", "Facilities identified - excavator responsibility")[1])
    _assert("Recognized: 3A No Access → False", False, classify("Current", "3A: Unmarked - Could Not Gain Access to Property")[1])
    # IL/JULIE scheduling entries (recognized, not unrecognized)
    _assert("IL schedule: DECLINED CODE 50", "Pending", classify("950", "DECLINED CODE 50")[0])
    _assert("IL schedule: DECLINED CODE 50 recognized", False, classify("950", "DECLINED CODE 50")[1])
    _assert("IL schedule: ACCEPTED CODE 50", "Pending", classify("850", "ACCEPTED CODE 50")[0])
    _assert("IL schedule: ACCEPTED CODE 50 recognized", False, classify("850", "ACCEPTED CODE 50")[1])
    _assert("IL schedule: ALTERNATE SCHEDULE", "Pending", classify("50", "LOCATOR AND EXCAVATOR AGREED TO A DOCUMENTED ALTERNATE MARKING SCHEDULE")[0])
    _assert("IL schedule: ALTERNATE SCHEDULE recognized", False, classify("50", "LOCATOR AND EXCAVATOR AGREED TO A DOCUMENTED ALTERNATE MARKING SCHEDULE")[1])

    # ── needs_private_locator() tests ──
    print("\nneeds_private_locator():")
    _assert("3H detected", True, needs_private_locator("3H - Privately owned facility"))
    _assert("privately owned", True, needs_private_locator("This is a privately owned line"))
    _assert("private facility owner", True, needs_private_locator("Contact private facility owner"))
    _assert("normal response", False, needs_private_locator("1 - Marked"))
    _assert("empty string", False, needs_private_locator(""))

    # ── needs_watch_and_protect() tests ──
    print("\nneeds_watch_and_protect():")
    _assert("W&P detected", True, needs_watch_and_protect("WATCH AND PROTECT"))
    _assert("W&P in sentence", True, needs_watch_and_protect("60 - Watch and Protect required"))
    _assert("normal response", False, needs_watch_and_protect("1: Marked"))
    _assert("empty string", False, needs_watch_and_protect(""))
    # classify: code 60 W&P → Clear, recognized
    _assert("Clear: W&P code 60", "Clear", classify("60", "WATCH AND PROTECT")[0])
    _assert("Clear: W&P recognized", False, classify("60", "WATCH AND PROTECT")[1])
    # classify: RE-MARK NOT NEEDED (Code 21) → Pending (ack extensão, NÃO é clear real)
    _assert("Pending: RE-MARK NOT NEEDED", "Pending", classify("RE-MARK NOT NEEDED", "")[0])
    _assert("Pending: RE-MARK NOT NEEDED recognized", False, classify("RE-MARK NOT NEEDED", "")[1])
    _assert("Pending: REMARK NOT NEEDED variant", "Pending", classify("REMARK NOT NEEDED", "")[0])
    # classify: RE-MARK NEEDED (Code 22) → Pending (reset, precisa remarcar)
    _assert("Pending: RE-MARK NEEDED", "Pending", classify("RE-MARK NEEDED", "")[0])
    _assert("Pending: RE-MARK NEEDED recognized", False, classify("RE-MARK NEEDED", "")[1])
    _assert("Pending: REMARK NEEDED variant", "Pending", classify("REMARK NEEDED", "")[0])
    _assert("Pending: RE - MARK NEEDED", "Pending", classify("RE - MARK NEEDED", "")[0])

    # ── is_ticket_canceled() tests ──
    print("\nis_ticket_canceled():")
    _assert("CANCEL header", True, is_ticket_canceled("CANCEL\nSome body text"))
    _assert("FUNCTION: CANCEL", True, is_ticket_canceled("Header\nFunction: CANCEL\nMore text"))
    _assert("replaced by", True, is_ticket_canceled("CANCELED ticket\nREPLACED BY TICKET NUMBER 12345678"))
    _assert("normal ticket", False, is_ticket_canceled("Normal ticket body\nLocation: Main St"))
    _assert("empty", False, is_ticket_canceled(""))

    # ── is_in_renewal_grace() tests ──
    from datetime import date
    print("\nis_in_renewal_grace():")
    today = date(2026, 5, 21)
    _assert("sem old_ticket2", (False, ""), is_in_renewal_grace({}, ref_date=today))
    _assert("com old mas sem expire", (False, "12345678"), is_in_renewal_grace({"old_ticket2": "12345678", "expire_old": ""}, ref_date=today))
    _assert("expire_old = traço", (False, "12345678"), is_in_renewal_grace({"old_ticket2": "12345678", "expire_old": "—"}, ref_date=today))
    _assert("em grace (futuro)", (True, "12345678"), is_in_renewal_grace({"old_ticket2": "12345678", "expire_old": "06/15/2026"}, ref_date=today))
    _assert("em grace (hoje)", (True, "12345678"), is_in_renewal_grace({"old_ticket2": "12345678", "expire_old": "05/21/2026"}, ref_date=today))
    _assert("fora de grace (ontem)", (False, "12345678"), is_in_renewal_grace({"old_ticket2": "12345678", "expire_old": "05/20/2026"}, ref_date=today))
    _assert("fora de grace (passado)", (False, "12345678"), is_in_renewal_grace({"old_ticket2": "12345678", "expire_old": "01/01/2026"}, ref_date=today))
    _assert("expire poluído Time:", (True, "12345678"), is_in_renewal_grace({"old_ticket2": "12345678", "expire_old": "06/15/26 Time: 23:59"}, ref_date=today))
    _assert("chain com seta", (True, "11111111"), is_in_renewal_grace({"old_ticket2": "11111111 → 22222222", "expire_old": "06/15/2026"}, ref_date=today))
    _assert("expire lixo", (False, "12345678"), is_in_renewal_grace({"old_ticket2": "12345678", "expire_old": "nao e data"}, ref_date=today))

    # ── _pdf_query_number() tests ──
    print("\n_pdf_query_number():")
    _assert("sem renovação", ("NEW123", False), _pdf_query_number({"ticket": "NEW123"}))
    _assert("renovado em grace", ("OLD123", True), _pdf_query_number({"ticket": "NEW123", "old_ticket2": "OLD123", "expire_old": "12/31/2030"}))
    _assert("renovado fora de grace", ("NEW123", False), _pdf_query_number({"ticket": "NEW123", "old_ticket2": "OLD123", "expire_old": "01/01/2020"}))
    _assert("renovado sem expire", ("NEW123", False), _pdf_query_number({"ticket": "NEW123", "old_ticket2": "OLD123", "expire_old": ""}))

    # ── append_auto_note() tests ──
    print("\nappend_auto_note():")
    _assert("adds to empty", "[AUTO 811] Clear", append_auto_note("", "[AUTO 811] Clear"))
    _assert("dedup exact", "existing\n[AUTO 811] Clear", append_auto_note("existing\n[AUTO 811] Clear", "[AUTO 811] Clear"))
    _assert("preserves manual", True, "manual note" in append_auto_note("manual note", "[AUTO 811] New"))
    _assert("truncates excess", True, len(append_auto_note(
        "\n".join(f"[AUTO 811] Note {i}" for i in range(15)),
        "[AUTO 811] Note 15"
    ).split("\n")) <= MAX_AUTO_NOTES + 1)

    # ── _get_latest_response_date() tests ──
    print("\n_get_latest_response_date():")
    _assert("ISO dates: picks latest", True,
        _get_latest_response_date([
            {"utility": "A", "responded_date": "2026-03-23T10:11:00"},
            {"utility": "B", "responded_date": "2026-03-25T09:30:00"},
            {"utility": "C", "responded_date": "2026-03-24T08:50:00"},
        ])[0].strftime('%m/%d/%Y') == "03/25/2026")
    _assert("ISO dates: not fallback", False,
        _get_latest_response_date([
            {"utility": "A", "responded_date": "2026-03-25T09:30:00"},
        ])[1])
    _assert("mixed formats", True,
        _get_latest_response_date([
            {"utility": "A", "responded_date": "03/28/2026"},
            {"utility": "B", "responded_date": "2026-03-25T09:30:00"},
        ])[0].strftime('%m/%d/%Y') == "03/28/2026")
    _assert("no dates: returns now (fallback)", True,
        _get_latest_response_date([
            {"utility": "A", "status": "Clear"},
        ])[0].date() == datetime.now().date())
    _assert("no dates: is_fallback=True", True,
        _get_latest_response_date([
            {"utility": "A", "status": "Clear"},
        ])[1])
    _assert("empty list: returns now", True,
        _get_latest_response_date([])[0].date() == datetime.now().date())
    _assert("empty list: is_fallback=True", True,
        _get_latest_response_date([])[1])

    # ── extract_expire_date() tests ──
    print("\nextract_expire_date():")
    _assert("mm/dd/yyyy", "04/15/2026", extract_expire_date("Expire on: 04/15/2026\nNext line"))
    # Due Date ISOLADO deve ser IGNORADO (era bug antigo) — retorna vazio
    _assert("due date isolado: IGNORADO", "", extract_expire_date("Due Date: 03/20/2026\nBody"))
    _assert("no date", "", extract_expire_date("No expiration info here"))
    # Formato Sunshine FL (Exp Date + Due Date na mesma linha)
    sunshine_body = "Due Date : 04/15/26 Time: 23:59ET  Exp Date : 05/13/26 Time: 23:59ET"
    _assert("sunshine FL: Exp Date", "05/13/2026", extract_expire_date(sunshine_body))
    # Ticket Expires com hora AM/PM
    _assert("Indiana: Ticket Expires PM", "05/13/2026",
            extract_expire_date("Ticket Expires: 05/13/2026 11:59 PM"))
    # Formato com "at"
    _assert("expires at", "05/13/2026", extract_expire_date("Expires: 05/13/2026 at 11:59 PM"))

    # ── normalize_expire() tests ──
    print("\nnormalize_expire():")
    _assert("formato poluído legado",     "04/15/2026", normalize_expire("04/15/26 Time: 23:59"))
    _assert("já normalizado",             "05/13/2026", normalize_expire("05/13/2026"))
    _assert("2 dígitos de ano",           "05/13/2026", normalize_expire("05/13/26"))
    _assert("com sufixo ET",              "04/15/2026", normalize_expire("04/15/26 Time: 23:59ET"))
    _assert("com AM/PM",                  "05/13/2026", normalize_expire("05/13/2026 11:59 PM"))
    _assert("vazio",                      "",           normalize_expire(""))
    _assert("traço",                      "",           normalize_expire("—"))
    _assert("None",                       "",           normalize_expire(None))
    # Fix bug #24: casos de paridade JS ↔ Python.
    # Se alguém mudar Python sem mudar JS (ou vice-versa), aqui disparamos.
    # Os mesmos 23 casos estão em testNormalizeExpireParity() no app.js.
    _assert("parity: 5/13/2026",          "05/13/2026", normalize_expire("5/13/2026"))
    _assert("parity: 5/13/26",            "05/13/2026", normalize_expire("5/13/26"))
    _assert("parity: ISO 2026-05-13",     "05/13/2026", normalize_expire("2026-05-13"))
    _assert("parity: ISO 2026-5-13",      "05/13/2026", normalize_expire("2026-5-13"))
    _assert("parity: May 13, 2026",       "05/13/2026", normalize_expire("May 13, 2026"))
    _assert("parity: Jan 5, 2026",        "01/05/2026", normalize_expire("Jan 5, 2026"))
    _assert("parity: December 31, 2025",  "12/31/2025", normalize_expire("December 31, 2025"))
    _assert("parity: Feb 29, 2024",       "02/29/2024", normalize_expire("Feb 29, 2024"))
    _assert("parity: 05/13/2026 23:59",   "05/13/2026", normalize_expire("05/13/2026 23:59"))
    _assert("parity: 05/13/26 23:59ET",   "05/13/2026", normalize_expire("05/13/26 23:59ET"))
    _assert("parity: at 23:59",           "05/13/2026", normalize_expire("05/13/2026 at 23:59"))
    _assert("parity: N/A → vazio",        "",           normalize_expire("N/A"))
    _assert("parity: None str → vazio",   "",           normalize_expire("None"))
    _assert("parity: null str → vazio",   "",           normalize_expire("null"))
    _assert("parity: -  → vazio",         "",           normalize_expire("-"))
    _assert("parity: lixo → vazio",       "",           normalize_expire("lixo que não é data"))
    _assert("parity: pt-BR ambíguo",      "",           normalize_expire("13/05/2026"))

    # ── _is_polluted_expire() tests ──
    print("\n_is_polluted_expire():")
    _assert("formato legado 'Time:'", True,  _is_polluted_expire("04/15/26 Time: 23:59"))
    _assert("formato limpo",          False, _is_polluted_expire("05/13/2026"))
    _assert("ano 2 dígitos",          True,  _is_polluted_expire("05/13/26"))
    _assert("vazio",                  False, _is_polluted_expire(""))
    _assert("com hora bruta",         True,  _is_polluted_expire("05/13/2026 11:59"))
    _assert("com ET",                 True,  _is_polluted_expire("05/13/2026 11:59ET"))

    # ── _is_valid_utility_name() tests ──
    print("\n_is_valid_utility_name():")
    # Válidos (utilities reais)
    _assert("DUKE ENERGY",              True,  _is_valid_utility_name("DUKE ENERGY"))
    _assert("COMCAST NORTH",            True,  _is_valid_utility_name("COMCAST NORTH"))
    _assert("IN AMERICAN WATER",        True,  _is_valid_utility_name("IN AMERICAN WATER"))
    _assert("NIPSCO GAS & ELECTRIC",    True,  _is_valid_utility_name("NIPSCO GAS & ELECTRIC (VALPARAISO)"))
    _assert("FRONTIER",                 True,  _is_valid_utility_name("FRONTIER"))
    _assert("AT&T",                     True,  _is_valid_utility_name("AT&T"))
    # Siglas curtas reais (bug fix — MCI era rejeitada por ter ≤5 chars)
    _assert("MCI",                      True,  _is_valid_utility_name("MCI"))
    _assert("AEP",                      True,  _is_valid_utility_name("AEP"))
    _assert("TECO",                     True,  _is_valid_utility_name("TECO"))
    # Inválidos (lixo de UI — era o bug)
    _assert("All (6)",                  False, _is_valid_utility_name("All (6)"))
    _assert("All (7)",                  False, _is_valid_utility_name("All (7)"))
    _assert("All (9)",                  False, _is_valid_utility_name("All (9)"))
    _assert("Current (3)",              False, _is_valid_utility_name("Current (3)"))
    _assert("Show all (9)",             False, _is_valid_utility_name("Show all (9)"))
    _assert("All",                      False, _is_valid_utility_name("All"))
    _assert("Event",                    False, _is_valid_utility_name("Event"))
    _assert("Positive Response",        False, _is_valid_utility_name("Positive Response"))
    _assert("No Response",              False, _is_valid_utility_name("No Response"))
    _assert("Status",                   False, _is_valid_utility_name("Status"))
    _assert("Service Area",             False, _is_valid_utility_name("Service Area"))
    # Inválidos (códigos/IDs)
    _assert("ID2227",                   False, _is_valid_utility_name("ID2227"))
    _assert("NI0005",                   False, _is_valid_utility_name("NI0005"))
    _assert("COMCN (sigla sem dígito)",  True,  _is_valid_utility_name("COMCN"))
    # Inválidos (vazio/curto)
    _assert("vazio",                    False, _is_valid_utility_name(""))
    _assert("None",                     False, _is_valid_utility_name(None))
    _assert("muito curto",              False, _is_valid_utility_name("AB"))

    print(f"\n{'='*40}")
    print(f"Resultado: {passed} passou, {failed} falhou")
    if failed == 0:
        print("✅ Todos os testes passaram!")
    else:
        print("❌ Alguns testes falharam.")
        sys.exit(1)
    print()


# ── │ SECTION: FIX_RENEWALS │ FIX RENEWALS: corrigir expire_old de tickets renovados 
async def fix_renewals():
    """Corrige expire_old de tickets renovados buscando a data real no portal.

    Bug: no merge de renovação, o código antigo sobrescrevia t.expire com o
    expire do ticket novo ANTES de capturar oldExpire, fazendo expire_old
    ficar igual ao expire (data do ticket novo, não do antigo).

    Este comando:
    1. Busca todos os tickets com old_ticket2 (renovados)
    2. Para cada um, extrai o número do ticket antigo
    3. Scrapa o ticket antigo no portal 811 para obter a data de expiração real
    4. Se expire_old estiver errado, corrige
    """
    all_renewed = sb_get("tickets", "&old_ticket2=not.is.null&select=id,ticket,state,old_ticket2,expire,expire_old")
    if not all_renewed:
        log.info("[FixRenewals] Nenhum ticket renovado encontrado")
        return

    log.info(f"[FixRenewals] {len(all_renewed)} tickets renovados encontrados")

    # Agrupa por estado
    by_state = {}
    for t in all_renewed:
        state = t.get("state", "")
        if state not in by_state:
            by_state[state] = []
        # Extrai número do ticket antigo (primeiro da cadeia)
        old_chain = (t.get("old_ticket2") or "").strip()
        old_num = old_chain.split(" → ")[0].strip() if old_chain else ""
        if not old_num:
            continue
        by_state[state].append({
            "id": t["id"],
            "ticket": t["ticket"],
            "old_num": old_num,
            "expire": t.get("expire", ""),
            "expire_old": t.get("expire_old", ""),
        })

    fixed = 0
    skipped = 0
    errors = 0

    for state, items in by_state.items():
        if state not in PORTALS:
            log.warning(f"[FixRenewals] Estado {state} sem portal configurado — pulando {len(items)} tickets")
            skipped += len(items)
            continue

        log.info(f"[FixRenewals] [{state}] Verificando {len(items)} tickets renovados...")
        perfil = _profile_path(state)

        async with async_playwright() as p:
            ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
            page = ctx.pages[0] if ctx.pages else await ctx.new_page()
            page.set_default_timeout(TIMEOUT_PAGE)

            await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
            await wait_stable(page)

            if "login" in page.url.lower():
                await ctx.close()
                ok = await auto_login_silent(state)

                if not ok:

                    log.warning(f"[{state}] auto_login_silent falhou, tentando manual...")

                    ok = await auto_login(state)
                if not ok:
                    log.error(f"[FixRenewals] [{state}] Login falhou — pulando")
                    errors += len(items)
                    continue
                ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
                page = ctx.pages[0] if ctx.pages else await ctx.new_page()
                page.set_default_timeout(TIMEOUT_PAGE)

            ok = await goto_dashboard(page, state)
            if not ok:
                log.error(f"[FixRenewals] [{state}] Dashboard inacessível — pulando")
                await ctx.close()
                errors += len(items)
                continue

            for idx, item in enumerate(items):
                old_num = item["old_num"]
                ticket = item["ticket"]
                tid = item["id"]
                current_expire_old = item["expire_old"]

                log.info(f"[FixRenewals] [{state}] ({idx+1}/{len(items)}) {ticket} → antigo {old_num} (expire_old atual: {current_expire_old})")

                try:
                    await filter_ticket(page, old_num)

                    if not await page.get_by_text(old_num, exact=True).count():
                        log.warning(f"[FixRenewals] {old_num}: não encontrado no portal")
                        skipped += 1
                        continue

                    await page.get_by_text(old_num, exact=True).first.click()
                    await wait_stable(page)

                    # Vai na aba Text pra pegar o body completo
                    tt = page.locator('[role="tab"]:has-text("Text")').first
                    if await tt.count():
                        await click_and_wait(page, tt, "tab")

                    body = await page.locator("body").inner_text()
                    real_expire = normalize_expire(extract_expire_date(body))

                    if not real_expire:
                        log.warning(f"[FixRenewals] {old_num}: expire não encontrado no body")
                        skipped += 1
                        await back_to_dashboard(page, state)
                        continue

                    # Compara (normaliza também o que está no banco pra comparação justa)
                    current_normalized = normalize_expire(current_expire_old)
                    if real_expire == current_normalized:
                        log.info(f"[FixRenewals] {ticket}: expire_old OK ({real_expire})")
                        skipped += 1
                    else:
                        log.warning(f"[FixRenewals] {ticket}: CORRIGINDO expire_old {current_expire_old} → {real_expire}")
                        sb_patch("tickets", tid, {"expire_old": real_expire})
                        fixed += 1

                    await back_to_dashboard(page, state)

                except Exception as e:
                    log.error(f"[FixRenewals] {old_num}: ERRO → {e}")
                    errors += 1
                    try:
                        await back_to_dashboard(page, state)
                    except Exception:
                        pass

            await ctx.close()

    log.info(f"[FixRenewals] === CONCLUÍDO: {fixed} corrigidos, {skipped} OK/pulados, {errors} erros ===")


# ── │ SECTION: FIX_EXPIRES │ FIX EXPIRES: corrigir data de vencimento de tickets com formato antigo 
async def fix_expires(target_ticket=None):
    """Re-scrapa a data de vencimento (expire) de tickets ativos no portal.

    Usado pra corrigir tickets que:
      - Foram importados por versão antiga do parser (formato "MM/DD/YY Time: HH:MM")
      - Tiveram expire capturada errada (pegou Due Date ou Work Date por engano)
      - Têm expire vazia

    Se `target_ticket` for informado, corrige apenas esse ticket.
    """
    log.info("=" * 55)
    log.info(f"  FIX-EXPIRES: Corrigindo data de vencimento")
    if target_ticket:
        log.info(f"  Ticket alvo: {target_ticket}")
    log.info("=" * 55)

    qs = "&status=in.(Open,Damage,Clear)&select=id,ticket,state,expire"
    if target_ticket:
        qs += f"&ticket=eq.{target_ticket}"
    all_active = sb_get("tickets", qs)

    if not all_active:
        log.info("[FixExpires] Nenhum ticket ativo encontrado")
        return

    # Agrupa por estado
    by_state = {}
    for t in all_active:
        state = t.get("state", "")
        if state not in PORTALS:
            continue  # IL (JULIE) não tem expire no body
        by_state.setdefault(state, []).append(t)

    if not by_state:
        log.warning("[FixExpires] Nenhum ticket em estado com portal (FL/IN)")
        return

    fixed = 0
    unchanged = 0
    errors = 0

    for state, items in by_state.items():
        log.info(f"[FixExpires] [{state}] {len(items)} tickets a verificar")
        perfil = _profile_path(state)

        async with async_playwright() as p:
            ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
            page = ctx.pages[0] if ctx.pages else await ctx.new_page()
            page.set_default_timeout(TIMEOUT_PAGE)

            await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
            await wait_stable(page)
            page, ctx = await ensure_login(page, ctx, p, state)
            if not page:
                log.error(f"[FixExpires] [{state}] Login falhou — pulando estado")
                errors += len(items)
                continue

            ok = await goto_dashboard(page, state)
            if not ok:
                log.error(f"[FixExpires] [{state}] Dashboard inacessível — pulando")
                await ctx.close()
                errors += len(items)
                continue

            for idx, t in enumerate(items):
                tnum = t["ticket"]
                tid = t["id"]
                current = (t.get("expire") or "").strip()

                try:
                    await filter_ticket(page, tnum)
                    if not await page.get_by_text(tnum, exact=True).count():
                        log.warning(f"[FixExpires] {tnum}: não encontrado no portal")
                        errors += 1
                        continue

                    await page.get_by_text(tnum, exact=True).first.click()
                    await wait_stable(page)

                    tt = page.locator('[role="tab"]:has-text("Text")').first
                    if await tt.count():
                        await click_and_wait(page, tt, "tab")

                    body = await page.locator("body").inner_text()
                    new_expire = normalize_expire(extract_expire_date(body, ticket_num=tnum, debug=True))
                    current_normalized = normalize_expire(current)

                    if not new_expire and current_normalized and _is_polluted_expire(current):
                        # Parser não achou no portal, mas dá pra limpar o valor do banco
                        log.warning(f"[FixExpires] {tnum}: scraper falhou, mas NORMALIZANDO {current} → {current_normalized}")
                        sb_patch("tickets", tid, {"expire": current_normalized})
                        fixed += 1
                    elif not new_expire:
                        log.warning(f"[FixExpires] {tnum}: novo expire vazio — mantendo atual ({current})")
                        errors += 1
                    elif new_expire == current:
                        log.info(f"[FixExpires] {tnum}: OK ({current})")
                        unchanged += 1
                    else:
                        log.warning(f"[FixExpires] {tnum}: CORRIGINDO {current} → {new_expire}")
                        sb_patch("tickets", tid, {"expire": new_expire})
                        fixed += 1

                    await back_to_dashboard(page, state)
                except Exception as e:
                    log.error(f"[FixExpires] {tnum}: ERRO → {e}")
                    errors += 1
                    try:
                        await back_to_dashboard(page, state)
                    except Exception:
                        pass

            await ctx.close()

    log.info(f"[FixExpires] === CONCLUÍDO: {fixed} corrigidos, {unchanged} OK, {errors} erros ===")


# ── │ SECTION: POPULATE_COUNTIES │ POPULATE COUNTIES: backfill de county em tickets antigos ──
async def populate_counties(target_state=None, force=False):
    """Popula o campo `county` em tickets que ainda não têm.

    Usa resolve_county() com estratégia 3-em-cascata (base estática → variantes → Nominatim).
    Por default processa só tickets com county IS NULL/vazio. Use force=True pra re-resolver todos.

    Args:
        target_state: se informado ('FL'/'IN'/'IL'), só processa aquele estado. None = todos.
        force: se True, re-processa tickets que já têm county (útil se base foi atualizada).
    """
    log.info("=" * 55)
    log.info(f"  POPULATE-COUNTIES: backfill de county")
    log.info(f"  Estado: {target_state or 'TODOS'} | Force: {force}")
    log.info("=" * 55)

    # Query com filtro
    qs = "&select=id,ticket,state,location,geocoded_lat,geocoded_lon,county"
    if target_state:
        qs += f"&state=eq.{_qv(target_state.upper())}"
    # Sem force, só pega os que precisam (county vazio ou null)
    if not force:
        qs += "&or=(county.is.null,county.eq.)"

    tickets_list = sb_get("tickets", qs)
    if not tickets_list:
        log.info("[PopCounties] Nenhum ticket pra processar.")
        return

    log.info(f"[PopCounties] {len(tickets_list)} tickets a processar")

    resolved = 0
    unchanged = 0
    unresolved = 0
    by_source = {"static": 0, "nominatim": 0, "none": 0}

    for i, t in enumerate(tickets_list, 1):
        tid = t["id"]
        tnum = t.get("ticket", "?")
        state = t.get("state", "")
        location = t.get("location", "")
        lat = t.get("geocoded_lat")
        lon = t.get("geocoded_lon")
        current = (t.get("county") or "").strip()

        # Pra contabilizar qual estratégia resolveu, fazemos lookup estático primeiro
        # direto (sem o fallback Nominatim) pra saber se veio da base.
        db = _load_counties_db()
        static_hit = ""
        city_raw = (location or "").split(",")[0].strip()
        if city_raw and state:
            for variant in _city_variants(city_raw):
                if variant in db.get(state.upper(), {}):
                    static_hit = db[state.upper()][variant]
                    break

        if static_hit:
            new_county = static_hit
            by_source["static"] += 1
        else:
            # Fallback Nominatim (só se tem lat/lon)
            new_county = await resolve_county(location, state, lat, lon)
            if new_county:
                by_source["nominatim"] += 1
            else:
                by_source["none"] += 1

        if new_county and new_county != current:
            try:
                sb_patch("tickets", tid, {"county": new_county})
                resolved += 1
                log.info(f"[PopCounties] [{i}/{len(tickets_list)}] {tnum} ({state}): {location!r} → {new_county}")
            except Exception as e:
                log.error(f"[PopCounties] Erro salvando {tnum}: {e}")
        elif new_county == current and current:
            unchanged += 1
        else:
            unresolved += 1
            log.debug(f"[PopCounties] [{i}/{len(tickets_list)}] {tnum} ({state}): {location!r} → não resolvido")

    log.info("=" * 55)
    log.info(f"[PopCounties] CONCLUÍDO")
    log.info(f"  Resolvidos:   {resolved}")
    log.info(f"  Já corretos:  {unchanged}")
    log.info(f"  Sem match:    {unresolved}")
    log.info(f"  Origem — base estática: {by_source['static']} | Nominatim: {by_source['nominatim']} | none: {by_source['none']}")
    log.info("=" * 55)


# ── │ SECTION: REBUILD_COVERAGE │ REBUILD COVERAGE: auto-derive utility × county ──
# Fase 2 do filtro por county: analisa ticket_811_responses × tickets.county
# e constroi a tabela utility_county_coverage (qual utility atende qual county).

# Threshold mínimo de respostas pra uma utility ser considerada "atende" um county.
# Evita falsos positivos de utility que respondeu 1-2x por engano em county adjacente.
COVERAGE_MIN_RESPONSES = 5

# Status que NÃO contam pra inferência (são eventos administrativos, não cobertura real):
#   "Private Locator" = sinalização de que precisa locator privado (código 3H)
#   "Watch and Protect" = representante obrigatório (código 60 IL)
#   "Unrecognized" = parser não entendeu a resposta (não usar pra inferir)
COVERAGE_EXCLUDED_STATUSES = {"Private Locator", "Watch and Protect", "Unrecognized"}


def rebuild_utility_county_coverage():
    """Reconstroi a tabela utility_county_coverage a partir das respostas históricas.

    Estratégia (pull-based — pega tudo do banco e agrega localmente pra evitar queries N+1):
      1) Busca TODOS os tickets com county preenchido (select: ticket, county, state)
      2) Busca TODAS as respostas válidas (select: ticket_num, utility_name, status)
      3) Cruza em memória: pra cada (utility, county, state) conta quantas respostas
      4) Filtra pelos que atingiram threshold
      5) Upsert em utility_county_coverage (sobrescreve, não acumula entre execuções)

    Safety: antes de substituir, deleta só as linhas cobertas pela nova análise —
    evita apagar dados se a query retornar vazio por bug.
    """
    log.info("=" * 55)
    log.info(f"  REBUILD-COVERAGE: auto-derive utility × county")
    log.info(f"  Threshold: ≥{COVERAGE_MIN_RESPONSES} respostas | Excluídos: {COVERAGE_EXCLUDED_STATUSES}")
    log.info("=" * 55)

    # 1. Tickets com county válido
    tickets_with_county = sb_get("tickets", "&select=ticket,county,state&county=not.is.null&county=neq.")
    if not tickets_with_county:
        log.warning("[Coverage] Nenhum ticket com county preenchido — rodar --populate-counties primeiro")
        return

    # Map: ticket_num → (county, state)
    ticket_map = {}
    for t in tickets_with_county:
        tn = (t.get("ticket") or "").strip()
        c = (t.get("county") or "").strip()
        s = (t.get("state") or "").strip().upper()
        if tn and c and s:
            ticket_map[tn] = (c, s)
    log.info(f"[Coverage] {len(ticket_map)} tickets com county válido carregados")

    # 2. Respostas em lotes (pagina pra evitar timeout em bases grandes)
    all_responses = []
    offset = 0
    page_size = 1000
    while True:
        page = sb_get(
            "ticket_811_responses",
            f"&select=ticket_num,utility_name,status&order=id&limit={page_size}&offset={offset}"
        )
        if not page:
            break
        all_responses.extend(page)
        if len(page) < page_size:
            break
        offset += page_size
    log.info(f"[Coverage] {len(all_responses)} respostas totais carregadas")

    # 3. Agregação em memória
    # Chave: (utility_name, county, state) → count
    counts = {}
    skipped_invalid_utility = 0
    skipped_excluded_status = 0
    skipped_no_county = 0

    for r in all_responses:
        utility = (r.get("utility_name") or "").strip()
        status = (r.get("status") or "").strip()
        ticket_num = (r.get("ticket_num") or "").strip()

        # Filtra utility lixo (UI, códigos puros, etc)
        if not _is_valid_utility_name(utility):
            skipped_invalid_utility += 1
            continue

        # Filtra status administrativos
        if status in COVERAGE_EXCLUDED_STATUSES:
            skipped_excluded_status += 1
            continue

        # Liga ticket → county/state
        if ticket_num not in ticket_map:
            skipped_no_county += 1
            continue
        county, state = ticket_map[ticket_num]

        key = (utility, county, state)
        counts[key] = counts.get(key, 0) + 1

    log.info(f"[Coverage] Agregação concluída:")
    log.info(f"  Combinações (utility × county × state): {len(counts)}")
    log.info(f"  Respostas skipadas — utility inválida:  {skipped_invalid_utility}")
    log.info(f"  Respostas skipadas — status excluído:   {skipped_excluded_status}")
    log.info(f"  Respostas skipadas — ticket sem county: {skipped_no_county}")

    # 4. Filtra pelo threshold
    qualified = [(u, c, s, cnt) for (u, c, s), cnt in counts.items() if cnt >= COVERAGE_MIN_RESPONSES]
    log.info(f"[Coverage] {len(qualified)} combinações atingiram threshold de {COVERAGE_MIN_RESPONSES}+ respostas")

    if not qualified:
        log.warning("[Coverage] Nenhuma combinação atingiu o threshold. Nada a gravar.")
        return

    # 5. Upsert em batches
    # Nota: como a unique key é (utility_name, county, state), upsert com on_conflict
    # sobrescreve o response_count e last_seen — comportamento desejado.
    records = []
    from datetime import datetime as _dt, timezone as _tz
    now_iso = _dt.now(_tz.utc).isoformat().replace("+00:00", "Z")
    for u, c, s, cnt in qualified:
        records.append({
            "utility_name": u,
            "county": c,
            "state": s,
            "response_count": cnt,
            "last_seen": now_iso,
        })

    # Upsert em lotes de 100
    batch_size = 100
    inserted = 0
    errors = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        try:
            # Bug fix: _sb_request espera função (requests.post), não string "POST".
            # on_conflict via query string (padrão do resto do código em sb_upsert).
            url = f"{SB_URL}/rest/v1/utility_county_coverage?on_conflict=utility_name,county,state"
            headers = {**SB_H, "Prefer": "resolution=merge-duplicates,return=minimal"}
            resp = _sb_request(requests.post, url, headers=headers, json=batch, timeout=30)
            if resp.status_code in (200, 201, 204):
                inserted += len(batch)
            else:
                log.error(f"[Coverage] Erro batch {i}-{i+len(batch)}: HTTP {resp.status_code} — {resp.text[:200]}")
                errors += len(batch)
        except Exception as e:
            log.error(f"[Coverage] Exceção batch {i}: {e}")
            errors += len(batch)

    # Top 10 coberturas pra amostra no log
    qualified_sorted = sorted(qualified, key=lambda x: -x[3])
    log.info("[Coverage] Top 10 coberturas (por nº de respostas):")
    for u, c, s, cnt in qualified_sorted[:10]:
        log.info(f"  {cnt:>4}x  {u} → {c} County, {s}")

    log.info("=" * 55)
    log.info(f"[Coverage] CONCLUÍDO: {inserted} gravados, {errors} erros")
    log.info("=" * 55)


# ── │ SECTION: AUDIT │ CLEAN GHOST UTILITIES: limpa respostas fantasmas do banco 
def _audit_collect():
    """Coleta anomalias do banco sem apresentar.

    Retorna dict com 6 listas (A..F), contagens e contexto.
    Reutilizável pra print e pra email.
    """
    tickets = sb_get("tickets", "&status=in.(Open,Damage,Clear)&select=id,ticket,state,status,expire,expire_old,old_ticket2,status_old")

    all_utils = []
    offset = 0
    while True:
        page = sb_get("ticket_811_responses", f"&select=id,ticket_num,utility_name,status&limit=1000&offset={offset}")
        if not page:
            break
        all_utils.extend(page)
        if len(page) < 1000:
            break
        offset += 1000

    pending_by_ticket = {}
    for u in all_utils:
        if u.get("status") == "Pending":
            tn = str(u.get("ticket_num", "")).strip()
            if not tn:
                continue
            pending_by_ticket.setdefault(tn, []).append(u.get("utility_name", ""))

    a_expire_poluted = []
    b_expire_old_poluted = []
    c_expire_empty = []
    d_false_clear = []
    e_ghost_utils = []
    f_renewed_no_expire_old = []

    for t in tickets:
        tnum = t.get("ticket", "")
        exp = (t.get("expire") or "").strip()
        exp_old = (t.get("expire_old") or "").strip()
        is_renewed_t = bool(t.get("old_ticket2"))
        status = t.get("status", "")

        if exp and _is_polluted_expire(exp):
            a_expire_poluted.append({"ticket": tnum, "state": t.get("state"), "expire": exp, "normalized": normalize_expire(exp)})
        if is_renewed_t and exp_old and _is_polluted_expire(exp_old):
            b_expire_old_poluted.append({"ticket": tnum, "state": t.get("state"), "expire_old": exp_old, "normalized": normalize_expire(exp_old)})
        if not exp and status in ("Open", "Damage"):
            c_expire_empty.append({"ticket": tnum, "state": t.get("state"), "status": status})
        if status == "Clear":
            pend = pending_by_ticket.get(str(tnum).strip(), [])
            if pend:
                d_false_clear.append({"ticket": tnum, "state": t.get("state"), "pending_utils": pend})
        if is_renewed_t and not exp_old:
            f_renewed_no_expire_old.append({"ticket": tnum, "state": t.get("state"), "old_ticket2": t.get("old_ticket2")})

    for u in all_utils:
        name = u.get("utility_name", "")
        if not _is_valid_utility_name(name):
            e_ghost_utils.append({
                "id": u.get("id"),
                "ticket_num": u.get("ticket_num"),
                "utility_name": name,
                "status": u.get("status"),
            })

    total_affected = len(a_expire_poluted) + len(b_expire_old_poluted) + len(c_expire_empty) + len(d_false_clear) + len(f_renewed_no_expire_old)
    total_issues = total_affected + len(e_ghost_utils)

    return {
        "tickets_checked": len(tickets),
        "responses_checked": len(all_utils),
        "A_expire_poluted": a_expire_poluted,
        "B_expire_old_poluted": b_expire_old_poluted,
        "C_expire_empty": c_expire_empty,
        "D_false_clear": d_false_clear,
        "E_ghost_utils": e_ghost_utils,
        "F_renewed_no_expire_old": f_renewed_no_expire_old,
        "total_affected_tickets": total_affected,
        "total_issues": total_issues,
    }


def _audit_format_report(data, max_show=20):
    """Formata o resultado de _audit_collect como string texto (pra print ou email).

    max_show = 0 retorna só o resumo (sem listagens).
    """
    lines = []
    lines.append("=" * 60)
    lines.append("           RELATÓRIO DE AUDITORIA DO BANCO")
    lines.append("=" * 60)
    lines.append(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    lines.append("")

    sections = [
        ("A. EXPIRE POLUÍDO (formato antigo tipo '04/15/26 Time: 23:59')",
            data["A_expire_poluted"], "python 811_sync.py --fix-expires"),
        ("B. EXPIRE_OLD POLUÍDO (renovados com formato antigo)",
            data["B_expire_old_poluted"], "python 811_sync.py --fix-renewals"),
        ("C. EXPIRE VAZIO em ticket Open/Damage (scraper não achou data)",
            data["C_expire_empty"], "editar manualmente no app OU aguardar próximo sync"),
        ("D. FALSO CLEAR (status=Clear mas utilities Pending - perigoso!)",
            data["D_false_clear"], "scraper reverte automaticamente no próximo sync"),
        ("E. GHOST UTILITIES (lixo de UI: 'All (N)', 'Event', etc)",
            data["E_ghost_utils"], "python 811_sync.py --clean-ghost-utilities"),
        ("F. RENOVADO SEM expire_old (cadeia de renovação quebrada)",
            data["F_renewed_no_expire_old"], "python 811_sync.py --fix-renewals"),
    ]

    for label, items, fix in sections:
        lines.append(f"── {label} ──")
        lines.append(f"  Total: {len(items)}")
        if items and max_show > 0:
            lines.append(f"  Como corrigir: {fix}")
            lines.append(f"  Primeiros {min(max_show, len(items))}:")
            for it in items[:max_show]:
                bits = [f"{k}={v!r}" for k, v in it.items() if k != "id"]
                lines.append(f"    - {' · '.join(bits)}")
            if len(items) > max_show:
                lines.append(f"    ... + {len(items) - max_show} outros")
        lines.append("")

    lines.append("=" * 60)
    lines.append(f"RESUMO: {data['tickets_checked']} tickets ativos auditados")
    lines.append(f"  A. expire poluído:              {len(data['A_expire_poluted']):4d}")
    lines.append(f"  B. expire_old poluído:          {len(data['B_expire_old_poluted']):4d}")
    lines.append(f"  C. expire vazio (Open/Damage):  {len(data['C_expire_empty']):4d}")
    lines.append(f"  D. falso Clear:                 {len(data['D_false_clear']):4d}")
    lines.append(f"  E. ghost utilities:             {len(data['E_ghost_utils']):4d}")
    lines.append(f"  F. renovado sem expire_old:     {len(data['F_renewed_no_expire_old']):4d}")
    lines.append(f"  ─────────────────────────────────────")
    lines.append(f"  Tickets com algum problema:     {data['total_affected_tickets']} (pode haver sobreposição)")
    lines.append(f"  Respostas utilities ghost:      {len(data['E_ghost_utils'])}")
    lines.append("=" * 60)
    return "\n".join(lines)


def audit_health():
    """Audita o banco e imprime relatório de anomalias SEM alterar nada.

    Detecta 6 problemas comuns em tickets ativos (Open/Damage/Clear).
    Só LEITURA. Pra corrigir, usar --fix-expires, --fix-renewals, --clean-ghost-utilities.
    """
    log.info("=" * 55)
    log.info(f"  AUDIT-HEALTH: diagnóstico de anomalias (read-only)")
    log.info("=" * 55)
    data = _audit_collect()
    log.info(f"[Audit] {data['tickets_checked']} tickets ativos, {data['responses_checked']} respostas verificadas")
    print()
    print(_audit_format_report(data))
    print()


def audit_health_and_email():
    """Roda auditoria e envia email SE houver anomalia.

    Pensado pra agendamento diário (Task Scheduler).
    Se total_issues == 0, não envia email — silêncio é bom.
    """
    log.info("=" * 55)
    log.info(f"  AUDIT-HEALTH-EMAIL: verificação diária com alerta")
    log.info("=" * 55)
    data = _audit_collect()
    log.info(f"[AuditEmail] {data['tickets_checked']} tickets, {data['responses_checked']} respostas")
    log.info(f"[AuditEmail] Total de anomalias: {data['total_issues']}")

    if data["total_issues"] == 0:
        log.info("[AuditEmail] ✅ Banco limpo — nenhum email necessário")
        return

    log.warning(f"[AuditEmail] ⚠ {data['total_issues']} anomalia(s) detectada(s) — enviando email")

    import smtplib
    from email.mime.text import MIMEText

    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_PASS")
    alert_to = os.getenv("AUDIT_EMAIL") or os.getenv("ALERT_EMAIL") or "engineering@onedrill.us"

    if not all([gmail_user, gmail_pass]):
        log.error("[AuditEmail] GMAIL_USER/GMAIL_PASS não configurados no .env — email não enviado")
        log.info("[AuditEmail] Relatório (sem enviar):")
        print(_audit_format_report(data))
        return

    # Prioridade por severidade
    critical = len(data["D_false_clear"]) + len(data["F_renewed_no_expire_old"])
    high = len(data["A_expire_poluted"]) + len(data["B_expire_old_poluted"])
    medium = len(data["E_ghost_utils"]) + len(data["C_expire_empty"])

    prefix = "🔴 CRÍTICO" if critical > 0 else ("🟡 ALTO" if high > 0 else "🟢 MÉDIO")

    subject = f"[OneDrill 811] {prefix} — {data['total_issues']} anomalia(s) no banco ({datetime.now().strftime('%d/%m/%Y')})"

    body = _audit_format_report(data)
    body += "\n\n"
    body += "=" * 60 + "\n"
    body += "AÇÕES RECOMENDADAS (em ordem):\n"
    body += "=" * 60 + "\n"
    if data["E_ghost_utils"]:
        body += f"1. Limpar ghost utilities ({len(data['E_ghost_utils'])} respostas):\n"
        body += "   python 811_sync.py --clean-ghost-utilities\n\n"
    if data["A_expire_poluted"]:
        body += f"2. Corrigir expire poluído ({len(data['A_expire_poluted'])} tickets):\n"
        body += "   python 811_sync.py --fix-expires\n\n"
    if data["B_expire_old_poluted"] or data["F_renewed_no_expire_old"]:
        n = len(data["B_expire_old_poluted"]) + len(data["F_renewed_no_expire_old"])
        body += f"3. Corrigir renovações quebradas ({n} tickets):\n"
        body += "   python 811_sync.py --fix-renewals\n\n"
    if data["D_false_clear"]:
        body += f"4. Falsos Clear ({len(data['D_false_clear'])} tickets): o scraper novo reverte automaticamente no próximo ciclo. Se persistir após 2 dias, investigar.\n\n"
    if data["C_expire_empty"]:
        body += f"5. Expire vazio ({len(data['C_expire_empty'])} tickets): editar manualmente no app ou aguardar sync.\n\n"

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = gmail_user
    msg["To"] = alert_to

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(gmail_user, gmail_pass)
            s.send_message(msg)
        log.info(f"[AuditEmail] ✅ Email enviado pra {alert_to}")
    except Exception as e:
        log.error(f"[AuditEmail] Erro ao enviar email: {e}")
        log.info("[AuditEmail] Relatório que seria enviado:")
        print(body)


def check_expiring_tickets(days_ahead=5, target_state=None):
    """Alerta por email sobre tickets Open/Damage que vencem nos próximos N dias.

    Pensado pra rodar diário (Task Scheduler) ou antes de cada sync.
    Se nenhum ticket expirando, silêncio total.
    """
    log.info("=" * 55)
    log.info(f"  CHECK-EXPIRING: tickets vencendo em {days_ahead} dias")
    log.info("=" * 55)

    query = "&status=in.(Open,Damage)&order=expire"
    if target_state:
        query += f"&state=eq.{_qv(target_state)}"

    tickets_list = sb_get("tickets", query)
    if not tickets_list:
        log.info("[Expiring] Nenhum ticket Open/Damage no banco")
        return

    today = datetime.now().date()
    cutoff = today + timedelta(days=days_ahead)

    expiring = []
    already_expired = []

    for t in tickets_list:
        raw_expire = (t.get("expire") or "").strip()
        if not raw_expire:
            continue
        norm = normalize_expire(raw_expire)
        if not norm:
            continue
        try:
            m = re.match(r"(\d{2})/(\d{2})/(\d{4})", norm)
            if not m:
                continue
            exp_date = datetime(int(m.group(3)), int(m.group(1)), int(m.group(2))).date()
        except (ValueError, IndexError):
            continue

        if exp_date < today:
            already_expired.append((t, exp_date))
        elif exp_date <= cutoff:
            expiring.append((t, exp_date))

    total = len(expiring) + len(already_expired)
    if total == 0:
        log.info(f"[Expiring] ✅ Nenhum ticket vencendo até {cutoff.strftime('%m/%d/%Y')}")
        return

    log.warning(f"[Expiring] ⚠ {len(expiring)} vencendo em {days_ahead}d + {len(already_expired)} já vencido(s)")

    lines = []
    lines.append(f"OneDrill 811 — Alerta de Vencimento ({datetime.now().strftime('%d/%m/%Y %H:%M')})")
    lines.append("=" * 60)

    if already_expired:
        lines.append(f"\n⛔ JÁ VENCIDOS ({len(already_expired)}):")
        lines.append("-" * 40)
        for t, d in sorted(already_expired, key=lambda x: x[1]):
            days_ago = (today - d).days
            lines.append(f"  [{t.get('state','?')}] {t['ticket']}  venceu {d.strftime('%m/%d/%Y')} ({days_ago}d atrás)  — {t.get('status')} — {t.get('location','')}")

    if expiring:
        lines.append(f"\n⚠ VENCENDO EM {days_ahead} DIAS ({len(expiring)}):")
        lines.append("-" * 40)
        for t, d in sorted(expiring, key=lambda x: x[1]):
            days_left = (d - today).days
            lines.append(f"  [{t.get('state','?')}] {t['ticket']}  vence {d.strftime('%m/%d/%Y')} ({days_left}d)  — {t.get('status')} — {t.get('location','')}")

    lines.append("\n" + "=" * 60)
    lines.append("Ação: renovar tickets antes do vencimento no portal 811.")
    report = "\n".join(lines)

    print(report)

    import smtplib
    from email.mime.text import MIMEText

    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_PASS")
    alert_to = os.getenv("AUDIT_EMAIL") or os.getenv("ALERT_EMAIL")

    if not all([gmail_user, gmail_pass, alert_to]):
        log.info("[Expiring] Email não configurado — alerta só no console")
        return

    prefix = "⛔" if already_expired else "⚠"
    subject = f"[OneDrill] {prefix} {total} ticket(s) vencendo/vencido(s) — {datetime.now().strftime('%d/%m/%Y')}"

    msg = MIMEText(report, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = gmail_user
    msg["To"] = alert_to

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(gmail_user, gmail_pass)
            s.send_message(msg)
        log.info(f"[Expiring] ✅ Alerta enviado pra {alert_to}")
    except Exception as e:
        log.warning(f"[Expiring] Erro ao enviar email: {e}")


def clean_ghost_utilities():
    """Remove respostas de utilities fantasmas (lixo de UI) do Supabase.

    Um bug antigo no parser captava strings de UI do portal Indiana
    (ex: "All (6)", "Current (3)", "Event") como se fossem utilities reais.
    Essas linhas foram gravadas em ticket_811_responses e aparecem no app
    como utilities pendentes inexistentes.

    Esta função:
      - Lista todas as utilities no banco
      - Filtra pelas que FALHAM no _is_valid_utility_name (lixo de UI)
      - Mostra quantas vai deletar e de quais tickets
      - Deleta em batch

    SEGURO: só apaga respostas cujo utility_name casa com padrões de UI.
    Utilities reais (DUKE ENERGY, COMCAST, etc) nunca são tocadas.
    """
    log.info("=" * 55)
    log.info(f"  CLEAN GHOST UTILITIES: limpando respostas fantasmas")
    log.info("=" * 55)

    # Puxa todas as respostas em páginas de 1000
    all_resps = []
    offset = 0
    while True:
        page = sb_get("ticket_811_responses", f"&select=id,ticket_num,utility_name,status,state&order=id&limit=1000&offset={offset}")
        if not page:
            break
        all_resps.extend(page)
        if len(page) < 1000:
            break
        offset += 1000

    log.info(f"[CleanGhost] {len(all_resps)} respostas totais no banco")

    # Identifica fantasmas
    ghosts = []
    for r in all_resps:
        name = r.get("utility_name", "")
        if not _is_valid_utility_name(name):
            ghosts.append(r)

    if not ghosts:
        log.info("[CleanGhost] ✅ Nenhuma utility fantasma encontrada — banco limpo")
        return

    # Mostra preview
    log.info(f"[CleanGhost] ⚠ {len(ghosts)} respostas fantasmas encontradas:")
    by_name = {}
    by_ticket = {}
    for r in ghosts:
        name = r.get("utility_name", "(vazio)")
        tnum = r.get("ticket_num", "?")
        by_name[name] = by_name.get(name, 0) + 1
        by_ticket.setdefault(tnum, []).append(name)

    log.info(f"[CleanGhost] Nomes fantasmas encontrados:")
    for name, count in sorted(by_name.items(), key=lambda x: -x[1]):
        log.info(f"  - '{name}' ({count}x)")

    log.info(f"[CleanGhost] {len(by_ticket)} tickets afetados")

    # Confirmação
    print()
    print(f"⚠ Esta ação vai DELETAR {len(ghosts)} respostas do banco.")
    print(f"   Tickets afetados: {len(by_ticket)}")
    print(f"   Exemplos de nomes a deletar: {list(by_name.keys())[:5]}")
    print()
    resp = input("Confirma exclusão? (digite 'SIM' pra continuar): ").strip()
    if resp != "SIM":
        log.info("[CleanGhost] Cancelado pelo usuário")
        return

    # Deleta em batch por ID
    deleted = 0
    errors = 0
    batch_size = 50
    for i in range(0, len(ghosts), batch_size):
        batch = ghosts[i:i+batch_size]
        ids = [r["id"] for r in batch]
        # Supabase DELETE com filtro in.()
        id_list = ",".join(str(x) for x in ids)
        try:
            sb_delete("ticket_811_responses", f"&id=in.({id_list})")
            deleted += len(batch)
            log.info(f"[CleanGhost] {deleted}/{len(ghosts)} deletados")
        except Exception as e:
            log.error(f"[CleanGhost] Erro no batch {i}-{i+batch_size}: {e}")
            errors += len(batch)

    log.info(f"[CleanGhost] === CONCLUÍDO: {deleted} deletados, {errors} erros ===")


# ── │ SECTION: DEBUG_EXPIRE │ DEBUG EXPIRE: mostra o body e testa todos os patterns 
async def debug_expire(target_ticket, state=None):
    """Abre o ticket no portal, mostra o body cru e testa todos os patterns.

    Útil pra diagnosticar quando extract_expire_date está pegando data errada.
    """
    if not state:
        # Auto-detecta o estado consultando o banco
        rows = sb_get("tickets", f"&ticket=eq.{_qv(target_ticket)}&select=state&limit=1")
        if not rows:
            log.error(f"[DebugExpire] Ticket {target_ticket} não encontrado no banco")
            return
        state = rows[0].get("state", "")

    if state not in PORTALS:
        log.error(f"[DebugExpire] Estado '{state}' sem portal (FL/IN suportados)")
        return

    log.info("=" * 55)
    log.info(f"  DEBUG-EXPIRE: {target_ticket} ({state})")
    log.info("=" * 55)

    perfil = _profile_path(state)

    async with async_playwright() as p:
        ctx = await p.chromium.launch_persistent_context(perfil, headless=True, args=["--no-sandbox"])
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()
        page.set_default_timeout(TIMEOUT_PAGE)

        await page.goto(PORTALS[state]["home"], wait_until="domcontentloaded")
        await wait_stable(page)
        page, ctx = await ensure_login(page, ctx, p, state)
        if not page:
            log.error(f"[DebugExpire] Login falhou")
            return

        ok = await goto_dashboard(page, state)
        if not ok:
            log.error(f"[DebugExpire] Dashboard inacessível")
            await ctx.close()
            return

        try:
            await filter_ticket(page, target_ticket)
            if not await page.get_by_text(target_ticket, exact=True).count():
                log.error(f"[DebugExpire] Ticket {target_ticket} não encontrado")
                await ctx.close()
                return

            await page.get_by_text(target_ticket, exact=True).first.click()
            await wait_stable(page)

            tt = page.locator('[role="tab"]:has-text("Text")').first
            if await tt.count():
                await click_and_wait(page, tt, "tab")

            body = await page.locator("body").inner_text()

            # Salva body cru num arquivo pra inspeção
            dump_path = os.path.join(BASE_DIR, f"debug_expire_{target_ticket}.txt")
            with open(dump_path, "w", encoding="utf-8") as f:
                f.write(body)
            log.info(f"Body cru salvo em: {dump_path}")

            # Mostra todas as linhas que contêm "expir", "due", "date", "work"
            print("\n" + "=" * 55)
            print("LINHAS RELEVANTES DO BODY:")
            print("=" * 55)
            keywords = ["expir", "due", "work date", "legal", "response", "start", "renewal"]
            for line_num, line in enumerate(body.split("\n"), 1):
                l = line.lower()
                if any(kw in l for kw in keywords):
                    print(f"  L{line_num:4d}: {line.strip()[:100]}")

            # Testa cada pattern e mostra o que bateu
            print("\n" + "=" * 55)
            print("TESTE DE PATTERNS:")
            print("=" * 55)
            test_patterns = [
                (r"Ticket\s+Expires?\s*(?:on)?\s*:\s*([^\n]+)",    "✓ Ticket Expires"),
                (r"Expiration\s+Date\s*:\s*([^\n]+)",               "✓ Expiration Date"),
                (r"Expiration\s*:\s*([^\n]+)",                       "✓ Expiration"),
                (r"(?<!\w)Expires?\s+on\s*:\s*([^\n]+)",             "✓ Expires on"),
                (r"(?<!\w)Expires\s*:\s*([^\n]+)",                   "✓ Expires"),
                (r"(?<!\w)Expire\s*:\s*([^\n]+)",                    "✓ Expire"),
                (r"Due\s*Date\s*:\s*([^\n]+)",                       "✗ Due Date (IGNORADO — data errada)"),
                (r"Work\s*[Dd]ate\s*:\s*([^\n]+)",                   "✗ Work Date (IGNORADO — data errada)"),
                (r"Legal\s*Date\s*:\s*([^\n]+)",                     "✗ Legal Date (IGNORADO — data errada)"),
            ]
            for pat, label in test_patterns:
                m = re.search(pat, body, re.IGNORECASE)
                val = m.group(1).strip()[:80] if m else "—"
                print(f"  {label:50s}: {val}")

            # Resultado final
            print("\n" + "=" * 55)
            print("RESULTADO DO PARSER:")
            print("=" * 55)
            result = extract_expire_date(body, ticket_num=target_ticket, debug=True)
            print(f"  extract_expire_date() → '{result}'")

            # Compara com banco
            rows = sb_get("tickets", f"&ticket=eq.{_qv(target_ticket)}&select=expire&limit=1")
            if rows:
                db_val = rows[0].get("expire", "")
                print(f"  Valor no banco        → '{db_val}'")
                if result != db_val:
                    print(f"  ⚠ DIFERENTE — rode '--fix-expires --ticket {target_ticket}' pra corrigir")

            print("=" * 55 + "\n")

        except Exception as e:
            log.error(f"[DebugExpire] ERRO: {e}")
        finally:
            try:
                await ctx.close()
            except Exception:
                pass


# ── │ SECTION: CLI │ CLI ──────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="OneDrill - 811 Sync v2")
    parser.add_argument("--state",      choices=["FL", "IN", "IL", "WI"], help="Estado a sincronizar")
    parser.add_argument("--all",        action="store_true", help="Sincronizar FL + IN")
    parser.add_argument("--imp",        action="store_true", help="Importar tickets novos + respostas")
    parser.add_argument("--imp_all",    action="store_true", help="Importar FL + IN")
    parser.add_argument("--excel",      action="store_true", help="Exportar Excel")
    parser.add_argument("--debug",      action="store_true", help="Debug: screenshots")
    parser.add_argument("--rescrape",   action="store_true", help="Re-scrape notes")
    parser.add_argument("--cleanup",    action="store_true", help="Excluir tickets cancelados")
    parser.add_argument("--cleanup_all", action="store_true", help="Cleanup IN + FL em paralelo")
    parser.add_argument("--cleanup-wi-dhl", action="store_true", help="Cleanup retroativo: tickets WI Clear com todas responses Closed by DHL → Cancel")
    parser.add_argument("--reclassify-wi", action="store_true", help="Re-aplica classify nas responses WI existentes (sem ir ao portal) + re-checa auto-cancel")
    parser.add_argument("--scan-email-wi", action="store_true", help="Scan emails da pasta Winsconsin (Outlook 365) — confirmações de utilities. Sem --commit é dry run.")
    parser.add_argument("--list-outlook-folders", action="store_true", help="Lista pastas IMAP da conta Outlook 365 (debug)")
    parser.add_argument("--compare-wi-excel", type=str, help="Compara planilha .xlsx com tickets WI no banco. Ex: --compare-wi-excel \"C:\\path\\to\\file.xlsx\"")
    parser.add_argument("--contacts",   action="store_true", help="Scrape contatos de utilities (FL)")
    parser.add_argument("--force",      action="store_true", help="Forçar re-scrape de TODOS")
    parser.add_argument("--backfill",   action="store_true", help="Backfill: adicionar eventos de clear no histórico")
    parser.add_argument("--fix-dates",  action="store_true", help="Corrigir datas de clear usando data real da última resposta 811")
    parser.add_argument("--fix-clear-ts", action="store_true", help="Fix 2026-05-14: ajusta ts dos AUTO 811 Clear pra synced_at")
    parser.add_argument("--scan-email", action="store_true", help="Varre Gmail buscando status change de utilities (dry run por default)")
    parser.add_argument("--scan-email-debug", action="store_true", help="Lista subject/sender/body de TODOS emails (sem filtros) — pra entender o formato e ajustar o parser do --scan-email")
    parser.add_argument("--apply-overrides", action="store_true", help="Aplica LOCAL_OVERRIDES + auto-clear")
    parser.add_argument("--undo-fake-overrides", action="store_true", help="Remove entries (override local) com ts=hoje (cleanup do bug RE-CONFIRMA)")
    parser.add_argument("--debug-history", type=str, help="Mostra historico completo de um ticket. Ex: --debug-history 26031001348")
    parser.add_argument("--apply-today-clears", action="store_true", help="Adiciona entry no historico de tickets ja Clear cujas utilities responderam HOJE")
    parser.add_argument("--target-date", type=str, default=None, help="YYYY-MM-DD pra apply-today-clears (default: hoje)")
    parser.add_argument("--commit", action="store_true", help="Usado com --scan-email: atualiza banco (sem ele eh dry run)")
    parser.add_argument("--days-back", type=int, default=7, help="Dias pra tras pra varrer emails (default 7)")
    parser.add_argument("--fix-renewals", action="store_true", help="Corrigir expire_old de tickets renovados (busca data real no portal)")
    parser.add_argument("--fix-expires", action="store_true", help="Corrigir data de vencimento (expire) de tickets com formato antigo ou errado")
    parser.add_argument("--populate-counties", action="store_true", help="Backfill de county em tickets antigos (usa base cidade→county + Nominatim fallback)")
    parser.add_argument("--rebuild-coverage", action="store_true", help="Reconstroi tabela utility_county_coverage a partir de ticket_811_responses × tickets.county")
    parser.add_argument("--clean-ghost-utilities", action="store_true", help="Remove respostas de utilities fantasmas (lixo 'All (N)', 'Event', etc)")
    parser.add_argument("--audit-health", action="store_true", help="Relatório de anomalias do banco (read-only, não altera nada)")
    parser.add_argument("--audit-health-email", action="store_true", help="Auditoria + email se anomalias > 0 (pra agendador)")
    parser.add_argument("--check-expiring", action="store_true", help="Alerta de tickets Open/Damage vencendo nos próximos N dias")
    parser.add_argument("--days-ahead", type=int, default=5, help="Dias pra frente pra checar vencimento (default 5)")
    parser.add_argument("--debug-expire", action="store_true", help="Diagnosticar extração de expire de um ticket — requer --ticket")
    parser.add_argument("--ticket",     type=str, help="Número do ticket alvo (usado com --fix-expires ou --debug-expire)")
    parser.add_argument("--no-cache",   action="store_true", help="Forçar re-scrape de todos (incluindo Clear em cache)")
    parser.add_argument("--selftest",   action="store_true", help="Rodar testes internos (classify, parser)")
    parser.add_argument("--backup",     action="store_true", help="Backup completo do banco de dados (JSON)")
    parser.add_argument("--save-pdf",   action="store_true", help="Salvar PDF de tickets Clear/Damage/Completed (FL/IN via impressora, IL/WI via headless)")
    parser.add_argument("--sync-il",    action="store_true", help="Sincronizar respostas JULIE (Illinois)")
    parser.add_argument("--sync-wi",    action="store_true", help="Sincronizar respostas Diggers Hotline (Wisconsin)")
    parser.add_argument("--imp-wi",     action="store_true", help="Importar tickets novos WI (Diggers Hotline) + sync respostas")
    parser.add_argument("--imp-il",     action="store_true", help="Importar tickets novos IL (JULIE Ticket Entry) + sync respostas")
    parser.add_argument("--no-lock",    action="store_true", help="Pular verificação de lock file")
    args = parser.parse_args()

    # Lock file — pula pra debug e selftest
    use_lock = not args.no_lock and not args.debug and not args.selftest

    def _run():
        if args.selftest:
            run_self_tests()
        elif args.backup:
            backup_database()
        elif getattr(args, 'save_pdf', False):
            if args.state == "IL":
                asyncio.run(save_ticket_pdfs_il(force=args.force))
            elif args.state == "WI":
                asyncio.run(save_ticket_pdfs_wi(force=args.force))
            elif args.state:
                asyncio.run(save_ticket_pdfs(args.state, force=args.force))
            else:
                asyncio.run(save_ticket_pdfs("FL", force=args.force))
                asyncio.run(save_ticket_pdfs("IN", force=args.force))
                asyncio.run(save_ticket_pdfs_il(force=args.force))
                asyncio.run(save_ticket_pdfs_wi(force=args.force))
        elif getattr(args, 'sync_il', False):
            asyncio.run(sync_il())
        elif getattr(args, 'imp_il', False):
            asyncio.run(sync_and_import_il())
        elif getattr(args, 'imp_wi', False):
            asyncio.run(sync_and_import_wi())
        elif getattr(args, 'sync_wi', False):
            asyncio.run(sync_wi())
        elif getattr(args, 'fix_renewals', False):
            asyncio.run(fix_renewals())
        elif getattr(args, 'fix_expires', False):
            asyncio.run(fix_expires(target_ticket=args.ticket))
        elif getattr(args, 'populate_counties', False):
            asyncio.run(populate_counties(target_state=args.state, force=args.force))
        elif getattr(args, 'rebuild_coverage', False):
            rebuild_utility_county_coverage()
        elif getattr(args, 'clean_ghost_utilities', False):
            clean_ghost_utilities()
        elif getattr(args, 'audit_health', False):
            audit_health()
        elif getattr(args, 'audit_health_email', False):
            audit_health_and_email()
        elif getattr(args, 'check_expiring', False):
            check_expiring_tickets(
                days_ahead=getattr(args, 'days_ahead', 5),
                target_state=args.state
            )
        elif getattr(args, 'debug_expire', False):
            if not args.ticket:
                log.error("--debug-expire requer --ticket NNNNNN")
                sys.exit(1)
            asyncio.run(debug_expire(args.ticket, state=args.state))
        elif args.debug:
            asyncio.run(debug_screenshot(args.state or "IN"))
        elif args.excel:
            export_excel()
        elif args.contacts:
            asyncio.run(scrape_contacts(args.state or "FL", force=args.force))
        elif args.cleanup_all:
            asyncio.run(cleanup_all())
        elif args.cleanup:
            asyncio.run(cleanup_canceled(args.state or "IN"))
        elif getattr(args, 'cleanup_wi_dhl', False):
            cleanup_wi_dhl_clears()
        elif getattr(args, 'reclassify_wi', False):
            reclassify_wi_responses()
        elif getattr(args, 'list_outlook_folders', False):
            list_outlook_folders()
        elif getattr(args, 'scan_email_wi', False):
            scan_emails_wi(commit=args.commit, days_back=getattr(args, 'days_back', 14))
        elif getattr(args, 'compare_wi_excel', None):
            compare_wi_excel(args.compare_wi_excel)
        elif args.rescrape:
            asyncio.run(rescrape_notes(args.state or "FL", force=args.force))
        elif args.backfill:
            backfill_history()
        elif getattr(args, 'fix_dates', False):
            fix_clear_dates()
        elif getattr(args, 'fix_clear_ts', False):
            fix_clear_ts(target_state=args.state)
        elif getattr(args, 'apply_overrides', False):
            apply_overrides_now(target_state=args.state)
        elif getattr(args, 'undo_fake_overrides', False):
            undo_fake_overrides_today(target_state=args.state)
        elif getattr(args, 'debug_history', None):
            debug_ticket_history(args.debug_history)
        elif getattr(args, 'apply_today_clears', False):
            apply_today_clears(
                target_state=args.state,
                target_date=getattr(args, 'target_date', None)
            )
        elif getattr(args, 'scan_email', False):
            scan_emails_for_responses(
                commit=getattr(args, 'commit', False),
                state_filter=args.state,
                days_back=getattr(args, 'days_back', 7)
            )
        elif getattr(args, 'scan_email_debug', False):
            scan_emails_debug(days_back=getattr(args, 'days_back', 7))
        elif args.all:
            asyncio.run(sync_all())
        elif args.imp_all:
            asyncio.run(sync_and_import_all())
        elif args.imp:
            asyncio.run(sync_and_import(args.state or "IN"))
        elif args.state == "IL":
            asyncio.run(sync_il())
        elif args.state == "WI":
            asyncio.run(sync_wi())
        elif args.state:
            asyncio.run(sync_state(args.state))
        else:
            parser.print_help()

    if use_lock:
        try:
            with ProcessLock():
                _run()
        except RuntimeError as e:
            log.error(str(e))
            sys.exit(1)
    else:
        _run()
